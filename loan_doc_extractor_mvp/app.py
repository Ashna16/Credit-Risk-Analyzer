from __future__ import annotations

import io
import json
import os
import re
import html
import uuid
import hashlib
import difflib
import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple
from urllib.parse import quote, unquote

import altair as alt
import pandas as pd
import streamlit as st
import fitz
import requests

try:
    from extractor import extract_fields
except ModuleNotFoundError:
    from loan_doc_extractor_mvp.extractor import extract_fields

parse_financial_statements = None
try:
    PROJECT_ROOT = Path(__file__).resolve().parents[1]
    BACKEND_DIR = PROJECT_ROOT / "backend"
    if BACKEND_DIR.exists() and str(BACKEND_DIR) not in sys.path:
        sys.path.append(str(BACKEND_DIR))
    from financial_parser import parse_financial_statements as _parse_financial_statements

    parse_financial_statements = _parse_financial_statements
except Exception:
    parse_financial_statements = None

st.set_page_config(page_title="Credit Document Portal", page_icon="🏦", layout="wide", initial_sidebar_state="expanded")

APP_BASE_URL = os.getenv("APP_BASE_URL", "http://localhost:9000").rstrip("/")
BACKEND_BASE_URL = os.getenv("BACKEND_BASE_URL", "http://localhost:8000").rstrip("/")
EXTRACTION_CACHE_VERSION = "v20260409_12"
APP_BUILD_ID = "build-2026-04-09-1225"
RISK_BUNDLE_LOGIC_VERSION = "v20260224_1"

# Extraction feature flags (debug-friendly toggles).
ENABLE_FUZZY_MATCHING = True
ENABLE_REGEX_FALLBACK = True
ENABLE_IFRS_DICTIONARY = True
ENABLE_MULTI_YEAR_ALIGNMENT = True
ANALYSIS_YEAR = os.getenv("ANALYSIS_YEAR", "latest")
ANALYSIS_YEAR_MODE = os.getenv("ANALYSIS_YEAR_MODE", "latest_available")
ANALYSIS_SPECIFIC_YEAR = os.getenv("ANALYSIS_SPECIFIC_YEAR", "")


def _max_extractable_fiscal_year() -> int:
    """Latest FY columns in 10-K/10-Q often reference year-end in the next calendar year."""
    return datetime.now().year + 1


st.markdown(
    """
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Material+Symbols+Outlined:opsz,wght,FILL,GRAD@20..48,500,0,0&display=swap');

    html, body, [class*="css"] { font-family: 'Manrope', sans-serif; }
    .stApp { background: #f4f7fb; color: #0f172a; }
    .block-container {
      padding-top: 0.7rem;
      max-width: calc(100vw - 1.2rem);
      padding-left: 0.4rem;
      padding-right: 0.4rem;
    }

    header[data-testid="stHeader"] { height: 0; background: transparent; }
    div[data-testid="stToolbar"] { display: none; }
    #MainMenu { visibility: hidden; }
    footer { visibility: hidden; }

    [data-testid="stSidebar"] {
      background: linear-gradient(180deg, #0a2239 0%, #163a5d 100%);
      border-right: 1px solid #2a547c;
      min-width: 300px !important;
      min-height: 100vh !important;
    }
    [data-testid="stSidebar"] > div:first-child {
      padding-top: 0.7rem !important;
      padding-left: 0.65rem !important;
      padding-right: 0.65rem !important;
      padding-bottom: 0.9rem !important;
      min-height: 100vh !important;
    }
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3,
    [data-testid="stSidebar"] .stMarkdown p,
    [data-testid="stSidebar"] label {
      color: #e6f0ff !important;
      font-weight: 600;
    }
    [data-testid="stSidebar"] .stSelectbox div[data-baseweb="select"] > div {
      background: #0f2f4f !important;
      border: 1px solid #3b638b !important;
      color: #f4f8ff !important;
      border-radius: 10px !important;
    }
    [data-testid="stSidebar"] .stTextInput input {
      background: #0f2f4f !important;
      border: 1px solid #3b638b !important;
      color: #f4f8ff !important;
      border-radius: 10px !important;
    }
    [data-testid="stSidebar"] .stFileUploader > div {
      background: #0f2f4f !important;
      border: 1px dashed #4f7ca8 !important;
      border-radius: 10px !important;
      color: #dcecff !important;
    }
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label {
      background: #0f2f4f !important;
      border: 1px solid #3b638b !important;
      border-radius: 8px;
      padding: 0.15rem 0.4rem;
      margin-bottom: 0.1rem;
      color: #e6f0ff !important;
    }
    [data-testid="stSidebar"] [data-testid="stVerticalBlock"] {
      gap: 0.3rem !important;
    }
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label p,
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label span,
    [data-testid="stSidebar"] .stRadio [role="radiogroup"] label div {
      color: #e6f0ff !important;
    }
    [data-testid="stSidebar"] .stToggle label,
    [data-testid="stSidebar"] .stToggle div,
    [data-testid="stSidebar"] .stSlider label {
      color: #d9e9ff !important;
    }
    [data-testid="stSidebar"] .stSlider > div[data-baseweb="slider"] > div > div {
      background: #2f6fa3 !important;
    }
    .left-nav-panel {
      background: linear-gradient(180deg, #0a2239 0%, #163a5d 100%);
      border: 1px solid #2a547c;
      border-radius: 14px;
      min-height: calc(100vh - 1rem);
      height: 100%;
      padding: 0.55rem 0.6rem;
      position: static;
      top: auto;
      z-index: 5;
      overflow: visible;
      display: flex;
      flex-direction: column;
      gap: 0.45rem;
      justify-content: flex-start;
      align-items: stretch;
    }
    .left-nav-panel * {
      visibility: visible !important;
      opacity: 1 !important;
    }
    .left-nav-top {
      display: grid;
      gap: 0.5rem;
    }
    .left-nav-brand {
      border: 0;
      border-radius: 12px;
      padding: 0.1rem 0.1rem 0.35rem;
      background: transparent;
    }
    .left-nav-brand h3 {
      margin: 0;
      color: #ffffff !important;
      font-size: 1.38rem;
      font-weight: 800;
      letter-spacing: 0;
    }
    .left-nav-brand p {
      margin: 0.08rem 0 0;
      color: #b5c8ff !important;
      font-size: 0.9rem;
      font-weight: 700;
    }
    .left-nav-modules {
      margin-top: 0;
      padding: 0.15rem 0.15rem 0.08rem;
      color: #e8efff !important;
      font-size: 0.98rem;
      letter-spacing: 0.06em;
      font-weight: 800;
      text-transform: uppercase;
    }
    .left-nav-links {
      display: grid;
      gap: 0.34rem;
      margin-top: 0.08rem;
      align-content: start;
    }
    .left-nav-link {
      display: flex;
      align-items: center;
      gap: 0.58rem;
      width: 100%;
      border-radius: 14px;
      border: 1px solid transparent;
      padding: 0.65rem 0.9rem;
      min-height: 3rem;
      color: #d5e3ff !important;
      text-decoration: none !important;
      font-size: 1rem;
      font-weight: 700;
      line-height: 1.2;
      background: transparent;
      transition: all 0.18s ease;
    }
    .left-nav-link:hover {
      background: rgba(255,255,255,0.08);
      color: #eaf2ff !important;
      text-decoration: none !important;
      transform: translateY(-0.5px);
    }
    .left-nav-link.active {
      background: #f8fbff;
      color: #2141a6 !important;
      border-color: #ffffff;
      box-shadow: 0 2px 10px rgba(15, 23, 42, 0.10);
    }
    .left-nav-link .nav-icon {
      font-size: 1.05rem;
      line-height: 1;
      width: 1.2rem;
      text-align: center;
      flex: 0 0 1.2rem;
    }
    .left-nav-link .nav-label {
      flex: 1;
    }
    .warn-pop {
      position: fixed;
      left: 50%;
      top: 22%;
      transform: translateX(-50%);
      z-index: 1000;
      width: min(560px, 92vw);
      border: 1px solid #f5d78e;
      background: #fff9e8;
      border-radius: 12px;
      box-shadow: 0 12px 32px rgba(15, 23, 42, 0.25);
      padding: 0.95rem 1rem;
    }
    .warn-pop-title {
      margin: 0;
      font-size: 1.02rem;
      font-weight: 800;
      color: #9a5a00;
    }
    .warn-pop-body {
      margin: 0.25rem 0 0;
      font-size: 0.92rem;
      color: #7a4a00;
      font-weight: 600;
    }
    .warn-pop-actions {
      margin-top: 0.75rem;
      display: flex;
      gap: 0.55rem;
      justify-content: flex-end;
    }
    .warn-btn {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      min-width: 122px;
      border-radius: 9px;
      padding: 0.46rem 0.72rem;
      text-decoration: none !important;
      font-size: 0.88rem;
      font-weight: 700;
      border: 1px solid transparent;
    }
    .warn-btn-yes {
      background: #1f5eff;
      color: #ffffff !important;
      border-color: #1f5eff;
    }
    .warn-btn-no {
      background: #fff;
      color: #7a4a00 !important;
      border-color: #efcf85;
    }
    .graph-menu-dot button[aria-label="⋮"] {
      background: transparent !important;
      border: none !important;
      box-shadow: none !important;
      color: #334155 !important;
      font-size: 1.25rem !important;
      padding: 0 !important;
      min-height: 1.2rem !important;
      min-width: 1.2rem !important;
    }
    .graph-menu-panel {
      border: 1px solid #d8e1ec;
      border-radius: 10px;
      padding: 8px 10px;
      background: #f8fafc;
      margin-bottom: 8px;
    }
    .side-modules-wrap {
      display: grid;
      gap: 0.3rem;
      margin-top: 0.15rem;
    }
    .side-module {
      display: flex;
      align-items: center;
      gap: 0.45rem;
      text-decoration: none !important;
      background: rgba(15, 47, 79, 0.45);
      border: 1px solid rgba(110, 164, 214, 0.35);
      border-radius: 8px;
      padding: 0.34rem 0.48rem;
      color: #ffffff !important;
      font-size: 0.93rem;
      font-weight: 700;
      line-height: 1.1;
    }
    .side-module,
    .side-module:link,
    .side-module:visited,
    .side-module:hover,
    .side-module:active {
      color: #ffffff !important;
      text-decoration: none !important;
    }
    .side-module * {
      color: #ffffff !important;
      text-decoration: none !important;
    }
    .side-module.active {
      background: #f7fbff;
      border-color: #d7e9fd;
      color: #123f67 !important;
    }
    .icon-glyph {
      font-family: "Material Symbols Outlined", sans-serif;
      font-size: 1.18rem;
      line-height: 1;
      font-variation-settings: "FILL" 0, "wght" 500, "GRAD" 0, "opsz" 24;
      margin-right: 0.1rem;
    }
    .icon-active { color: #2563eb !important; }
    .icon-muted { color: #94a3b8 !important; }
    .side-module.active .icon-glyph { color: #2563eb !important; }
    .side-module.active .side-chevron { color: #2563eb !important; }
    .left-nav-foot {
      border-top: 1px solid rgba(191, 219, 254, 0.28);
      padding-top: 0.4rem;
      margin-top: auto;
    }
    .left-nav-foot-card {
      border: 0;
      border-radius: 12px;
      background: #2b54d9;
      padding: 0.85rem 0.9rem;
    }
    .left-nav-foot-card p {
      margin: 0;
      color: #dce8ff !important;
      font-size: 0.82rem;
      line-height: 1.25;
      font-weight: 700;
    }
    .side-nav-btns [data-testid="stButton"] button {
      width: 100%;
      text-align: left !important;
      justify-content: flex-start !important;
      border-radius: 14px !important;
      border: 1px solid transparent !important;
      padding: 0.65rem 0.9rem !important;
      font-size: 1rem !important;
      font-weight: 700 !important;
      min-height: 3rem !important;
      box-shadow: none !important;
      transition: all 0.18s ease !important;
    }
    .side-nav-btns [data-testid="stButton"] button[kind="secondary"] {
      background: transparent !important;
      color: #d5e3ff !important;
      border-color: transparent !important;
    }
    .side-nav-btns [data-testid="stButton"] button[kind="primary"] {
      background: #f8fbff !important;
      color: #2141a6 !important;
      border-color: #ffffff !important;
      box-shadow: 0 2px 10px rgba(15, 23, 42, 0.10) !important;
    }
    .side-nav-btns [data-testid="stButton"] button:hover {
      transform: translateY(-0.5px);
      background: rgba(255,255,255,0.08) !important;
    }
    .side-module .side-chevron {
      margin-left: auto;
      color: #9ec3e8;
      font-size: 0.95rem;
      font-weight: 800;
    }
    .mode-row { margin-top: 0.55rem; }

    .app-header {
      display: flex;
      align-items: center;
      justify-content: space-between;
      gap: 0.85rem;
      margin-bottom: 0.75rem;
      background: linear-gradient(180deg, #0a2239 0%, #163a5d 100%);
      border: 1px solid #2a547c;
      border-radius: 14px;
      padding: 0.9rem 1rem;
    }
    .app-head-left {
      display: flex;
      align-items: center;
      gap: 0.85rem;
    }
    .app-icon {
      width: 42px;
      height: 42px;
      border-radius: 12px;
      background: linear-gradient(140deg, #2f6bff 0%, #2554d8 100%);
      display: flex;
      align-items: center;
      justify-content: center;
      color: white;
      font-size: 1.25rem;
      box-shadow: 0 10px 20px rgba(47, 107, 255, 0.28);
    }
    .app-icon .icon-glyph {
      color: #ffffff !important;
      font-size: 1.45rem;
    }
    .app-header .app-title {
      margin: 0;
      font-size: 2rem;
      color: #ffffff !important;
      font-weight: 800;
      line-height: 1.05;
    }
    .app-header .app-subtitle {
      margin: 0.2rem 0 0;
      color: #e2e8f0 !important;
      font-size: 1rem;
    }
    .app-header * {
      color: #ffffff !important;
    }
    .app-head-right {
      color: #b9d5f4 !important;
      font-size: 0.95rem;
      font-weight: 600;
      white-space: nowrap;
    }

    .upload-shell {
      background: #ffffff;
      border: 1px solid #d8e1ec;
      border-radius: 16px;
      padding: 1.2rem 1.2rem 0.95rem;
      margin-bottom: 0.95rem;
      box-shadow: 0 4px 16px rgba(15, 23, 42, 0.05);
    }
    .upload-shell h2 {
      margin: 0;
      text-align: center;
      color: #101f39;
      font-size: 2rem;
      font-weight: 800;
    }
    .upload-shell p {
      margin: 0.45rem auto 0.9rem;
      color: #4b617d;
      text-align: center;
      max-width: 900px;
      font-size: 1.04rem;
    }
    .supported-box {
      margin-top: 0.85rem;
      background: linear-gradient(180deg, #eef4ff 0%, #e8f0ff 100%);
      border: 1px solid #c8d8f7;
      border-radius: 12px;
      padding: 0.75rem 0.9rem;
      color: #274870;
      font-size: 0.95rem;
    }
    .cap-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 0.8rem;
      margin: 0.65rem 0 0.95rem;
    }
    .cap-card {
      background: #fff;
      border: 1px solid #d8e1ec;
      border-radius: 14px;
      padding: 0.9rem 0.95rem;
      box-shadow: 0 4px 14px rgba(15, 23, 42, 0.03);
    }
    .cap-title {
      margin: 0;
      color: #0f2740;
      font-weight: 800;
      font-size: 1.1rem;
    }
    .cap-text {
      margin: 0.35rem 0 0;
      color: #4e6178;
      font-size: 0.95rem;
      line-height: 1.35;
    }
    .flow-kicker {
      color: #3d5777;
      font-size: 0.9rem;
      margin-bottom: 0.35rem;
      font-weight: 700;
    }

    .panel {
      background: #ffffff;
      border: 1px solid #d8e1ec;
      border-radius: 14px;
      padding: 0.86rem 0.98rem;
      margin-bottom: 0.75rem;
      box-shadow: 0 4px 14px rgba(15, 23, 42, 0.04);
    }

    .selected-file {
      border: 1px solid #88b7d8;
      background: linear-gradient(120deg, #d9ecfb 0%, #cbe4f7 100%);
      color: #0f2e4c !important;
      border-radius: 8px;
      padding: 0.45rem 0.6rem;
      font-size: 0.9rem;
      font-weight: 700;
      margin-top: 0.35rem;
    }
    [data-testid="stSidebar"] .selected-file {
      color: #0f2e4c !important;
    }

    .recommendation {
      background: #edf7f6;
      border: 1px solid #bfdedb;
      border-radius: 12px;
      color: #144b58;
      padding: 0.75rem 0.9rem;
      margin-bottom: 0.7rem;
      font-size: 0.93rem;
      line-height: 1.42;
    }

    .info-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 0.7rem;
      margin-top: 0.2rem;
    }
    .info-card {
      border: 1px solid #d9e3ef;
      border-radius: 10px;
      padding: 0.55rem 0.7rem;
      background: #f8fbff;
    }
    .info-label {
      color: #4e6178;
      font-size: 0.78rem;
      font-weight: 600;
      margin-bottom: 0.1rem;
    }
    .info-value {
      color: #0f2740;
      font-size: 0.96rem;
      font-weight: 600;
      word-break: break-word;
    }

    .stButton button[kind="primary"] {
      background: linear-gradient(120deg, #123457 0%, #1f4f7f 100%) !important;
      color: white !important;
      border: none !important;
      border-radius: 10px !important;
      font-weight: 700 !important;
    }

    .stDownloadButton button {
      border-radius: 10px !important;
      font-weight: 700 !important;
      border: 1px solid #c7d7e8 !important;
    }
    .export-wrap {
      background: #fff;
      border: 1px solid #d8e1ec;
      border-radius: 14px;
      padding: 0.9rem;
      margin-top: 0.5rem;
      box-shadow: 0 4px 14px rgba(15,23,42,0.04);
    }
    .viewer-label {
      margin-top: 0.45rem;
      font-size: 0.88rem;
      color: #3a4d63;
      font-weight: 600;
    }
    @media (max-width: 1200px) {
      .cap-grid { grid-template-columns: 1fr; }
      .upload-shell h2 { font-size: 1.6rem; }
      .app-title { font-size: 1.6rem; }
    }
    div[data-testid="stDataFrame"] * {
      color: #0f172a !important;
    }
    div[data-testid="stDataFrame"] a {
      color: #0d4e8b !important;
      font-weight: 700 !important;
      text-decoration: underline !important;
    }
    .risk-table-wrap {
      border: 1px solid #d8e1ec;
      border-radius: 12px;
      overflow: hidden;
      background: #fff;
      margin-bottom: 0.7rem;
    }
    .risk-table {
      width: 100%;
      border-collapse: collapse;
      font-size: 0.9rem;
      overflow: visible;
    }
    .risk-table th {
      background: #f7fafc;
      color: #1e293b;
      font-weight: 800;
      padding: 10px 10px;
      border-bottom: 1px solid #e2e8f0;
      text-align: left;
      letter-spacing: 0.01em;
    }
    .risk-table td {
      padding: 9px 10px;
      border-bottom: 1px solid #eef2f7;
      color: #0f172a !important;
      vertical-align: top;
      font-weight: 600;
    }
    .metric-help {
      position: relative;
      cursor: help;
      border-bottom: 1px dotted rgba(15, 23, 42, 0.3);
    }
    .metric-help:hover::after {
      content: attr(data-tip);
      position: absolute;
      left: 0;
      bottom: calc(100% + 6px);
      background: #eef6ff;
      color: #16324f;
      border: 1px solid #bfdbfe;
      padding: 6px 8px;
      border-radius: 8px;
      font-size: 0.78rem;
      font-weight: 600;
      line-height: 1.25;
      width: max-content;
      max-width: 320px;
      white-space: normal;
      z-index: 9999;
      box-shadow: 0 6px 20px rgba(2, 6, 23, 0.35);
      pointer-events: none;
    }
    .hover-help {
      display: inline-flex;
      align-items: center;
      justify-content: center;
      width: 18px;
      height: 18px;
      border-radius: 50%;
      border: 1px solid #c7d2fe;
      color: #1d4ed8;
      font-weight: 800;
      font-size: 0.72rem;
      margin-left: 6px;
      cursor: help;
      position: relative;
      top: -1px;
    }
    .risk-top-actions .stButton button {
      border-radius: 10px !important;
      font-weight: 700 !important;
      min-height: 2.65rem !important;
    }
    .risk-summary-card {
      background: #ffffff;
      border: 1px solid #cfdbea;
      border-radius: 16px;
      padding: 1rem 1rem 0.9rem;
      box-shadow: 0 8px 24px rgba(15, 23, 42, 0.05);
      margin-bottom: 0.75rem;
    }
    .risk-summary-row {
      display: grid;
      grid-template-columns: repeat(6, minmax(0, 1fr));
      gap: 0.8rem;
      margin-bottom: 0.85rem;
    }
    .risk-summary-item {
      min-width: 0;
    }
    .risk-summary-label {
      color: #111827;
      font-size: 0.92rem;
      font-weight: 700;
      margin-bottom: 0.25rem;
    }
    .risk-summary-value {
      background: #f3f6fb;
      border: 1px solid #e0e7f2;
      border-radius: 10px;
      padding: 0.5rem 0.62rem;
      color: #1f2937;
      font-weight: 700;
      min-height: 2.45rem;
      display: flex;
      align-items: center;
      overflow: hidden;
      text-overflow: ellipsis;
      white-space: nowrap;
    }
    .risk-summary-value.blue-soft {
      background: #eaf3ff;
      border-color: #bfdbfe;
      color: #1e3a8a;
    }
    /* Cleaner inline editor surfaces in risk dashboard (non-sidebar) */
    [data-testid="stAppViewContainer"] .risk-summary-card [data-baseweb="select"] > div {
      background: #ffffff !important;
      border: 1px solid #cdd9e6 !important;
      border-radius: 12px !important;
      box-shadow: inset 0 0 0 1px rgba(255,255,255,0.35);
      min-height: 2.65rem !important;
    }
    [data-testid="stAppViewContainer"] .risk-summary-card [data-baseweb="select"] span,
    [data-testid="stAppViewContainer"] .risk-summary-card [data-baseweb="select"] div {
      color: #1f2b3d !important;
      font-weight: 700 !important;
    }
    [data-testid="stAppViewContainer"] .risk-summary-card [data-testid="stNumberInput"] input {
      background: #ffffff !important;
      border: 1px solid #cdd9e6 !important;
      border-radius: 12px !important;
      color: #1f2b3d !important;
      font-weight: 700 !important;
      min-height: 2.65rem !important;
    }
    [data-testid="stAppViewContainer"] .risk-summary-card label {
      color: #111827 !important;
      font-weight: 800 !important;
    }
    [data-testid="stAppViewContainer"] .risk-summary-card [data-testid="stButton"] button {
      min-height: 2.65rem !important;
      border-radius: 12px !important;
    }
    .provisional-box {
      border: 1px solid #f59e0b;
      background: #fffbeb;
      color: #7c2d12;
      border-radius: 12px;
      padding: 0.9rem 1rem;
      margin-bottom: 0.8rem;
    }
    .provisional-box strong {
      color: #92400e;
    }
    .risk-score-grid {
      display: grid;
      grid-template-columns: repeat(2, minmax(0, 1fr));
      gap: 0.8rem;
      margin-bottom: 0.8rem;
    }
    .risk-score-card {
      border-radius: 14px;
      padding: 1rem 1.05rem;
      border: 2px solid #bfdbfe;
      background: #eff6ff;
    }
    .risk-score-card.provisional {
      background: linear-gradient(140deg, #dcecff 0%, #d7e7ff 100%);
      border-color: #7eb5ff;
      box-shadow: inset 0 1px 0 rgba(255,255,255,0.7);
    }
    .risk-inline-editor {
      background: #ffffff;
      border: 1px solid #d6e1ef;
      border-radius: 14px;
      padding: 0.9rem 1rem 0.7rem;
      margin-bottom: 0.75rem;
      box-shadow: 0 6px 18px rgba(15, 23, 42, 0.04);
    }
    .risk-inline-editor > div,
    .risk-inline-editor [data-testid="stVerticalBlock"] {
      background: #ffffff !important;
    }
    .risk-inline-editor [data-testid="stNumberInput"],
    .risk-inline-editor [data-testid="stSelectbox"],
    .risk-inline-editor [data-testid="stCaptionContainer"],
    .risk-inline-editor [data-testid="stMarkdownContainer"] {
      background: transparent !important;
    }
    .risk-inline-wrap [data-testid="stVerticalBlockBorderWrapper"] {
      background: linear-gradient(180deg, #f8fbff 0%, #f3f8ff 100%) !important;
      border: 1px solid #cfe0f5 !important;
      border-radius: 14px !important;
      box-shadow: 0 6px 18px rgba(15, 23, 42, 0.04) !important;
    }
    .risk-inline-wrap [data-testid="stVerticalBlock"] {
      background: transparent !important;
    }
    .limit-card {
      border-radius: 14px;
      padding: 1rem 1.05rem;
      border: 2px solid #86efac;
      background: #f0fdf4;
    }
    .memo-card {
      background: #ffffff;
      border: 1px solid #d8e1ec;
      border-radius: 14px;
      padding: 1rem 1.05rem;
      box-shadow: 0 4px 14px rgba(15, 23, 42, 0.04);
      margin-bottom: 0.8rem;
    }
    .memo-sec {
      border-radius: 11px;
      border: 1px solid #dbe5f4;
      padding: 0.72rem 0.84rem;
      margin-top: 0.62rem;
    }
    .memo-sec-title {
      margin: 0 0 0.36rem;
      font-weight: 800;
      font-size: 1.01rem;
      color: #1e3a8a;
    }
    .memo-list {
      margin: 0;
      padding-left: 1.05rem;
      color: #1f2937;
      line-height: 1.45;
      font-weight: 600;
    }
    .memo-exec { background: #eff6ff; border-color: #bfdbfe; }
    .memo-strength { background: #f0fdf4; border-color: #bbf7d0; }
    .memo-risk { background: #fef2f2; border-color: #fecaca; }
    .memo-cushion { background: #eef2ff; border-color: #c7d2fe; }
    .memo-rec { background: #faf5ff; border-color: #ddd6fe; }
    @media (max-width: 1280px) {
      .risk-summary-row { grid-template-columns: repeat(3, minmax(0, 1fr)); }
      .risk-score-grid { grid-template-columns: 1fr; }
    }
    @media (max-width: 820px) {
      .risk-summary-row { grid-template-columns: repeat(2, minmax(0, 1fr)); }
    }
    </style>
    """,
    unsafe_allow_html=True,
)

st.markdown(
    """
    <style>
    .panel, .upload-shell, .risk-table-wrap, .memo-card, .risk-summary-card, .limit-card, .risk-score-card {
      border-radius: 16px !important;
      border: 1px solid #d7e0ea !important;
      box-shadow: 0 10px 28px rgba(15,23,42,0.06) !important;
    }
    .panel { padding: 0.95rem 1rem !important; margin-bottom: 0.72rem !important; }
    .cap-card {
      border-radius: 14px !important;
      border: 1px solid #d8e1ec !important;
      background: linear-gradient(180deg, #ffffff 0%, #f8fbff 100%) !important;
    }
    .stButton > button {
      border-radius: 11px !important;
      border: 1px solid #c6d4e5 !important;
      font-weight: 700 !important;
      transition: all .16s ease;
    }
    .stButton > button:hover { transform: translateY(-1px); box-shadow: 0 7px 16px rgba(15,23,42,0.10) !important; }
    .risk-table th {
      background: linear-gradient(180deg, #f1f6fc 0%, #e8f0fa 100%) !important;
      border-bottom: 1px solid #d4dfec !important;
    }
    [data-testid="stDataFrame"] > div {
      border-radius: 14px !important;
      border: 1px solid #d8e1ec !important;
      box-shadow: 0 8px 18px rgba(15,23,42,0.05) !important;
    }
    .memo-rec { background: #f3f8ff !important; border-color: #c7d9ef !important; }
    .memo-cushion { background: #eef6ff !important; border-color: #c8dcf4 !important; }
    .hover-help {
      border-color: #c7d2fe !important;
      color: #1d4ed8 !important;
      background: #f3f7ff !important;
    }
    @keyframes liftIn { from { opacity: .0; transform: translateY(7px); } to { opacity: 1; transform: translateY(0); } }
    .panel, .upload-shell, .risk-table-wrap, .memo-card { animation: liftIn .2s ease-out both; }
    </style>
    """,
    unsafe_allow_html=True,
)

BASE_DIR = Path("/Users/ashnaparekh/workspace/ashna_finance_project")
LIB_DIR = BASE_DIR / "loan_doc_extractor_mvp" / "uploaded_docs"
LIB_DIR.mkdir(parents=True, exist_ok=True)
EXTRACT_CACHE_DIR = LIB_DIR / ".extract_cache"
EXTRACT_CACHE_DIR.mkdir(parents=True, exist_ok=True)

WORKSPACE_PDFS = [
    BASE_DIR / "EX-10.59.pdf",
    BASE_DIR / "Document.pdf",
    BASE_DIR / "aple-ex101_50.htm.pdf",
]

DOC_TYPE_CONFIG = {
    "Credit Agreement": {
        "sheets": ["Covenant Definitions", "Financial Covenants", "Testing Terms", "Default Triggers"],
        "message": "Core legal covenant and trigger extraction for monitoring setup.",
    },
    "Term Sheet": {
        "sheets": ["Facility Overview", "Pricing", "Covenant Summary"],
        "message": "Pre-close terms, pricing and covenant signal capture.",
    },
    "Amendment": {
        "sheets": ["Covenant Changes", "EBITDA Adjustments", "Maturity Changes"],
        "message": "Amendment deltas and revised covenant posture.",
    },
    "Compliance Certificate": {
        "sheets": ["Reported Financials", "Calculated Ratios", "Certification"],
        "message": "Borrower reporting and compliance validation tabs.",
    },
    "Financial Statements": {
        "sheets": ["Income Statement", "Balance Sheet", "Cash Flow"],
        "message": "Financial statement extraction and ratio validation.",
    },
    "Forecast Model": {
        "sheets": ["Projected Income", "Projected Debt", "Scenario Assumptions"],
        "message": "Forward-looking covenant and stress metrics.",
    },
    "Security Agreement": {
        "sheets": ["Collateral Details", "Guarantee Info"],
        "message": "Collateral, lien and guarantor recovery structure.",
    },
    "Borrowing Base": {
        "sheets": ["Eligible A/R", "Inventory"],
        "message": "Asset-based lending availability and coverage.",
    },
    "Fee Letter": {
        "sheets": ["Fee Structure", "Pricing Grid"],
        "message": "Fee income and leverage-linked step-up terms.",
    },
    # Standalone + filing-level financial upload modes.
    "Income Statement": {
        "sheets": ["Income Statement"],
        "message": "Standalone income statement extraction.",
    },
    "Balance Sheet (Standalone)": {
        "sheets": ["Balance Sheet"],
        "message": "Standalone balance sheet extraction.",
    },
    "Cash Flow Statement (Standalone)": {
        "sheets": ["Cash Flow"],
        "message": "Standalone cash flow extraction.",
    },
    "10-K": {
        "sheets": ["Income Statement", "Balance Sheet", "Cash Flow"],
        "message": "Full annual filing extraction (multi-statement).",
    },
    "10-Q": {
        "sheets": ["Income Statement", "Balance Sheet", "Cash Flow"],
        "message": "Full quarterly filing extraction (multi-statement).",
    },
    "Annual Report": {
        "sheets": ["Income Statement", "Balance Sheet", "Cash Flow"],
        "message": "Annual report extraction (multi-statement).",
    },
    "Other Financial Filing": {
        "sheets": ["Income Statement", "Balance Sheet", "Cash Flow"],
        "message": "Generic financial filing extraction (multi-statement).",
    },
}

MANUAL_DOC_TYPE_OPTIONS = [
    "Income Statement",
    "Balance Sheet",
    "Cash Flow",
    "10-K",
    "10-Q",
    "Annual Report",
    "Other Financial Filing",
    "Credit Agreement",
    "Term Sheet",
    "Amendment",
    "Compliance Certificate",
    "Forecast Model",
    "Security Agreement",
    "Borrowing Base",
    "Fee Letter",
]

FILING_DOC_TYPES = {"10-K", "10-Q", "Annual Report", "Other Financial Filing"}
STANDALONE_DOC_TYPES = {"Income Statement", "Balance Sheet", "Cash Flow"}

TEMPLATE_COLUMNS = {
    "Covenant Definitions": [
        "EBITDA Add-backs", "EBITDA Exclusions", "Interest Expense Definition", "Total Debt Definition",
        "Normalized EBITDA", "Adjusted Interest", "Net Debt Calculation", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Financial Covenants": [
        "Minimum Interest Coverage", "Maximum Leverage Ratio", "Fixed Charge Coverage", "Minimum Liquidity",
        "Required Covenant Thresholds", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Testing Terms": [
        "Testing Frequency", "Test Dates", "Equity Cure Rights", "Covenant Monitoring Schedule",
        "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Default Triggers": [
        "Cross-default Clauses", "Material Adverse Change Clauses", "Risk Escalation Triggers",
        "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Facility Overview": [
        "Loan Amount", "Revolver Size", "Tenor", "Amortization", "Debt Maturity Schedule", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Pricing": [
        "Margin", "Base Rate Type", "Margin Grid", "Effective Interest Rate", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Covenant Summary": [
        "Headline Covenant Levels", "Preliminary Covenant Thresholds", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Covenant Changes": [
        "Updated Leverage Limits", "Reset Coverage Ratios", "Revised Covenant Thresholds", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "EBITDA Adjustments": [
        "Temporary Add-backs", "Covenant Holiday Terms", "Adjusted EBITDA Logic", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Maturity Changes": [
        "Extended Maturity Dates", "Updated Maturity Ladder", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Reported Financials": [
        "EBITDA", "Interest Expense", "Total Debt", "Cash", "Reported Coverage Ratio", "Reported Leverage Ratio",
        "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Calculated Ratios": [
        "Borrower Interest Coverage", "Borrower Leverage", "Required Interest Coverage", "Required Leverage",
        "Covenant Cushion", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Certification": [
        "Officer Sign-off", "Legal Attestation Record", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Income Statement": [
        "Revenue", "COGS", "EBITDA", "EBIT", "Net Income", "Interest Expense",
        "EBITDA (Adjusted)", "EBIT (Adjusted)", "Net Income (Adjusted)",
        "Total Expense", "Employee Compensation", "Non-Interest Expense",
        "Operating Income", "Operating Margin", "Income Before Taxes", "Tax Expense", "Effective Tax Rate",
        "EPS (Basic)", "EPS (Diluted)",
        "EBITDA Margin", "Interest Coverage Ratio", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Balance Sheet": [
        "Total Assets", "Total Liabilities", "Shareholders' Equity",
        "Total Debt", "Short-term Debt", "Current Portion of Long-term Debt", "Long-term Debt",
        "Cash", "Current Assets", "Inventory", "Accounts Receivable", "Current Liabilities",
        "Net Debt", "Current Ratio", "Leverage Ratio", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Cash Flow": [
        "Operating Cash Flow", "CapEx", "Free Cash Flow", "Free Cash Flow Coverage", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Projected Income": [
        "Projected EBITDA", "Projected Revenue", "Projected Interest Expense", "Forward Interest Coverage", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Projected Debt": [
        "Future Debt Balances", "Projected EBITDA", "Refinancing Assumptions", "Projected Leverage Ratio", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Scenario Assumptions": [
        "Revenue Growth %", "Stress Case", "Stress Test Ratios", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Collateral Details": [
        "Asset Type", "Lien Priority", "Collateral Value", "Loan Amount", "Loan-to-Value (LTV)", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Guarantee Info": [
        "Guarantor Entities", "Structural Subordination Risk", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Eligible A/R": [
        "A/R Aging", "Ineligible Receivables", "Borrowing Base Availability", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Inventory": [
        "Eligible Inventory Value", "Collateral Coverage Ratio", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Fee Structure": [
        "Upfront Fee %", "Commitment Fee %", "Ticking Fee", "Total Fee Income", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
    "Pricing Grid": [
        "Margin Step-ups", "Dynamic Interest Cost", "Confidence", "Page", "Open PDF Page", "Snippet",
    ],
}

SHEET_FIELD_HINTS: Dict[str, Dict[str, List[str]]] = {
    "Covenant Definitions": {
        "EBITDA Add-backs": ["ebitda", "add", "back"],
        "EBITDA Exclusions": ["ebitda", "exclusion"],
        "Interest Expense Definition": ["interest", "expense"],
        "Total Debt Definition": ["total", "debt"],
    },
    "Financial Covenants": {
        "Minimum Interest Coverage": ["minimum", "interest", "coverage"],
        "Maximum Leverage Ratio": ["maximum", "leverage"],
        "Fixed Charge Coverage": ["fixed", "charge", "coverage"],
        "Minimum Liquidity": ["minimum", "liquidity"],
    },
    "Testing Terms": {
        "Testing Frequency": ["test", "frequency"],
        "Test Dates": ["test", "date"],
        "Equity Cure Rights": ["equity", "cure"],
    },
    "Default Triggers": {
        "Cross-default Clauses": ["cross", "default"],
        "Material Adverse Change Clauses": ["material", "adverse", "change"],
    },
    "Facility Overview": {
        "Loan Amount": ["loan", "amount"],
        "Revolver Size": ["revolver", "size"],
        "Tenor": ["tenor"],
        "Amortization": ["amortization"],
    },
    "Pricing": {
        "Margin": ["margin"],
        "Base Rate Type": ["base", "rate", "type"],
        "Margin Grid": ["margin", "grid"],
    },
    "Covenant Summary": {"Headline Covenant Levels": ["headline", "covenant", "level"]},
    "Covenant Changes": {
        "Updated Leverage Limits": ["updated", "leverage", "limit"],
        "Reset Coverage Ratios": ["reset", "coverage", "ratio"],
    },
    "EBITDA Adjustments": {
        "Temporary Add-backs": ["temporary", "add", "back"],
        "Covenant Holiday Terms": ["covenant", "holiday"],
    },
    "Maturity Changes": {"Extended Maturity Dates": ["extended", "maturity", "date"]},
    "Reported Financials": {
        "EBITDA": ["ebitda"],
        "Interest Expense": ["interest", "expense"],
        "Total Debt": ["total", "debt"],
        "Cash": ["cash"],
    },
    "Calculated Ratios": {
        "Borrower Interest Coverage": ["borrower", "interest", "coverage"],
        "Borrower Leverage": ["borrower", "leverage"],
        "Required Interest Coverage": ["required", "interest", "coverage"],
        "Required Leverage": ["required", "leverage"],
    },
    "Certification": {"Officer Sign-off": ["officer", "sign"]},
    "Income Statement": {
        "Revenue": ["revenue"],
        "COGS": ["cogs", "cost of goods sold", "cost of revenue", "cost of revenues", "cost of sales"],
        "EBITDA": ["ebitda"],
        "EBIT": ["ebit"],
        "Net Income": ["net", "income"],
        "Interest Expense": ["interest", "expense"],
        "Non-Interest Expense": ["non-interest expense", "noninterest expense", "non interest expense"],
        "Operating Income": ["operating", "income", "income from operations", "operating profit"],
        "Income Before Taxes": ["income before taxes", "income before tax", "pretax income", "income before income taxes"],
        "Tax Expense": ["income tax expense", "provision for income taxes", "tax expense"],
        "EPS (Basic)": ["basic", "earnings per share", "eps"],
        "EPS (Diluted)": ["diluted", "earnings per share", "eps"],
    },
    "Balance Sheet": {
        "Total Assets": ["total", "assets"],
        "Total Liabilities": ["total", "liabilities"],
        "Shareholders' Equity": ["shareholders", "equity", "stockholders", "equity"],
        "Total Debt": ["total", "debt"],
        "Short-term Debt": ["short", "term", "debt"],
        "Current Portion of Long-term Debt": ["current", "portion", "long", "term", "debt"],
        "Long-term Debt": ["long", "term", "debt"],
        "Cash": ["cash"],
        "Current Assets": ["current", "assets"],
        "Inventory": ["inventory"],
        "Accounts Receivable": ["accounts", "receivable", "trade", "receivable"],
        "Current Liabilities": ["current", "liabilities"],
    },
    "Cash Flow": {
        "Operating Cash Flow": ["operating", "cash", "flow"],
        "CapEx": ["capex"],
        "Free Cash Flow": ["free", "cash", "flow"],
    },
    "Projected Income": {
        "Projected EBITDA": ["projected", "ebitda"],
        "Projected Revenue": ["projected", "revenue"],
        "Projected Interest Expense": ["projected", "interest", "expense"],
    },
    "Projected Debt": {
        "Future Debt Balances": ["future", "debt", "balance"],
        "Projected EBITDA": ["projected", "ebitda"],
        "Refinancing Assumptions": ["refinancing", "assumption"],
    },
    "Scenario Assumptions": {"Revenue Growth %": ["revenue", "growth"], "Stress Case": ["stress", "case"]},
    "Collateral Details": {
        "Asset Type": ["asset", "type"],
        "Lien Priority": ["lien", "priority"],
        "Collateral Value": ["collateral", "value"],
        "Loan Amount": ["loan", "amount"],
    },
    "Guarantee Info": {"Guarantor Entities": ["guarantor", "entities"]},
    "Eligible A/R": {"A/R Aging": ["ar", "aging"], "Ineligible Receivables": ["ineligible", "receivable"]},
    "Inventory": {"Eligible Inventory Value": ["eligible", "inventory", "value"]},
    "Fee Structure": {"Upfront Fee %": ["upfront", "fee"], "Commitment Fee %": ["commitment", "fee"], "Ticking Fee": ["ticking", "fee"]},
    "Pricing Grid": {"Margin Step-ups": ["margin", "step"]},
}

META_COLUMNS = {"Confidence", "Page", "Pages", "Open PDF Page", "Snippet"}
META_TAIL_ORDER = ["Open PDF Page", "Snippet", "Confidence", "Page", "Pages"]
YEAR_META_COLUMNS = ["Selected Year", "Fiscal Period", "Analysis Mode", "Detected Unit", "Unit Multiplier", "Available Years"]
DERIVED_COLUMNS = {
    "Normalized EBITDA", "Adjusted Interest", "Net Debt Calculation", "Required Covenant Thresholds",
    "Covenant Monitoring Schedule", "Risk Escalation Triggers", "Debt Maturity Schedule", "Effective Interest Rate",
    "Preliminary Covenant Thresholds", "Revised Covenant Thresholds", "Adjusted EBITDA Logic", "Updated Maturity Ladder",
    "Reported Coverage Ratio", "Reported Leverage Ratio", "Covenant Cushion", "Legal Attestation Record",
    "EBITDA Margin", "Interest Coverage Ratio", "Net Debt", "Current Ratio", "Leverage Ratio", "Free Cash Flow Coverage",
    "Forward Interest Coverage", "Projected Leverage Ratio", "Stress Test Ratios", "Loan-to-Value (LTV)",
    "Structural Subordination Risk", "Borrowing Base Availability", "Collateral Coverage Ratio", "Total Fee Income", "Dynamic Interest Cost",
    "Operating Margin", "Effective Tax Rate",
}

DOC_KEYWORDS = {
    "Credit Agreement": ["credit", "agreement"],
    "Term Sheet": ["term", "sheet"],
    "Amendment": ["amendment"],
    "Compliance Certificate": ["compliance", "certificate"],
    "Financial Statements": ["financial", "statement"],
    "Forecast Model": ["forecast", "projection", "model"],
    "Security Agreement": ["security", "collateral"],
    "Borrowing Base": ["borrowing", "base"],
    "Fee Letter": ["fee", "letter"],
}


def _file_page_url(pdf_path: Path, page_number: Optional[int]) -> Optional[str]:
    if page_number is None:
        return None
    return f"{APP_BASE_URL}/?open_doc={quote(str(pdf_path.resolve()), safe='')}&open_page={int(page_number)}"


def _normalize_doc_type(value: Optional[str]) -> str:
    v = re.sub(r"\s+", " ", (value or "").strip().lower())
    if not v:
        return "Other Financial Filing"

    # Explicit canonical aliases first.
    exact_aliases = {
        "10-k": "10-K",
        "10k": "10-K",
        "10-q": "10-Q",
        "10q": "10-Q",
        "annual report": "Annual Report",
        "other financial filing": "Other Financial Filing",
        "financial statements": "Financial Statements",
        "income statement": "Income Statement",
        "balance sheet": "Balance Sheet",
        "cash flow statement": "Cash Flow",
        "credit agreement": "Credit Agreement",
        "term sheet": "Term Sheet",
        "amendment": "Amendment",
        "compliance certificate": "Compliance Certificate",
        "covenant compliance certificate": "Compliance Certificate",
        "security agreement": "Security Agreement",
        "borrowing base": "Borrowing Base",
        "fee letter": "Fee Letter",
        "forecast model": "Forecast Model",
    }
    if v in exact_aliases:
        return exact_aliases[v]

    # Filing / financial precedence by semantic keywords.
    if ("10-k" in v) or ("10k" in v) or ("form 10-k" in v):
        return "10-K"
    if ("10-q" in v) or ("10q" in v) or ("form 10-q" in v):
        return "10-Q"
    if "annual report" in v:
        return "Annual Report"
    if "income statement" in v or "statement of operations" in v or "statement of earnings" in v:
        return "Income Statement"
    if "balance sheet" in v or "statement of financial position" in v:
        return "Balance Sheet"
    if "cash flow statement" in v or "statement of cash flows" in v or "cash flow" in v:
        return "Cash Flow"
    if "financial statement" in v:
        return "Financial Statements"

    if "term" in v and "sheet" in v:
        return "Term Sheet"
    if "amend" in v:
        return "Amendment"
    if "compliance" in v and "certificate" in v:
        return "Compliance Certificate"
    if "security" in v and "agreement" in v:
        return "Security Agreement"
    if "borrowing" in v and "base" in v:
        return "Borrowing Base"
    if "fee" in v and "letter" in v:
        return "Fee Letter"
    if "credit" in v and "agreement" in v:
        return "Credit Agreement"
    if "forecast" in v or "projection" in v:
        return "Forecast Model"

    # Keep previous broad keyword mapping as fallback.
    for doc_type, kws in DOC_KEYWORDS.items():
        if all(k in v for k in kws):
            return doc_type
    return "Other Financial Filing"


def _guess_type_from_filename(path: Path) -> str:
    n = path.name.lower()
    income_syn = [
        "income statement",
        "income_statement",
        "statement of operations",
        "statements of operations",
        "statement of earnings",
        "profit and loss",
        "p&l",
        "comprehensive income",
        "consolidated statements of operations",
        "consolidated statement of operations",
    ]
    balance_syn = [
        "balance sheet",
        "balance_sheet",
        "statement of financial position",
        "consolidated balance sheets",
    ]
    cashflow_syn = [
        "cash flow statement",
        "cash flow",
        "cash_flow",
        "statement of cash flows",
        "consolidated statements of cash flows",
    ]
    debt_schedule_syn = [
        "debt schedule",
        "schedule of indebtedness",
        "maturities of debt",
        "debt obligations",
        "notes payable",
        "long term debt",
        "long-term debt",
    ]
    if any(k in n for k in income_syn):
        return "Income Statement"
    if any(k in n for k in balance_syn):
        return "Balance Sheet"
    if any(k in n for k in cashflow_syn):
        return "Cash Flow"
    if any(k in n for k in debt_schedule_syn):
        return "Term Sheet"
    if "10-k" in n or "10k" in n or "form 10-k" in n:
        return "10-K"
    if "10-q" in n or "10q" in n or "quarterly report" in n or "form 10-q" in n:
        return "10-Q"
    if "annual report" in n:
        return "Annual Report"
    if "income" in n and "statement" in n:
        return "Income Statement"
    if "balance" in n and "sheet" in n:
        return "Balance Sheet"
    if "term" in n and "sheet" in n:
        return "Term Sheet"
    if "amend" in n:
        return "Amendment"
    if "compliance" in n or "certificate" in n:
        return "Compliance Certificate"
    if "financial" in n and ("statement" in n or "fs" in n):
        return "Financial Statements"
    if "forecast" in n or "projection" in n:
        return "Forecast Model"
    if "security" in n:
        return "Security Agreement"
    if "borrowing" in n and "base" in n:
        return "Borrowing Base"
    if "fee" in n and "letter" in n:
        return "Fee Letter"
    return "Other Financial Filing"


def _guess_type_from_pdf_content(path: Path) -> str:
    try:
        with fitz.open(path) as doc:
            parts: List[str] = []
            for i, page in enumerate(doc):
                if i >= 12:
                    break
                parts.append((page.get_text("text") or "").lower())
        text = " ".join(parts)
    except Exception:
        return "Other Financial Filing"

    income_syn = [
        "income statement",
        "statement of operations",
        "statements of operations",
        "statement of earnings",
        "profit and loss",
        "p&l",
        "comprehensive income",
        "consolidated statements of operations",
        "consolidated statement of operations",
    ]
    balance_syn = [
        "balance sheet",
        "statement of financial position",
        "consolidated balance sheets",
    ]
    cashflow_syn = [
        "statement of cash flows",
        "consolidated statements of cash flows",
        "cash flow statement",
    ]
    debt_schedule_syn = [
        "notes payable",
        "long term debt",
        "long-term debt",
        "maturities of debt",
        "debt obligations",
        "schedule of indebtedness",
        "debt schedule",
    ]

    # SEC filing detection: prefer explicit form cues and common "10-K" variants.
    if "form 10-k" in text or re.search(r"\b10[\s-]?k\b", text):
        return "10-K"
    if "form 10-q" in text or re.search(r"\b10[\s-]?q\b", text):
        return "10-Q"

    # Robust semantic scoring for filing-style docs with non-obvious titles.
    income_hits = sum(1 for k in income_syn if k in text)
    balance_hits = sum(1 for k in balance_syn if k in text)
    cashflow_hits = sum(1 for k in cashflow_syn if k in text)
    fin_statement_hits = income_hits + balance_hits + cashflow_hits
    if "net income" in text:
        income_hits += 1
        fin_statement_hits += 1
    if "total assets" in text and "total liabilities" in text:
        balance_hits += 2
        fin_statement_hits += 2
    if "operating activities" in text and "investing activities" in text:
        cashflow_hits += 2
        fin_statement_hits += 2
    statement_count = sum(1 for x in [income_hits, balance_hits, cashflow_hits] if x > 0)
    if statement_count >= 2:
        return "Financial Statements"
    if balance_hits > 0 and income_hits == 0 and cashflow_hits == 0:
        return "Balance Sheet"
    if income_hits > 0 and balance_hits == 0 and cashflow_hits == 0:
        return "Income Statement"
    if cashflow_hits > 0 and income_hits == 0 and balance_hits == 0:
        return "Cash Flow"
    # Credit-document semantic scoring.
    credit_terms = [
        "credit agreement",
        "borrower",
        "lender",
        "administrative agent",
        "term loan",
        "revolving credit",
        "commitment",
        "covenant",
        "event of default",
        "borrowing base",
        "security interest",
    ]
    credit_hits = sum(1 for k in credit_terms if k in text)
    if "term sheet" in text:
        return "Term Sheet"
    if "compliance certificate" in text:
        return "Compliance Certificate"
    if "amendment no." in text or ("first amendment" in text) or ("amendment to credit agreement" in text):
        return "Amendment"
    if "security agreement" in text:
        return "Security Agreement"
    if "fee letter" in text:
        return "Fee Letter"
    if credit_hits >= 3:
        return "Credit Agreement"
    if any(k in text for k in debt_schedule_syn):
        return "Term Sheet"
    if "annual report" in text:
        return "Annual Report"
    if "borrowing base" in text:
        return "Borrowing Base"
    return "Other Financial Filing"


def _resolve_detected_type(selected_pdf: Path, extracted_type: Optional[str]) -> str:
    extracted_norm = _normalize_doc_type(extracted_type)
    filename_hint = _guess_type_from_filename(selected_pdf)
    content_hint = _guess_type_from_pdf_content(selected_pdf)
    hints = [extracted_norm, filename_hint, content_hint]

    # Treat these as generic classes that should not override specific doc evidence.
    generic_types = {"Other Financial Filing", "Financial Statements"}
    specific_hints = [h for h in hints if h and h not in generic_types]
    if specific_hints:
        # Prefer explicit, specific types first.
        priority = {
            "10-K": 100,
            "10-Q": 98,
            "Annual Report": 96,
            "Income Statement": 95,
            "Balance Sheet": 95,
            "Cash Flow": 95,
            "Credit Agreement": 95,
            "Term Sheet": 92,
            "Amendment": 90,
            "Compliance Certificate": 88,
            "Security Agreement": 86,
            "Borrowing Base": 84,
            "Fee Letter": 82,
            "Forecast Model": 80,
        }
        source_weight = [6, 3, 2]  # extracted, filename, content
        best = specific_hints[0]
        best_score = -1
        for idx, hint in enumerate(hints):
            if hint not in specific_hints:
                continue
            score = priority.get(hint, 0) * 10 + source_weight[idx]
            if score > best_score:
                best_score = score
                best = hint
        return best

    # Generic fallback: prefer Financial Statements over Other Financial Filing.
    if "Financial Statements" in hints:
        return "Financial Statements"
    if "Other Financial Filing" in hints:
        return "Other Financial Filing"
    return "Financial Statements"


def _normalize_manual_doc_type(selected: str) -> str:
    v = str(selected or "").strip()
    if v == "Balance Sheet":
        return "Balance Sheet (Standalone)"
    if v == "Cash Flow":
        return "Cash Flow Statement (Standalone)"
    return v


def _display_doc_type_label(applied_type: str) -> str:
    if applied_type == "Balance Sheet (Standalone)":
        return "Balance Sheet"
    if applied_type == "Cash Flow Statement (Standalone)":
        return "Cash Flow"
    return applied_type


def _active_sheets_for_doc_type(applied_type: str) -> List[str]:
    cfg = DOC_TYPE_CONFIG.get(applied_type, DOC_TYPE_CONFIG["Financial Statements"])
    return cfg["sheets"]


def _is_filing_doc_type(applied_type: str) -> bool:
    return applied_type in FILING_DOC_TYPES


def _is_standalone_financial_doc_type(applied_type: str) -> bool:
    return applied_type in STANDALONE_DOC_TYPES or applied_type in {"Balance Sheet (Standalone)", "Cash Flow Statement (Standalone)"}


def _build_derived_metrics_table(sheet_map: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    income = sheet_map.get("Income Statement", pd.DataFrame())
    balance = sheet_map.get("Balance Sheet", pd.DataFrame())
    cashflow = sheet_map.get("Cash Flow", pd.DataFrame())

    def _get(df: pd.DataFrame, col: str) -> Optional[float]:
        if df is None or df.empty or col not in df.columns:
            return None
        return _to_numeric_financial(df.iloc[0].get(col))

    revenue = _get(income, "Revenue")
    net_income = _get(income, "Net Income")
    ebit = _get(income, "EBIT")
    interest = _get(income, "Interest Expense")
    total_debt = _get(balance, "Total Debt")
    equity = _get(balance, "Shareholders' Equity")
    curr_assets = _get(balance, "Current Assets")
    curr_liab = _get(balance, "Current Liabilities")
    cash = _get(balance, "Cash")
    receivables = _extract_num(_build_consolidated_view(sheet_map), ["Accounts Receivable", "Trade Receivables", "Receivables, net"])
    ocf = _get(cashflow, "Operating Cash Flow")
    capex = _get(cashflow, "CapEx")
    fcf = _get(cashflow, "Free Cash Flow")
    if fcf is None and ocf is not None and capex is not None:
        fcf = ocf - capex

    rows = [
        {"Metric": "Net Margin", "Value": _safe_div(net_income, revenue), "Formula": "Net Income / Revenue"},
        {"Metric": "Interest Coverage", "Value": _safe_div(ebit, interest), "Formula": "EBIT / Interest Expense"},
        {"Metric": "Debt to Equity", "Value": _safe_div(total_debt, equity), "Formula": "Total Debt / Equity"},
        {"Metric": "Current Ratio", "Value": _safe_div(curr_assets, curr_liab), "Formula": "Current Assets / Current Liabilities"},
        {"Metric": "Quick Ratio", "Value": _safe_div((cash or 0.0) + (receivables or 0.0), curr_liab), "Formula": "(Cash + Receivables) / Current Liabilities"},
        {"Metric": "Free Cash Flow", "Value": fcf, "Formula": "Operating Cash Flow - CapEx"},
    ]
    out = pd.DataFrame(rows)
    out["Status"] = out["Value"].apply(lambda x: "Incomplete" if x is None or pd.isna(x) else "Available")
    return out


def _build_extraction_confidence_table(sheet_map: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for sheet_name, df in sheet_map.items():
        if df is None or df.empty:
            rows.append({"Statement": sheet_name, "Detected": "No", "Confidence": None, "Status": "Statement not detected in document."})
            continue
        conf = None
        if "Confidence" in df.columns:
            vals = [_to_numeric_financial(v) for v in df["Confidence"].tolist()]
            vals = [v for v in vals if v is not None]
            conf = (sum(vals) / len(vals)) if vals else None
        has_values = False
        for col in df.columns:
            if col in META_COLUMNS:
                continue
            if any(v not in (None, "", "None", "null") and not pd.isna(v) for v in df[col].tolist()):
                has_values = True
                break
        rows.append(
            {
                "Statement": sheet_name,
                "Detected": "Yes" if has_values else "No",
                "Confidence": conf,
                "Status": "OK" if has_values else "Statement not detected in document.",
            }
        )
    return pd.DataFrame(rows)


def _recommendation(applied_type: str, mode: str, detected_type: Optional[str]) -> str:
    sheets = ", ".join(DOC_TYPE_CONFIG[applied_type]["sheets"])
    if mode == "Auto-detect":
        if detected_type:
            return (
                f"We identified this as a {detected_type}. "
                f"Using the {applied_type} template with these output tabs: {sheets}. "
                "If this looks wrong, switch to Manual mode."
            )
        return (
            f"Auto-detect is on. Using the {applied_type} template with these output tabs: {sheets}. "
            "You can switch to Manual mode anytime."
        )
    return (
        f"Manual mode is on. You selected {applied_type}. "
        f"Your output tabs will be: {sheets}. "
        "Change document type anytime and columns update instantly."
    )


def _list_library_files() -> List[Path]:
    lib = sorted(
        [p for p in LIB_DIR.glob("*.pdf") if p.exists() and p.is_file() and p.stat().st_size > 0],
        key=lambda p: p.stat().st_mtime,
        reverse=True,
    )
    ws = [p for p in WORKSPACE_PDFS if p.exists()]
    return lib + ws


def _pdf_paths_equal(a: Any, b: Any) -> bool:
    """True if two paths refer to the same file (resolved). Avoids false mismatches from string vs Path or relative vs absolute."""
    if a is None or b is None:
        return a is b
    try:
        return Path(a).resolve() == Path(b).resolve()
    except Exception:
        return os.path.normpath(str(a)) == os.path.normpath(str(b))


def _is_allowed_pdf_path(path: Path) -> bool:
    try:
        rp = path.resolve()
    except Exception:
        return False
    if rp.suffix.lower() != ".pdf":
        return False
    if rp.exists() and rp.is_file():
        if rp in [p.resolve() for p in WORKSPACE_PDFS if p.exists()]:
            return True
        try:
            return LIB_DIR.resolve() in rp.parents
        except Exception:
            return False
    return False


def _should_block_selected_context(selected_pdf: Optional[Path], session_pdf_path: Optional[str], has_cache: bool) -> bool:
    if not selected_pdf or not session_pdf_path:
        return False
    return str(Path(session_pdf_path)) != str(selected_pdf) and not has_cache


def _uploaded_file_bytes(uploaded_file) -> bytes:
    # Streamlit UploadedFile supports getvalue()/getbuffer() and is safe for repeated reads.
    data: bytes = b""
    try:
        if hasattr(uploaded_file, "getvalue"):
            raw = uploaded_file.getvalue()
            data = raw if isinstance(raw, (bytes, bytearray)) else bytes(raw or b"")
        elif hasattr(uploaded_file, "getbuffer"):
            data = bytes(uploaded_file.getbuffer())
        else:
            if hasattr(uploaded_file, "seek"):
                uploaded_file.seek(0)
            data = uploaded_file.read() or b""
            if hasattr(uploaded_file, "seek"):
                uploaded_file.seek(0)
    except Exception:
        data = b""
    if not data:
        raise ValueError(f"Uploaded file is empty or unreadable: {getattr(uploaded_file, 'name', 'unknown.pdf')}")
    return bytes(data)


def _save_uploaded(uploaded_file) -> Path:
    name = Path(str(getattr(uploaded_file, "name", "document.pdf"))).name
    stem = Path(name).stem
    suffix = Path(name).suffix or ".pdf"
    out = LIB_DIR / f"{stem}{suffix}"
    idx = 2
    while out.exists():
        out = LIB_DIR / f"{stem} ({idx}){suffix}"
        idx += 1
    payload = _uploaded_file_bytes(uploaded_file)
    with out.open("wb") as f:
        f.write(payload)
    if not out.exists() or out.stat().st_size <= 0:
        try:
            out.unlink(missing_ok=True)
        except Exception:
            pass
        raise ValueError(f"Failed to persist uploaded file: {name}")
    return out


def _snake_key(text: str) -> str:
    cleaned = re.sub(r"[\[\]\.]+", "_", text or "")
    cleaned = re.sub(r"[^0-9a-zA-Z_]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned.lower()


def _backend_doc_type(raw_doc_type: Optional[str]) -> str:
    if not raw_doc_type:
        return "other"
    norm = str(raw_doc_type).lower()
    if "10-k" in norm:
        return "10-K"
    if "10-q" in norm:
        return "10-Q"
    if "balance" in norm and "sheet" in norm:
        return "balance_sheet"
    if "credit agreement" in norm or "credit" in norm:
        return "credit_agreement"
    return "other"


def _flatten_extractor_fields(result: Dict[str, Any]) -> Dict[str, Any]:
    payload: Dict[str, Any] = {}
    flat_fields = result.get("flat_fields") or []
    for row in flat_fields:
        path = row.get("field_path") or ""
        key = _snake_key(path)
        if not key:
            continue
        payload[key] = row.get("value")
        payload[f"{key}__page_number"] = row.get("page_number")
        payload[f"{key}__confidence"] = row.get("confidence")
        payload[f"{key}__source_snippet"] = row.get("source_snippet")
    summary = result.get("summary") or {}
    if isinstance(summary, dict):
        for k, v in summary.items():
            payload[f"summary__{_snake_key(k)}"] = v
    return payload


# Display column (from _build_sheet / TEMPLATE_COLUMNS) -> ExtractedFinancial API field name
_BACKEND_STMT_MAP: List[tuple[str, str, Dict[str, str]]] = [
    (
        "Income Statement",
        "income_statement",
        {
            "Revenue": "revenue",
            "COGS": "cogs",
            "EBITDA": "ebitda",
            "EBIT": "ebit",
            "Net Income": "net_income",
            "Interest Expense": "interest_expense",
        },
    ),
    (
        "Balance Sheet",
        "balance_sheet",
        {
            "Total Assets": "total_assets",
            "Shareholders' Equity": "total_equity",
            "Current Assets": "current_assets",
            "Current Liabilities": "current_liabilities",
            "Inventory": "inventory",
            "Short-term Debt": "st_debt",
            "Long-term Debt": "lt_debt",
        },
    ),
    (
        "Cash Flow",
        "cash_flow",
        {
            "Operating Cash Flow": "operating_cf",
            "CapEx": "capex",
        },
    ),
]


def _statement_payloads_for_backend_sync(pdf_path: Path, extraction: Dict[str, Any]) -> List[Dict[str, Any]]:
    """Map MVP statement sheets to backend ExtractedFinancialCreate-shaped dicts (typed columns, not raw_fields)."""
    out: List[Dict[str, Any]] = []
    for sheet_name, statement_type, colmap in _BACKEND_STMT_MAP:
        try:
            df = _build_sheet(extraction, sheet_name, pdf_path)
        except Exception:
            continue
        if df is None or df.empty:
            continue
        row = df.iloc[0]
        payload: Dict[str, Any] = {"statement_type": statement_type}
        has_number = False
        for disp, api_key in colmap.items():
            if disp not in row.index:
                continue
            val = _to_numeric_financial(row.get(disp))
            if val is not None:
                payload[api_key] = float(val)
                has_number = True
        if "Selected Year" in row.index:
            sy = _to_numeric_financial(row.get("Selected Year"))
            if sy is not None:
                payload["selected_year"] = int(sy)
        conf = row.get("Confidence")
        if conf is not None and not (isinstance(conf, float) and pd.isna(conf)):
            if isinstance(conf, (int, float)):
                payload["extraction_confidence"] = f"{float(conf):.4f}".rstrip("0").rstrip(".")
            else:
                payload["extraction_confidence"] = str(conf)
        if has_number:
            out.append(payload)
    return out


def _api_risk_band_for_credit_api(grade: str, final_score: float) -> str:
    """Map MVP grade/score to FastAPI ALLOWED_RISK_BANDS."""
    g = str(grade or "").strip()
    if "Validation" in g:
        return "Elevated"
    if g == "Provisional":
        return "Moderate"
    if "Low Risk" in g:
        return "Low"
    if "Moderate Risk" in g:
        return "Moderate"
    if "Elevated Risk" in g:
        return "Elevated"
    if "High Risk" in g:
        return "High"
    fs = float(final_score or 0.0)
    if fs > 80:
        return "Low"
    if fs >= 60:
        return "Moderate"
    if fs >= 40:
        return "Elevated"
    return "High"


def _metric_scores_for_credit_api(risk_table: pd.DataFrame) -> List[Dict[str, Any]]:
    """Build exactly 12 metric rows for POST /api/persist/credit-analysis-with-metrics."""

    def _fo(v: Any) -> Optional[float]:
        if v is None:
            return None
        if isinstance(v, float) and pd.isna(v):
            return None
        try:
            return float(v)
        except Exception:
            return None

    out: List[Dict[str, Any]] = []
    for _, r in risk_table.iterrows():
        risk_cell = str(r.get("Risk") or "Medium").strip()
        api_risk = "High" if risk_cell == "High" else "Low"
        st_cell = str(r.get("Status") or "Incomplete")
        api_status = st_cell if st_cell in ("Calculated", "Incomplete") else "Incomplete"
        out.append(
            {
                "metric_name": str(r.get("Metric") or "Metric"),
                "calculated_value": _fo(r.get("Calculated Value")),
                "industry_threshold": _fo(r.get("Industry Threshold")),
                "base_score": _fo(r.get("Base Score")),
                "adjusted_score": _fo(r.get("Adjusted Score")),
                "status": api_status,
                "risk_level": api_risk,
            }
        )
    while len(out) < 12:
        out.append(
            {
                "metric_name": f"_pad_{len(out)}",
                "calculated_value": None,
                "industry_threshold": None,
                "base_score": None,
                "adjusted_score": None,
                "status": "Incomplete",
                "risk_level": "Low",
            }
        )
    return out[:12]


def _fetch_backend_document_id_by_filename(pdf_path: Path) -> Optional[int]:
    """Match uploaded document in API by exact file_name (list is newest-first)."""
    if not BACKEND_BASE_URL:
        return None
    try:
        r = requests.get(f"{BACKEND_BASE_URL}/api/credit/documents", timeout=30)
        if not r.ok:
            return None
        name = pdf_path.name
        for doc in r.json():
            if doc.get("file_name") == name:
                return int(doc["id"])
    except Exception:
        return None
    return None


def _ensure_backend_document_id_for_session_pdf() -> Optional[int]:
    """Ensure session has backend document id (upload sync, map, or API lookup)."""
    raw = st.session_state.get("backend_document_id")
    if raw is not None:
        return int(raw)
    pdf_ctx = st.session_state.get("pdf_path")
    if not pdf_ctx:
        return None
    try:
        p = Path(str(pdf_ctx))
    except Exception:
        return None
    bid = _fetch_backend_document_id_by_filename(p)
    if bid is not None:
        st.session_state["backend_document_id"] = bid
        st.session_state.pop("credit_analysis_missing_doc_warned", None)
        m = dict(st.session_state.get("backend_document_id_by_pdf") or {})
        try:
            m[str(p.resolve())] = bid
        except Exception:
            m[str(p)] = bid
        st.session_state["backend_document_id_by_pdf"] = m
        return bid
    return None


def _maybe_persist_credit_analysis_to_backend(
    model: Dict[str, Any],
    applied_inputs: Dict[str, Any],
    risk_compute_key: str,
) -> None:
    """Persist latest risk run to Postgres via FastAPI (credit_analyses + metric_scores)."""
    if not BACKEND_BASE_URL:
        return
    raw_id = _ensure_backend_document_id_for_session_pdf()
    if raw_id is None:
        if not st.session_state.get("credit_analysis_missing_doc_warned"):
            st.info(
                "Credit analysis was not saved: no backend document id. "
                "Ensure the API is running at **BACKEND_BASE_URL**, then run **Extract Data** once so the PDF uploads."
            )
            st.session_state["credit_analysis_missing_doc_warned"] = True
        return
    st.session_state.pop("credit_analysis_missing_doc_warned", None)
    dedupe = f"{int(raw_id)}|{risk_compute_key}|credit_analysis"
    if st.session_state.get("credit_analysis_backend_dedupe") == dedupe:
        return
    try:
        tbl = model.get("table")
        if tbl is None or not isinstance(tbl, pd.DataFrame) or tbl.empty:
            return
        metric_scores = _metric_scores_for_credit_api(tbl)
        final_score = float(model.get("final_score") or 0.0)
        body: Dict[str, Any] = {
            "document_id": int(raw_id),
            "industry": (str(applied_inputs.get("industry", ""))[:500] or None),
            "geographic_risk": (str(applied_inputs.get("geography", ""))[:500] or None),
            "business_stage": (str(applied_inputs.get("business_stage", ""))[:200] or None),
            "company_size": (str(applied_inputs.get("company_size", ""))[:100] or None),
            "loan_type": str(applied_inputs.get("loan_type", "Term Loan"))[:200],
            "years_in_operation": int(applied_inputs.get("years_in_operation", 0)),
            "requested_amount": float(applied_inputs.get("requested_amount", 0.0)),
            "currency_scale": str(applied_inputs.get("currency_scale", "Units"))[:100],
            "risk_score": final_score,
            "risk_band": _api_risk_band_for_credit_api(str(model.get("grade", "")), final_score),
            "policy_limit": float(model.get("policy_approved_limit") or model.get("approved_limit") or 0.0),
            "approval_status": (str(model.get("recommendation", ""))[:200] or None),
            "weighted_score": final_score,
            "metric_scores": metric_scores,
        }
        resp = requests.post(
            f"{BACKEND_BASE_URL}/api/persist/credit-analysis-with-metrics",
            json=body,
            timeout=120,
        )
        if resp.ok:
            st.session_state["credit_analysis_backend_dedupe"] = dedupe
            st.session_state.pop("credit_analysis_persist_warned", None)
        else:
            wk = f"cap_{dedupe}_{resp.status_code}"
            if st.session_state.get("credit_analysis_persist_warned") != wk:
                st.warning(
                    f"Credit analysis was not saved to Postgres (HTTP {resp.status_code}). "
                    f"{(resp.text or '')[:320]}"
                )
                st.session_state["credit_analysis_persist_warned"] = wk
    except Exception as exc:
        wk = f"cap_exc_{dedupe}"
        if st.session_state.get("credit_analysis_persist_warned") != wk:
            st.warning(f"Credit analysis save failed: {exc!s}"[:500])
            st.session_state["credit_analysis_persist_warned"] = wk


def _backend_post_dedupe_key(pdf_path: Path) -> str:
    """Stable per extraction artifact so we sync once per cached PDF, not every Streamlit rerun."""
    try:
        cp = _cache_path_for_pdf(pdf_path)
        if cp.exists():
            return f"{pdf_path.resolve()}|cache|{cp.stat().st_mtime_ns}"
        if pdf_path.exists():
            return f"{pdf_path.resolve()}|pdf|{pdf_path.stat().st_mtime_ns}"
    except OSError:
        pass
    return f"{pdf_path.resolve()}|nostat"


def _post_extraction_to_backend(pdf_path: Path, result: Dict[str, Any]) -> None:
    if not BACKEND_BASE_URL:
        return
    try:
        dedupe = _backend_post_dedupe_key(pdf_path)
        if st.session_state.get("backend_post_dedupe") == dedupe:
            byp = dict(st.session_state.get("backend_document_id_by_pdf") or {})
            bid = byp.get(str(pdf_path.resolve()))
            if bid is None:
                bid = _fetch_backend_document_id_by_filename(pdf_path)
                if bid is not None:
                    byp[str(pdf_path.resolve())] = bid
                    st.session_state["backend_document_id_by_pdf"] = byp
            if bid is not None:
                st.session_state["backend_document_id"] = int(bid)
            return
        detected_type = result.get("summary", {}).get("document_type")
        doc_type = _backend_doc_type(detected_type)
        company_name = _extract_company_name_from_pdf(pdf_path)
        with pdf_path.open("rb") as f:
            files = {"file": (pdf_path.name, f, "application/pdf")}
            data = {"doc_type": doc_type, "company_name": company_name or ""}
            resp = requests.post(f"{BACKEND_BASE_URL}/api/credit/documents/upload-and-parse", data=data, files=files, timeout=60)
        if not resp.ok:
            st.warning("Backend upload failed; document saved locally only.")
            return
        doc_id = resp.json().get("id")
        if not doc_id:
            st.warning("Backend upload response missing document ID.")
            return
        bid = int(doc_id)
        st.session_state["backend_document_id"] = bid
        st.session_state.pop("credit_analysis_missing_doc_warned", None)
        byp = dict(st.session_state.get("backend_document_id_by_pdf") or {})
        byp[str(pdf_path.resolve())] = bid
        st.session_state["backend_document_id_by_pdf"] = byp
        extraction = result.get("extraction") or {}
        payloads = _statement_payloads_for_backend_sync(pdf_path, extraction)
        if not payloads:
            st.warning(
                "Backend sync: no mapped statement numbers to save (extract data first, or check filing sheets)."
            )
            st.session_state["backend_post_dedupe"] = dedupe
            st.session_state.pop("risk_compute_cache", None)
            st.session_state.pop("credit_analysis_backend_dedupe", None)
            return
        last_err = None
        for body in payloads:
            persist = requests.post(
                f"{BACKEND_BASE_URL}/api/credit/documents/{doc_id}/financial-rows",
                json=body,
                timeout=60,
            )
            if not persist.ok:
                last_err = persist.status_code
        if last_err is not None:
            st.warning(
                f"Backend persist partially failed (last HTTP {last_err}); "
                "document uploaded but some statement rows may be missing."
            )
            return
        st.session_state["backend_post_dedupe"] = dedupe
        st.session_state.pop("risk_compute_cache", None)
        st.session_state.pop("credit_analysis_backend_dedupe", None)
    except Exception:
        st.warning("Backend sync failed; document saved locally only.")


def _is_probably_financial_document_name(path: Path) -> bool:
    name = path.name.lower()
    hints = [
        "credit",
        "term",
        "amend",
        "compliance",
        "certificate",
        "financial",
        "statement",
        "forecast",
        "projection",
        "security",
        "collateral",
        "borrowing",
        "fee",
        "covenant",
        "loan",
        "10-k",
        "10k",
        "10-q",
        "10q",
        "annual report",
        "quarterly report",
        "income statement",
        "statement of operations",
        "balance sheet",
        "cash flow",
    ]
    return any(h in name for h in hints)


def _is_probably_financial_document_content(path: Path) -> bool:
    keywords = [
        "credit agreement",
        "term sheet",
        "covenant",
        "leverage",
        "interest coverage",
        "ebitda",
        "balance sheet",
        "cash flow",
        "security agreement",
        "fee letter",
        "amendment",
        "compliance certificate",
        "borrowing base",
        "loan amount",
        "maturity date",
        "form 10-k",
        "form 10-q",
        "consolidated statements of operations",
        "statement of operations",
        "income statement",
        "balance sheet",
        "statement of financial position",
        "statement of cash flows",
        "net cash provided by operating activities",
        "total assets",
        "total liabilities",
        "net income",
        "ebitda",
    ]
    try:
        with fitz.open(path) as doc:
            sample = []
            for i, page in enumerate(doc):
                if i >= 12:
                    break
                sample.append((page.get_text("text") or "").lower())
            text = " ".join(sample)
        if any(k in text for k in keywords):
            return True
        # Fallback: SEC filing signatures often have sparse cover text on page 1.
        sec_markers = ["sec.gov", "united states securities and exchange commission", "annual report", "quarterly report"]
        financial_markers = ["consolidated", "operations", "assets", "liabilities", "cash flows", "income"]
        if any(m in text for m in sec_markers) and any(m in text for m in financial_markers):
            return True
        # Last-pass semantic type inference.
        inferred = _guess_type_from_pdf_content(path)
        if inferred in {"Financial Statements", "10-K", "10-Q", "Annual Report", "Other Financial Filing", "Credit Agreement", "Term Sheet", "Amendment", "Compliance Certificate", "Security Agreement", "Borrowing Base", "Fee Letter"}:
            return True
        return False
    except Exception:
        return True


def _is_probably_financial_document(path: Path) -> bool:
    return _is_probably_financial_document_name(path) or _is_probably_financial_document_content(path)


def _non_financial_dialog_needed_before_extract(pdf_path: Path, mode: str, manual: str) -> bool:
    """True when Extract should show the non-financial confirmation dialog instead of running."""
    manual_financial_selected = mode == "Manual" and _normalize_manual_doc_type(manual) in DOC_TYPE_CONFIG
    inferred_type_for_gate = _guess_type_from_pdf_content(pdf_path)
    inferred_financial = inferred_type_for_gate in {
        "Financial Statements",
        "10-K",
        "10-Q",
        "Annual Report",
        "Other Financial Filing",
        "Credit Agreement",
        "Term Sheet",
        "Amendment",
        "Compliance Certificate",
        "Security Agreement",
        "Borrowing Base",
        "Fee Letter",
    }
    return (not manual_financial_selected) and (not inferred_financial) and (not _is_probably_financial_document(pdf_path))


def _cache_path_for_pdf(pdf_path: Path) -> Path:
    stamp = int(pdf_path.stat().st_mtime) if pdf_path.exists() else 0
    safe_name = re.sub(r"[^a-zA-Z0-9_.-]", "_", pdf_path.name)
    return EXTRACT_CACHE_DIR / f"{safe_name}.{stamp}.{EXTRACTION_CACHE_VERSION}.json"


def _load_cached_extraction(pdf_path: Path) -> Optional[Dict[str, Any]]:
    cache_path = _cache_path_for_pdf(pdf_path)
    if not cache_path.exists():
        return None
    try:
        return json.loads(cache_path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _save_cached_extraction(pdf_path: Path, result: Dict[str, Any]) -> None:
    cache_path = _cache_path_for_pdf(pdf_path)
    cache_path.write_text(json.dumps(result), encoding="utf-8")


def _field(obj: Dict[str, Any], pdf_path: Path) -> Dict[str, Any]:
    return {
        "value": obj.get("value"),
        "confidence": obj.get("confidence"),
        "page": obj.get("page_number"),
        "link": _file_page_url(pdf_path, obj.get("page_number")),
        "snippet": obj.get("source_snippet"),
    }


def _coalesce(*vals):
    for v in vals:
        if v not in (None, ""):
            return v
    return None


def _normalize_entity_name(value: Optional[str]) -> str:
    if value is None:
        return ""
    text = str(value)
    text = "".join(ch for ch in text if ch.isprintable())
    text = re.sub(r"[\u200b-\u200f\u202a-\u202e\ufeff]", "", text)
    text = re.sub(r"\s+", " ", text.strip().lower())
    text = re.sub(r"[^a-z0-9 ]+", " ", text).strip()
    if not text:
        return ""
    legal_suffixes = {
        "inc", "incorporated", "corp", "corporation", "llc", "l l c", "ltd", "limited", "plc",
        "lp", "l p", "llp", "l l p", "co", "company", "holdings", "group",
    }
    tokens = [t for t in text.split() if t not in legal_suffixes]
    return " ".join(tokens).strip()


def _borrower_similarity(a: str, b: str) -> float:
    na = _normalize_entity_name(a)
    nb = _normalize_entity_name(b)
    if not na or not nb:
        return 0.0
    if na == nb:
        return 1.0
    return float(difflib.SequenceMatcher(None, na, nb).ratio())


def _clean_entity_candidate(value: Optional[str]) -> Optional[str]:
    if value in (None, "", "null", "None"):
        return None
    text = str(value).strip()
    cut_markers = [
        " commitment fee",
        " upfront fee",
        " interest rate",
        " covenant",
        " shall ",
        " will ",
        ". ",
        "; ",
        "\n",
    ]
    lowered = text.lower()
    cut_pos = len(text)
    for marker in cut_markers:
        idx = lowered.find(marker)
        if idx != -1:
            cut_pos = min(cut_pos, idx)
    text = text[:cut_pos].strip(" ,.;:-")
    if len(text.split()) > 12:
        text = " ".join(text.split()[:12]).strip(" ,.;:-")
    return text or None


def _extract_borrower_from_extraction(extraction: Dict[str, Any]) -> Optional[str]:
    parties = extraction.get("parties", {}) if isinstance(extraction, dict) else {}
    borrower_obj = parties.get("borrower_name", {}) if isinstance(parties, dict) else {}
    if isinstance(borrower_obj, dict):
        v = borrower_obj.get("value")
        if v not in (None, "", "null", "None"):
            return _clean_entity_candidate(v)
    info = _find_field_info(extraction, ["borrower", "name"])
    if info.get("value") not in (None, "", "null", "None"):
        return _clean_entity_candidate(info["value"])
    return None


def _borrower_from_filename(file_name: str) -> Optional[str]:
    name = re.sub(r"\.pdf$", "", file_name, flags=re.IGNORECASE)
    cleaned = re.sub(r"[_\-]+", " ", name)
    cleaned = re.sub(
        r"\b(credit|agreement|term|sheet|amendment|compliance|certificate|financial|statement|forecast|security|fee|letter|sample|valid|v\d+)\b",
        "",
        cleaned,
        flags=re.IGNORECASE,
    )
    cleaned = re.sub(r"\s+", " ", cleaned).strip(" ,.;:-")
    return _clean_entity_candidate(cleaned) if cleaned else None


def _is_same_borrower_group(names: List[str]) -> bool:
    normalized = [_normalize_entity_name(n) for n in names if n not in (None, "", "null", "None")]
    normalized = [n for n in normalized if n]
    if len(normalized) <= 1:
        return True
    uniq = sorted(set(normalized))
    base = uniq[0].strip()
    for n in uniq[1:]:
        n = n.strip()
        if base in n or n in base:
            continue
        if _borrower_similarity(base, n) >= 0.85:
            continue
        return False
    return True


def _open_page_label(page_number: Optional[int]) -> Optional[str]:
    if page_number is None:
        return None
    return f"Open page {int(page_number)}"


def _parse_page_values(value: Any) -> List[int]:
    if value is None:
        return []
    if isinstance(value, (int, float)) and not pd.isna(value):
        try:
            return [int(value)]
        except Exception:
            return []
    out: List[int] = []
    text = str(value)
    for token in re.split(r"[,\s|;/]+", text):
        token = token.strip()
        if not token:
            continue
        if token.isdigit():
            out.append(int(token))
    return out


def _all_cited_pages(df: pd.DataFrame) -> List[int]:
    if df is None or df.empty:
        return []
    pages: Set[int] = set()
    if "Page" in df.columns:
        for p in df["Page"].dropna().tolist():
            for parsed in _parse_page_values(p):
                pages.add(parsed)
    if "Pages" in df.columns:
        for raw in df["Pages"].dropna().tolist():
            for parsed in _parse_page_values(raw):
                pages.add(parsed)
    # Include per-field metadata pages so all cited table pages are available
    # in the page selector (not just the top-level Page/Pages columns).
    for col in df.columns:
        if not str(col).startswith("__meta__"):
            continue
        for raw in df[col].dropna().tolist():
            meta_obj = None
            if isinstance(raw, dict):
                meta_obj = raw
            else:
                try:
                    if isinstance(raw, str) and raw.strip().startswith("{"):
                        meta_obj = json.loads(raw)
                except Exception:
                    meta_obj = None
            if not isinstance(meta_obj, dict):
                continue
            for parsed in _parse_page_values(meta_obj.get("page")):
                pages.add(parsed)
            for parsed in _parse_page_values(meta_obj.get("pages")):
                pages.add(parsed)
    return sorted(p for p in pages if p > 0)


def _recommended_cited_page(df: pd.DataFrame, pages: List[int]) -> Optional[int]:
    if df is None or df.empty or "Page" not in df.columns:
        return pages[0] if pages else None
    work = df.copy()
    work["_page_num"] = pd.to_numeric(work["Page"], errors="coerce")
    work = work[work["_page_num"].notna()]
    if work.empty:
        return pages[0] if pages else None

    null_like = {"", "none", "null", "n/a", "na", "-", "–", "—", "nan", "undefined"}

    def _has_value(v: Any) -> bool:
        if v is None:
            return False
        try:
            if isinstance(v, float) and pd.isna(v):
                return False
        except Exception:
            pass
        s = str(v).strip().lower()
        return s not in null_like

    business_cols = [c for c in work.columns if c not in {"Sheet", "Confidence", "Page", "Open PDF Page", "Snippet", "_page_num"} and not str(c).startswith("__meta__")]
    if business_cols:
        work["_value_hits"] = work[business_cols].apply(lambda r: sum(1 for v in r.tolist() if _has_value(v)), axis=1)
    else:
        work["_value_hits"] = 0

    if "Confidence" in work.columns:
        work["_conf_num"] = pd.to_numeric(work["Confidence"], errors="coerce").fillna(-1.0)
    else:
        work["_conf_num"] = -1.0

    # Recommend the dominant table page:
    # page with the highest aggregate value density (sum of populated business fields),
    # then highest average confidence, then most cited rows.
    page_rollup = (
        work.groupby("_page_num", as_index=False)
        .agg(
            total_value_hits=("_value_hits", "sum"),
            avg_conf=("_conf_num", "mean"),
            row_count=("_page_num", "size"),
        )
        .sort_values(
            by=["total_value_hits", "avg_conf", "row_count", "_page_num"],
            ascending=[False, False, False, True],
        )
    )
    if not page_rollup.empty:
        rec = int(page_rollup.iloc[0]["_page_num"])
        return rec if rec in pages else (pages[0] if pages else None)

    ranked = work.sort_values(by=["_value_hits", "_conf_num", "_page_num"], ascending=[False, False, True])
    if ranked.empty:
        return pages[0] if pages else None
    rec = int(ranked.iloc[0]["_page_num"])
    return rec if rec in pages else (pages[0] if pages else None)


def _render_page_jump_controls(df: pd.DataFrame, key_prefix: str) -> None:
    if ("Page" not in df.columns) and ("Pages" not in df.columns):
        return
    pages = _all_cited_pages(df)
    if not pages:
        st.caption("No page citations available in this view.")
        return
    recommended_page = _recommended_cited_page(df, pages)
    extract_ctx = str(st.session_state.get("extract_token") or "")
    widget_key = f"{key_prefix}_page_select_{extract_ctx}"
    open_btn_key = f"{key_prefix}_open_page_btn_{extract_ctx}"

    c1, c2 = st.columns([4, 1], vertical_alignment="bottom")
    page_help = f"Recommended page: {recommended_page}" if recommended_page is not None else None
    # Session state only — do not combine key + index= (Streamlit warns and can desync widgets).
    init_page = int(recommended_page) if recommended_page in pages else int(pages[0])
    if widget_key not in st.session_state or st.session_state.get(widget_key) not in pages:
        st.session_state[widget_key] = init_page
    sel_page = c1.selectbox(
        "Open cited page in viewer",
        pages,
        key=widget_key,
        help=page_help,
    )
    if c2.button("Open Page", key=open_btn_key, use_container_width=True):
        st.session_state["viewer_page"] = int(sel_page)
        st.rerun()


def _table_height(df: pd.DataFrame, min_h: int = 140, max_h: int = 540, row_h: int = 36, header_h: int = 44) -> int:
    rows = len(df.index) if df is not None else 0
    return max(min_h, min(max_h, header_h + rows * row_h))


def _iter_field_nodes(node: Any, path: str = ""):
    if isinstance(node, dict):
        if "value" in node:
            yield path.lower(), node
        for k, v in node.items():
            child_path = f"{path}.{k}" if path else str(k)
            yield from _iter_field_nodes(v, child_path)
    elif isinstance(node, list):
        for i, v in enumerate(node):
            yield from _iter_field_nodes(v, f"{path}[{i}]")


def _first_non_empty(*vals):
    for v in vals:
        if v not in (None, "", "null", "None"):
            return v
    return None


def _find_field_info(extraction: Dict[str, Any], hints: List[str]) -> Dict[str, Any]:
    hints_l = [h.lower() for h in hints if str(h).strip()]
    best: Optional[Dict[str, Any]] = None
    best_score = -1
    for path, node in _iter_field_nodes(extraction):
        value = node.get("value")
        if value in (None, "", "null", "None"):
            continue
        path_l = str(path).lower()
        val_l = str(value).lower()
        snip_l = str(node.get("source_snippet") or "").lower()
        hay = f"{path_l} {val_l} {snip_l}"
        score = 0
        for h in hints_l:
            if h and h in hay:
                score += 1
        if score <= 0:
            continue
        conf = node.get("confidence")
        conf_num = float(conf) if isinstance(conf, (int, float)) else 0.0
        if score > best_score or (score == best_score and conf_num > float(best.get("confidence") or 0.0) if best else True):
            best_score = score
            best = {
                "value": value,
                "confidence": conf,
                "page": node.get("page_number"),
                "snippet": node.get("source_snippet"),
            }
    if best:
        return best
    return {"value": None, "confidence": None, "page": None, "snippet": None}


def _node_by_path(extraction: Dict[str, Any], path: str) -> Optional[Dict[str, Any]]:
    cur: Any = extraction
    for tok in path.split("."):
        if tok.endswith("]") and "[" in tok:
            base = tok[: tok.index("[")]
            idx = int(tok[tok.index("[") + 1 : -1])
            if not isinstance(cur, dict) or base not in cur or not isinstance(cur[base], list) or idx >= len(cur[base]):
                return None
            cur = cur[base][idx]
        else:
            if not isinstance(cur, dict) or tok not in cur:
                return None
            cur = cur[tok]
    return cur if isinstance(cur, dict) and "value" in cur else None


def _info_from_path(extraction: Dict[str, Any], path: str) -> Dict[str, Any]:
    node = _node_by_path(extraction, path)
    if not node:
        return {"value": None, "confidence": None, "page": None, "snippet": None}
    return {
        "value": node.get("value"),
        "confidence": node.get("confidence"),
        "page": node.get("page_number"),
        "snippet": node.get("source_snippet"),
    }


def _first_covenant_info(extraction: Dict[str, Any], terms: List[str], value_key: str = "threshold_value") -> Dict[str, Any]:
    covs = extraction.get("financial_covenants", []) if isinstance(extraction, dict) else []
    if not isinstance(covs, list):
        return {"value": None, "confidence": None, "page": None, "snippet": None}
    for cov in covs:
        if not isinstance(cov, dict):
            continue
        ctype = str((cov.get("covenant_type") or {}).get("value") or "").lower()
        if not ctype:
            continue
        if terms and not any(t.lower() in ctype for t in terms):
            continue
        tgt = cov.get(value_key) if isinstance(cov.get(value_key), dict) else None
        if not tgt:
            continue
        val = tgt.get("value")
        if val in (None, "", "null", "None"):
            continue
        return {
            "value": val,
            "confidence": tgt.get("confidence"),
            "page": tgt.get("page_number"),
            "snippet": tgt.get("source_snippet"),
        }
    return {"value": None, "confidence": None, "page": None, "snippet": None}


def _custom_sheet_field_info(extraction: Dict[str, Any], sheet: str, col: str, pdf: Optional[Path] = None) -> Optional[Dict[str, Any]]:
    direct_paths = {
        "Loan Amount": "facility_overview[0].facility_amount_total_commitment",
        "Revolver Size": "facility_overview[0].facility_amount_total_commitment",
        "Tenor": "dates_tenor.maturity_date",
        "Amortization": "dates_tenor.amortization_schedule",
        "Margin": "pricing.margin",
        "Base Rate Type": "pricing.base_rate_type",
        "Margin Grid": "pricing.spread_grid",
        "Extended Maturity Dates": "amendment_terms.revised_maturity_date",
        "Temporary Add-backs": "amendment_terms.covenant_terms_amended",
        "Covenant Holiday Terms": "amendment_terms.waiver_included",
        "Asset Type": "collateral_security.collateral_type",
        "Lien Priority": "collateral_security.lien_priority",
        "Collateral Value": "collateral_security.collateral_type",
        "Guarantor Entities": "parties.guarantors",
        "Upfront Fee %": "fees.upfront_fee_percent",
        "Commitment Fee %": "fees.commitment_fee_percent",
        "Ticking Fee": "fees.letter_of_credit_fee",
        "Margin Step-ups": "pricing.spread_grid",
        "Cross-default Clauses": "events_of_default.cross_default",
        "Material Adverse Change Clauses": "events_of_default.change_of_control",
        "Officer Sign-off": "amendment_terms.parties_consenting",
        "A/R Aging": "facility_overview[0].facility_id",
        "Ineligible Receivables": "facility_overview[0].facility_id",
        "Eligible Inventory Value": "facility_overview[0].facility_amount_total_commitment",
    }
    if col in direct_paths:
        info = _info_from_path(extraction, direct_paths[col])
        if info.get("value") not in (None, "", "null", "None"):
            return info

    if col in {"Minimum Interest Coverage", "Required Interest Coverage", "Borrower Interest Coverage"}:
        return _first_covenant_info(extraction, ["interest", "coverage", "debt service"], "threshold_value")
    if col in {"Maximum Leverage Ratio", "Updated Leverage Limits", "Required Leverage", "Borrower Leverage"}:
        return _first_covenant_info(extraction, ["leverage"], "threshold_value")
    if col in {"Fixed Charge Coverage", "Reset Coverage Ratios"}:
        return _first_covenant_info(extraction, ["fixed", "coverage", "debt service"], "threshold_value")
    if col in {"Testing Frequency"}:
        return _first_covenant_info(extraction, [], "test_frequency")
    if col in {"Equity Cure Rights"}:
        return _first_covenant_info(extraction, [], "equity_cure_allowed")
    if col in {"Headline Covenant Levels", "Preliminary Covenant Thresholds"}:
        info = _first_covenant_info(extraction, [], "threshold_value")
        if info.get("value") not in (None, "", "null", "None"):
            return info
        info = _first_covenant_info(extraction, [], "covenant_type")
        if info.get("value") not in (None, "", "null", "None"):
            return info

    regex_tokens = {
        "Minimum Interest Coverage": [["minimum", "interest", "coverage"], ["debt", "service", "coverage"]],
        "Maximum Leverage Ratio": [["maximum", "leverage"], ["total", "leverage"]],
        "Fixed Charge Coverage": [["fixed", "charge", "coverage"]],
        "Minimum Liquidity": [["minimum", "liquidity"]],
        "Updated Leverage Limits": [["updated", "leverage"], ["revised", "leverage"]],
        "Reset Coverage Ratios": [["reset", "coverage"], ["revised", "coverage"]],
        "Headline Covenant Levels": [["covenant", "compliance"], ["covenant", "level"]],
        "Upfront Fee %": [["upfront", "fee"]],
        "Commitment Fee %": [["commitment", "fee"]],
        "Margin": [["margin"], ["spread"]],
        "Loan Amount": [["loan", "amount"], ["facility", "amount"], ["commitment"]],
    }
    if pdf and col in regex_tokens:
        info = _regex_info_from_pdf(pdf, regex_tokens[col])
        if info.get("value") not in (None, "", "null", "None"):
            return info
    return None


def _to_numeric_financial(value: Any) -> Optional[float]:
    if value is None:
        return None
    try:
        if pd.api.types.is_scalar(value) and pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, (int, float)):
        if pd.isna(value):
            return None
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    mult = 1.0
    rate_divisor = 1.0
    is_negative_paren = s.startswith("(") and s.endswith(")")
    lower = s.lower()
    has_percent = "%" in lower or "percent" in lower or "percentage" in lower
    has_bps = "bps" in lower or "basis point" in lower
    if has_percent:
        rate_divisor = 100.0
    if has_bps:
        rate_divisor = 10_000.0
    if "billion" in lower or re.search(r"(?:^|[^a-z])bn(?:$|[^a-z])", lower) or re.search(r"\d\s*b\b", lower):
        mult = 1_000_000_000.0
    elif "million" in lower or re.search(r"\d\s*m(?:$|[^a-z])", lower):
        mult = 1_000_000.0
    elif "thousand" in lower or re.search(r"\d\s*k(?:$|[^a-z])", lower):
        mult = 1_000.0
    clean = re.sub(r"[^0-9.\-]", "", s.replace(",", ""))
    if is_negative_paren and clean and not clean.startswith("-"):
        clean = "-" + clean
    if clean in {"", "-", ".", "-."}:
        m = re.search(r"(-?\d[\d,]*(?:\.\d+)?)", s)
        if not m:
            return None
        clean = m.group(1).replace(",", "")
    try:
        return (float(clean) * mult) / rate_divisor
    except ValueError:
        return None


def _safe_div(n: Optional[float], d: Optional[float]) -> Optional[float]:
    if n is None or d is None:
        return None
    if pd.isna(n) or pd.isna(d):
        return None
    if d == 0:
        return None
    return n / d


def _apply_derived_fields(row: Dict[str, Any], sheet: str) -> None:
    num = {k: _to_numeric_financial(v) for k, v in row.items()}
    if sheet == "Covenant Definitions":
        row["Normalized EBITDA"] = _first_non_empty(row.get("EBITDA Add-backs"), row.get("EBITDA Exclusions"))
        row["Adjusted Interest"] = row.get("Interest Expense Definition")
        if num.get("Total Debt Definition") is not None:
            row["Net Debt Calculation"] = num.get("Total Debt Definition")
    elif sheet == "Financial Covenants":
        row["Required Covenant Thresholds"] = _first_non_empty(
            row.get("Minimum Interest Coverage"), row.get("Maximum Leverage Ratio"),
            row.get("Fixed Charge Coverage"), row.get("Minimum Liquidity")
        )
    elif sheet == "Testing Terms":
        row["Covenant Monitoring Schedule"] = _first_non_empty(row.get("Testing Frequency"), row.get("Test Dates"))
    elif sheet == "Default Triggers":
        row["Risk Escalation Triggers"] = _first_non_empty(row.get("Cross-default Clauses"), row.get("Material Adverse Change Clauses"))
    elif sheet == "Facility Overview":
        row["Debt Maturity Schedule"] = _first_non_empty(row.get("Tenor"), row.get("Amortization"))
    elif sheet == "Pricing":
        row["Effective Interest Rate"] = row.get("Margin")
    elif sheet == "Covenant Summary":
        row["Preliminary Covenant Thresholds"] = row.get("Headline Covenant Levels")
    elif sheet == "Covenant Changes":
        row["Revised Covenant Thresholds"] = _first_non_empty(row.get("Updated Leverage Limits"), row.get("Reset Coverage Ratios"))
    elif sheet == "EBITDA Adjustments":
        row["Adjusted EBITDA Logic"] = _first_non_empty(row.get("Temporary Add-backs"), row.get("Covenant Holiday Terms"))
    elif sheet == "Maturity Changes":
        row["Updated Maturity Ladder"] = row.get("Extended Maturity Dates")
    elif sheet == "Reported Financials":
        row["Reported Coverage Ratio"] = _safe_div(num.get("EBITDA"), num.get("Interest Expense"))
        row["Reported Leverage Ratio"] = _safe_div(num.get("Total Debt"), num.get("EBITDA"))
    elif sheet == "Calculated Ratios":
        cov = num.get("Borrower Interest Coverage")
        req = num.get("Required Interest Coverage")
        if cov is not None and req is not None:
            row["Covenant Cushion"] = cov - req
    elif sheet == "Certification":
        row["Legal Attestation Record"] = row.get("Officer Sign-off")
    elif sheet == "Income Statement":
        row["EBITDA Margin"] = _safe_div(num.get("EBITDA"), num.get("Revenue"))
        ebit_for_cov = num.get("EBIT")
        if ebit_for_cov is None:
            ebit_for_cov = num.get("EBITDA")
        # Normalize expense sign for ratio display: coverage uses absolute interest expense.
        interest_for_cov = abs(float(num.get("Interest Expense"))) if num.get("Interest Expense") is not None else None
        row["Interest Coverage Ratio"] = _safe_div(ebit_for_cov, interest_for_cov)
    elif sheet == "Balance Sheet":
        if num.get("Total Debt") is not None and num.get("Cash") is not None:
            row["Net Debt"] = num.get("Total Debt") - num.get("Cash")
        row["Current Ratio"] = _safe_div(num.get("Current Assets"), num.get("Current Liabilities"))
        row["Leverage Ratio"] = _safe_div(num.get("Total Debt"), num.get("EBITDA"))
    elif sheet == "Cash Flow":
        row["Free Cash Flow Coverage"] = _safe_div(num.get("Free Cash Flow"), num.get("Operating Cash Flow"))
    elif sheet == "Projected Income":
        row["Forward Interest Coverage"] = _safe_div(num.get("Projected EBITDA"), num.get("Projected Interest Expense"))
    elif sheet == "Projected Debt":
        denom = _to_numeric_financial(row.get("Projected EBITDA"))
        if denom is None:
            denom = _to_numeric_financial(row.get("EBITDA"))
        row["Projected Leverage Ratio"] = _safe_div(num.get("Future Debt Balances"), denom)
    elif sheet == "Scenario Assumptions":
        row["Stress Test Ratios"] = row.get("Stress Case")
    elif sheet == "Collateral Details":
        loan_amt = _to_numeric_financial(row.get("Loan Amount"))
        coll_val = _to_numeric_financial(row.get("Collateral Value"))
        row["Loan-to-Value (LTV)"] = _safe_div(loan_amt, coll_val)
    elif sheet == "Guarantee Info":
        row["Structural Subordination Risk"] = row.get("Guarantor Entities")
    elif sheet == "Eligible A/R":
        row["Borrowing Base Availability"] = row.get("Ineligible Receivables")
    elif sheet == "Inventory":
        row["Collateral Coverage Ratio"] = row.get("Eligible Inventory Value")
    elif sheet == "Fee Structure":
        vals = [_to_numeric_financial(row.get("Upfront Fee %")), _to_numeric_financial(row.get("Commitment Fee %")), _to_numeric_financial(row.get("Ticking Fee"))]
        row["Total Fee Income"] = sum(v for v in vals if v is not None) if any(v is not None for v in vals) else None
    elif sheet == "Pricing Grid":
        row["Dynamic Interest Cost"] = row.get("Margin Step-ups")


def _column_name_calendar_year(col: Any) -> Optional[int]:
    """Match year in headers like '2022', 'FY 2023', 'FY2024' (strict 19xx/20xx)."""
    s = str(col).strip()
    m = re.search(r"\b((?:19|20)\d{2})\b", s)
    if not m:
        return None
    try:
        y = int(m.group(1))
        return y if 1900 <= y <= 2100 else None
    except ValueError:
        return None


def _build_sheet_financial_chart_df(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["Field", "Value"])

    # Long-form tables first (Metric | Value). Column-wise scan would wrongly sum the Value column alone.
    metric_col = next((c for c in ("Metric", "Line Item", "Item", "Field", "Name") if c in df.columns), None)
    value_col = next((c for c in ("Value", "Calculated Value", "Amount", "FY Total") if c in df.columns), None)
    year_cols_all = [c for c in df.columns if _column_name_calendar_year(c) is not None]
    # Common 10-K shape: Metric + fiscal year columns, while "Value"/"Amount" exists but is empty — do not
    # take the long-form branch or we return no rows and skip the year-column fallback below.
    skip_long_form = False
    if metric_col and value_col:
        probe = [_to_numeric_financial(v) for v in df[value_col].tolist()]
        if year_cols_all and not any(v is not None for v in probe):
            skip_long_form = True

    if metric_col and value_col and not skip_long_form:
        rows_long: List[Dict[str, Any]] = []
        for _, r in df.iterrows():
            label = r.get(metric_col)
            if label is None or (isinstance(label, float) and pd.isna(label)):
                continue
            fv = _to_numeric_financial(r.get(value_col))
            if fv is None:
                continue
            rows_long.append({"Field": str(label).strip(), "Value": float(fv)})
        if rows_long:
            dedup_l: Dict[str, float] = {}
            for r in rows_long:
                if r["Field"] not in dedup_l:
                    dedup_l[r["Field"]] = r["Value"]
            return pd.DataFrame([{"Field": k, "Value": v} for k, v in dedup_l.items()], columns=["Field", "Value"])

    rows: List[Dict[str, Any]] = []
    ignore_cols = {
        "Sheet",
        "Metric",
        "Line Item",
        "Description",
        "Label",
        "Confidence",
        "Page",
        "Open PDF Page",
        "Snippet",
        "Currency",
        "Selected Year",
        "Fiscal Period",
        "Analysis Mode",
        "Detected Unit",
        "Available Years",
        "Unit Multiplier",
    }
    # Keep visualization deterministic: include numeric columns except metadata / label columns.
    # Use _to_numeric_financial so formatted values ($, commas, parentheses) still plot.
    for col in df.columns:
        if col in ignore_cols:
            continue
        col_s = str(col).strip()
        col_l = col_s.lower()
        if col_s.startswith("__meta__") or "__meta__" in col_s:
            continue
        if col_l in {"unit multiplier", "selected year"}:
            continue
        if any(
            k in col_l
            for k in {
                "confidence",
                "page",
                "snippet",
                "selected year",
                "fiscal period",
                "analysis mode",
                "available years",
                "detected unit",
                "detected type",
                "document type",
            }
        ):
            continue
        # Skip derived % / ratio columns by name (raw dollar lines stay plottable).
        if any(k in col_l for k in {"ratio", "margin", "coverage", "percent", "%", "bps"}):
            continue
        vals = [_to_numeric_financial(v) for v in df[col].tolist()]
        vals = [v for v in vals if v is not None]
        if not vals:
            continue
        if re.search(r"\b(19|20)\d{2}\b", col_l):
            continue
        rows.append({"Field": col_s, "Value": float(sum(vals))})

    # Wide statements with only calendar-year columns (no Revenue/EBITDA named columns).
    if not rows:
        year_cols = [c for c in df.columns if _column_name_calendar_year(c) is not None]
        for yc in year_cols:
            vals = [_to_numeric_financial(v) for v in df[yc].tolist()]
            vals = [v for v in vals if v is not None]
            if vals:
                ynum = _column_name_calendar_year(yc)
                label = str(ynum) if ynum is not None else str(yc).strip()
                rows.append({"Field": label, "Value": float(sum(vals))})
    if rows:
        dedup: Dict[str, float] = {}
        for r in rows:
            if r["Field"] not in dedup:
                dedup[r["Field"]] = r["Value"]
        return pd.DataFrame([{"Field": k, "Value": v} for k, v in dedup.items()], columns=["Field", "Value"])
    return pd.DataFrame(columns=["Field", "Value"])


def _coerce_object_cols_for_chart(df: pd.DataFrame) -> pd.DataFrame:
    """Coerce object/string cells (formatted $, commas) so column-wise chart math sees numbers."""
    if df is None or df.empty:
        return df
    skip = {
        "Sheet",
        "Metric",
        "Line Item",
        "Description",
        "Label",
        "Currency",
        "Snippet",
        "Open PDF Page",
        "Fiscal Period",
        "Analysis Mode",
        "Detected Unit",
        "Available Years",
        "Detected Type",
    }
    out = df.copy()
    for c in out.columns:
        if c in skip or str(c).startswith("__meta__"):
            continue
        s = out[c]
        if pd.api.types.is_numeric_dtype(s):
            continue
        parsed = s.apply(lambda v: _to_numeric_financial(v))
        if parsed.notna().any():
            out[c] = parsed
    return out


def _auto_scale_financial_values(chart_df: pd.DataFrame) -> tuple[pd.DataFrame, str, float]:
    if chart_df.empty or "Value" not in chart_df.columns:
        return chart_df.copy(), "", 1.0
    max_abs = float(chart_df["Value"].abs().max())
    if max_abs >= 1_000_000_000:
        unit, factor = "B", 1_000_000_000.0
    elif max_abs >= 1_000_000:
        unit, factor = "M", 1_000_000.0
    elif max_abs >= 1_000:
        unit, factor = "K", 1_000.0
    else:
        unit, factor = "", 1.0
    scaled = chart_df.copy()
    scaled["Value"] = scaled["Value"] / factor
    return scaled, unit, factor


def _scale_factor_from_label(scale_label: str) -> tuple[str, float]:
    s = str(scale_label or "").strip().lower()
    if s in {"k", "thousand", "thousands"}:
        return "K", 1_000.0
    if s in {"m", "million", "millions"}:
        return "M", 1_000_000.0
    if s in {"b", "bn", "billion", "billions"}:
        return "B", 1_000_000_000.0
    if s in {"t", "trillion", "trillions"}:
        return "T", 1_000_000_000_000.0
    return "", 1.0


def _apply_preferred_scale(chart_df: pd.DataFrame, scale_label: str) -> tuple[pd.DataFrame, str, float]:
    if chart_df.empty or "Value" not in chart_df.columns:
        return chart_df.copy(), "", 1.0
    unit, factor = _scale_factor_from_label(scale_label)
    if factor <= 1.0:
        return _auto_scale_financial_values(chart_df)
    # If values are already expressed in the detected display unit (e.g., filing says
    # "in millions" and extracted revenue is 402,836), do not divide again.
    try:
        max_abs = float(chart_df["Value"].abs().max())
    except Exception:
        max_abs = 0.0
    if max_abs > 0 and max_abs < factor:
        scaled = chart_df.copy()
        return scaled, unit, 1.0
    scaled = chart_df.copy()
    scaled["Value"] = scaled["Value"] / factor
    return scaled, unit, factor


def _render_scaled_xy_chart(chart_df: pd.DataFrame, chart_type: str, unit_label: str) -> None:
    if chart_df.empty:
        st.info("No numeric financial values available for graph.")
        return
    mark = "bar" if chart_type == "Bar" else "line"
    axis_cfg = {"labelAngle": -35, "labelLimit": 1000, "labelPadding": 8, "labelOverlap": False}
    spec = {
        "mark": {"type": mark, "point": True},
        "encoding": {
            "x": {"field": "Field", "type": "nominal", "axis": axis_cfg},
            "y": {"field": "Value", "type": "quantitative", "title": f"Value ({unit_label})" if unit_label else "Value"},
            "tooltip": [{"field": "Field", "type": "nominal"}, {"field": "Value", "type": "quantitative", "format": ",.0f"}],
        },
    }
    if chart_type == "Pie":
        spec = {
            "mark": {"type": "arc"},
            "encoding": {
                "theta": {"field": "Value", "type": "quantitative"},
                "color": {"field": "Field", "type": "nominal"},
                "tooltip": [{"field": "Field", "type": "nominal"}, {"field": "Value", "type": "quantitative", "format": ",.0f"}],
            },
        }
    st.vega_lite_chart(chart_df[["Field", "Value"]], spec, use_container_width=True)


def _field_palette(n: int) -> List[str]:
    palette = [
        "#2563eb", "#14b8a6", "#f59e0b", "#334155", "#0ea5e9", "#ef4444",
        "#22c55e", "#f97316", "#0d9488", "#84cc16", "#0f766e", "#64748b",
    ]
    out = []
    for i in range(n):
        out.append(palette[i % len(palette)])
    return out


def _graph_blue_palette(n: int) -> List[str]:
    """Blue-forward categorical colors for Data Visualization (high contrast on white)."""
    blues = [
        "#172554", "#1e3a8a", "#1e40af", "#1d4ed8", "#2563eb", "#3b82f6",
        "#60a5fa", "#93c5fd", "#0c4a6e", "#0369a1", "#0ea5e9", "#38bdf8",
    ]
    return [blues[i % len(blues)] for i in range(n)]


def _render_aura_legend(df: pd.DataFrame, colors: List[str]) -> None:
    if df.empty or "Field" not in df.columns:
        return
    uniq = list(dict.fromkeys([str(v) for v in df["Field"].dropna().tolist()]))
    chips = []
    for i, field_name in enumerate(uniq):
        color = colors[i % len(colors)]
        chips.append(
            f"<span style='display:inline-flex;align-items:center;gap:6px;padding:4px 10px;border:1px solid #d8e1ec;border-radius:999px;background:#fff;margin-right:8px;margin-top:6px;'>"
            f"<span style='width:10px;height:10px;background:{color};border-radius:2px;display:inline-block;'></span>{field_name}"
            f"</span>"
        )
    st.markdown("".join(chips), unsafe_allow_html=True)


def _render_aura_grid_chart(chart_df: pd.DataFrame, chart_type: str, unit_label: str) -> None:
    if chart_df.empty:
        st.info("No numeric financial values available for graph.")
        return
    plot = chart_df.copy()
    plot = plot.dropna(subset=["Field", "Value"]).copy()
    plot["Field"] = plot["Field"].astype(str)
    plot = plot.groupby("Field", as_index=False)["Value"].sum()
    plot = plot.sort_values("Value", ascending=False, kind="stable")
    fields = plot["Field"].tolist()
    if not fields:
        st.info("No numeric financial values available for graph.")
        return
    colors = _graph_blue_palette(len(fields))
    chart_data = plot.set_index("Field")[["Value"]]
    if chart_type == "Pie":
        spec = {
            "mark": {"type": "arc", "outerRadius": 120},
            "encoding": {
                "theta": {"field": "Value", "type": "quantitative"},
                "color": {
                    "field": "Field",
                    "type": "nominal",
                    "scale": {"domain": fields, "range": colors},
                    "legend": None,
                },
                "tooltip": [{"field": "Field"}, {"field": "Value", "type": "quantitative", "format": ",.0f"}],
            },
        }
        try:
            st.vega_lite_chart(plot[["Field", "Value"]], spec, use_container_width=True)
        except Exception:
            st.dataframe(plot[["Field", "Value"]], width="stretch", hide_index=True)
            st.caption("Pie view unavailable; showing table.")
    else:
        y_title = f"Value ({unit_label})" if unit_label else "Value"
        try:
            alt.data_transformers.disable_max_rows()
        except Exception:
            pass
        color_scale = alt.Scale(domain=fields, range=colors)
        try:
            base = alt.Chart(plot).encode(
                x=alt.X(
                    "Field",
                    sort=fields,
                    title="",
                    axis=alt.Axis(labelAngle=-35, labelLimit=1000, labelPadding=8),
                ),
                y=alt.Y("Value", title=y_title),
                color=alt.Color("Field", scale=color_scale, legend=None),
                tooltip=[
                    alt.Tooltip("Field", title="Field"),
                    alt.Tooltip("Value", format=",.0f", title="Value"),
                ],
            )
            if chart_type == "Line":
                chart = base.mark_line(point=True, strokeWidth=2).properties(height=380)
            else:
                chart = base.mark_bar().properties(height=380)
            st.altair_chart(chart, use_container_width=True)
        except Exception:
            if chart_type == "Line":
                st.line_chart(chart_data, height=380)
            else:
                st.bar_chart(chart_data, height=380)
            st.caption("Using built-in chart (Altair unavailable in this session).")
    if unit_label:
        st.caption(f"Scale: {unit_label}")
    _render_aura_legend(plot, colors)


def _facility_summary_rows(extraction: Dict[str, Any], pdf: Path) -> pd.DataFrame:
    rows = []
    borrower = _field(extraction["parties"]["borrower_name"], pdf)
    base_rate = _field(extraction["pricing"]["base_rate_type"], pdf)
    spread = _field(extraction["pricing"]["margin"], pdf)
    secured = _field(extraction["collateral_security"]["secured_or_unsecured"], pdf)
    maturity = _field(extraction["dates_tenor"]["maturity_date"], pdf)

    for fac in extraction.get("facility_overview", []):
        ftype = _field(fac.get("facility_type", {}), pdf)
        famt = _field(fac.get("facility_amount_total_commitment", {}), pdf)
        cur = _field(fac.get("currency", {}), pdf)
        close = _field(fac.get("origination_date", {}), pdf)
        agent = _field(fac.get("agent_bank", {}), pdf)
        gov = _field(fac.get("governing_law", {}), pdf)
        stat = _field(fac.get("status", {}), pdf)

        rows.append(
            {
                "Borrower": borrower["value"],
                "Facility Type": ftype["value"],
                "Facility Amount": famt["value"],
                "Currency": cur["value"],
                "Closing Date": close["value"],
                "Maturity Date": maturity["value"],
                "Base Rate": base_rate["value"],
                "Spread (bps)": spread["value"],
                "Secured?": secured["value"],
                "Agent Bank": agent["value"],
                "Governing Law": gov["value"],
                "Status": stat["value"],
                "Confidence": _coalesce(ftype["confidence"], famt["confidence"], base_rate["confidence"], borrower["confidence"]),
                "Page": _coalesce(ftype["page"], famt["page"], base_rate["page"], borrower["page"]),
                "Open PDF Page": _open_page_label(_coalesce(ftype["page"], famt["page"], base_rate["page"], borrower["page"])),
                "Snippet": _coalesce(ftype["snippet"], famt["snippet"], base_rate["snippet"], borrower["snippet"]),
            }
        )
    return pd.DataFrame(rows if rows else [{col: None for col in TEMPLATE_COLUMNS["Facility Summary"]}], columns=TEMPLATE_COLUMNS["Facility Summary"])


def _covenant_rows(extraction: Dict[str, Any], pdf: Path) -> pd.DataFrame:
    rows = []
    borrower = _field(extraction["parties"]["borrower_name"], pdf)

    for cov in extraction.get("financial_covenants", []):
        ctype = _field(cov.get("covenant_type", {}), pdf)
        thr = _field(cov.get("threshold_value", {}), pdf)
        freq = _field(cov.get("test_frequency", {}), pdf)
        eq = _field(cov.get("equity_cure_allowed", {}), pdf)
        src = _field(cov.get("calculation_definition_page", {}), pdf)

        rows.append(
            {
                "Borrower": borrower["value"],
                "Covenant Type": ctype["value"],
                "Threshold": thr["value"],
                "Actual": None,
                "Headroom": None,
                "Test Frequency": freq["value"],
                "Test Date": None,
                "Breach?": None,
                "Equity Cure Allowed?": eq["value"],
                "Source Section": None,
                "Confidence": _coalesce(ctype["confidence"], thr["confidence"], freq["confidence"], eq["confidence"]),
                "Page": _coalesce(ctype["page"], thr["page"], freq["page"], src["page"]),
                "Open PDF Page": _open_page_label(_coalesce(ctype["page"], thr["page"], freq["page"], src["page"])),
                "Snippet": _coalesce(ctype["snippet"], thr["snippet"], freq["snippet"], src["snippet"]),
            }
        )

    return pd.DataFrame(rows if rows else [{col: None for col in TEMPLATE_COLUMNS["Covenant Monitoring"]}], columns=TEMPLATE_COLUMNS["Covenant Monitoring"])


def _fees_rows(extraction: Dict[str, Any], pdf: Path) -> pd.DataFrame:
    borrower = _field(extraction["parties"]["borrower_name"], pdf)
    fac = extraction.get("facility_overview", [{}])[0]
    facility = _field(fac.get("facility_type", {}), pdf)
    cm = _field(extraction["fees"].get("commitment_fee_percent", {}), pdf)
    ag = _field(extraction["fees"].get("agency_fee", {}), pdf)
    up = _field(extraction["fees"].get("upfront_fee_percent", {}), pdf)
    pp = _field(extraction["fees"].get("prepayment_premium", {}), pdf)
    ds = _field(extraction["pricing"].get("default_interest_rate", {}), pdf)

    row = {
        "Borrower": borrower["value"],
        "Facility": facility["value"],
        "Commitment Fee (bps)": cm["value"],
        "Agency Fee": ag["value"],
        "Upfront Fee": up["value"],
        "Prepayment Penalty": pp["value"],
        "Default Spread": ds["value"],
        "Confidence": _coalesce(cm["confidence"], ag["confidence"], up["confidence"], pp["confidence"], ds["confidence"]),
        "Page": _coalesce(cm["page"], ag["page"], up["page"], pp["page"], ds["page"]),
        "Open PDF Page": _open_page_label(_coalesce(cm["page"], ag["page"], up["page"], pp["page"], ds["page"])),
        "Snippet": _coalesce(cm["snippet"], ag["snippet"], up["snippet"], pp["snippet"], ds["snippet"]),
    }
    return pd.DataFrame([row], columns=TEMPLATE_COLUMNS["Fees & Pricing"])


def _collateral_rows(extraction: Dict[str, Any], pdf: Path) -> pd.DataFrame:
    borrower = _field(extraction["parties"]["borrower_name"], pdf)
    sec = _field(extraction["collateral_security"].get("secured_or_unsecured", {}), pdf)
    lien = _field(extraction["collateral_security"].get("lien_priority", {}), pdf)
    ctype = _field(extraction["collateral_security"].get("collateral_type", {}), pdf)
    guar = _field(extraction["parties"].get("guarantors", {}), pdf)

    row = {
        "Borrower": borrower["value"],
        "Secured?": sec["value"],
        "Lien Type": lien["value"],
        "Collateral Type": ctype["value"],
        "Guarantors": guar["value"],
        "UCC Filed?": None,
        "Collateral Coverage": None,
        "Confidence": _coalesce(sec["confidence"], lien["confidence"], ctype["confidence"], guar["confidence"]),
        "Page": _coalesce(sec["page"], lien["page"], ctype["page"], guar["page"]),
        "Open PDF Page": _open_page_label(_coalesce(sec["page"], lien["page"], ctype["page"], guar["page"])),
        "Snippet": _coalesce(sec["snippet"], lien["snippet"], ctype["snippet"], guar["snippet"]),
    }
    return pd.DataFrame([row], columns=TEMPLATE_COLUMNS["Collateral Tracking"])


def _amendment_rows(extraction: Dict[str, Any], pdf: Path) -> pd.DataFrame:
    borrower = _field(extraction["parties"]["borrower_name"], pdf)
    amend = extraction.get("amendment_terms", {})

    num = _field(amend.get("amendment_number", {}), pdf)
    eff = _field(amend.get("amendment_effective_date", {}), pdf)
    pm = _field(amend.get("prior_maturity_date", {}), pdf)
    rm = _field(amend.get("revised_maturity_date", {}), pdf)
    pc = _field(amend.get("prior_commitment_amount", {}), pdf)
    rc = _field(amend.get("revised_commitment_amount", {}), pdf)
    sp = _field(amend.get("spread_change", {}), pdf)
    cf = _field(amend.get("consent_fee", {}), pdf)
    wa = _field(amend.get("waiver_included", {}), pdf)
    ca = _field(amend.get("covenant_terms_amended", {}), pdf)
    ps = _field(amend.get("parties_consenting", {}), pdf)
    ac = _field(amend.get("agent_consent_required", {}), pdf)

    row = {
        "Borrower": borrower["value"],
        "Amendment Number": num["value"],
        "Amendment Effective Date": eff["value"],
        "Prior Maturity Date": pm["value"],
        "Revised Maturity Date": rm["value"],
        "Prior Commitment": pc["value"],
        "Revised Commitment": rc["value"],
        "Spread Change": sp["value"],
        "Consent Fee": cf["value"],
        "Waiver Included?": wa["value"],
        "Covenant Terms Amended?": ca["value"],
        "Parties Consenting": ps["value"],
        "Agent Consent Required?": ac["value"],
        "Confidence": _coalesce(num["confidence"], eff["confidence"], rm["confidence"], rc["confidence"], cf["confidence"]),
        "Page": _coalesce(num["page"], eff["page"], rm["page"], rc["page"], cf["page"]),
        "Open PDF Page": _open_page_label(_coalesce(num["page"], eff["page"], rm["page"], rc["page"], cf["page"])),
        "Snippet": _coalesce(num["snippet"], eff["snippet"], rm["snippet"], rc["snippet"], cf["snippet"]),
    }
    return pd.DataFrame([row], columns=TEMPLATE_COLUMNS["Amendment Tracker"])


@st.cache_data(show_spinner=False, ttl=600)
def _build_sheet_cached(extraction: Dict[str, Any], sheet: str, pdf_path_str: str) -> pd.DataFrame:
    return _build_sheet(extraction, sheet, Path(pdf_path_str))


def _build_sheet(extraction: Dict[str, Any], sheet: str, pdf: Path) -> pd.DataFrame:
    columns = TEMPLATE_COLUMNS.get(sheet, [])
    # Financial-services filings should not show COGS; replace with Total Expense and Employee Compensation.
    by_file = dict(st.session_state.get("last_applied_doc_type_by_file", {}))
    pdf_key = str(pdf.resolve())
    doc_type = by_file.get(pdf_key, st.session_state.get("last_applied_doc_type"))
    fin_services_types = {"Annual Report", "Form 10-K", "10-K", "Investment Manager", "Asset Manager", "Financial Institution"}
    if sheet == "Income Statement" and doc_type in fin_services_types:
        columns = [c for c in columns if c != "COGS"]
        for extra in [
            "Total Expense",
            "Employee Compensation",
            "Non-Interest Expense",
            "Operating Income",
            "Operating Margin",
            "Income Before Taxes",
            "Tax Expense",
            "Effective Tax Rate",
            "EPS (Basic)",
            "EPS (Diluted)",
        ]:
            if extra not in columns:
                columns.insert(1, extra)
    if not columns:
        return pd.DataFrame()
    row: Dict[str, Any] = {c: None for c in columns}
    confs: List[float] = []
    pages: List[int] = []
    snippets: List[str] = []
    hints = SHEET_FIELD_HINTS.get(sheet, {})

    # Deterministic extraction for financial statements (strict labels, latest-year preference, no-guess policy).
    if sheet in {"Income Statement", "Balance Sheet", "Cash Flow"}:
        strict_row, strict_errors = _strict_financial_sheet_row_cached(sheet, pdf)
        metric_meta = strict_row.get("_metric_meta", {}) if isinstance(strict_row, dict) else {}
        for col in columns:
            if col in META_COLUMNS or col in DERIVED_COLUMNS:
                continue
            if col in strict_row:
                row[col] = strict_row[col]
        for mcol in YEAR_META_COLUMNS:
            if mcol in strict_row:
                row[mcol] = strict_row[mcol]
        _apply_derived_fields(row, sheet)
        has_any_value = any(row.get(c) not in (None, "", "null", "None") for c in columns if c not in META_COLUMNS and c not in DERIVED_COLUMNS)
        row["Confidence"] = 0.95 if has_any_value else None
        metric_pages = [
            int(v.get("page")) for v in metric_meta.values()
            if isinstance(v, dict) and isinstance(v.get("page"), int)
        ]
        metric_pages_unique = sorted(set(metric_pages))
        if has_any_value and metric_pages:
            page_freq: Dict[int, int] = {}
            for p in metric_pages:
                page_freq[p] = page_freq.get(p, 0) + 1
            # Prefer dominant cited page; tie-break to earlier page.
            row["Page"] = sorted(page_freq.items(), key=lambda it: (-it[1], it[0]))[0][0]
            row["Pages"] = ", ".join(str(p) for p in metric_pages_unique)
        else:
            row["Page"] = None
            row["Pages"] = None
        row["Open PDF Page"] = _open_page_label(row["Page"])
        selected_year = strict_row.get("Selected Year")
        detected_unit = strict_row.get("Detected Unit")
        if strict_errors:
            row["Snippet"] = strict_errors[0]
        elif selected_year:
            match_counts = {"exact": 0, "fuzzy": 0, "regex": 0, "derived": 0}
            for meta in metric_meta.values():
                if isinstance(meta, dict):
                    mt = str(meta.get("match_type") or "").lower()
                    if mt in match_counts:
                        match_counts[mt] += 1
            match_info = f" | match: exact={match_counts['exact']}, fuzzy={match_counts['fuzzy']}, regex={match_counts['regex']}, derived={match_counts['derived']}"
            unit_part = f" | unit: {detected_unit}" if detected_unit else ""
            row["Snippet"] = f"selected_year: {int(selected_year)}{unit_part}{match_info}"
        else:
            row["Snippet"] = None
        # Attach per-field confidence metadata for UI badges.
        for col in columns:
            meta = metric_meta.get(col)
            if isinstance(meta, dict):
                row[f"__meta__{col}"] = {
                    "match_type": meta.get("match_type") or meta.get("source"),
                    "confidence_score": meta.get("confidence_score"),
                }
        ordered_cols = list(columns)
        if "Pages" in row and "Pages" not in ordered_cols:
            ordered_cols.append("Pages")
        for mcol in YEAR_META_COLUMNS:
            if mcol in row and mcol not in ordered_cols:
                ordered_cols.append(mcol)
        # Keep meta columns at the end so they can be stripped for display.
        for col in list(row.keys()):
            if col.startswith("__meta__") and col not in ordered_cols:
                ordered_cols.append(col)
        return pd.DataFrame([row], columns=ordered_cols)

    for col in columns:
        if col in META_COLUMNS or col in DERIVED_COLUMNS:
            continue
        info = _custom_sheet_field_info(extraction, sheet, col, pdf)
        if not info:
            info = _find_field_info(extraction, hints.get(col, re.findall(r"[A-Za-z0-9]+", col)))
        if (not info) or info.get("value") in (None, "", "null", "None"):
            token_groups = _fallback_token_groups_for_column(col)
            if token_groups:
                pdf_info = _regex_info_from_pdf(pdf, token_groups)
                if pdf_info.get("value") not in (None, "", "null", "None"):
                    info = pdf_info
        if not info:
            info = {"value": None, "confidence": None, "page": None, "snippet": None}
        row[col] = info["value"]
        if info.get("confidence") is not None:
            confs.append(float(info["confidence"]))
        if info.get("page") is not None:
            pages.append(int(info["page"]))
        if info.get("snippet"):
            snippets.append(str(info["snippet"]))

    _apply_derived_fields(row, sheet)

    if confs:
        row["Confidence"] = round(sum(confs) / len(confs), 3)
    else:
        row["Confidence"] = None
    if pages:
        page_freq: Dict[int, int] = {}
        for p in pages:
            page_freq[int(p)] = page_freq.get(int(p), 0) + 1
        row["Page"] = sorted(page_freq.items(), key=lambda it: (-it[1], it[0]))[0][0]
        row["Pages"] = ", ".join(str(p) for p in sorted(set(int(x) for x in pages)))
    else:
        row["Page"] = None
        row["Pages"] = None
    row["Open PDF Page"] = _open_page_label(row["Page"])
    row["Snippet"] = snippets[0] if snippets else None
    output_cols = list(columns)
    if "Pages" in row and "Pages" not in output_cols:
        output_cols.append("Pages")
    return pd.DataFrame([row], columns=output_cols)


def _to_excel_bytes(sheet_map: Dict[str, pd.DataFrame]) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheet_map.items():
            df.to_excel(writer, index=False, sheet_name=name[:31])
    return output.getvalue()


def _to_credit_risk_package_excel_bytes(
    tables: Dict[str, pd.DataFrame],
    model: Dict[str, Any],
) -> bytes:
    from openpyxl.styles import Font, PatternFill

    output = io.BytesIO()
    table_df = model.get("table", pd.DataFrame()).copy()
    memo = str(model.get("memo", "") or "")

    fin_df = tables.get("financial_actuals", pd.DataFrame()).copy()
    income_df = pd.DataFrame()
    balance_df = pd.DataFrame()
    cashflow_df = pd.DataFrame()
    if not fin_df.empty:
        r0 = fin_df.iloc[0].to_dict()
        income_df = pd.DataFrame(
            [
                {
                    "Revenue": r0.get("revenue"),
                    "EBITDA": r0.get("ebitda"),
                    "Interest Expense": r0.get("interest_expense"),
                    "Net Income": r0.get("net_income"),
                }
            ]
        )
        balance_df = pd.DataFrame(
            [
                {
                    "Total Assets": r0.get("total_assets"),
                    "Total Liabilities": r0.get("total_liabilities"),
                    "Equity": r0.get("equity"),
                    "Current Assets": r0.get("current_assets"),
                    "Current Liabilities": r0.get("current_liabilities"),
                    "Total Debt": r0.get("total_debt"),
                    "Cash": r0.get("cash"),
                }
            ]
        )
        cashflow_df = pd.DataFrame(
            [
                {
                    "Operating Cash Flow": r0.get("operating_cash_flow"),
                    "CapEx": r0.get("capital_expenditures"),
                    "Free Cash Flow": r0.get("free_cash_flow"),
                }
            ]
        )

    sheet_map: Dict[str, pd.DataFrame] = {
        "Income Statement": income_df,
        "Balance Sheet": balance_df,
        "Cash Flow Statement": cashflow_df,
        "Derived Metrics": pd.DataFrame(
            model.get("ratios", {}) if isinstance(model.get("ratios", {}), dict) else {}
            , index=[0]
        ),
        "Credit Risk Scoring Table": table_df.copy(),
    }

    memo_rows = []
    memo_sections = [
        ("Company Overview", "Auto generated from extracted dataset and underwriting inputs."),
        ("Financial Highlights", ", ".join((table_df[table_df["Risk"] == "Low"]["Metric"].head(5).tolist() if "Risk" in table_df.columns else [])) or "N/A"),
        ("Risk Analysis", f"Final Risk Score: {model.get('final_score', 'N/A')} | Grade: {model.get('grade', 'N/A')}"),
        ("Industry Adjustments", json.dumps(model.get("policy_limit_breakdown", {}), default=str)),
        ("Final Recommended Limit", f"{model.get('policy_approved_limit', 'N/A')}"),
        ("Confidence Score", f"{model.get('data_completeness_pct', 'N/A')}%"),
        ("Narrative", memo),
    ]
    for sec, txt in memo_sections:
        memo_rows.append({"Section": sec, "Details": txt})
    sheet_map["Credit Memo Summary"] = pd.DataFrame(memo_rows)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        for name, df in sheet_map.items():
            out_df = df.copy().where(pd.notna(df), None)
            out_df.to_excel(writer, index=False, sheet_name=name[:31])
            ws = writer.book[name[:31]]
            ws.freeze_panes = "A2"
            for cell in ws[1]:
                cell.font = Font(bold=True)
            if name == "Credit Risk Scoring Table" and "Risk" in out_df.columns:
                risk_col_idx = list(out_df.columns).index("Risk") + 1
                for r in range(2, ws.max_row + 1):
                    v = str(ws.cell(row=r, column=risk_col_idx).value or "").strip().lower()
                    fill = None
                    if v == "low":
                        fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                    elif v == "medium":
                        fill = PatternFill(start_color="FFF4E5", end_color="FFF4E5", fill_type="solid")
                    elif v == "high":
                        fill = PatternFill(start_color="FDECEA", end_color="FDECEA", fill_type="solid")
                    elif v in {"incomplete", "provisional"}:
                        fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                    if fill:
                        ws.cell(row=r, column=risk_col_idx).fill = fill

    return output.getvalue()


@st.cache_data(show_spinner=False)
def _pdf_page_png(pdf_path_str: str, page_number: int, zoom: float = 1.5) -> Optional[bytes]:
    pdf_path = Path(pdf_path_str)
    if not pdf_path.exists():
        return None
    doc = fitz.open(pdf_path)
    idx = max(0, min(page_number - 1, doc.page_count - 1))
    page = doc[idx]
    mat = fitz.Matrix(zoom, zoom)
    pix = page.get_pixmap(matrix=mat, alpha=False)
    return pix.tobytes("png")


@st.cache_data(show_spinner=False, ttl=3600)
def _pdf_text_lines(
    pdf_path_str: str, max_pages: int = 20, cache_version: str = EXTRACTION_CACHE_VERSION
) -> List[Dict[str, Any]]:
    _ = cache_version  # cache key only; bump EXTRACTION_CACHE_VERSION to invalidate after parser/year fixes
    pdf_path = Path(pdf_path_str)
    if not pdf_path.exists():
        return []
    lines: List[Dict[str, Any]] = []
    try:
        # Prefer table extraction when available (handles dense concatenation better than raw text).
        try:
            import pdfplumber  # type: ignore

            with pdfplumber.open(str(pdf_path)) as doc:
                for i, page in enumerate(doc.pages):
                    if i >= max_pages:
                        break
                    tables = page.extract_tables() or []
                    for tbl in tables:
                        for row in tbl:
                            if not row:
                                continue
                            cells = [str(c).strip() for c in row if c is not None and str(c).strip()]
                            if not cells:
                                continue
                            # Join table row to preserve column order.
                            lines.append({"page": i + 1, "text": " | ".join(cells)})
        except Exception:
            # If pdfplumber isn't available or fails, fall back to raw text below.
            pass
        with fitz.open(pdf_path) as doc:
            for i, page in enumerate(doc):
                if i >= max_pages:
                    break
                text = page.get_text("text") or ""
                for ln in text.splitlines():
                    t = ln.strip()
                    if t:
                        lines.append({"page": i + 1, "text": t})
    except Exception:
        return []
    return lines


def _looks_like_search_results_pdf(pdf: Path) -> bool:
    lines = _pdf_text_lines(str(pdf), max_pages=3)
    if not lines:
        return False
    blob = " ".join(str(x.get("text") or "").lower() for x in lines)
    markers = [
        "search results",
        "refine search results",
        "show columns",
        "filing category",
        "company name, ticker, cik",
        "10,000 search results",
    ]
    hits = sum(1 for m in markers if m in blob)
    return hits >= 2


def _strict_scan_page_limit(pdf: Path, cap: int = 220) -> int:
    try:
        with fitz.open(str(pdf)) as d:
            return max(1, min(int(d.page_count), cap))
    except Exception:
        return min(cap, 40)


STRICT_STATEMENT_HEADERS: Dict[str, List[str]] = {
    "Income Statement": [
        "consolidated statements of operations",
        "consolidated statement of operations",
        "consolidated statements of income",
        "consolidated statement of income",
        "condensed consolidated statements of operations",
        "condensed consolidated statement of operations",
        "condensed consolidated interim statements of operations",
        "condensed consolidated interim statement of operations",
        "condensed consolidated interim statements of operations and comprehensive loss",
        "condensed consolidated interim statements of operations and comprehensive income",
        "statements of operations and comprehensive loss",
        "statements of operations and comprehensive income",
        "income statement",
        "statements of earnings",
        "statement of operations",
    ],
    "Balance Sheet": [
        "balance sheet",
        "consolidated balance sheet",
        "statement of financial position",
        "consolidated balance sheets",
        "condensed consolidated balance sheets",
        "condensed consolidated interim balance sheets",
    ],
    "Cash Flow": [
        "statement of cash flows",
        "statement of cash flow",
        "consolidated statements of cash flows",
        "consolidated statement of cash flows",
        "condensed consolidated statements of cash flows",
        "condensed consolidated interim statements of cash flows",
        "cash flow statement",
    ],
    "Debt Schedule": [
        "debt",
        "borrowings",
        "notes payable",
    ],
}

STRICT_LABEL_PRIORITY: Dict[str, List[str]] = {
    "Revenue": ["total revenue", "net revenue", "revenue", "total sales", "net sales"],
    "COGS": [
        "cost of goods sold",
        "cost of revenue",
        "cost of revenues",
        "cost of sales",
        "cost of products sold",
        "cost of product revenue",
    ],
    "Total Expense": [
        "total expense",
        "total expenses",
        "total operating expenses",
        "operating expenses",
    ],
    "Employee Compensation": [
        "employee compensation",
        "compensation and benefits",
        "salaries and benefits",
        "personnel expense",
        "compensation expense",
    ],
    "Non-Interest Expense": [
        "non-interest expense",
        "noninterest expense",
        "non interest expense",
    ],
    "Net Income": [
        "net income",
        "net loss",
        "net income (loss)",
        "net earnings",
        "profit (loss)",
        "net income attributable to",
        "net income attributable to owners",
        "net income attributable to shareholders",
        "net income for the year",
        "net income attributable to company",
    ],
    "EBIT": [
        "operating income",
        "operating income (loss)",
        "income from operations",
        "loss from operations",
        "operating profit",
        "operating loss",
    ],
    "Operating Income": [
        "operating income",
        "income from operations",
        "operating profit",
    ],
    "EBITDA": [
        "ebitda",
        "adjusted ebitda",
        "operating profit before d&a",
        "operating income before depreciation",
        "operating income before depreciation and amortization",
    ],
    "Income Before Taxes": [
        "income before taxes",
        "income before tax",
        "pretax income",
        "income before income taxes",
    ],
    "Tax Expense": [
        "income tax expense",
        "provision for income taxes",
        "tax expense",
    ],
    "EPS (Basic)": [
        "basic earnings per share",
        "eps basic",
        "earnings per share basic",
    ],
    "EPS (Diluted)": [
        "diluted earnings per share",
        "eps diluted",
        "earnings per share diluted",
    ],
    "Interest Expense": [
        "interest expense",
        "finance cost",
        "finance costs",
        "borrowing cost",
        "borrowing costs",
        "interest and finance charges",
    ],
    "Total Assets": ["total assets"],
    "Total Liabilities": ["total liabilities"],
    "Shareholders' Equity": ["total stockholders equity", "total equity", "shareholders equity"],
    "Current Assets": ["total current assets", "current assets"],
    "Current Liabilities": ["total current liabilities", "current liabilities"],
    "Inventory": ["inventory", "inventories", "total inventory"],
    "Accounts Receivable": [
        "accounts receivable",
        "trade receivables",
        "receivables, net",
        "receivables net",
        "trade and other receivables",
    ],
    "Cash": ["cash and cash equivalents", "cash"],
    "Accounts Receivable": [
        "accounts receivable",
        "trade receivables",
        "receivables, net",
        "receivables net",
        "trade and other receivables",
    ],
    "Total Debt": ["total debt", "long-term debt", "short-term borrowings", "current portion of long-term debt", "notes payable"],
    "Short-term Debt": ["short-term borrowings", "short term borrowings", "short-term debt", "short term debt"],
    "Current Portion LT Debt": [
        "current portion of long-term debt",
        "current maturities of long-term debt",
        "current portion long term debt",
        "current portion of long-term obligations",
        "current portion of long term obligations",
    ],
    "Long-term Debt": [
        "long-term debt",
        "long term debt",
        "long-term debt non-current",
        "long term debt non current",
        "long-term debt, net",
        "long term debt net",
        "non-current borrowings",
    ],
    "Operating Cash Flow": [
        "operating cash flow",
        "net cash provided by operating activities",
        "net cash used in operating activities",
        "net cash provided by used in operating activities",
        "net cash provided by (used in) operating activities",
        "net cash from operating activities",
        "cash flow from operations",
        "cash from operations",
        "net cash from operations",
        "cash from operating activities",
    ],
    "CapEx": [
        "capital expenditures",
        "capital expenditure",
        "purchase of property and equipment",
        "purchases of property and equipment",
        "purchase of property and equipment",
        "purchase of property plant and equipment",
        "purchases of property plant and equipment",
        "purchase of property, plant and equipment",
        "purchases of property, plant and equipment",
        "acquisition of property and equipment",
        "acquisitions of property and equipment",
        "acquisition of pp&e",
        "additions to property and equipment",
        "additions to property, plant and equipment",
    ],
    "Principal Repayment": [
        "principal repayment",
        "principal repayments",
        "debt repayment",
        "debt repayments",
        "repayment of debt",
        "repayments of long-term debt",
        "net debt issued/(repaid)",
        "net debt issued repaid",
        "net borrowings (repayments)",
    ],
}

# Enterprise financial field map (GAAP + IFRS variants).
FINANCIAL_FIELD_MAP: Dict[str, List[str]] = {
    "revenue": [
        "revenue",
        "total revenue",
        "net sales",
        "sales",
        "revenues",
        "turnover",
        "gross revenue",
    ],
    "operating_income": [
        "operating income",
        "income from operations",
        "operating profit",
        "profit from operations",
        "ebit",
        "operating profit before d a",
        "operating income before depreciation",
        "operating income before depreciation and amortization",
    ],
    "cogs": [
        "cost of goods sold",
        "cost of revenue",
        "cost of revenues",
        "cost of sales",
        "cost of products sold",
        "cost of product revenue",
    ],
    "total_expense": [
        "total expense",
        "total expenses",
        "total operating expenses",
        "operating expenses",
    ],
    "employee_compensation": [
        "employee compensation",
        "compensation and benefits",
        "salaries and benefits",
        "personnel expense",
        "compensation expense",
    ],
    "ebitda": [
        "ebitda",
        "adjusted ebitda",
        "operating profit before d a",
        "operating income before depreciation",
        "operating income before depreciation and amortization",
    ],
    "net_income": [
        "net income",
        "net earnings",
        "net income attributable to",
        "net income attributable to shareholders",
        "net income attributable to owners",
        "net income for the year",
        "profit for the year",
        "profit attributable to owners",
        "profit after tax",
    ],
    "interest_expense": [
        "interest expense",
        "interest and other expense",
        "interest expense, net",
        "finance cost",
        "finance costs",
        "borrowing cost",
        "borrowing costs",
    ],
    "current_assets": ["total current assets", "current assets"],
    "current_liabilities": ["total current liabilities", "current liabilities"],
    "accounts_receivable": [
        "accounts receivable",
        "trade receivables",
        "receivables net",
        "receivables, net",
        "trade and other receivables",
        "a r",
    ],
    "operating_cash_flow": [
        "operating cash flow",
        "net cash provided by operating activities",
        "net cash used in operating activities",
        "net cash provided by used in operating activities",
        "net cash provided by (used in) operating activities",
        "cash generated from operations",
        "net cash from operating activities",
    ],
    "capex": [
        "capital expenditures",
        "capital expenditure",
        "additions to property and equipment",
        "additions to property, plant and equipment",
        "purchase of property and equipment",
        "purchases of property and equipment",
        "purchase of property plant and equipment",
        "purchases of property plant and equipment",
        "purchase of property, plant and equipment",
        "purchases of property, plant and equipment",
        "acquisition of property plant and equipment",
        "acquisition of property and equipment",
        "acquisitions of property and equipment",
    ],
    "total_assets": ["total assets"],
    "total_liabilities": ["total liabilities"],
    "equity": ["total stockholders equity", "total equity", "shareholders equity", "stockholders equity"],
    "short_term_debt": ["short-term borrowings", "short term borrowings", "short-term debt", "short term debt"],
    "current_portion_long_term_debt": [
        "current portion of long-term debt",
        "current maturities of long-term debt",
        "current portion long term debt",
    ],
    "long_term_debt": ["long-term debt", "long term debt"],
    "cash": ["cash and cash equivalents", "cash"],
    "depreciation_amortization": ["depreciation", "amortization", "depreciation and amortization"],
    "principal_repayment": [
        "principal repayment",
        "principal repayments",
        "debt repayment",
        "debt repayments",
        "repayment of debt",
        "repayments of long-term debt",
        "net debt issued repaid",
        "net debt issued repaid",
    ],
}

REGEX_FIELD_PATTERNS: Dict[str, str] = {
    "revenue": r"^(net\s+)?(total\s+)?(sales|revenue|turnover)(\s+\w+){0,2}$",
    "net_income": r"^(net\s+)?(income|earnings|profit)(\s+\w+){0,3}$",
    "interest_expense": r"^(interest|finance)\s+(expense|cost)(\s+\w+){0,2}$",
    "operating_cash_flow": r"^net\s+cash.*operating(\s+\w+){0,3}$",
}

TABLE_CLASSIFICATION_KEYWORDS: Dict[str, List[str]] = {
    "Income Statement": ["revenue", "net income", "operating income", "earnings per share"],
    "Balance Sheet": ["total assets", "total liabilities", "equity", "current assets"],
    "Cash Flow": ["net cash provided", "operating activities", "investing activities", "financing activities"],
    "Notes Section": ["notes to financial statements", "interest expense", "finance cost", "borrowing cost"],
}


def _norm_label(text: str) -> str:
    clean = re.sub(r"[^a-z0-9 ]+", " ", (text or "").lower())
    return re.sub(r"\s+", " ", clean).strip()


def _normalize_label_for_match(text: str) -> str:
    s = str(text or "")
    # Remove hidden control characters and normalize whitespace/casing.
    s = "".join(ch for ch in s if ch.isprintable())
    s = re.sub(r"[\u200b-\u200f\u202a-\u202e\ufeff]", "", s)
    s = re.sub(r"\s+", " ", s.strip().lower())
    s = _norm_label(s)
    # Strip entity qualifiers and attribution tails to stabilize matching.
    s = re.sub(r"attributable to non controlling interest.*$", "", s).strip()
    s = re.sub(r"attributable to.*$", "attributable to", s).strip()
    s = re.sub(r"\bfor the year\b.*$", "for the year", s).strip()
    s = re.sub(r"\binc\b\.?$", "", s).strip()
    s = re.sub(r"\bcorp\b\.?$", "", s).strip()
    s = re.sub(r"\bltd\b\.?$", "", s).strip()
    return s[:80]


def _levenshtein_distance(a: str, b: str, max_dist: int = 2) -> int:
    if a == b:
        return 0
    if not a:
        return len(b)
    if not b:
        return len(a)
    if abs(len(a) - len(b)) > max_dist:
        return max_dist + 1
    prev = list(range(len(b) + 1))
    for i, ca in enumerate(a, 1):
        curr = [i]
        row_min = curr[0]
        for j, cb in enumerate(b, 1):
            ins = curr[j - 1] + 1
            dele = prev[j] + 1
            rep = prev[j - 1] + (0 if ca == cb else 1)
            curr.append(min(ins, dele, rep))
            row_min = min(row_min, curr[-1])
        if row_min > max_dist:
            return max_dist + 1
        prev = curr
    return prev[-1]


def _match_financial_label(label_text: str, aliases: List[str], allow_regex: bool = True) -> Tuple[bool, str, int]:
    # Step 1 semantic mapping layer: normalize alternative labels to canonical keys.
    FINANCIAL_LABEL_NORMALIZATION: Dict[str, str] = {
        "operating profit before d a": "ebitda",
        "operating income before depreciation": "ebitda",
        "operating income before depreciation and amortization": "ebitda",
        "ebitda": "ebitda",
        "finance costs": "interest expense",
        "finance cost": "interest expense",
        "borrowing costs": "interest expense",
        "borrowing cost": "interest expense",
        "interest expense": "interest expense",
        "trade receivables": "accounts receivable",
        "receivables net": "accounts receivable",
        "receivables net of allowance": "accounts receivable",
        "accounts receivable": "accounts receivable",
        # Keep debt labels distinct to prevent cross-mapping between short-term,
        # current portion, and long-term debt rows.
        "short term borrowings": "short term borrowings",
        "short term debt": "short term debt",
        "current portion of long term obligations": "current portion of long term obligations",
        "current portion of long term debt": "current portion of long term debt",
        "long term debt non current": "long term debt non current",
        "long term debt": "long term debt",
    }

    def _canonical_label(raw: str) -> str:
        normalized = _normalize_label_for_match(raw)
        return FINANCIAL_LABEL_NORMALIZATION.get(normalized, normalized)

    label_n = _canonical_label(label_text)
    if not label_n or len(label_n.split()) > 10:
        return False, "none", 0
    alias_n = [_canonical_label(x) for x in aliases]
    if any(label_n == a for a in alias_n):
        return True, "exact", 100
    # In strict extraction calls we pass allow_regex=False; keep those exact-only
    # to avoid false row captures (e.g., note references/per-share rows).
    if ENABLE_FUZZY_MATCHING and allow_regex:
        for a in alias_n:
            if _levenshtein_distance(label_n, a, max_dist=2) <= 2:
                return True, "fuzzy", 85
    if ENABLE_REGEX_FALLBACK and allow_regex:
        for key, pattern in REGEX_FIELD_PATTERNS.items():
            if any(_normalize_label_for_match(x) in key.replace("_", " ") or key.replace("_", " ") in _normalize_label_for_match(x) for x in aliases):
                if re.fullmatch(pattern, label_n, flags=re.IGNORECASE):
                    return True, "regex", 70
    return False, "none", 0


def _extract_years_from_line(text: str) -> List[int]:
    years = re.findall(r"\b(19\d{2}|20\d{2})\b", text or "")
    return [int(y) for y in years]


def _candidate_fiscal_years_from_lines(lines: List[Dict[str, Any]]) -> List[int]:
    # Strict guard: do not allow far-future noise (e.g., maturity 2040) to override
    # statement-year locking. Allow +1 calendar year for current FY columns.
    upper_year = _max_extractable_fiscal_year()
    out: List[int] = []
    strict_context_rx = re.compile(
        r"(fiscal year ended|for the year ended|years ended|year ended)",
        re.IGNORECASE,
    )
    loose_context_rx = re.compile(r"(december|fiscal|as of|as at)", re.IGNORECASE)
    for ln in lines:
        raw = str(ln.get("text") or "")
        ys = [y for y in _extract_years_from_line(raw) if 1990 <= int(y) <= upper_year]
        if not ys:
            continue
        if 2 <= len(ys) <= 4:
            diffs = [abs(ys[i] - ys[i + 1]) for i in range(len(ys) - 1)]
            alpha_chars = len(re.findall(r"[A-Za-z]", raw))
            if all(d <= 2 for d in diffs) and (
                strict_context_rx.search(raw) or (len(ys) >= 3 and (alpha_chars <= 38 or loose_context_rx.search(raw)))
            ):
                out.extend(ys)
                continue
        if len(ys) == 1 and strict_context_rx.search(raw):
            out.extend(ys)
    return sorted(set(out))


def _extract_fiscal_years_detected(pdf: Path, max_pages: int = 6) -> List[int]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    # Prefer statement-anchor windows to avoid non-statement years (e.g., debt maturities).
    statement_order = ["Income Statement", "Cash Flow", "Balance Sheet"]
    years_by_statement: Dict[str, List[int]] = {}
    for st in statement_order:
        ap = _sheet_anchor_page(st, pdf)
        if ap is None:
            continue
        windowed = [ln for ln in lines if isinstance(ln.get("page"), int) and (ap - 2) <= int(ln.get("page")) <= (ap + 4)]
        # Use strict context-aware year extraction to avoid footnote/date noise.
        yrs = _candidate_fiscal_years_from_lines(windowed)
        years_by_statement[st] = yrs

    if years_by_statement:
        income_years = set(years_by_statement.get("Income Statement", []))
        cash_years = set(years_by_statement.get("Cash Flow", []))
        if income_years and cash_years:
            common_years = sorted(income_years.intersection(cash_years))
            if common_years:
                return common_years
        if income_years:
            return sorted(income_years)
        if cash_years:
            return sorted(cash_years)
        counts: Dict[int, int] = {}
        for ys in years_by_statement.values():
            for y in ys:
                counts[y] = counts.get(y, 0) + 1
        common_years = sorted([y for y, c in counts.items() if c >= 2])
        if common_years:
            return common_years
        for st in statement_order:
            ys = years_by_statement.get(st, [])
            if ys:
                return ys

    lines = lines[: max_pages * 120]
    years = _candidate_fiscal_years_from_lines(lines)
    if years:
        return years
    # Fallback if structured year headers are not detected.
    raw_years: List[int] = []
    for ln in lines:
        raw_years.extend(_extract_years_from_line(str(ln.get("text") or "")))
    return sorted({y for y in raw_years if 1990 <= y <= _max_extractable_fiscal_year()})


def _extract_company_name_from_pdf(pdf: Path, max_pages: int = 2) -> Optional[str]:
    lines = _pdf_text_lines(str(pdf), max_pages=max_pages)
    if not lines:
        return None
    for ln in lines[:200]:
        txt = str(ln.get("text") or "").strip()
        if not txt:
            continue
        low = txt.lower()
        if "inc" in low or "corporation" in low or "company" in low or "corp" in low:
            cleaned = re.sub(r"\s+", " ", txt).strip(" ,.;:-")
            if 3 <= len(cleaned) <= 120:
                return cleaned
    return None


def _line_label_prefix(text: str) -> str:
    if not text:
        return ""
    # Treat text before first numeric cell as the row label.
    m = re.search(r"[\(\-\u2013\u2014]?\$?\d", text)
    if not m:
        return text.strip()
    return text[: m.start()].strip(" :.-")


def _extract_numeric_tokens(text: str) -> List[str]:
    if not text:
        return []
    # Keep accounting signs and preserve negatives like (2,242).
    # Also support space-grouped thousands (e.g., "100 000" from OCR/PDF extraction).
    grouped_num = r"\d{1,3}(?:[,\s]\d{3})+(?:\.\d+)?"
    plain_num = r"\d+(?:\.\d+)?"
    rx = re.compile(
        rf"\(\s*\$?(?:{grouped_num}|{plain_num})\s*\)|-?\$?(?:{grouped_num}|{plain_num})|[—–]"
    )
    vals = [m.group(0).strip() for m in rx.finditer(text)]
    out: List[str] = []
    for v in vals:
        if v in {"—", "–"}:
            out.append("")
        else:
            out.append(v)
    return out


def _classify_statement_blocks(pdf: Path) -> Dict[str, Optional[int]]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    by_page: Dict[int, str] = {}
    for ln in lines:
        pg = ln.get("page")
        if isinstance(pg, int):
            by_page[pg] = by_page.get(pg, "") + " " + str(ln.get("text") or "").lower()
    out: Dict[str, Optional[int]] = {"Income Statement": None, "Balance Sheet": None, "Cash Flow": None, "Notes Section": None}
    for block_name, kws in TABLE_CLASSIFICATION_KEYWORDS.items():
        best_page: Optional[int] = None
        best_score = -1
        for pg, txt in by_page.items():
            score = sum(txt.count(k) for k in kws)
            if score > best_score:
                best_score = score
                best_page = pg
        out[block_name] = best_page if best_score > 0 else None
    return out


def _sheet_header_present(sheet: str, pdf: Path) -> bool:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    joined = " ".join([str(x.get("text") or "").lower() for x in lines])
    keys = STRICT_STATEMENT_HEADERS.get(sheet, [])
    return any(k in joined for k in keys)


def _sheet_anchor_page(sheet: str, pdf: Path) -> Optional[int]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    if sheet in {"Income Statement", "Balance Sheet", "Cash Flow"}:
        # Financial statements are usually in the main filing body, not in tail exhibits.
        lines = [ln for ln in lines if isinstance(ln.get("page"), int) and int(ln.get("page")) <= 140]
    keys = [k.lower() for k in STRICT_STATEMENT_HEADERS.get(sheet, [])]
    signal_terms: List[str] = []
    if sheet == "Income Statement":
        signal_terms = ["revenue", "net income", "operating income", "income from operations", "cost of revenues"]
    elif sheet == "Balance Sheet":
        signal_terms = ["total assets", "total liabilities", "stockholders", "equity", "current assets", "current liabilities"]
    elif sheet == "Cash Flow":
        signal_terms = ["net cash provided by operating activities", "operating activities", "investing activities", "financing activities", "capital expenditures"]

    by_page: Dict[int, str] = {}
    for ln in lines:
        pg = ln.get("page")
        if isinstance(pg, int):
            by_page[pg] = by_page.get(pg, "") + " " + str(ln.get("text") or "").lower()

    candidate_scores: Dict[int, int] = {}
    strict_table_scores: Dict[int, int] = {}
    for pg, txt in by_page.items():
        header_hits = sum(1 for k in keys if k in txt)
        year_hits = len(re.findall(r"\b(19|20)\d{2}\b", txt))
        signal_hits = sum(txt.count(t) for t in signal_terms)

        # Strict table-page scoring: require statement header + dense table cues.
        if header_hits > 0 and year_hits >= 2:
            strict_markers: List[str] = []
            if sheet == "Income Statement":
                strict_markers = [
                    "total revenue",
                    "cost of goods sold",
                    "gross profit",
                    "operating expenses",
                    "loss from operations",
                    "income from operations",
                    "interest expense",
                    "net loss",
                    "net income",
                ]
            elif sheet == "Balance Sheet":
                strict_markers = [
                    "total assets",
                    "total liabilities",
                    "stockholders",
                    "current assets",
                    "current liabilities",
                ]
            elif sheet == "Cash Flow":
                strict_markers = [
                    "cash flows from operating activities",
                    "cash flows from investing activities",
                    "cash flows from financing activities",
                    "net cash",
                ]
            marker_hits = sum(1 for m in strict_markers if m in txt)
            if marker_hits >= 2:
                strict_table_scores[pg] = (header_hits * 6) + (marker_hits * 5) + min(year_hits, 6)

        if sheet == "Income Statement":
            if "percentage of revenue" in txt or "expressed as a percentage" in txt:
                continue
            # Strictly require statement-title context to avoid picking narrative note pages.
            has_income_title = (
                ("consolidated statements" in txt and ("income" in txt or "operations" in txt or "earnings" in txt))
                or ("income statement" in txt)
                or ("statements of operations" in txt)
                or ("statement of operations" in txt)
            )
            if not has_income_title:
                continue
            # Prefer actual statement pages that contain fiscal-year headers.
            if year_hits < 2 and "year ended" not in txt and "years ended" not in txt:
                continue
            # Penalize note/disclosure sections that often mention revenue repeatedly.
            if (
                "product revenue and service and other revenue" in txt
                or "revenue recognition" in txt
                or "notes to consolidated financial statements" in txt
            ):
                continue
            # Require statement-row evidence to exclude note pages that merely reference
            # "statements of operations" in explanatory prose.
            income_row_markers = [
                "total revenue",
                "cost of goods sold",
                "cost of revenue",
                "gross profit",
                "operating expenses",
                "income from operations",
                "loss from operations",
                "operating income",
                "net income",
                "net loss",
            ]
            marker_hits = sum(1 for t in income_row_markers if t in txt)
            if marker_hits < 3:
                continue
        elif sheet == "Balance Sheet":
            # Require core balance sheet structure; notes often mention one metric only.
            bs_markers = [
                "total assets",
                "total liabilities",
                "total equity",
                "stockholders",
                "current assets",
                "current liabilities",
            ]
            marker_hits = sum(1 for t in bs_markers if t in txt)
            if marker_hits < 3:
                continue
        elif sheet == "Cash Flow":
            # Require cash flow section structure, not isolated line-item mentions.
            cf_markers = [
                "operating activities",
                "investing activities",
                "financing activities",
                "net cash",
            ]
            marker_hits = sum(1 for t in cf_markers if t in txt)
            if marker_hits < 2:
                continue
        if header_hits <= 0:
            continue
        # Weight signal terms heavily; header phrases alone are not sufficient.
        score = (header_hits * 5) + (signal_hits * 3)
        if score > 0:
            candidate_scores[pg] = score

    if strict_table_scores:
        return sorted(strict_table_scores.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
    if candidate_scores:
        return sorted(candidate_scores.items(), key=lambda kv: (-kv[1], kv[0]))[0][0]
    classified = _classify_statement_blocks(pdf)
    if sheet in {"Income Statement", "Balance Sheet", "Cash Flow"} and classified.get(sheet) is not None:
        return classified.get(sheet)
    return None


def _selected_fiscal_year_and_order(
    pdf: Path,
    page_min: Optional[int] = None,
    page_max: Optional[int] = None,
    analysis_year: str = ANALYSIS_YEAR,
    allow_fallback_to_latest: bool = True,
) -> Tuple[Optional[int], List[int], List[int]]:
    # Global scan for all available fiscal years (source of truth for locking).
    all_lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    upper_year = _max_extractable_fiscal_year()
    statement_order = ["Income Statement", "Cash Flow", "Balance Sheet"]
    years_by_statement: Dict[str, List[int]] = {}
    for st in statement_order:
        ap = _sheet_anchor_page(st, pdf)
        if ap is None:
            continue
        windowed = [ln for ln in all_lines if isinstance(ln.get("page"), int) and (ap - 2) <= int(ln.get("page")) <= (ap + 4)]
        years_by_statement[st] = _candidate_fiscal_years_from_lines(windowed)
    global_years: List[int] = []
    union_years: List[int] = []
    if years_by_statement:
        income_years = set(years_by_statement.get("Income Statement", []))
        cash_years = set(years_by_statement.get("Cash Flow", []))
        union_years = sorted({int(y) for ys in years_by_statement.values() for y in ys if ys})
        if income_years and cash_years:
            common_years = sorted(income_years.intersection(cash_years))
            if common_years:
                global_years = common_years
        if not global_years and income_years:
            global_years = sorted(income_years)
        if not global_years and cash_years:
            global_years = sorted(cash_years)
        if not global_years:
            counts: Dict[int, int] = {}
            for ys in years_by_statement.values():
                for y in ys:
                    counts[y] = counts.get(y, 0) + 1
            common_years = sorted([y for y, c in counts.items() if c >= 2])
            if common_years:
                global_years = common_years
        if not global_years:
            for st in statement_order:
                ys = years_by_statement.get(st, [])
                if ys:
                    global_years = ys
                    break
        if union_years and global_years and max(union_years) > max(global_years):
            global_years = sorted(set(global_years) | set(union_years))
    if not global_years:
        global_years = sorted(
            {
                y
                for ln in all_lines[:1200]
                for y in _extract_years_from_line(str(ln.get("text") or ""))
                if 1990 <= int(y) <= upper_year
            }
        )
    if not union_years and global_years:
        union_years = list(global_years)
    # Remove obvious outlier years (e.g., historical references) so statement mapping
    # is anchored to the current filing period.
    if global_years:
        max_y = max(global_years)
        recent_floor = max(1990, max_y - 6)
        pruned = sorted([y for y in global_years if y >= recent_floor])
        if pruned:
            global_years = pruned

    lines = list(all_lines)
    if page_min is not None or page_max is not None:
        lines = [
            ln for ln in lines
            if isinstance(ln.get("page"), int)
            and (page_min is None or ln["page"] >= page_min)
            and (page_max is None or ln["page"] <= page_max)
        ]
    local_years: List[int] = []
    year_counts: Dict[int, int] = {}
    year_order: List[int] = []
    multi_year_sequences: List[List[int]] = []
    for line in lines:
        raw = str(line.get("text") or "")
        ys = _extract_years_from_line(raw)
        ys = [y for y in ys if 1990 <= y <= upper_year]
        if ys:
            local_years.extend(ys)
            for y in ys:
                year_counts[y] = year_counts.get(y, 0) + 1
            if len(ys) >= 2:
                alpha_chars = len(re.findall(r"[A-Za-z]", raw))
                # Year header rows are usually compact and mostly numeric.
                if alpha_chars <= 28 or len(ys) >= 3 or ("december" in raw.lower() and "year ended" in raw.lower()):
                    # Ignore year lists that include old outliers relative to current filing.
                    if global_years:
                        max_y = max(global_years)
                        recent_floor = max(1990, max_y - 6)
                        ys = [y for y in ys if y >= recent_floor]
                    if len(ys) >= 2:
                        multi_year_sequences.append(ys)
                    if not year_order:
                        year_order = ys
    if not global_years:
        return None, [], []
    # Prefer explicit multi-year column headers from statement pages over global counts.
    # Only trust 3+ year headers for stable column-position mapping.
    if multi_year_sequences:
        seq_scores: List[Tuple[int, List[int]]] = []
        for seq in multi_year_sequences:
            # Bias toward realistic statement year headers (descending year columns).
            uniq_seq = list(dict.fromkeys([int(y) for y in seq if 1990 <= int(y) <= upper_year]))
            if len(uniq_seq) < 3 or len(uniq_seq) > 4:
                continue
            diffs = [abs(uniq_seq[i] - uniq_seq[i + 1]) for i in range(len(uniq_seq) - 1)]
            if any(d > 2 for d in diffs):
                continue
            asc = uniq_seq == sorted(uniq_seq)
            desc = uniq_seq == sorted(uniq_seq, reverse=True)
            # Reject mixed-order year lists (e.g., 2024,2025,2023,2022) since they
            # are typically non-header artifacts and break column alignment.
            if not (asc or desc):
                continue
            score = len(uniq_seq)
            if desc:
                score += 2
            if asc:
                score += 1
            seq_scores.append((score, uniq_seq))
        if seq_scores:
            seq_scores.sort(key=lambda x: (x[0], max(x[1]), min(x[1])), reverse=True)
            year_order = seq_scores[0][1]
        else:
            # Fall back to sorted global years to avoid mixed-order artifacts.
            year_order = sorted(set(int(y) for y in global_years))

    selected_year: Optional[int]
    if str(analysis_year).lower() != "latest":
        try:
            y = int(str(analysis_year))
            selected_year = y if y in (union_years or global_years) else None
        except Exception:
            selected_year = None
    else:
        # Non-negotiable latest-year rule: strict numeric max across detected years.
        selected_year = max(union_years or global_years)
    if selected_year is None and allow_fallback_to_latest:
        selected_year = max(union_years or global_years)
    if year_order and selected_year not in year_order:
        # Keep strict behavior: if selected year is not present in statement-local year headers,
        # no positional fallback is allowed. Caller will mark metrics incomplete.
        return selected_year, year_order, global_years
    if year_order:
        max_order_y = max(year_order)
        recent_floor = max(1990, max_order_y - 6)
        filtered_order = [y for y in year_order if y >= recent_floor]
        if filtered_order:
            year_order = filtered_order
    if not year_order:
        # Fallback must be stable and monotonic. Some OCR layouts surface mixed local
        # year sightings (e.g., 2024,2025,2023,2022) that are not true column order.
        local_order = list(dict.fromkeys(int(y) for y in (local_years or [])))
        local_asc = local_order == sorted(local_order)
        local_desc = local_order == sorted(local_order, reverse=True)
        if local_order and (local_asc or local_desc):
            year_order = local_order
        else:
            year_order = sorted(set(int(y) for y in global_years))
    # Final guard: keep only the recent contiguous filing years to prevent old
    # historical references (e.g., 2004) from contaminating column alignment.
    if year_order:
        max_y = max(int(y) for y in year_order)
        year_order = [int(y) for y in year_order if int(y) >= (max_y - 3)]
        year_order = list(dict.fromkeys(year_order))
        # Final safety: never return mixed-order year vectors.
        y_asc = year_order == sorted(year_order)
        y_desc = year_order == sorted(year_order, reverse=True)
        if not (y_asc or y_desc):
            year_order = sorted(set(int(y) for y in global_years if int(y) >= (max_y - 3)))
    return selected_year, year_order, global_years


@st.cache_data(show_spinner=False, ttl=1800)
def _statement_unit_factor_cached(pdf_path_str: str, cache_version: str = EXTRACTION_CACHE_VERSION) -> float:
    _ = cache_version
    pdf = Path(pdf_path_str)
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    unit_rx = re.compile(
        r"\b(in|amounts? (?:are )?in|dollars? in|\$ in)\s+(millions?|thousands?|billions?|trillions?)\b",
        re.IGNORECASE,
    )
    usd_000_rx = re.compile(r"\b(?:usd|us\$|\$)?\s*0{3}s\b", re.IGNORECASE)
    usd_m_rx = re.compile(r"\b(?:usd|us\$|\$)?\s*(millions?|m)\b", re.IGNORECASE)
    usd_b_rx = re.compile(r"\b(?:usd|us\$|\$)?\s*(billions?|bn)\b", re.IGNORECASE)

    def _factor_from_text(text: str) -> Optional[float]:
        m = unit_rx.search(text or "")
        if not m:
            # Support USD 000s / $ millions / $ billions shorthand.
            if usd_000_rx.search(text or ""):
                return 1_000.0
            if usd_b_rx.search(text or ""):
                return 1_000_000_000.0
            if usd_m_rx.search(text or ""):
                return 1_000_000.0
            return None
        u = m.group(2).lower()
        if "million" in u:
            return 1_000_000.0
        if "thousand" in u:
            return 1_000.0
        if "billion" in u:
            return 1_000_000_000.0
        if "trillion" in u:
            return 1_000_000_000_000.0
        return None

    # Prefer unit declarations around statement anchor pages (reduces false global matches).
    anchors = []
    for st_name in ("Income Statement", "Balance Sheet", "Cash Flow"):
        ap = _sheet_anchor_page(st_name, pdf)
        if isinstance(ap, int):
            anchors.append(ap)
    if anchors:
        anchor_min = max(1, min(anchors) - 2)
        anchor_max = max(anchors) + 4
        scoped = [
            str(x.get("text") or "").lower()
            for x in lines
            if isinstance(x.get("page"), int) and anchor_min <= int(x.get("page")) <= anchor_max
        ]
        scoped_text = re.sub(r"\s+", " ", " ".join(scoped))
        scoped_factor = _factor_from_text(scoped_text)
        if scoped_factor is not None:
            return scoped_factor

    # Fallback to whole document.
    joined = re.sub(r"\s+", " ", " ".join([str(x.get("text") or "").lower() for x in lines]))
    full_factor = _factor_from_text(joined)
    return full_factor if full_factor is not None else 1.0


def _statement_unit_factor(pdf: Path) -> float:
    try:
        return float(_statement_unit_factor_cached(str(pdf.resolve())))
    except Exception:
        return float(_statement_unit_factor_cached(str(pdf)))


def _analysis_year_config() -> Dict[str, Any]:
    mode_raw = str(ANALYSIS_YEAR_MODE or "latest_available").strip().lower()
    if mode_raw in {"latest", "latest_available"}:
        mode = "latest_available"
    elif mode_raw in {"specific", "specific_year"}:
        mode = "specific_year"
    elif mode_raw in {"t12m", "trailing_12_months", "trailing12m"}:
        mode = "t12m"
    else:
        mode = "latest_available"
    specific_year: Optional[int] = None
    if ANALYSIS_SPECIFIC_YEAR:
        try:
            specific_year = int(str(ANALYSIS_SPECIFIC_YEAR).strip())
        except Exception:
            specific_year = None
    # For non-risk modules, always default to latest available year.
    active_mod = str(st.session_state.get("active_module") or "").lower()
    if active_mod and active_mod != "risk":
        return {"mode": "latest_available", "specific_year": None}
    try:
        smode = st.session_state.get("analysis_year_mode")
        if isinstance(smode, str) and smode.strip():
            sm = smode.strip().lower()
            if sm in {"latest", "latest_available"}:
                mode = "latest_available"
            elif sm in {"specific", "specific_year"}:
                mode = "specific_year"
            elif sm in {"t12m", "trailing_12_months", "trailing12m"}:
                mode = "t12m"
        syear = st.session_state.get("analysis_specific_year")
        if syear not in (None, ""):
            specific_year = int(syear)
    except Exception:
        pass
    return {"mode": mode, "specific_year": specific_year}


def _scale_label_from_factor(factor: float) -> str:
    if factor >= 1_000_000_000_000.0:
        return "Trillions"
    if factor >= 1_000_000_000.0:
        return "Billions"
    if factor >= 1_000_000.0:
        return "Millions"
    if factor >= 1_000.0:
        return "Thousands"
    return "Units"


def _detect_scale_from_combined_df(df: Optional[pd.DataFrame]) -> str:
    if df is None or df.empty:
        return "Units"
    # Priority 1: explicit parsed unit from strict extraction.
    if "Detected Unit" in df.columns:
        unit_vals = [str(x).strip().lower() for x in df["Detected Unit"].dropna().tolist() if str(x).strip()]
        if any("trillion" in u for u in unit_vals):
            return "Trillions"
        if any("billion" in u for u in unit_vals):
            return "Billions"
        if any("million" in u for u in unit_vals):
            return "Millions"
        if any("thousand" in u for u in unit_vals):
            return "Thousands"
    blobs: List[str] = []
    for c in ["Snippet", "Source Document"]:
        if c in df.columns:
            blobs.extend([str(x).lower() for x in df[c].dropna().tolist()])
    text = re.sub(r"\s+", " ", " ".join(blobs))
    if ("unit: trillions" in text) or re.search(r"\b(in|amounts? (?:are )?in|dollars? in|\$ in)\s+trillions?\b", text):
        return "Trillions"
    if ("unit: billions" in text) or re.search(r"\b(in|amounts? (?:are )?in|dollars? in|\$ in)\s+billions?\b", text):
        return "Billions"
    if ("unit: millions" in text) or re.search(r"\b(in|amounts? (?:are )?in|dollars? in|\$ in)\s+millions?\b", text):
        return "Millions"
    if ("unit: thousands" in text) or re.search(r"\b(in|amounts? (?:are )?in|dollars? in|\$ in)\s+thousands?\b", text):
        return "Thousands"
    return "Units"


def _display_scale_from_df(df: Optional[pd.DataFrame]) -> str:
    if df is None or df.empty:
        return "Units"
    ignore_cols = {"Sheet", "Confidence", "Page", "Open PDF Page", "Snippet", "Detected Unit", "Selected Year", "Currency"}
    non_amount_keywords = {"ratio", "coverage", "margin", "leverage", "%", "percent", "bps", "basis"}
    vals: List[float] = []
    for col in df.columns:
        if col in ignore_cols:
            continue
        col_l = str(col).lower()
        if any(k in col_l for k in non_amount_keywords):
            continue
        for v in df[col].tolist():
            n = _to_numeric_financial(v)
            if n is not None:
                vals.append(abs(float(n)))
    if not vals:
        # fallback to detected unit metadata when no numeric payload available
        return _detect_scale_from_combined_df(df)
    max_abs = max(vals)
    if max_abs >= 1_000_000_000_000:
        return "Trillions"
    if max_abs >= 1_000_000_000:
        return "Billions"
    if max_abs >= 1_000_000:
        return "Millions"
    if max_abs >= 1_000:
        return "Thousands"
    return "Units"


def _preferred_scale_from_df(df: Optional[pd.DataFrame]) -> str:
    # Prefer reported document unit when available; fallback to magnitude-based display scale.
    detected = _detect_scale_from_combined_df(df)
    if detected in {"Units", "Thousands", "Millions", "Billions", "Trillions"}:
        return detected
    return _display_scale_from_df(df)


def _currency_scale_display_label(scale: Any) -> str:
    s = str(scale or "").strip().lower()
    if s == "units":
        return "USD (Whole Dollars)"
    if s == "thousands":
        return "USD (Thousands)"
    if s == "millions":
        return "USD (Millions)"
    if s == "billions":
        return "USD (Billions)"
    if s == "trillions":
        return "USD (Trillions)"
    return "USD"


def _strict_pick_from_pdf(
    pdf: Path,
    labels: List[str],
    selected_year: Optional[int],
    year_order: List[int],
    *,
    choose_last: bool = False,
    reject_tokens: Optional[List[str]] = None,
    require_tokens: Optional[List[str]] = None,
    page_min: Optional[int] = None,
    page_max: Optional[int] = None,
    field_key: Optional[str] = None,
    allow_regex: bool = True,
) -> Dict[str, Any]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    if page_min is not None or page_max is not None:
        filtered: List[Dict[str, Any]] = []
        for it in lines:
            pg = it.get("page")
            if not isinstance(pg, int):
                continue
            if page_min is not None and pg < page_min:
                continue
            if page_max is not None and pg > page_max:
                continue
            filtered.append(it)
        if filtered:
            lines = filtered
    reject_tokens = [t.lower() for t in (reject_tokens or [])]
    require_tokens = [t.lower() for t in (require_tokens or [])]
    aliases = labels[:]
    if ENABLE_IFRS_DICTIONARY and field_key and field_key in FINANCIAL_FIELD_MAP:
        aliases = list(dict.fromkeys(labels + FINANCIAL_FIELD_MAP[field_key]))
    selected_idx = 0
    if selected_year is not None and year_order and selected_year in year_order:
        selected_idx = year_order.index(selected_year)
    matched_hits: List[Dict[str, Any]] = []

    def _context_year_order_for_row(row_idx: int, fallback_order: List[int]) -> List[int]:
        # Prefer table-local header order near the matched row (works for headers split across lines,
        # e.g., "Jan 25, 2026" / "Jan 26, 2025" / "Jan 28, 2024").
        target_page = lines[row_idx].get("page")
        start = max(0, row_idx - 260)
        nearby = lines[start:row_idx]
        same_page = [ln for ln in nearby if ln.get("page") == target_page]
        context_lines = same_page if same_page else nearby

        def _extract_seen(cands: List[Dict[str, Any]]) -> List[int]:
            seen_local: List[int] = []
            upper = _max_extractable_fiscal_year()
            for ln in cands:
                txt = str(ln.get("text") or "")
                ys = [int(y) for y in _extract_years_from_line(txt) if 1990 <= int(y) <= upper]
                if not ys:
                    continue
                is_years_only_line = bool(
                    re.fullmatch(
                        r"[\s$\(\)\-–—,]*(?:19\d{2}|20\d{2})(?:[\s$\(\)\-–—,]+(?:19\d{2}|20\d{2}))*[\s$\(\)\-–—,]*",
                        txt or "",
                    )
                )
                # Keep date/header-like lines and standalone year-only lines; skip narrative references.
                if not is_years_only_line and not re.search(r"(year ended|as of|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)", txt, re.IGNORECASE):
                    continue
                alpha_chars = len(re.findall(r"[A-Za-z]", txt))
                if alpha_chars > 120:
                    continue
                for y in ys:
                    if y not in seen_local:
                        seen_local.append(y)
            return seen_local

        seen: List[int] = _extract_seen(context_lines)
        if len(seen) < 2 and same_page:
            # Small fallback to previous page when table headers and values are split.
            prev_page_lines = [ln for ln in nearby if isinstance(target_page, int) and ln.get("page") == (target_page - 1)]
            seen = _extract_seen(prev_page_lines + same_page)
        if len(seen) >= 2:
            asc = seen == sorted(seen)
            desc = seen == sorted(seen, reverse=True)
            if asc or desc:
                max_seen = max(seen)
                filtered = [y for y in seen if y >= (max_seen - 3)]
                if len(filtered) >= 2:
                    return filtered
        return list(fallback_order or [])

    def _audited_context_score(row_idx: int) -> tuple[int, int]:
        # Prefer audited statement pages over summary pages.
        target_page = lines[row_idx].get("page")
        start = max(0, row_idx - 120)
        window = lines[start : row_idx + 6]
        same_page = [ln for ln in window if ln.get("page") == target_page]
        context_text = " ".join(str(ln.get("text") or "") for ln in (same_page or window)).lower()
        audited = 1 if re.search(r"\baudited\b", context_text) else 0
        stmt = 1 if re.search(r"consolidated statements", context_text) else 0
        # 10-K audited financials often use F-page numbering (F-1, F-2, ...).
        if re.search(r"\bf-\s?\d+\b", context_text):
            audited = 1
            stmt = 1
        if re.search(r"independent registered public accounting firm", context_text):
            audited = 1
            stmt = 1
        return audited, stmt
    for idx_line, item in enumerate(lines):
        raw = str(item.get("text") or "")
        label = _line_label_prefix(raw)
        label_n = _norm_label(label)
        if not label_n:
            continue
        raw_l = raw.lower()
        # Context guard: reject percentage-of-revenue presentation tables and nearby rows.
        context_text = raw_l
        if idx_line > 0:
            context_text = f"{str(lines[idx_line - 1].get('text') or '').lower()} {context_text}"
        if idx_line + 1 < len(lines):
            context_text = f"{context_text} {str(lines[idx_line + 1].get('text') or '').lower()}"
        if field_key in {"revenue", "net_income", "operating_income", "interest_expense"}:
            if "percentage of revenue" in context_text or "expressed as a percentage" in context_text:
                continue
            if "%" in raw_l:
                continue
        if any(tok in label_n for tok in reject_tokens):
            continue
        if require_tokens and not any(tok in label_n for tok in require_tokens):
            # Require adjusted/non-GAAP qualifier when requested.
            continue
        matched, match_type, conf_score = _match_financial_label(label_n, aliases, allow_regex=allow_regex)
        if not matched:
            continue

        row_year_order = _context_year_order_for_row(idx_line, year_order)
        expected_cols = len(row_year_order) if row_year_order else (len(year_order) if year_order else 1)
        expected_cols = max(1, expected_cols)
        # Value can be on the same line or in following lines.
        nums = _extract_numeric_tokens(raw)
        cursor = idx_line + 1
        while len(nums) < expected_cols and cursor < min(len(lines), idx_line + 10):
            next_raw = str(lines[cursor].get("text") or "")
            # Stop at next labeled row to prevent cross-row value bleed.
            next_label = _line_label_prefix(next_raw)
            if nums and next_label and re.search(r"[A-Za-z]", next_label):
                break
            nums.extend(_extract_numeric_tokens(next_raw))
            cursor += 1
        nums = [n for n in nums if n != ""]
        # Skip narrative matches that don't look like table rows.
        if not nums and len(label_n.split()) > 4:
            continue
        if not nums:
            continue
        # Prefer year-column aligned values from the trailing numeric group to avoid
        # footnote/index tokens at the beginning of statement rows.
        year_values: List[Dict[str, Any]] = []
        if row_year_order and nums:
            if len(nums) >= len(row_year_order):
                effective_order = list(row_year_order)
                year_vals = nums[-len(effective_order):]
            else:
                # Row contains a subset of year columns; align to right-most years.
                effective_order = list(row_year_order[-len(nums):])
                year_vals = nums[-len(effective_order):]
            # Statement tables usually list values left-to-right as newest→oldest. If the
            # nearby header scan returned ascending years (oldest→newest), flip to
            # descending so column order matches the numeric token order.
            if len(effective_order) >= 2 and effective_order == sorted(effective_order):
                effective_order = sorted(effective_order, reverse=True)
            # Build full year-value array for all detected columns.
            for y, tok in zip(effective_order, year_vals):
                yv = _to_numeric_financial(tok)
                if yv is not None:
                    year_values.append({"year": int(y), "metric": field_key or (aliases[0] if aliases else label), "value": yv})
            if selected_year is not None and selected_year in effective_order:
                pick_idx = effective_order.index(selected_year)
            else:
                pick_idx = selected_idx if selected_idx < len(year_vals) else 0
            token = year_vals[pick_idx]
        else:
            pick_idx = selected_idx if selected_idx < len(nums) else 0
            token = nums[pick_idx]
        value = _to_numeric_financial(token)
        if value is None:
            continue
        audited_score, stmt_score = _audited_context_score(idx_line)
        matched_hits.append(
            {
                "value": value,
                "confidence": conf_score / 100.0,
                "page": item.get("page"),
                "snippet": raw[:260],
                "label": label,
                "match_type": match_type,
                "confidence_score": conf_score,
                "year": selected_year,
                "year_values": year_values,
                "audited_score": audited_score,
                "statement_score": stmt_score,
            }
        )
    if matched_hits:
        audited_hits = [h for h in matched_hits if h.get("audited_score")]
        candidate_hits = audited_hits if audited_hits else matched_hits
        if field_key == "revenue":
            # Prefer consolidated top-line revenue over segment sub-lines.
            return max(candidate_hits, key=lambda h: abs(float(h.get("value") or 0.0)))
        return candidate_hits[-1] if choose_last else candidate_hits[0]
    return {"value": None, "confidence": None, "page": None, "snippet": None}


def _strict_pick_t12m_from_pdf(
    pdf: Path,
    labels: List[str],
    *,
    reject_tokens: Optional[List[str]] = None,
    page_min: Optional[int] = None,
    page_max: Optional[int] = None,
    field_key: Optional[str] = None,
    allow_regex: bool = True,
) -> Dict[str, Any]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    if page_min is not None or page_max is not None:
        filtered: List[Dict[str, Any]] = []
        for it in lines:
            pg = it.get("page")
            if not isinstance(pg, int):
                continue
            if page_min is not None and pg < page_min:
                continue
            if page_max is not None and pg > page_max:
                continue
            filtered.append(it)
        if filtered:
            lines = filtered
    reject_tokens = [t.lower() for t in (reject_tokens or [])]
    aliases = labels[:]
    if ENABLE_IFRS_DICTIONARY and field_key and field_key in FINANCIAL_FIELD_MAP:
        aliases = list(dict.fromkeys(labels + FINANCIAL_FIELD_MAP[field_key]))

    hits: List[Dict[str, Any]] = []
    quarter_rx = re.compile(r"\b(q[1-4]|quarter|three months ended|quarter ended)\b", re.IGNORECASE)
    for idx_line, item in enumerate(lines):
        raw = str(item.get("text") or "")
        label = _line_label_prefix(raw)
        label_n = _norm_label(label)
        if not label_n:
            continue
        if any(tok in label_n for tok in reject_tokens):
            continue
        matched, match_type, conf_score = _match_financial_label(label_n, aliases, allow_regex=allow_regex)
        if not matched:
            continue
        # T12M must be built from quarterly context, not annual columns.
        quarter_context = bool(quarter_rx.search(raw))
        nums = _extract_numeric_tokens(raw)
        cursor = idx_line + 1
        while len(nums) < 4 and cursor < min(len(lines), idx_line + 10):
            next_raw = str(lines[cursor].get("text") or "")
            if quarter_rx.search(next_raw):
                quarter_context = True
            nums.extend(_extract_numeric_tokens(next_raw))
            cursor += 1
        if not quarter_context:
            continue
        nums = [n for n in nums if n != ""]
        if len(nums) < 4:
            continue
        quarter_vals = nums[-4:]
        parsed_vals: List[float] = []
        for tok in quarter_vals:
            val = _to_numeric_financial(tok)
            if val is None:
                parsed_vals = []
                break
            parsed_vals.append(float(val))
        if len(parsed_vals) != 4:
            continue
        hits.append(
            {
                "value": sum(parsed_vals),
                "confidence": 0.70,
                "page": item.get("page"),
                "snippet": raw[:260],
                "label": label,
                "match_type": "t12m",
                "confidence_score": 70,
                "year": None,
            }
        )
    if hits:
        return hits[0]
    return {"value": None, "confidence": None, "page": None, "snippet": None}


def _strict_pick_capex_from_cashflow_context(
    pdf: Path,
    selected_year: Optional[int],
    year_order: List[int],
    *,
    page_min: Optional[int] = None,
    page_max: Optional[int] = None,
) -> Dict[str, Any]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    if page_min is not None or page_max is not None:
        filtered: List[Dict[str, Any]] = []
        for it in lines:
            pg = it.get("page")
            if not isinstance(pg, int):
                continue
            if page_min is not None and pg < page_min:
                continue
            if page_max is not None and pg > page_max:
                continue
            filtered.append(it)
        if filtered:
            lines = filtered
    selected_idx = 0
    if selected_year is not None and year_order and selected_year in year_order:
        selected_idx = year_order.index(selected_year)

    def _context_year_order_for_row(row_idx: int, fallback_order: List[int]) -> List[int]:
        target_page = lines[row_idx].get("page")
        start = max(0, row_idx - 260)
        nearby = lines[start:row_idx]
        same_page = [ln for ln in nearby if ln.get("page") == target_page]
        context_lines = same_page if same_page else nearby

        def _extract_seen(cands: List[Dict[str, Any]]) -> List[int]:
            seen_local: List[int] = []
            upper = _max_extractable_fiscal_year()
            for ln in cands:
                txt = str(ln.get("text") or "")
                ys = [int(y) for y in _extract_years_from_line(txt) if 1990 <= int(y) <= upper]
                if not ys:
                    continue
                is_years_only_line = bool(
                    re.fullmatch(
                        r"[\s$\(\)\-–—,]*(?:19\d{2}|20\d{2})(?:[\s$\(\)\-–—,]+(?:19\d{2}|20\d{2}))*[\s$\(\)\-–—,]*",
                        txt or "",
                    )
                )
                if not is_years_only_line and not re.search(r"(year ended|as of|jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)", txt, re.IGNORECASE):
                    continue
                alpha_chars = len(re.findall(r"[A-Za-z]", txt))
                if alpha_chars > 120:
                    continue
                for y in ys:
                    if y not in seen_local:
                        seen_local.append(y)
            return seen_local

        seen: List[int] = _extract_seen(context_lines)
        if len(seen) < 2 and same_page:
            prev_page_lines = [ln for ln in nearby if isinstance(target_page, int) and ln.get("page") == (target_page - 1)]
            seen = _extract_seen(prev_page_lines + same_page)
        if len(seen) >= 2:
            asc = seen == sorted(seen)
            desc = seen == sorted(seen, reverse=True)
            if asc or desc:
                max_seen = max(seen)
                filtered = [y for y in seen if y >= (max_seen - 3)]
                if len(filtered) >= 2:
                    return filtered
        return list(fallback_order or [])
    for idx_line, item in enumerate(lines):
        raw = str(item.get("text") or "")
        label = _line_label_prefix(raw)
        label_n = _norm_label(label)
        if not label_n:
            continue
        is_capex_label = (
            ("capital" in label_n and "expend" in label_n)
            or ("property" in label_n and "equipment" in label_n and any(tok in label_n for tok in ["purchase", "purchases", "acquisition", "acquisitions", "addition", "additions"]))
            or ("pp&e" in raw.lower())
        )
        if not is_capex_label:
            continue
        row_year_order = _context_year_order_for_row(idx_line, year_order)
        expected_cols = len(row_year_order) if row_year_order else (len(year_order) if year_order else 1)
        expected_cols = max(1, expected_cols)
        nums = _extract_numeric_tokens(raw)
        cursor = idx_line + 1
        while len(nums) < expected_cols and cursor < min(len(lines), idx_line + 10):
            next_raw = str(lines[cursor].get("text") or "")
            next_label = _line_label_prefix(next_raw)
            if nums and next_label and re.search(r"[A-Za-z]", next_label):
                break
            nums.extend(_extract_numeric_tokens(next_raw))
            cursor += 1
        nums = [n for n in nums if n != ""]
        if not nums:
            continue
        if row_year_order and nums:
            if len(nums) >= len(row_year_order):
                effective_order = list(row_year_order)
                year_vals = nums[-len(effective_order):]
            else:
                effective_order = list(row_year_order[-len(nums):])
                year_vals = nums[-len(effective_order):]
            if selected_year is not None and selected_year in effective_order:
                pick_idx = effective_order.index(selected_year)
            else:
                pick_idx = selected_idx if selected_idx < len(year_vals) else 0
            token = year_vals[pick_idx]
        else:
            pick_idx = selected_idx if selected_idx < len(nums) else 0
            token = nums[pick_idx]
        value = _to_numeric_financial(token)
        if value is None:
            continue
        return {
            "value": value,
            "confidence": 0.74,
            "page": item.get("page"),
            "snippet": raw[:260],
            "match_type": "cashflow-capex-fallback",
            "confidence_score": 74,
            "year": selected_year,
        }
    return {"value": None, "confidence": None, "page": None, "snippet": None}


def _strict_financial_sheet_row(sheet: str, pdf: Path) -> Tuple[Dict[str, Any], List[str]]:
    row: Dict[str, Any] = {}
    errors: List[str] = []
    metric_meta: Dict[str, Dict[str, Any]] = {}
    if _looks_like_search_results_pdf(pdf):
        return row, [
            "Uploaded PDF appears to be a search-results page, not the actual filing statement. "
            "Please open the filing document (10-K/10-Q report page) and upload that PDF."
        ]
    anchor_page = _sheet_anchor_page(sheet, pdf)
    page_min = max(1, (anchor_page or 1) - 2)
    page_max = (anchor_page + 4) if anchor_page else None
    year_cfg = _analysis_year_config()
    analysis_mode = year_cfg.get("mode", "latest_available")
    specific_year = year_cfg.get("specific_year")
    year_request = "latest"
    allow_fallback = True
    if analysis_mode == "specific_year":
        year_request = str(specific_year) if specific_year else "latest"
        allow_fallback = False
    selected_year, year_order, global_years = _selected_fiscal_year_and_order(
        pdf,
        page_min=page_min,
        page_max=page_max,
        analysis_year=year_request,
        allow_fallback_to_latest=allow_fallback,
    )
    available_years_detected = _extract_fiscal_years_detected(pdf, max_pages=_strict_scan_page_limit(pdf))
    if year_order and len(year_order) >= 2:
        available_years_detected = sorted(set(int(y) for y in year_order))
    elif global_years and len(global_years) >= 2:
        available_years_detected = sorted(set(int(y) for y in global_years))

    parser_rows: List[Dict[str, Any]] = []
    if parse_financial_statements is not None and sheet in {"Income Statement", "Balance Sheet", "Cash Flow"}:
        try:
            parser_rows = _financial_parser_rows_cached(str(pdf), EXTRACTION_CACHE_VERSION)
        except Exception:
            parser_rows = []
    statement_key = {
        "Income Statement": "income_statement",
        "Balance Sheet": "balance_sheet",
        "Cash Flow": "cash_flow",
    }.get(sheet)
    available_years_for_sheet = list(available_years_detected)
    if parser_rows:
        parser_years = sorted({int(r.get("selected_year")) for r in parser_rows if r.get("selected_year") is not None})
        parser_years_for_statement = sorted(
            {
                int(r.get("selected_year"))
                for r in parser_rows
                if r.get("selected_year") is not None and (statement_key is None or r.get("statement_type") == statement_key)
            }
        )
        if parser_years:
            available_years_detected = parser_years
            available_years_for_sheet = parser_years_for_statement or parser_years
            if analysis_mode == "latest_available":
                selected_year = max(available_years_for_sheet)
            elif analysis_mode == "specific_year":
                if specific_year is None:
                    return row, ["Specific Year mode requires a selected year."]
                if int(specific_year) not in available_years_for_sheet:
                    return row, [f"Selected year {int(specific_year)} not found in document. Metrics marked Incomplete."]
                selected_year = int(specific_year)
            year_order = available_years_for_sheet
    if not _sheet_header_present(sheet, pdf) and anchor_page is None and not parser_rows:
        return row, ["Unable to identify document type."]
    if analysis_mode == "latest_available" and available_years_for_sheet:
        max_year = max(available_years_for_sheet)
        if selected_year != max_year:
            return row, [f"Year selection engine failure: latest available year is {max_year}, but selected {selected_year}."]
    if analysis_mode == "specific_year":
        if specific_year is None:
            return row, ["Specific Year mode requires a selected year."]
        if selected_year is None:
            return row, [f"Selected year {specific_year} not found in document. Metrics marked Incomplete."]
    if selected_year is not None and year_order and selected_year not in year_order:
        return row, [f"Selected fiscal year not found in {sheet}. Metrics marked Incomplete."]
    unit_label = "units"
    uf = _statement_unit_factor(pdf)
    if uf == 1_000.0:
        unit_label = "thousands"
    elif uf == 1_000_000.0:
        unit_label = "millions"
    elif uf == 1_000_000_000.0:
        unit_label = "billions"
    row["Selected Year"] = selected_year
    row["Detected Unit"] = unit_label
    row["Unit Multiplier"] = uf
    row["Analysis Mode"] = analysis_mode
    row["Fiscal Period"] = "FY"
    row["Available Years"] = ", ".join(str(y) for y in available_years_detected)

    if parser_rows and sheet in {"Income Statement", "Balance Sheet", "Cash Flow"} and selected_year is not None:
        parser_match = next(
            (
                r for r in parser_rows
                if r.get("statement_type") == statement_key and int(r.get("selected_year")) == int(selected_year)
            ),
            None,
        )
        if parser_match:
            scale = float(uf) if uf not in (None, 0) else 1.0
            def _scaled(v: Any) -> Optional[float]:
                if v is None:
                    return None
                try:
                    return float(v) / scale if scale else float(v)
                except Exception:
                    return None

            if sheet == "Income Statement":
                row["Revenue"] = _scaled(parser_match.get("revenue"))
                row["COGS"] = _scaled(parser_match.get("cogs"))
                row["EBIT"] = _scaled(parser_match.get("ebit"))
                row["EBITDA"] = _scaled(parser_match.get("ebitda"))
                row["Net Income"] = _scaled(parser_match.get("net_income"))
                row["Interest Expense"] = _scaled(parser_match.get("interest_expense"))
                row["Operating Income"] = row.get("EBIT")
                row["Operating Margin"] = _safe_div(row.get("Operating Income"), row.get("Revenue"))
            elif sheet == "Balance Sheet":
                row["Total Assets"] = _scaled(parser_match.get("total_assets"))
                row["Shareholders' Equity"] = _scaled(parser_match.get("total_equity"))
                row["Current Assets"] = _scaled(parser_match.get("current_assets"))
                row["Current Liabilities"] = _scaled(parser_match.get("current_liabilities"))
                row["Inventory"] = _scaled(parser_match.get("inventory"))
                row["Short-term Debt"] = _scaled(parser_match.get("st_debt"))
                row["Long-term Debt"] = _scaled(parser_match.get("lt_debt"))
                if row.get("Short-term Debt") is not None or row.get("Long-term Debt") is not None:
                    row["Total Debt"] = (row.get("Short-term Debt") or 0.0) + (row.get("Long-term Debt") or 0.0)
            elif sheet == "Cash Flow":
                row["Operating Cash Flow"] = _scaled(parser_match.get("operating_cf"))
                capex_raw = _scaled(parser_match.get("capex"))
                # Filings usually show CapEx in parentheses (outflow). Store magnitude like strict-path.
                row["CapEx"] = abs(float(capex_raw)) if capex_raw is not None else None
                row["Free Cash Flow"] = (
                    row.get("Operating Cash Flow") - row["CapEx"]
                    if row.get("Operating Cash Flow") is not None and row["CapEx"] is not None
                    else None
                )

            _apply_derived_fields(row, sheet)
            row["Confidence"] = "high" if parser_match.get("extraction_confidence") == "high" else None
            page_map = parser_match.get("__page_map") if isinstance(parser_match.get("__page_map"), dict) else {}
            all_pages = sorted(
                {
                    int(p)
                    for p in (
                        parser_match.get("__pages") if isinstance(parser_match.get("__pages"), list) else page_map.values()
                    )
                    if isinstance(p, (int, float)) and int(p) > 0
                }
            )
            row["Page"] = min(all_pages) if all_pages else None
            row["Pages"] = ", ".join(str(p) for p in all_pages) if all_pages else None
            row["Open PDF Page"] = _open_page_label(row["Page"])
            row["Snippet"] = (
                f"parsed via pdfplumber tables | cited pages: {row['Pages']}"
                if row.get("Pages")
                else "parsed via pdfplumber tables"
            )
            parser_key_for_col = {
                "Revenue": "revenue",
                "COGS": "cogs",
                "EBIT": "ebit",
                "EBITDA": "ebitda",
                "Net Income": "net_income",
                "Interest Expense": "interest_expense",
                "Operating Income": "ebit",
                "Total Assets": "total_assets",
                "Shareholders' Equity": "total_equity",
                "Current Assets": "current_assets",
                "Current Liabilities": "current_liabilities",
                "Inventory": "inventory",
                "Short-term Debt": "st_debt",
                "Long-term Debt": "lt_debt",
                "Total Debt": "total_debt",
                "Operating Cash Flow": "operating_cf",
                "CapEx": "capex",
                "Free Cash Flow": "free_cash_flow",
            }
            for col in list(row.keys()):
                if col in META_COLUMNS or col in DERIVED_COLUMNS or col in YEAR_META_COLUMNS:
                    continue
                meta = {}
                parser_key = parser_key_for_col.get(col)
                page_val = page_map.get(parser_key) if parser_key else None
                if isinstance(page_val, int):
                    meta["page"] = int(page_val)
                if meta:
                    metric_meta[col] = {**metric_meta.get(col, {}), **meta}
            row["_metric_meta"] = metric_meta
            return row, errors
        # Parser found document-year data, but this specific statement/year is missing.
        # Keep strict fallback enabled for this sheet to avoid all-null collapse.
        errors.append(f"Selected fiscal year not found in {sheet}. Using strict fallback extraction.")

    if sheet == "Income Statement":
        if analysis_mode == "t12m":
            row["Fiscal Period"] = "T12M"
        total_exp = {"value": None}
        emp_comp = {"value": None}
        if analysis_mode == "t12m":
            revenue = _strict_pick_t12m_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Revenue"],
                reject_tokens=["cloud", "product", "segment", "service revenue"],
                page_min=page_min,
                page_max=page_max,
                field_key="revenue",
                allow_regex=False,
            )
            net_income = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Net Income"], page_min=page_min, page_max=page_max, field_key="net_income", allow_regex=False
            )
            cogs = _strict_pick_t12m_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["COGS"],
                page_min=page_min,
                page_max=page_max,
                field_key="cogs",
                allow_regex=False,
            )
            ebit = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["EBIT"], page_min=page_min, page_max=page_max, field_key="operating_income", allow_regex=False
            )
            ebitda = _strict_pick_t12m_from_pdf(pdf, STRICT_LABEL_PRIORITY["EBITDA"], page_min=page_min, page_max=page_max)
            interest_exp = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Interest Expense"], page_min=page_min, page_max=page_max, field_key="interest_expense"
            )
            ebit_adj = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["EBIT"], page_min=page_min, page_max=page_max, field_key="operating_income", allow_regex=False
            )
            ebitda_adj = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["EBITDA"], page_min=page_min, page_max=page_max
            )
            net_income_adj = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Net Income"], page_min=page_min, page_max=page_max, field_key="net_income", allow_regex=False
            )
            total_exp = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Total Expense"], page_min=page_min, page_max=page_max, field_key="total_expense"
            )
            emp_comp = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Employee Compensation"], page_min=page_min, page_max=page_max, field_key="employee_compensation"
            )
            non_interest_exp = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Non-Interest Expense"], page_min=page_min, page_max=page_max, field_key="non_interest_expense"
            )
            income_before_taxes = {"value": None}
            tax_expense = {"value": None}
            eps_basic = {"value": None}
            eps_diluted = {"value": None}
        else:
            revenue = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Revenue"],
                selected_year,
                year_order,
                reject_tokens=["cloud", "product", "segment", "service revenue", "cost of revenue", "cost of revenues"],
                page_min=page_min,
                page_max=page_max,
                field_key="revenue",
                allow_regex=False,
            )
            net_income = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Net Income"],
                selected_year,
                year_order,
                choose_last=False,
                page_min=page_min,
                page_max=page_max,
                field_key="net_income",
                allow_regex=False,
                reject_tokens=["per share", "basic", "diluted", "other income", "comprehensive income"],
            )
            cogs = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["COGS"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="cogs",
                allow_regex=False,
                reject_tokens=["gross margin", "gross profit", "percent", "%", "operating expenses", "sg&a", "research and development"],
            )
            ebit = _strict_pick_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["EBIT"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="operating_income", allow_regex=False
            )
            ebitda = _strict_pick_from_pdf(pdf, STRICT_LABEL_PRIORITY["EBITDA"], selected_year, year_order, page_min=page_min, page_max=page_max)
            interest_exp = _strict_pick_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Interest Expense"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="interest_expense"
            )
            ebit_adj = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["EBIT"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="operating_income",
                allow_regex=False,
                require_tokens=["adjusted", "non-gaap", "as adjusted"],
            )
            ebitda_adj = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["EBITDA"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                allow_regex=False,
                require_tokens=["adjusted", "non-gaap", "as adjusted"],
            )
            net_income_adj = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Net Income"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="net_income",
                allow_regex=False,
                require_tokens=["adjusted", "non-gaap", "as adjusted"],
            )
            total_exp = _strict_pick_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Total Expense"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="total_expense"
            )
            emp_comp = _strict_pick_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Employee Compensation"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="employee_compensation"
            )
            non_interest_exp = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Non-Interest Expense"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="non_interest_expense",
                allow_regex=False,
            )
            income_before_taxes = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Income Before Taxes"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="income_before_taxes",
                allow_regex=False,
            )
            tax_expense = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["Tax Expense"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="tax_expense",
                allow_regex=False,
            )
            eps_basic = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["EPS (Basic)"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="eps_basic",
                allow_regex=False,
            )
            eps_diluted = _strict_pick_from_pdf(
                pdf,
                STRICT_LABEL_PRIORITY["EPS (Diluted)"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="eps_diluted",
                allow_regex=False,
            )
        if interest_exp.get("value") is None:
            notes_anchor = _classify_statement_blocks(pdf).get("Notes Section")
            if notes_anchor is not None:
                if analysis_mode == "t12m":
                    interest_exp = _strict_pick_t12m_from_pdf(
                        pdf,
                        STRICT_LABEL_PRIORITY["Interest Expense"],
                        page_min=max(1, notes_anchor - 8),
                        page_max=notes_anchor + 22,
                        field_key="interest_expense",
                    )
                else:
                    interest_exp = _strict_pick_from_pdf(
                        pdf,
                        STRICT_LABEL_PRIORITY["Interest Expense"],
                        selected_year,
                        year_order,
                        page_min=max(1, notes_anchor - 8),
                        page_max=notes_anchor + 22,
                        field_key="interest_expense",
                    )
        row["Revenue"] = revenue.get("value")
        cogs_val = _to_numeric_financial(cogs.get("value"))
        row["COGS"] = abs(float(cogs_val)) if cogs_val is not None else None
        row["Net Income"] = net_income.get("value")
        row["EBIT"] = ebit.get("value")
        row["EBITDA"] = ebitda.get("value")
        row["EBIT (Adjusted)"] = ebit_adj.get("value")
        row["EBITDA (Adjusted)"] = ebitda_adj.get("value")
        row["Net Income (Adjusted)"] = net_income_adj.get("value")
        row["Total Expense"] = total_exp.get("value")
        row["Employee Compensation"] = emp_comp.get("value")
        row["Non-Interest Expense"] = non_interest_exp.get("value")
        # Financial services add-ons (when present).
        row["Operating Income"] = row.get("EBIT")
        op_margin = _safe_div(row.get("Operating Income"), row.get("Revenue"))
        row["Operating Margin"] = op_margin
        row["Income Before Taxes"] = income_before_taxes.get("value")
        row["Tax Expense"] = tax_expense.get("value")
        eff_tax = _safe_div(row.get("Tax Expense"), row.get("Income Before Taxes"))
        row["Effective Tax Rate"] = eff_tax
        row["EPS (Basic)"] = eps_basic.get("value")
        row["EPS (Diluted)"] = eps_diluted.get("value")
        # Normalize expense sign for ratio inputs: (200) -> 200
        ie_val = _to_numeric_financial(interest_exp.get("value"))
        row["Interest Expense"] = abs(float(ie_val)) if ie_val is not None else None
        metric_meta["Revenue"] = revenue
        metric_meta["COGS"] = cogs
        metric_meta["Net Income"] = net_income
        metric_meta["EBIT"] = ebit
        metric_meta["EBITDA"] = ebitda
        metric_meta["Interest Expense"] = interest_exp
        metric_meta["EBIT (Adjusted)"] = {**ebit_adj, "basis": "adjusted"} if isinstance(ebit_adj, dict) else ebit_adj
        metric_meta["EBITDA (Adjusted)"] = {**ebitda_adj, "basis": "adjusted"} if isinstance(ebitda_adj, dict) else ebitda_adj
        metric_meta["Net Income (Adjusted)"] = {**net_income_adj, "basis": "adjusted"} if isinstance(net_income_adj, dict) else net_income_adj
        metric_meta["Total Expense"] = total_exp
        metric_meta["Employee Compensation"] = emp_comp
        metric_meta["Non-Interest Expense"] = non_interest_exp
        metric_meta["Operating Income"] = metric_meta.get("EBIT", ebit)
        if op_margin is not None:
            metric_meta["Operating Margin"] = {
                "value": op_margin,
                "source": "derived",
                "match_type": "derived",
                "confidence_score": 60,
                "year": selected_year,
            }
        metric_meta["Income Before Taxes"] = income_before_taxes
        metric_meta["Tax Expense"] = tax_expense
        if eff_tax is not None:
            metric_meta["Effective Tax Rate"] = {
                "value": eff_tax,
                "source": "derived",
                "match_type": "derived",
                "confidence_score": 60,
                "year": selected_year,
            }
        metric_meta["EPS (Basic)"] = eps_basic
        metric_meta["EPS (Diluted)"] = eps_diluted
        if analysis_mode == "t12m":
            if row["Revenue"] is None:
                errors.append("T12M mode selected but 4-quarter Income Statement series not found. Metrics marked Incomplete.")
            if row["Net Income"] is None:
                errors.append("T12M mode selected but 4-quarter Net Income series not found. Metrics marked Incomplete.")
        if row["EBITDA"] is None:
            # Derive EBITDA = EBIT + (Depreciation + Amortization) when explicit EBITDA missing.
            dep = _strict_pick_from_pdf(
                pdf,
                ["depreciation", "depreciation expense", "depreciation and amortization"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="depreciation",
            )
            amort = _strict_pick_from_pdf(
                pdf,
                ["amortization", "amortization of intangible assets", "amortization expense", "depreciation and amortization"],
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
                field_key="amortization",
            )
            dep_v = _to_numeric_financial(dep.get("value"))
            amort_v = _to_numeric_financial(amort.get("value"))
            da_sum = None
            if dep_v is not None and amort_v is not None:
                da_sum = float(dep_v) + float(amort_v)
            elif dep_v is not None:
                da_sum = float(dep_v)
            elif amort_v is not None:
                da_sum = float(amort_v)
            if da_sum is not None and row["EBIT"] is not None:
                row["EBITDA"] = row["EBIT"] + float(da_sum)
                metric_meta["EBITDA"] = {
                    "value": row["EBITDA"],
                    "source": "derived",
                    "match_type": "derived",
                    "confidence_score": 60,
                    "year": selected_year,
                }
        # Sanity guard: EBITDA should generally not be materially below EBIT.
        # If explicit EBITDA is suspiciously low vs EBIT, prefer derived EBITDA.
        if row.get("EBITDA") is not None and row.get("EBIT") is not None:
            try:
                ebitda_v = float(row["EBITDA"])
                ebit_v = float(row["EBIT"])
                if ebit_v > 0 and ebitda_v < (ebit_v * 0.90):
                    da_fix = _strict_pick_from_pdf(
                        pdf,
                        FINANCIAL_FIELD_MAP["depreciation_amortization"],
                        selected_year,
                        year_order,
                        page_min=page_min,
                        page_max=page_max,
                        field_key="depreciation_amortization",
                    )
                    da_val = _to_numeric_financial(da_fix.get("value"))
                    if da_val is not None:
                        row["EBITDA"] = float(row["EBIT"]) + float(da_val)
                        metric_meta["EBITDA"] = {
                            "value": row["EBITDA"],
                            "source": "derived_sanity_fix",
                            "match_type": "derived",
                            "confidence_score": 60,
                            "year": selected_year,
                        }
                    else:
                        # conservative fallback so coverage/leverage don't run on a bad EBITDA pick
                        row["EBITDA"] = float(row["EBIT"])
                        metric_meta["EBITDA"] = {
                            "value": row["EBITDA"],
                            "source": "ebit_fallback_sanity_fix",
                            "match_type": "derived",
                            "confidence_score": 55,
                            "year": selected_year,
                        }
            except Exception:
                pass
        if row["Revenue"] is None and analysis_mode != "t12m":
            errors.append("Income Statement document not detected. Please upload a valid Consolidated Statement of Operations.")
    elif sheet == "Balance Sheet":
        if analysis_mode == "t12m":
            row["Fiscal Period"] = "LatestQuarter"
        total_assets = _strict_pick_from_pdf(
            pdf, STRICT_LABEL_PRIORITY["Total Assets"], selected_year, year_order, choose_last=True, page_min=page_min, page_max=page_max, field_key="total_assets", allow_regex=False
        )
        total_liabilities = _strict_pick_from_pdf(
            pdf, STRICT_LABEL_PRIORITY["Total Liabilities"], selected_year, year_order, choose_last=True, page_min=page_min, page_max=page_max, field_key="total_liabilities", allow_regex=False
        )
        equity = _strict_pick_from_pdf(
            pdf,
            STRICT_LABEL_PRIORITY["Shareholders' Equity"],
            selected_year,
            year_order,
            choose_last=True,
            page_min=page_min,
            page_max=page_max,
            field_key="equity",
            allow_regex=False,
            reject_tokens=[
                "per share",
                "shares",
                "share count",
                "common stock",
                "par value",
                "basic",
                "diluted",
            ],
        )
        current_assets = _strict_pick_from_pdf(
            pdf, STRICT_LABEL_PRIORITY["Current Assets"], selected_year, year_order, choose_last=True, page_min=page_min, page_max=page_max, field_key="current_assets", allow_regex=False
        )
        current_liabilities = _strict_pick_from_pdf(
            pdf, STRICT_LABEL_PRIORITY["Current Liabilities"], selected_year, year_order, choose_last=True, page_min=page_min, page_max=page_max, field_key="current_liabilities", allow_regex=False
        )
        inventory = _strict_pick_from_pdf(
            pdf,
            STRICT_LABEL_PRIORITY["Inventory"],
            selected_year,
            year_order,
            choose_last=True,
            page_min=page_min,
            page_max=page_max,
            field_key="inventory",
            allow_regex=True,
        )
        accounts_receivable = _strict_pick_from_pdf(
            pdf,
            STRICT_LABEL_PRIORITY["Accounts Receivable"],
            selected_year,
            year_order,
            choose_last=True,
            page_min=page_min,
            page_max=page_max,
            field_key="accounts_receivable",
            allow_regex=True,
        )
        cash = _strict_pick_from_pdf(pdf, STRICT_LABEL_PRIORITY["Cash"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="cash")
        short_debt = _strict_pick_from_pdf(
            pdf,
            STRICT_LABEL_PRIORITY["Short-term Debt"],
            selected_year,
            year_order,
            page_min=page_min,
            page_max=page_max,
            field_key="short_term_debt",
            reject_tokens=["securities", "marketable", "investments", "assets"],
            allow_regex=True,
        )
        current_portion_debt = _strict_pick_from_pdf(
            pdf,
            STRICT_LABEL_PRIORITY["Current Portion LT Debt"],
            selected_year,
            year_order,
            page_min=page_min,
            page_max=page_max,
            field_key="current_portion_long_term_debt",
            reject_tokens=["securities", "marketable", "investments", "assets"],
            allow_regex=True,
        )
        long_debt = _strict_pick_from_pdf(
            pdf,
            STRICT_LABEL_PRIORITY["Long-term Debt"],
            selected_year,
            year_order,
            page_min=page_min,
            page_max=page_max,
            field_key="long_term_debt",
            reject_tokens=["securities", "marketable", "investments", "assets"],
            allow_regex=True,
        )
        total_debt_direct = _strict_pick_from_pdf(
            pdf,
            ["total debt", "total borrowings", "interest bearing liabilities"],
            selected_year,
            year_order,
            page_min=page_min,
            page_max=page_max,
            field_key="total_debt",
        )
        row["Total Assets"] = total_assets.get("value")
        row["Total Liabilities"] = total_liabilities.get("value")
        row["Shareholders' Equity"] = equity.get("value")
        row["Current Assets"] = current_assets.get("value")
        row["Current Liabilities"] = current_liabilities.get("value")
        row["Inventory"] = inventory.get("value")
        row["Accounts Receivable"] = accounts_receivable.get("value")
        row["Cash"] = cash.get("value")
        row["Short-term Debt"] = short_debt.get("value")
        row["Current Portion of Long-term Debt"] = current_portion_debt.get("value")
        row["Long-term Debt"] = long_debt.get("value")
        # Step 2: aggregate debt from all mapped interest-bearing components first.
        # Defensive dedupe: some statements only present "current portion of long-term debt",
        # which can be mistakenly matched as short-term debt. Treat equal values as one bucket.
        try:
            sd = _to_numeric_financial(row.get("Short-term Debt"))
            cd = _to_numeric_financial(row.get("Current Portion of Long-term Debt"))
            if sd is not None and cd is not None and abs(float(sd) - float(cd)) <= max(1.0, abs(float(cd)) * 0.001):
                row["Short-term Debt"] = None
                short_debt = {"value": None, "confidence": None, "page": None, "snippet": None}
        except Exception:
            pass
        debt_components = [
            row.get("Short-term Debt"),
            row.get("Current Portion of Long-term Debt"),
            row.get("Long-term Debt"),
        ]
        if any(v is not None for v in debt_components):
            comp_sum = sum(float(v or 0.0) for v in debt_components)
            direct_val = total_debt_direct.get("value")
            if direct_val is not None and comp_sum > 0 and direct_val > 0 and (comp_sum > direct_val * 1.75):
                # Guard against component false-positives (e.g., marketable securities rows).
                row["Total Debt"] = direct_val
            else:
                row["Total Debt"] = comp_sum
        else:
            # Fallback only when components are unavailable.
            row["Total Debt"] = total_debt_direct.get("value")
        metric_meta.update(
            {
                "Total Assets": total_assets,
                "Total Liabilities": total_liabilities,
                "Shareholders' Equity": equity,
                "Current Assets": current_assets,
                "Current Liabilities": current_liabilities,
                "Inventory": inventory,
                "Accounts Receivable": accounts_receivable,
                "Cash": cash,
                "Short-term Debt": short_debt,
                "Current Portion of Long-term Debt": current_portion_debt,
                "Long-term Debt": long_debt,
                "Total Debt": total_debt_direct,
            }
        )
        # Guard against obvious balance-sheet mis-mapping:
        # if equity captured from a per-share/share-count row (tiny value),
        # reconstruct from A=L+E relation when available.
        try:
            ta = _to_numeric_financial(row.get("Total Assets"))
            tl = _to_numeric_financial(row.get("Total Liabilities"))
            eq = _to_numeric_financial(row.get("Shareholders' Equity"))
            if ta is not None and tl is not None:
                derived_eq = float(ta) - float(tl)
                if eq is None or (abs(float(eq)) > 0 and abs(float(eq)) < max(1.0, abs(float(ta)) * 0.01)):
                    row["Shareholders' Equity"] = derived_eq
                    metric_meta["Shareholders' Equity"] = {
                        "value": derived_eq,
                        "source": "derived_balance_identity",
                        "match_type": "derived",
                        "confidence_score": 60,
                        "year": selected_year,
                    }
        except Exception:
            pass
        # Normalize core asset magnitudes to avoid parenthesis-driven sign flips
        # from OCR/table extraction artifacts (e.g., "(11,324)" inventory).
        for asset_field in ["Total Assets", "Current Assets", "Inventory", "Accounts Receivable", "Cash"]:
            v = _to_numeric_financial(row.get(asset_field))
            if v is not None:
                row[asset_field] = abs(float(v))
    elif sheet == "Cash Flow":
        if analysis_mode == "t12m":
            row["Fiscal Period"] = "T12M"
        if analysis_mode == "t12m":
            ocf = _strict_pick_t12m_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Operating Cash Flow"], page_min=page_min, page_max=page_max, field_key="operating_cash_flow"
            )
            capex = _strict_pick_t12m_from_pdf(pdf, STRICT_LABEL_PRIORITY["CapEx"], page_min=page_min, page_max=page_max, field_key="capex")
        else:
            ocf = _strict_pick_from_pdf(
                pdf, STRICT_LABEL_PRIORITY["Operating Cash Flow"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="operating_cash_flow"
            )
            capex = _strict_pick_from_pdf(pdf, STRICT_LABEL_PRIORITY["CapEx"], selected_year, year_order, page_min=page_min, page_max=page_max, field_key="capex")
        if capex.get("value") is None:
            capex = _strict_pick_capex_from_cashflow_context(
                pdf,
                selected_year,
                year_order,
                page_min=page_min,
                page_max=page_max,
            )
        if capex.get("value") is None:
            capex_regex = _regex_info_from_pdf(
                pdf,
                [
                    ["capital", "expenditures"],
                    ["property", "plant", "equipment"],
                    ["property", "equipment"],
                    ["additions", "property"],
                    ["purchase", "property"],
                    ["purchases", "property"],
                    ["acquisition", "property"],
                    ["acquisitions", "property"],
                ],
            )
            if capex_regex.get("value") not in (None, "", "null", "None"):
                capex = capex_regex
        row["Operating Cash Flow"] = ocf.get("value")
        capex_val = _to_numeric_financial(capex.get("value"))
        # Normalize expense sign for ratio inputs: (200) -> 200
        row["CapEx"] = abs(float(capex_val)) if capex_val is not None else None
        row["Free Cash Flow"] = (
            row["Operating Cash Flow"] - row["CapEx"]
            if row["Operating Cash Flow"] is not None and row["CapEx"] is not None
            else None
        )
        metric_meta["Operating Cash Flow"] = ocf
        metric_meta["CapEx"] = capex
        if analysis_mode == "t12m" and row["Operating Cash Flow"] is None:
            errors.append("T12M mode selected but 4-quarter Cash Flow series not found. Metrics marked Incomplete.")
        metric_meta["Free Cash Flow"] = {
            "value": row["Free Cash Flow"],
            "source": "derived",
            "match_type": "derived",
            "confidence_score": 60 if row["Free Cash Flow"] is not None else 0,
            "year": selected_year,
        }
    # Keep extracted values in raw statement units exactly as shown in filing tables.
    # Scaling is applied only at display layer.
    for k, v in list(row.items()):
        if k in {"Selected Year", "Detected Unit"}:
            continue
        if isinstance(v, (int, float)) and not pd.isna(v):
            row[k] = float(v)

    # Cross-statement validations (sheet-level partial checks).
    if sheet == "Income Statement" and row.get("Revenue") is not None and row.get("Net Income") is not None:
        if float(row["Revenue"]) < float(row["Net Income"]):
            errors.append("possible extraction error: Revenue is lower than Net Income.")
    if sheet == "Balance Sheet":
        ta, tl, eq = row.get("Total Assets"), row.get("Total Liabilities"), row.get("Shareholders' Equity")
        if ta is not None and tl is not None and eq is not None:
            if abs(float(ta) - (float(tl) + float(eq))) > max(1.0, abs(float(ta)) * 0.05):
                errors.append("balance sheet mismatch: Total Assets != Total Liabilities + Equity.")
    row["_metric_meta"] = metric_meta
    return row, errors


@st.cache_data(show_spinner=False, ttl=600, hash_funcs={Path: lambda p: str(p)})
def _strict_financial_sheet_row_cached(
    sheet: str, pdf: Path, cache_version: str = EXTRACTION_CACHE_VERSION
) -> Tuple[Dict[str, Any], List[str]]:
    _ = cache_version
    return _strict_financial_sheet_row(sheet, pdf)


@st.cache_data(show_spinner=False, ttl=600)
def _financial_parser_rows_cached(pdf_path_str: str, cache_version: str = EXTRACTION_CACHE_VERSION) -> List[Dict[str, Any]]:
    if parse_financial_statements is None:
        return []
    rows = parse_financial_statements(str(pdf_path_str))
    return rows if isinstance(rows, list) else []


def _regex_info_from_pdf(pdf: Path, token_groups: List[List[str]]) -> Dict[str, Any]:
    lines = _pdf_text_lines(str(pdf), max_pages=_strict_scan_page_limit(pdf))
    num_rx = re.compile(r"(\$?\d[\d,]*(?:\.\d+)?\s*(?:x|%|bps|bp|million|billion|m|bn)?)", re.IGNORECASE)
    for group in token_groups:
        tokens = [t.lower() for t in group]
        for idx, item in enumerate(lines):
            line_l = item["text"].lower()
            if all(t in line_l for t in tokens):
                candidates: List[Dict[str, Any]] = [item]
                candidates.extend(lines[idx + 1 : idx + 4])
                for cand in candidates:
                    m = num_rx.search(cand["text"])
                    if m:
                        snippet = item["text"] if cand is item else f"{item['text']} | {cand['text']}"
                        return {
                            "value": m.group(1).strip(),
                            "confidence": 0.42,
                            "page": cand["page"],
                            "snippet": snippet[:220],
                        }
    return {"value": None, "confidence": None, "page": None, "snippet": None}


def _fallback_token_groups_for_column(col: str) -> List[List[str]]:
    c = str(col).strip().lower()
    presets: Dict[str, List[List[str]]] = {
        "revenue": [["revenue"], ["sales"]],
        "ebitda": [["ebitda"]],
        "ebit": [["ebit"], ["operating", "income"]],
        "net income": [["net", "income"]],
        "interest expense": [["interest", "expense"]],
        "total debt": [["total", "debt"], ["debt"]],
        "cash": [["cash"]],
        "current assets": [["current", "assets"]],
        "current liabilities": [["current", "liabilities"]],
        "operating cash flow": [["operating", "cash", "flow"], ["cash", "flow", "from", "operations"]],
        "capex": [["capex"], ["capital", "expenditures"]],
        "free cash flow": [["free", "cash", "flow"]],
        "loan amount": [["loan", "amount"], ["facility", "amount"], ["commitment"]],
        "margin": [["margin"], ["spread"]],
        "minimum interest coverage": [["minimum", "interest", "coverage"], ["interest", "coverage"]],
        "maximum leverage ratio": [["maximum", "leverage"], ["leverage", "ratio"]],
    }
    for key, groups in presets.items():
        if key in c:
            return groups
    tokens = [t for t in re.findall(r"[a-z0-9]+", c) if len(t) > 2 and t not in {"ratio", "value", "type", "required", "reported", "calculated"}]
    if not tokens:
        return []
    groups: List[List[str]] = [tokens[:3]]
    if len(tokens) >= 2:
        groups.append(tokens[:2])
    groups.append([tokens[0]])
    return groups


def _build_consolidated_view(sheet_map: Dict[str, pd.DataFrame]) -> pd.DataFrame:
    all_cols: List[str] = ["Sheet"]
    for _, df in sheet_map.items():
        for col in df.columns:
            if col not in all_cols:
                all_cols.append(col)

    frames: List[pd.DataFrame] = []
    for sheet_name, df in sheet_map.items():
        t = df.copy()
        t.insert(0, "Sheet", sheet_name)
        frames.append(t.reindex(columns=all_cols))

    if not frames:
        return pd.DataFrame(columns=["Sheet"])
    return _order_display_columns(pd.concat(frames, ignore_index=True))


def _order_display_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    # Currency is denomination (USD). Reporting scale belongs in "Detected Unit".
    out["Currency"] = "USD"
    cols = list(df.columns)
    tail = [c for c in META_TAIL_ORDER if c in cols]
    body = [c for c in out.columns if c not in META_TAIL_ORDER]
    if "Currency" in body:
        body = [c for c in body if c != "Currency"] + ["Currency"]
    return out.reindex(columns=body + tail)


def _norm_key(s: str) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(s).strip().lower()).strip()


def _extract_first_value(
    df: pd.DataFrame,
    candidates: List[str],
    sheet_names: Optional[List[str]] = None,
    preferred_doc_types: Optional[List[str]] = None,
    selection_rule: str = "highest_confidence",
    preferred_year: Optional[int] = None,
) -> Any:
    if df is None or df.empty:
        return None
    work = df
    if sheet_names and "Sheet" in df.columns:
        allowed = {_norm_key(s) for s in sheet_names}
        work = df[df["Sheet"].astype(str).apply(lambda x: _norm_key(x) in allowed)]
    if preferred_doc_types and "Detected Type" in work.columns:
        allowed_types = {_norm_key(t) for t in preferred_doc_types}
        scoped = work[work["Detected Type"].astype(str).apply(lambda x: _norm_key(x) in allowed_types)]
        # Do not hard-fail extraction when doc-type labels mismatch between modules.
        # Keep working set unfiltered as fallback if strict type scope is empty.
        if not scoped.empty:
            work = scoped
    if "Selected Year" in work.columns:
        years = pd.to_numeric(work["Selected Year"], errors="coerce")
        # Avoid mixing year-scoped rows with unlabeled rows (often stale OCR).
        if years.notna().any():
            work = work[years.notna()].copy()
            years = pd.to_numeric(work["Selected Year"], errors="coerce")
        used_preferred_year = False
        if preferred_year is not None:
            scoped_year = work[years == float(int(preferred_year))]
            if not scoped_year.empty:
                work = scoped_year
                used_preferred_year = True
        if not used_preferred_year:
            valid_years = years.dropna()
            if not valid_years.empty:
                latest_year = int(valid_years.max())
                scoped_latest = work[pd.to_numeric(work["Selected Year"], errors="coerce") == float(latest_year)]
                if not scoped_latest.empty:
                    work = scoped_latest
    if work.empty:
        return None

    norm_to_col: Dict[str, str] = {_norm_key(c): c for c in work.columns}
    resolved_col: Optional[str] = None
    for c in candidates:
        nc = _norm_key(c)
        if nc in norm_to_col:
            resolved_col = norm_to_col[nc]
            break
    if resolved_col is None:
        for c in candidates:
            nc = _norm_key(c)
            for existing_norm, existing_col in norm_to_col.items():
                if nc and (nc in existing_norm or existing_norm in nc):
                    resolved_col = existing_col
                    break
            if resolved_col:
                break
    if not resolved_col:
        return None
    ranked = work.copy()
    if "Confidence" in ranked.columns:
        ranked["_conf"] = pd.to_numeric(ranked["Confidence"], errors="coerce").fillna(-1.0)
    else:
        ranked["_conf"] = -1.0
    if "Page" in ranked.columns:
        ranked["_page"] = pd.to_numeric(ranked["Page"], errors="coerce").fillna(-1.0)
    else:
        ranked["_page"] = -1.0
    if "Source Document" in ranked.columns:
        ranked["_doc"] = ranked["Source Document"].astype(str)
    else:
        ranked["_doc"] = ""
    period_cols = ["Reporting Period Date", "Test Dates", "Maturity Date", "Amendment Effective Date", "effective_date"]
    ranked["_period"] = pd.NaT
    for pc in period_cols:
        if pc in ranked.columns:
            parsed = pd.to_datetime(ranked[pc], errors="coerce")
            ranked["_period"] = ranked["_period"].fillna(parsed)
    ranked["_period_epoch"] = ranked["_period"].apply(lambda x: x.value if pd.notna(x) else -1)

    if selection_rule == "latest_page":
        ranked = ranked.sort_values(by=["_page", "_period_epoch", "_conf"], ascending=[False, False, False])
    elif selection_rule == "first_document":
        ranked = ranked.sort_values(by=["_doc", "_period_epoch", "_conf", "_page"], ascending=[True, False, False, False])
    else:
        ranked = ranked.sort_values(by=["_period_epoch", "_conf", "_page"], ascending=[False, False, False])
    vals = ranked[resolved_col].tolist()
    for v in vals:
        if v not in (None, "", "null", "None") and not (isinstance(v, float) and pd.isna(v)):
            return v
    return None


def _extract_num(
    df: pd.DataFrame,
    candidates: List[str],
    sheet_names: Optional[List[str]] = None,
    preferred_doc_types: Optional[List[str]] = None,
    selection_rule: str = "highest_confidence",
    preferred_year: Optional[int] = None,
) -> Optional[float]:
    return _to_numeric_financial(_extract_first_value(df, candidates, sheet_names, preferred_doc_types, selection_rule, preferred_year))


def _infer_principal_repayment_from_value(value: Any) -> Optional[float]:
    num = _to_numeric_financial(value)
    if num is None:
        return None
    # "Net Debt Issued/(Repaid)" style fields are negative when debt is repaid.
    if num < 0:
        return abs(float(num))
    return 0.0


def _extract_principal_repayment(
    df: pd.DataFrame,
    preferred_doc_types: Optional[List[str]] = None,
    selection_rule: str = "highest_confidence",
) -> Optional[float]:
    explicit = _extract_num(
        df,
        [
            "Principal Repayment",
            "Principal Repayments",
            "Debt Repayment",
            "Repayment of Debt",
            "Repayments of Long-term Debt",
        ],
        sheet_names=["Cash Flow", "Projected Debt"],
        preferred_doc_types=preferred_doc_types,
        selection_rule=selection_rule,
    )
    if explicit is not None:
        return max(float(explicit), 0.0)
    inferred_raw = _extract_first_value(
        df,
        [
            "Net Debt Issued/(Repaid)",
            "Net Debt Issued Repaid",
            "Net Borrowings (Repayments)",
            "Debt Issued/(Repaid)",
        ],
        sheet_names=["Cash Flow", "Projected Debt"],
        preferred_doc_types=preferred_doc_types,
        selection_rule=selection_rule,
    )
    return _infer_principal_repayment_from_value(inferred_raw)


def _extract_bool(
    df: pd.DataFrame,
    candidates: List[str],
    sheet_names: Optional[List[str]] = None,
    preferred_doc_types: Optional[List[str]] = None,
    selection_rule: str = "highest_confidence",
    preferred_year: Optional[int] = None,
) -> Optional[bool]:
    v = _extract_first_value(df, candidates, sheet_names, preferred_doc_types, selection_rule, preferred_year)
    if v is None:
        return None
    s = str(v).strip().lower()
    if s in {"yes", "y", "true", "1", "present"}:
        return True
    if s in {"no", "n", "false", "0", "absent"}:
        return False
    return None


def _build_credit_risk_tables(
    consolidated_df: pd.DataFrame,
    applied_doc_type: str,
    bundle_doc_types: Optional[List[str]] = None,
    selection_rule: str = "highest_confidence",
    locked_year: Optional[int] = None,
) -> Dict[str, pd.DataFrame]:
    df = consolidated_df.copy() if consolidated_df is not None else pd.DataFrame()
    today = datetime.now().strftime("%Y-%m-%d")
    dtypes = bundle_doc_types if bundle_doc_types else [applied_doc_type]
    dtype_label = ", ".join(sorted({str(x) for x in dtypes if x})) if dtypes else str(applied_doc_type)
    cov_docs = [x for x in dtypes if x in {"Credit Agreement", "Amendment", "Covenant Compliance Certificate", "Compliance Certificate"}]
    _fin_scope_types = {
        "Financial Statements",
        "10-K",
        "10-Q",
        "Annual Report",
        "Other Financial Filing",
        "Income Statement",
        "Balance Sheet",
        "Cash Flow",
        "Covenant Compliance Certificate",
        "Compliance Certificate",
    }
    fin_docs = [x for x in dtypes if x in _fin_scope_types]
    deal_docs = [x for x in dtypes if x in {"Term Sheet", "Credit Agreement", "Amendment"}]
    coll_docs = [x for x in dtypes if x in {"Security Agreement", "Borrowing Base", "Term Sheet"}]
    fee_docs = [x for x in dtypes if x in {"Fee Letter", "Term Sheet", "Amendment"}]
    fcast_docs = [x for x in dtypes if x in {"Forecast Model", "Forecast/Projections"}]
    if not cov_docs:
        cov_docs = dtypes
    # Keep risk financial pulls scoped to financial-capable documents when available.
    # If none were detected, gracefully fall back to bundle types instead of hard-NULL scope.
    if not fin_docs:
        fin_docs = dtypes
    if not deal_docs:
        deal_docs = dtypes
    if not coll_docs:
        coll_docs = dtypes
    if not fee_docs:
        fee_docs = dtypes
    if not fcast_docs:
        fcast_docs = dtypes

    def _extract_num_relaxed(
        candidates: List[str],
        sheet_names: Optional[List[str]] = None,
        preferred_doc_types: Optional[List[str]] = None,
    ) -> Optional[float]:
        val = _extract_num(
            df,
            candidates,
            sheet_names=sheet_names,
            preferred_doc_types=preferred_doc_types,
            selection_rule=selection_rule,
            preferred_year=locked_year,
        )
        if val is not None:
            return val
        # Fallback: when sheet labels are noisy, retry without sheet filter.
        return _extract_num(
            df,
            candidates,
            sheet_names=None,
            preferred_doc_types=preferred_doc_types,
            selection_rule=selection_rule,
            preferred_year=locked_year,
        )

    min_icr = _extract_num(
        df, ["Minimum Interest Coverage", "Borrower Interest Coverage", "Interest Coverage Ratio"],
        sheet_names=["Financial Covenants", "Calculated Ratios", "Reported Financials"],
        preferred_doc_types=cov_docs,
        selection_rule=selection_rule,
        preferred_year=locked_year,
    )
    max_tl = _extract_num(
        df, ["Maximum Leverage Ratio", "Total Leverage", "Updated Leverage Limits"],
        sheet_names=["Financial Covenants", "Covenant Changes", "Calculated Ratios", "Reported Financials"],
        preferred_doc_types=cov_docs,
        selection_rule=selection_rule,
        preferred_year=locked_year,
    )
    max_sl = _extract_num(df, ["Maximum Senior Leverage Ratio", "Senior Leverage Ratio"], sheet_names=["Financial Covenants"], preferred_doc_types=cov_docs, selection_rule=selection_rule, preferred_year=locked_year)
    min_fccr = _extract_num(
        df, ["Fixed Charge Coverage", "Fixed Charge Coverage Ratio", "Reset Coverage Ratios"],
        sheet_names=["Financial Covenants", "Covenant Changes"],
        preferred_doc_types=cov_docs,
        selection_rule=selection_rule,
        preferred_year=locked_year,
    )
    min_liq = _extract_num(df, ["Minimum Liquidity", "Liquidity"], sheet_names=["Financial Covenants", "Reported Financials"], preferred_doc_types=cov_docs, selection_rule=selection_rule, preferred_year=locked_year)
    ebitda_adj = _extract_first_value(df, ["EBITDA Add-backs", "Temporary Add-backs", "EBITDA Exclusions"], sheet_names=["Covenant Definitions", "EBITDA Adjustments"], preferred_doc_types=cov_docs, selection_rule=selection_rule, preferred_year=locked_year)
    cov_freq = _extract_first_value(df, ["Testing Frequency", "Test Frequency"], sheet_names=["Testing Terms"], preferred_doc_types=cov_docs, selection_rule=selection_rule, preferred_year=locked_year)
    default_clause = _extract_first_value(df, ["Cross-default Clauses", "Material Adverse Change Clauses"], sheet_names=["Default Triggers"], preferred_doc_types=cov_docs, selection_rule=selection_rule, preferred_year=locked_year)

    revised_lev = _extract_num(df, ["Updated Leverage Limits", "Revised Covenant Thresholds"], sheet_names=["Covenant Changes"], preferred_doc_types=["Amendment"], selection_rule=selection_rule, preferred_year=locked_year)
    revised_cov = _extract_num(df, ["Reset Coverage Ratios", "Revised Covenant Thresholds"], sheet_names=["Covenant Changes"], preferred_doc_types=["Amendment"], selection_rule=selection_rule, preferred_year=locked_year)
    covenant_holiday = _extract_bool(df, ["Covenant Holiday Terms", "Covenant Holiday"], sheet_names=["EBITDA Adjustments"], preferred_doc_types=["Amendment"], selection_rule=selection_rule, preferred_year=locked_year)
    maturity_extension = _extract_first_value(df, ["Extended Maturity Dates", "Maturity Changes"], sheet_names=["Maturity Changes"], preferred_doc_types=["Amendment"], selection_rule=selection_rule, preferred_year=locked_year)
    revised_pricing = _extract_first_value(df, ["Revised Pricing", "Margin Step-ups", "Margin"], preferred_doc_types=fee_docs, selection_rule=selection_rule, preferred_year=locked_year)
    modified_defs = _extract_first_value(df, ["Modified Definitions", "Adjusted EBITDA Logic"], preferred_doc_types=["Amendment"], selection_rule=selection_rule, preferred_year=locked_year)

    covenant_thresholds = pd.DataFrame(
        [
            {
                "effective_date": today,
                "document_type": dtype_label,
                "min_interest_coverage_ratio": revised_cov if revised_cov is not None else min_icr,
                "max_total_leverage_ratio": revised_lev if revised_lev is not None else max_tl,
                "max_senior_leverage_ratio": max_sl,
                "min_fixed_charge_coverage_ratio": min_fccr,
                "min_liquidity_requirement": min_liq,
                "ebitda_definition_adjustments": ebitda_adj,
                "covenant_testing_frequency": cov_freq,
                "default_trigger_clauses": default_clause,
                "is_amendment_override": applied_doc_type == "Amendment",
            }
        ]
    )

    deal_terms = pd.DataFrame(
        [
            {
                "effective_date": today,
                "loan_amount": _extract_num(df, ["Loan Amount", "Facility Amount", "Facility Amount Total Commitment"], sheet_names=["Facility Overview"], preferred_doc_types=deal_docs, preferred_year=locked_year),
                "interest_margin": _extract_num(df, ["Margin", "Spread (bps)"], sheet_names=["Pricing"], preferred_doc_types=deal_docs, preferred_year=locked_year),
                "tenor": _extract_first_value(df, ["Tenor"], sheet_names=["Facility Overview"], preferred_doc_types=deal_docs, preferred_year=locked_year),
                "amortization_schedule": _extract_first_value(df, ["Amortization"], sheet_names=["Facility Overview"], preferred_doc_types=deal_docs, preferred_year=locked_year),
                "covenant_summary_levels": _extract_first_value(df, ["Headline Covenant Levels", "Covenant Summary"], sheet_names=["Covenant Summary"], preferred_doc_types=deal_docs, preferred_year=locked_year),
                "pricing_grid": _extract_first_value(df, ["Pricing Grid", "Margin Grid"], sheet_names=["Pricing", "Pricing Grid"], preferred_doc_types=deal_docs, preferred_year=locked_year),
            }
        ]
    )

    reported_metrics = pd.DataFrame(
        [
            {
                "reporting_period_date": _extract_first_value(df, ["Test Dates", "Reporting Period Date"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year) or today,
                "reported_ebitda": _extract_num(df, ["EBITDA", "Reported EBITDA"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "reported_interest_expense": _extract_num(df, ["Interest Expense", "Reported Interest Expense"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "reported_total_debt": _extract_num(df, ["Total Debt", "Reported Total Debt"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "reported_net_debt": _extract_num(df, ["Net Debt", "Reported Net Debt"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "reported_liquidity": _extract_num(df, ["Liquidity", "Minimum Liquidity"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "reported_interest_coverage_ratio": _extract_num(df, ["Borrower Interest Coverage", "Reported Coverage Ratio", "Interest Coverage Ratio"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "reported_leverage_ratio": _extract_num(df, ["Borrower Leverage", "Reported Leverage Ratio", "Leverage Ratio"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
                "compliance_confirmation_status": _extract_first_value(df, ["Officer Sign-off", "Compliance Confirmation Status"], preferred_doc_types=["Covenant Compliance Certificate", "Compliance Certificate"], preferred_year=locked_year),
            }
        ]
    )

    fa_short_debt = _extract_num_relaxed(
        ["Short-term Debt", "Short Term Debt", "Short-term Borrowings"],
        sheet_names=["Balance Sheet", "Reported Financials"],
        preferred_doc_types=fin_docs,
    )
    fa_current_portion_debt = _extract_num_relaxed(
        ["Current Portion of Long-term Debt", "Current Portion of Long-Term Obligations", "Current Maturities of Long-term Debt"],
        sheet_names=["Balance Sheet", "Reported Financials"],
        preferred_doc_types=fin_docs,
    )
    fa_long_debt = _extract_num_relaxed(
        ["Long-term Debt", "Long Term Debt", "Long-term Debt Non-Current"],
        sheet_names=["Balance Sheet", "Reported Financials"],
        preferred_doc_types=fin_docs,
    )
    # Defensive dedupe for overlapping mappings between short-term debt and
    # current portion of long-term debt in noisy OCR tables.
    try:
        if (
            fa_short_debt is not None
            and fa_current_portion_debt is not None
            and abs(float(fa_short_debt) - float(fa_current_portion_debt)) <= max(1.0, abs(float(fa_current_portion_debt)) * 0.001)
        ):
            fa_short_debt = None
    except Exception:
        pass
    fa_total_debt_direct = _extract_num_relaxed(
        ["Total Debt", "Total Borrowings", "Notes Payable"],
        sheet_names=["Balance Sheet", "Reported Financials"],
        preferred_doc_types=fin_docs,
    )
    # Policy debt definition: ST financial debt + LT financial debt.
    # Use current portion of LT debt as primary ST bucket; fallback to short-term debt.
    fa_st_financial = fa_current_portion_debt if fa_current_portion_debt is not None else fa_short_debt
    debt_parts = [x for x in [fa_st_financial, fa_long_debt] if x is not None]
    fa_total_debt = sum(debt_parts) if debt_parts else fa_total_debt_direct

    financial_actuals = pd.DataFrame(
        [
            {
                # Keep financial_actuals strictly on actual/historical statements.
                "revenue": _extract_num(
                    df,
                    ["Revenue", "Total Revenue", "Net Sales"],
                    sheet_names=["Income Statement", "Reported Financials"],
                    preferred_doc_types=fin_docs,
                    preferred_year=locked_year,
                ),
                "cogs": _extract_num_relaxed(
                    ["COGS", "Cost of Goods Sold", "Cost of Revenue", "Cost of Revenues", "Cost of Sales"],
                    sheet_names=["Income Statement", "Reported Financials"],
                    preferred_doc_types=fin_docs,
                ),
                "ebitda": _extract_num_relaxed(
                    ["EBITDA"],
                    sheet_names=["Income Statement", "Reported Financials"],
                    preferred_doc_types=fin_docs,
                ),
                "ebit": _extract_num_relaxed(
                    ["EBIT", "Operating Income", "Income from Operations", "Loss from Operations"],
                    sheet_names=["Income Statement", "Reported Financials"],
                    preferred_doc_types=fin_docs,
                ),
                "interest_expense": _extract_num_relaxed(
                    ["Interest Expense"],
                    sheet_names=["Income Statement", "Reported Financials"],
                    preferred_doc_types=fin_docs,
                ),
                "net_income": _extract_num_relaxed(
                    ["Net Income", "Net Income (Loss)", "Net Loss"],
                    sheet_names=["Income Statement", "Reported Financials"],
                    preferred_doc_types=fin_docs,
                ),
                "total_debt": fa_total_debt,
                "short_term_debt": fa_short_debt,
                "current_portion_long_term_debt": fa_current_portion_debt,
                "long_term_debt": fa_long_debt,
                "cash": _extract_num_relaxed(["Cash"], sheet_names=["Balance Sheet", "Reported Financials"], preferred_doc_types=fin_docs),
                "total_assets": _extract_num_relaxed(["Total Assets"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "total_liabilities": _extract_num_relaxed(["Total Liabilities"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "equity": _extract_num_relaxed(["Shareholders' Equity", "Total Equity", "Total Stockholders' Equity"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "current_assets": _extract_num_relaxed(["Current Assets"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "current_liabilities": _extract_num_relaxed(["Current Liabilities"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "inventory": _extract_num_relaxed(["Inventory", "Inventories", "Total Inventory"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "accounts_receivable": _extract_num_relaxed(["Accounts Receivable", "Trade Receivables", "Receivables, net"], sheet_names=["Balance Sheet"], preferred_doc_types=fin_docs),
                "depreciation_amortization": _extract_num_relaxed(["Depreciation and Amortization", "Depreciation", "Amortization", "D&A"], sheet_names=["Income Statement", "Cash Flow"], preferred_doc_types=fin_docs),
                "operating_cash_flow": _extract_num_relaxed(["Operating Cash Flow"], sheet_names=["Cash Flow"], preferred_doc_types=fin_docs),
                "capital_expenditures": _extract_num_relaxed(["CapEx"], sheet_names=["Cash Flow"], preferred_doc_types=fin_docs),
                "free_cash_flow": _extract_num_relaxed(["Free Cash Flow"], sheet_names=["Cash Flow"], preferred_doc_types=fin_docs),
            }
        ]
    )

    forecast_metrics = pd.DataFrame(
        [
            {
                "projected_revenue": _extract_num(df, ["Projected Revenue"], preferred_doc_types=fcast_docs, preferred_year=locked_year),
                "projected_ebitda": _extract_num(df, ["Projected EBITDA"], preferred_doc_types=fcast_docs, preferred_year=locked_year),
                "projected_interest_expense": _extract_num(df, ["Projected Interest Expense", "Interest Expense"], preferred_doc_types=fcast_docs, preferred_year=locked_year),
                "projected_total_debt": _extract_num(df, ["Future Debt Balances", "Projected Total Debt"], preferred_doc_types=fcast_docs, preferred_year=locked_year),
            }
        ]
    )

    collateral_data = pd.DataFrame(
        [
            {
                "collateral_type": _extract_first_value(df, ["Asset Type", "Collateral Type"], preferred_doc_types=coll_docs, preferred_year=locked_year),
                "collateral_value": _extract_num(df, ["Collateral Value", "Eligible Inventory Value"], preferred_doc_types=coll_docs, preferred_year=locked_year),
                "lien_priority": _extract_first_value(df, ["Lien Priority"], preferred_doc_types=coll_docs, preferred_year=locked_year),
                "guarantee_presence": _extract_bool(df, ["Guarantor Entities", "Guarantee Presence"], preferred_doc_types=coll_docs, preferred_year=locked_year),
            }
        ]
    )

    pricing_data = pd.DataFrame(
        [
            {
                "upfront_fee": _extract_num(df, ["Upfront Fee %", "Upfront Fee"], preferred_doc_types=fee_docs, preferred_year=locked_year),
                "commitment_fee": _extract_num(df, ["Commitment Fee %", "Commitment Fee"], preferred_doc_types=fee_docs, preferred_year=locked_year),
                "margin_grid_levels": _extract_first_value(df, ["Margin Grid", "Margin Step-ups"], preferred_doc_types=fee_docs, preferred_year=locked_year),
                "amendment_fees": _extract_num(df, ["Amendment Fees", "Agency Fee"], preferred_doc_types=fee_docs, preferred_year=locked_year),
                "revised_pricing": revised_pricing,
                "maturity_extension": maturity_extension,
                "modified_definitions": modified_defs,
                "covenant_holiday": covenant_holiday,
            }
        ]
    )

    return {
        "covenant_thresholds": covenant_thresholds,
        "deal_terms": deal_terms,
        "reported_metrics": reported_metrics,
        "financial_actuals": financial_actuals,
        "forecast_metrics": forecast_metrics,
        "collateral_data": collateral_data,
        "pricing_data": pricing_data,
    }


def _risk_level_from_score(score: float) -> str:
    if score <= 40:
        return "Low"
    if score <= 70:
        return "Medium"
    return "High"


def _risk_style(level: str) -> Dict[str, str]:
    styles = {
        "Low": {"bg": "#E6F4EA", "text": "#1E7E34"},
        "Medium": {"bg": "#FFE8CC", "text": "#B45309"},
        "High": {"bg": "#FEE2E2", "text": "#B91C1C"},
        "Critical": {"bg": "#8B0000", "text": "#FFFFFF"},
        "Unknown": {"bg": "#F8FAFC", "text": "#475569"},
        "Incomplete": {"bg": "#FFFBE6", "text": "#8A6D3B"},
    }
    return styles.get(level, styles["Medium"])


def _score_from_level(level: str) -> int:
    return {"Low": 20, "Medium": 60, "High": 90, "Critical": 100}.get(level, 60)


def _render_risk_html_table(df: pd.DataFrame, risk_col: str, cols: List[str]) -> None:
    if df is None or df.empty:
        st.info("No rows to display.")
        return
    view = df.copy()
    use_cols = [c for c in cols if c in view.columns]
    if not use_cols:
        use_cols = list(view.columns)
    rows_html: List[str] = []
    metric_help = {
        "Current Ratio": "Current assets divided by current liabilities; short-term liquidity strength.",
        "Quick Ratio": "Cash plus receivables divided by current liabilities; excludes less liquid assets.",
        "Debt to Equity": "Total debt divided by shareholder equity; capital structure leverage.",
        "Debt to EBITDA": "Total debt divided by EBITDA; years of EBITDA to repay debt (approx.).",
        "DSCR": "Debt Service Coverage Ratio: EBITDA divided by interest plus principal debt service.",
        "Interest Coverage": "EBIT divided by interest expense; ability to service borrowing cost.",
        "Gross Margin": "Revenue minus COGS, divided by revenue; core profitability before operating expenses.",
        "Net Margin": "Net income divided by revenue; profit retained per unit of sales.",
        "ROA": "Return on Assets: net income divided by total assets.",
        "ROE": "Return on Equity: net income divided by shareholder equity.",
        "Operating Cash Flow Ratio": "Operating cash flow divided by current liabilities; cash liquidity coverage.",
        "Free Cash Flow": "Operating cash flow minus capital expenditure; discretionary cash after reinvestment.",
    }
    for _, row in view.iterrows():
        level = str(row.get(risk_col, "Medium"))
        s = _risk_style(level)
        bg = s["bg"]
        fg = s["text"]
        cells = []
        for c in use_cols:
            v = row.get(c)
            if c == "Color Code":
                color = str(v) if v else "#FFF4E5"
                txt = (
                    f"<span style='display:inline-flex;align-items:center;gap:8px;'>"
                    f"<span style='width:14px;height:14px;border-radius:3px;background:{color};border:1px solid #cbd5e1;display:inline-block;'></span>"
                    f"<span style='font-weight:700;'>Color</span></span>"
                )
                cells.append(f"<td style='background:{bg};color:{fg};'>{txt}</td>")
                continue
            if v is None or (isinstance(v, float) and pd.isna(v)):
                txt = "NULL"
            elif isinstance(v, float):
                if c.lower().endswith("%"):
                    txt = f"{v:.1f}%"
                else:
                    txt = (f"{int(v):,}" if float(v).is_integer() else f"{v:,.3f}".rstrip("0").rstrip("."))
            elif isinstance(v, int):
                txt = f"{v:,}"
            else:
                txt = html.escape(str(v))
            if c == "Metric":
                desc = metric_help.get(str(v), "")
                if desc:
                    txt = f"<span class='metric-help' data-tip='{html.escape(desc)}'>{txt}</span>"
            cells.append(f"<td style='background:{bg};color:{fg};'>{txt}</td>")
        rows_html.append(f"<tr>{''.join(cells)}</tr>")

    headers = "".join([f"<th>{html.escape(str(c))}</th>" for c in use_cols])
    st.markdown(
        f"""
        <div class="risk-table-wrap">
          <table class="risk-table">
            <thead><tr>{headers}</tr></thead>
            <tbody>{''.join(rows_html)}</tbody>
          </table>
        </div>
        """,
        unsafe_allow_html=True,
    )


def _canonical_doc_type(doc_type: str) -> str:
    d = (doc_type or "").strip().lower()
    aliases = {
        "compliance certificate": "Covenant Compliance Certificate",
        "covenant compliance certificate": "Covenant Compliance Certificate",
        "credit agreement": "Credit Agreement",
        "term sheet": "Term Sheet",
        "amendment": "Amendment",
        "financial statements": "Financial Statements",
        "income statement": "Financial Statements",
        "balance sheet (standalone)": "Financial Statements",
        "cash flow statement (standalone)": "Financial Statements",
        "10-k": "Financial Statements",
        "10-q": "Financial Statements",
        "annual report": "Financial Statements",
        "other financial filing": "Financial Statements",
        "forecast model": "Forecast/Projections",
        "security agreement": "Security Agreement",
        "fee letter": "Fee Letter",
    }
    return aliases.get(d, doc_type)


def _confidence_level_from_doc_types(doc_types: List[str]) -> str:
    s = {_canonical_doc_type(x) for x in doc_types}
    core = {
        "Credit Agreement",
        "Term Sheet",
        "Covenant Compliance Certificate",
        "Financial Statements",
    }
    if len(s & core) >= 4 and len(s) >= 6:
        return "Full Review"
    if len(s & core) >= 3:
        return "Enhanced"
    return "Basic"


REQUIRED_DOC_BUCKETS: Dict[str, List[str]] = {
    "financials": ["Financial Statements", "Covenant Compliance Certificate", "Compliance Certificate"],
    "agreements": ["Credit Agreement", "Term Sheet"],
    "collateral": ["Security Agreement", "Borrowing Base"],
    "debt_schedule": ["Term Sheet", "Credit Agreement", "Amendment"],
}


def _build_required_bucket_status(doc_types: List[str]) -> Dict[str, Any]:
    present = {_canonical_doc_type(x) for x in (doc_types or [])}
    missing: List[str] = []
    for bucket, options in REQUIRED_DOC_BUCKETS.items():
        if not any(opt in present for opt in options):
            missing.append(bucket)
    return {"missing_buckets": missing, "is_complete": len(missing) == 0}


def _stable_json_digest(obj: Any) -> str:
    try:
        payload = json.dumps(obj, sort_keys=True, default=str)
    except Exception:
        payload = str(obj)
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _single_context_hash(selected_file: str, extraction: Dict[str, Any], applied_doc_type: str) -> str:
    base = f"{selected_file}|{_stable_json_digest(extraction)}|{applied_doc_type}"
    return hashlib.sha256(base.encode("utf-8")).hexdigest()


def _detect_metric_conflicts(
    df: pd.DataFrame,
    preferred_doc_types: Optional[List[str]] = None,
    max_conflicts: int = 12,
) -> List[Dict[str, Any]]:
    if df is None or df.empty:
        return []
    work = df.copy()
    if preferred_doc_types and "Detected Type" in work.columns:
        allowed = {_norm_key(x) for x in preferred_doc_types}
        scoped = work[work["Detected Type"].astype(str).apply(lambda x: _norm_key(x) in allowed)]
        if not scoped.empty:
            work = scoped
    ignore_cols = {"Source Document", "Detected Type", "Sheet", "Confidence", "Page", "Open PDF Page", "Snippet"}
    conflicts: List[Dict[str, Any]] = []
    for col in work.columns:
        if col in ignore_cols:
            continue
        series = work[col].dropna()
        vals = [str(v).strip() for v in series.tolist() if str(v).strip() not in {"", "None", "null"}]
        uniq = sorted(set(vals))
        if len(uniq) > 1:
            conflicts.append({"metric": col, "values": uniq[:4], "count": len(uniq)})
        if len(conflicts) >= max_conflicts:
            break
    return conflicts


def _build_metric_map(df: pd.DataFrame, borrower_by_doc: Optional[Dict[str, str]] = None) -> Dict[str, List[Dict[str, Any]]]:
    if df is None or df.empty:
        return {}
    borrower_by_doc = borrower_by_doc or {}
    meta_cols = {"Source Document", "Detected Type", "Sheet", "Confidence", "Page", "Open PDF Page", "Snippet"}
    period_cols = ["Reporting Period Date", "Test Dates", "Maturity Date", "Amendment Effective Date"]
    metric_map: Dict[str, List[Dict[str, Any]]] = {}
    for _, row in df.iterrows():
        source_document = str(row.get("Source Document", "") or "")
        borrower = row.get("Borrower")
        if borrower in (None, "", "null", "None"):
            borrower = borrower_by_doc.get(source_document)
        statement_period = None
        for pc in period_cols:
            if pc in df.columns:
                pv = row.get(pc)
                if pv not in (None, "", "null", "None"):
                    statement_period = str(pv)
                    break
        for col in df.columns:
            if col in meta_cols:
                continue
            val = row.get(col)
            if val in (None, "", "null", "None"):
                continue
            key = f"{source_document}|{borrower or ''}|{statement_period or ''}|{col}"
            metric_map.setdefault(key, []).append(
                {
                    "source_document": source_document,
                    "borrower": borrower,
                    "statement_period": statement_period,
                    "metric_name": col,
                    "value": val,
                    "confidence": row.get("Confidence"),
                    "page": row.get("Page"),
                }
            )
    return metric_map


def _statement_type_from_sheet(sheet_name: str) -> Optional[str]:
    n = str(sheet_name or "").strip().lower()
    if n == "income statement":
        return "income_statement"
    if n == "balance sheet":
        return "balance_sheet"
    if n == "cash flow":
        return "cash_flow"
    return None


def _normalized_financial_metric_dataset(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame(columns=["metric_name", "statement_type", "fiscal_year", "fiscal_period", "value", "currency_scale"])
    meta_exclude = set(META_COLUMNS) | {"Sheet", "Source Document", "Detected Type", "Currency", "Selected Year", "Fiscal Period", "Analysis Mode", "Detected Unit", "Available Years"}
    out_rows: List[Dict[str, Any]] = []
    for _, row in df.iterrows():
        sheet = _statement_type_from_sheet(row.get("Sheet"))
        if sheet is None:
            continue
        fy = _to_numeric_financial(row.get("Selected Year"))
        if fy is None:
            continue
        fiscal_year = int(fy)
        fiscal_period = str(row.get("Fiscal Period") or "FY")
        currency_scale_raw = str(row.get("Detected Unit") or _detect_scale_from_combined_df(df) or "Units").strip().lower()
        if "thousand" in currency_scale_raw:
            currency_scale = "Thousands"
        elif "million" in currency_scale_raw:
            currency_scale = "Millions"
        elif "billion" in currency_scale_raw:
            currency_scale = "Billions"
        elif "trillion" in currency_scale_raw:
            currency_scale = "Trillions"
        else:
            currency_scale = "Units"
        for col in df.columns:
            if col in meta_exclude:
                continue
            val = _to_numeric_financial(row.get(col))
            if val is None:
                continue
            out_rows.append(
                {
                    "metric_name": str(col),
                    "statement_type": sheet,
                    "fiscal_year": fiscal_year,
                    "fiscal_period": fiscal_period,
                    "value": float(val),
                    "currency_scale": currency_scale,
                }
            )
    return pd.DataFrame(out_rows)


def _lock_analysis_dataset(df: pd.DataFrame, mode: str, specific_year: Optional[int]) -> Dict[str, Any]:
    out = {
        "locked_df": df.copy() if df is not None else pd.DataFrame(),
        "locked_year": None,
        "available_years": [],
        "warnings": [],
        "cross_year_error_flag": False,
        "t12m_status": "n/a",
        "years_per_statement": {},
        "critical_year_selection_failure": False,
    }
    if df is None or df.empty:
        out["warnings"].append("No extracted data found for analysis year locking.")
        return out

    work = df.copy()
    years: List[int] = []
    if "Selected Year" in work.columns:
        for v in work["Selected Year"].tolist():
            n = _to_numeric_financial(v)
            if n is not None:
                years.append(int(n))
    years = sorted(set(years))
    out["available_years"] = years
    if not years:
        out["warnings"].append("No fiscal years detected in extracted statements.")
        out["cross_year_error_flag"] = True
        return out

    mode_norm = (mode or "latest_available").strip().lower()
    if mode_norm == "specific_year":
        if specific_year is None:
            out["warnings"].append("Specific Year mode selected but no year provided.")
            out["cross_year_error_flag"] = True
            return out
        if int(specific_year) not in years:
            out["warnings"].append(f"Selected year {int(specific_year)} not found in document. Metrics marked Incomplete.")
            out["cross_year_error_flag"] = True
            out["locked_year"] = int(specific_year)
            out["locked_df"] = work.iloc[0:0].copy()
            return out
        locked_year = int(specific_year)
    else:
        locked_year = max(years)
        if years and locked_year != max(years):
            out["warnings"].append("Year selection engine failure.")
            out["critical_year_selection_failure"] = True
            out["cross_year_error_flag"] = True
    out["locked_year"] = locked_year

    def _is_statement_row(r: pd.Series) -> bool:
        return _statement_type_from_sheet(r.get("Sheet")) is not None

    statement_rows = work[work.apply(_is_statement_row, axis=1)]
    non_statement_rows = work[~work.apply(_is_statement_row, axis=1)]
    if "Selected Year" in statement_rows.columns:
        year_mask = statement_rows["Selected Year"].apply(lambda x: _to_numeric_financial(x) == float(locked_year))
        statement_rows = statement_rows[year_mask]
    else:
        statement_rows = statement_rows.iloc[0:0].copy()

    required = [("income_statement", "Income Statement"), ("balance_sheet", "Balance Sheet"), ("cash_flow", "Cash Flow")]
    years_per_statement: Dict[str, List[int]] = {}
    for stype, label in required:
        all_subset = work[work["Sheet"].astype(str).str.lower() == label.lower()] if not work.empty else pd.DataFrame()
        if not all_subset.empty and "Selected Year" in all_subset.columns:
            years_per_statement[stype] = sorted(
                {
                    int(_to_numeric_financial(v))
                    for v in all_subset["Selected Year"].tolist()
                    if _to_numeric_financial(v) is not None
                }
            )
        else:
            years_per_statement[stype] = []
        subset = statement_rows[statement_rows["Sheet"].astype(str).str.lower() == label.lower()] if not statement_rows.empty else pd.DataFrame()
        if subset.empty:
            out["warnings"].append(f"Selected fiscal year not found in {label}. Metrics marked Incomplete.")
    out["years_per_statement"] = years_per_statement

    if mode_norm == "t12m":
        out["t12m_status"] = "ok"
        if "Analysis Mode" in statement_rows.columns:
            is_cf = statement_rows["Sheet"].astype(str).str.lower().isin(["income statement", "cash flow"])
            t12m_cf = statement_rows[is_cf & (statement_rows["Analysis Mode"].astype(str).str.lower() == "t12m")]
            if t12m_cf.empty:
                out["warnings"].append("Insufficient quarterly data for T12M computation.")
                out["t12m_status"] = "insufficient"
                out["cross_year_error_flag"] = True
        else:
            out["warnings"].append("Insufficient quarterly data for T12M computation.")
            out["t12m_status"] = "insufficient"
            out["cross_year_error_flag"] = True

    if not statement_rows.empty and "Selected Year" in statement_rows.columns:
        used = sorted(
            {
                int(_to_numeric_financial(v))
                for v in statement_rows["Selected Year"].tolist()
                if _to_numeric_financial(v) is not None
            }
        )
        if len(used) > 1:
            out["warnings"].append("Cross-year data mismatch detected. Please reselect analysis year.")
            out["cross_year_error_flag"] = True

    out["locked_df"] = _order_display_columns(pd.concat([statement_rows, non_statement_rows], ignore_index=True))
    try:
        st.session_state["analysis_year_debug"] = {
            "analysis_mode": mode_norm,
            "locked_year": locked_year,
            "available_years_detected": years,
            "years_per_statement": years_per_statement,
            "cross_year_error_flag": out["cross_year_error_flag"],
            "t12m_status": out["t12m_status"],
        }
    except Exception:
        pass
    return out


def _should_invalidate_multi_bundle(bundle: Optional[Dict[str, Any]], current_upload_key: str, risk_mode: str) -> bool:
    if not bundle or risk_mode != "multi":
        return False
    return bundle.get("source") == "multi" and bundle.get("context_key") != current_upload_key


def _multi_block_reason(bundle: Dict[str, Any]) -> Optional[str]:
    if bundle.get("borrower_mismatch"):
        return "Documents appear to belong to different borrowers. Please upload documents for one borrower only."
    unresolved_required = bundle.get("unresolved_required_borrower_docs", [])
    if bundle.get("source") == "multi" and unresolved_required:
        return f"Borrower unresolved in required docs: {', '.join(unresolved_required)}. Resolve borrower for all required docs before scoring."
    unresolved_borrowers = bundle.get("unresolved_borrower_docs", [])
    if bundle.get("source") == "multi" and unresolved_borrowers:
        return f"Borrower unresolved in: {', '.join(unresolved_borrowers)}. Resolve borrower for all docs before scoring."
    completeness = bundle.get("completeness", {})
    missing = completeness.get("missing_buckets", [])
    if bundle.get("source") == "multi" and missing:
        return f"Missing required document buckets: {', '.join(missing)}. Upload required docs to continue."
    return None


def _build_credit_risk_bundle(files_uploaded: List[Any]) -> Dict[str, Any]:
    all_frames: List[pd.DataFrame] = []
    doc_types: List[str] = []
    processed_docs: List[str] = []
    failed_docs: List[str] = []
    borrower_by_doc: Dict[str, str] = {}
    unresolved_borrower_docs: List[str] = []
    doc_type_by_doc: Dict[str, str] = {}
    for up in files_uploaded:
        try:
            saved = _save_uploaded(up)
            res = extract_fields(saved)
            detected = _resolve_detected_type(saved, res["summary"].get("document_type"))
            canonical = _canonical_doc_type(detected)
            doc_types.append(canonical)
            processed_docs.append(saved.name)
            doc_type_by_doc[saved.name] = canonical
            mapped_type = "Compliance Certificate" if canonical == "Covenant Compliance Certificate" else detected
            mapped_type = mapped_type if mapped_type in DOC_TYPE_CONFIG else _guess_type_from_filename(saved)
            sheets = DOC_TYPE_CONFIG[mapped_type]["sheets"]
            sheet_map = {s: _build_sheet_cached(res["extraction"], s, str(saved)) for s in sheets}
            cons = _build_consolidated_view(sheet_map)
            cons.insert(0, "Source Document", saved.name)
            cons.insert(1, "Detected Type", canonical)
            all_frames.append(cons)
            borrower_name = _extract_borrower_from_extraction(res.get("extraction", {}))
            if borrower_name:
                borrower_by_doc[saved.name] = borrower_name
            else:
                fallback_name = _borrower_from_filename(saved.name)
                if fallback_name:
                    borrower_by_doc[saved.name] = fallback_name
                else:
                    unresolved_borrower_docs.append(saved.name)
        except Exception:
            failed_docs.append(getattr(up, "name", "unknown.pdf"))

    combined = _order_display_columns(pd.concat(all_frames, ignore_index=True)) if all_frames else pd.DataFrame()
    borrower_names = list(borrower_by_doc.values())
    borrower_mismatch = not _is_same_borrower_group(borrower_names) if len(borrower_names) > 1 else False
    required_type_set = set()
    for options in REQUIRED_DOC_BUCKETS.values():
        required_type_set.update({_canonical_doc_type(x) for x in options})
    unresolved_required_docs = [
        name for name in unresolved_borrower_docs if _canonical_doc_type(doc_type_by_doc.get(name, "")) in required_type_set
    ]
    if len(processed_docs) > 1 and len(unresolved_required_docs) > 0:
        borrower_mismatch = True
    completeness = _build_required_bucket_status(doc_types)
    metric_map = _build_metric_map(combined, borrower_by_doc)
    context_hash = hashlib.sha256(
        f"{'|'.join(sorted(processed_docs))}|{_stable_json_digest(doc_types)}|{_stable_json_digest(combined.to_dict(orient='records') if not combined.empty else [])}".encode("utf-8")
    ).hexdigest()
    return {
        "combined_df": combined,
        "doc_types": doc_types,
        "processed_docs": processed_docs,
        "confidence_level": _confidence_level_from_doc_types(doc_types),
        "failed_docs": failed_docs,
        "borrower_by_doc": borrower_by_doc,
        "doc_type_by_doc": doc_type_by_doc,
        "borrower_mismatch": borrower_mismatch,
        "unresolved_borrower_docs": unresolved_borrower_docs,
        "unresolved_required_borrower_docs": unresolved_required_docs,
        "completeness": completeness,
        "context_hash": context_hash,
        "metric_map": metric_map,
        "logic_version": RISK_BUNDLE_LOGIC_VERSION,
    }


def _risk_color_code(level: str) -> str:
    return {"Low": "#E6F4EA", "Medium": "#FFF4E5", "High": "#FDECEA", "Breach": "#8B0000", "Critical": "#8B0000"}.get(level, "#FFF4E5")


def _compute_credit_analysis_table(
    tables: Dict[str, pd.DataFrame], industry: str, region: str, stress_ebitda_pct: int
) -> Dict[str, Any]:
    cov = tables["covenant_thresholds"].iloc[0].to_dict()
    deal = tables["deal_terms"].iloc[0].to_dict()
    fin = tables["financial_actuals"].iloc[0].to_dict()
    col = tables["collateral_data"].iloc[0].to_dict()

    debt = _to_numeric_financial(fin.get("total_debt"))
    ebitda = _to_numeric_financial(fin.get("ebitda"))
    cash = _to_numeric_financial(fin.get("cash"))
    interest = _to_numeric_financial(fin.get("interest_expense"))
    capex = _to_numeric_financial(fin.get("capital_expenditures"))
    current_assets = _to_numeric_financial(fin.get("current_assets"))
    current_liabilities = _to_numeric_financial(fin.get("current_liabilities"))
    op_cf = _to_numeric_financial(fin.get("operating_cash_flow"))
    fcf = _to_numeric_financial(fin.get("free_cash_flow"))
    if fcf is None and op_cf is not None and capex is not None:
        fcf = op_cf - capex
    loan_amount = _to_numeric_financial(deal.get("loan_amount"))
    collateral_val = _to_numeric_financial(col.get("collateral_value"))
    principal = _to_numeric_financial(deal.get("amortization_schedule"))
    if principal is None and debt is not None:
        principal = debt * 0.05
    if principal is None:
        principal = 0.0

    max_lev = _to_numeric_financial(cov.get("max_total_leverage_ratio"))
    min_icr = _to_numeric_financial(cov.get("min_interest_coverage_ratio"))
    min_fccr = _to_numeric_financial(cov.get("min_fixed_charge_coverage_ratio"))
    min_liq = _to_numeric_financial(cov.get("min_liquidity_requirement"))

    lev_adj = 1.0
    region_penalty = 1.0
    extra_penalty = 0.0
    if industry == "Technology":
        lev_adj = 1.10
    elif industry == "Retail":
        lev_adj = 0.90
    elif industry == "Energy":
        extra_penalty += 5.0
    if region == "Emerging Markets":
        region_penalty = 1.10
    elif region == "Europe":
        region_penalty = 1.05

    adj_max_lev = (max_lev * lev_adj) if max_lev is not None else None
    stress_ebitda = ebitda * (1.0 - stress_ebitda_pct / 100.0) if ebitda is not None else None

    total_lev = _safe_div(debt, ebitda)
    net_lev = _safe_div((debt - cash) if debt is not None and cash is not None else None, ebitda)
    icr = _safe_div(ebitda, interest)
    fccr = _safe_div((ebitda - capex) if ebitda is not None and capex is not None else None, (interest + principal) if interest is not None else None)
    liq = _safe_div(current_assets, current_liabilities)
    fcf_cov = _safe_div(fcf, debt)
    ltv = _safe_div(loan_amount, collateral_val)

    stress_total_lev = _safe_div(debt, stress_ebitda)
    stress_net_lev = _safe_div((debt - cash) if debt is not None and cash is not None else None, stress_ebitda)
    stress_icr = _safe_div(stress_ebitda, interest)
    stress_fccr = _safe_div((stress_ebitda - capex) if stress_ebitda is not None and capex is not None else None, (interest + principal) if interest is not None else None)

    def lev_level(actual: Optional[float], thr: Optional[float]) -> str:
        if actual is None or thr is None:
            return "Medium"
        if actual > thr:
            return "Breach"
        pct = actual / thr if thr else 1.0
        if pct < 0.75:
            return "Low"
        if pct <= 0.90:
            return "Medium"
        return "High"

    def cov_level(actual: Optional[float], thr: Optional[float]) -> str:
        if actual is None or thr is None:
            return "Medium"
        if actual < thr:
            return "Breach"
        pct = actual / thr if thr else 1.0
        if pct > 1.25:
            return "Low"
        if pct >= 1.10:
            return "Medium"
        return "High"

    def liq_level(actual: Optional[float], thr: Optional[float]) -> str:
        if actual is None:
            return "Medium"
        if thr is not None:
            return cov_level(actual, thr)
        if actual > 1.5:
            return "Low"
        if actual >= 1.2:
            return "Medium"
        if actual >= 1.0:
            return "High"
        return "Critical"

    def fcf_level(actual: Optional[float]) -> str:
        if actual is None:
            return "Medium"
        if actual > 0.20:
            return "Low"
        if actual >= 0.10:
            return "Medium"
        if actual >= 0.0:
            return "High"
        return "Critical"

    def ltv_level(actual: Optional[float]) -> str:
        if actual is None:
            return "Medium"
        if actual < 0.50:
            return "Low"
        if actual <= 0.75:
            return "Medium"
        if actual <= 0.85:
            return "High"
        return "Critical"

    rows: List[Dict[str, Any]] = []
    rows.append(
        {
            "Metric Name": "Total Leverage Ratio",
            "Actual Value": total_lev,
            "Covenant Threshold (if applicable)": max_lev,
            "Cushion / Headroom": (adj_max_lev - total_lev) if (adj_max_lev is not None and total_lev is not None) else None,
            "Industry-Adjusted Threshold": adj_max_lev,
            "Stress Case Value (10% EBITDA reduction default)": stress_total_lev,
            "Risk Level (Low / Medium / High / Breach)": lev_level(total_lev, adj_max_lev),
        }
    )
    rows.append(
        {
            "Metric Name": "Net Leverage Ratio",
            "Actual Value": net_lev,
            "Covenant Threshold (if applicable)": max_lev,
            "Cushion / Headroom": (adj_max_lev - net_lev) if (adj_max_lev is not None and net_lev is not None) else None,
            "Industry-Adjusted Threshold": adj_max_lev,
            "Stress Case Value (10% EBITDA reduction default)": stress_net_lev,
            "Risk Level (Low / Medium / High / Breach)": lev_level(net_lev, adj_max_lev),
        }
    )
    rows.append(
        {
            "Metric Name": "Interest Coverage Ratio",
            "Actual Value": icr,
            "Covenant Threshold (if applicable)": min_icr,
            "Cushion / Headroom": (icr - min_icr) if (icr is not None and min_icr is not None) else None,
            "Industry-Adjusted Threshold": min_icr,
            "Stress Case Value (10% EBITDA reduction default)": stress_icr,
            "Risk Level (Low / Medium / High / Breach)": cov_level(icr, min_icr),
        }
    )
    rows.append(
        {
            "Metric Name": "Fixed Charge Coverage Ratio",
            "Actual Value": fccr,
            "Covenant Threshold (if applicable)": min_fccr,
            "Cushion / Headroom": (fccr - min_fccr) if (fccr is not None and min_fccr is not None) else None,
            "Industry-Adjusted Threshold": min_fccr,
            "Stress Case Value (10% EBITDA reduction default)": stress_fccr,
            "Risk Level (Low / Medium / High / Breach)": cov_level(fccr, min_fccr),
        }
    )
    rows.append(
        {
            "Metric Name": "Liquidity Ratio",
            "Actual Value": liq,
            "Covenant Threshold (if applicable)": min_liq,
            "Cushion / Headroom": (liq - min_liq) if (liq is not None and min_liq is not None) else None,
            "Industry-Adjusted Threshold": min_liq,
            "Stress Case Value (10% EBITDA reduction default)": liq,
            "Risk Level (Low / Medium / High / Breach)": liq_level(liq, min_liq),
        }
    )
    rows.append(
        {
            "Metric Name": "Free Cash Flow Coverage",
            "Actual Value": fcf_cov,
            "Covenant Threshold (if applicable)": 0.10,
            "Cushion / Headroom": (fcf_cov - 0.10) if fcf_cov is not None else None,
            "Industry-Adjusted Threshold": 0.10,
            "Stress Case Value (10% EBITDA reduction default)": fcf_cov,
            "Risk Level (Low / Medium / High / Breach)": fcf_level(fcf_cov),
        }
    )
    rows.append(
        {
            "Metric Name": "Loan-to-Value",
            "Actual Value": ltv,
            "Covenant Threshold (if applicable)": 0.75,
            "Cushion / Headroom": (0.75 - ltv) if ltv is not None else None,
            "Industry-Adjusted Threshold": 0.75,
            "Stress Case Value (10% EBITDA reduction default)": ltv,
            "Risk Level (Low / Medium / High / Breach)": ltv_level(ltv),
        }
    )
    cov_levels = [r["Risk Level (Low / Medium / High / Breach)"] for r in rows[:5] if r["Risk Level (Low / Medium / High / Breach)"] is not None]
    if any(x in {"Breach", "Critical"} for x in cov_levels):
        cov_status_level = "Breach"
    else:
        cov_cushions = [r["Cushion / Headroom"] for r in rows[:5] if r["Cushion / Headroom"] is not None]
        min_cush = min(cov_cushions) if cov_cushions else None
        if min_cush is None:
            cov_status_level = "Medium"
        elif min_cush < 0.10:
            cov_status_level = "High"
        elif min_cush <= 0.25:
            cov_status_level = "Medium"
        else:
            cov_status_level = "Low"
    rows.append(
        {
            "Metric Name": "Covenant Compliance Status",
            "Actual Value": cov_status_level,
            "Covenant Threshold (if applicable)": None,
            "Cushion / Headroom": None,
            "Industry-Adjusted Threshold": None,
            "Stress Case Value (10% EBITDA reduction default)": cov_status_level,
            "Risk Level (Low / Medium / High / Breach)": cov_status_level if cov_status_level in {"Low", "Medium", "High", "Breach"} else "Medium",
        }
    )

    table_df = pd.DataFrame(rows)
    table_df["Pass / Fail"] = table_df["Risk Level (Low / Medium / High / Breach)"].apply(
        lambda lvl: "Pass" if str(lvl) in {"Low", "Medium"} else "Fail"
    )

    def nscore(level: str) -> int:
        return {"Low": 25, "Medium": 55, "High": 80, "Critical": 100, "Breach": 100}.get(level, 55)

    metric_level = {r["Metric Name"]: r["Risk Level (Low / Medium / High / Breach)"] for r in rows}
    leverage_avg = (nscore(metric_level["Total Leverage Ratio"]) + nscore(metric_level["Net Leverage Ratio"])) / 2.0
    coverage_levels = [nscore(metric_level["Interest Coverage Ratio"])]
    if metric_level.get("Fixed Charge Coverage Ratio") is not None:
        coverage_levels.append(nscore(metric_level["Fixed Charge Coverage Ratio"]))
    coverage_avg = sum(coverage_levels) / len(coverage_levels)
    liquidity_score = nscore(metric_level["Liquidity Ratio"])
    cf_score = nscore(metric_level["Free Cash Flow Coverage"])
    collateral_score = nscore(metric_level["Loan-to-Value"])
    cov_score = nscore(metric_level["Covenant Compliance Status"])

    fs_liq_weight = 0.10
    fs_cov_weight = 0.20
    if industry == "Financial Services":
        fs_liq_weight = 0.15
        fs_cov_weight = 0.15

    total_score = (
        0.30 * cov_score
        + 0.20 * leverage_avg
        + fs_cov_weight * coverage_avg
        + fs_liq_weight * liquidity_score
        + 0.10 * cf_score
        + 0.10 * collateral_score
    )
    total_score = total_score + extra_penalty
    total_score = total_score * region_penalty
    if any(x in {"Breach", "Critical"} for x in cov_levels):
        total_score = max(total_score, 85.0)
    total_score = max(0.0, min(100.0, total_score))

    if total_score <= 40:
        final_class = "Low Risk"
    elif total_score <= 70:
        final_class = "Medium Risk"
    elif total_score <= 84:
        final_class = "High Risk"
    else:
        final_class = "Critical Risk"

    driver_df = pd.DataFrame(
        [
            {"Metric Name": r["Metric Name"], "Risk Level": r["Risk Level (Low / Medium / High / Breach)"], "Row Score": nscore(r["Risk Level (Low / Medium / High / Breach)"])}
            for r in rows
            if r["Metric Name"] != "Covenant Compliance Status"
        ]
    ).sort_values("Row Score", ascending=False)

    stress_breach = any(
        [
            (stress_total_lev is not None and adj_max_lev is not None and stress_total_lev > adj_max_lev),
            (stress_net_lev is not None and adj_max_lev is not None and stress_net_lev > adj_max_lev),
            (stress_icr is not None and min_icr is not None and stress_icr < min_icr),
            (stress_fccr is not None and min_fccr is not None and stress_fccr < min_fccr),
        ]
    )

    return {
        "table_df": table_df,
        "score": round(total_score, 1),
        "classification": final_class,
        "top3": driver_df.head(3),
        "stress_breach": stress_breach,
        "ratio_values": {
            "liquidity": liq,
            "fcf_cov": fcf_cov,
            "ltv": ltv,
            "loan": loan_amount,
            "tenor": deal.get("tenor"),
            "industry": industry,
        },
    }


def _compute_credit_risk(
    tables: Dict[str, pd.DataFrame],
    industry: str,
    region: str,
    use_reported: bool,
    stress_pct: int,
    tolerance_pct: int,
) -> Dict[str, Any]:
    cov = tables["covenant_thresholds"].iloc[0].to_dict()
    deal = tables["deal_terms"].iloc[0].to_dict()
    rep = tables["reported_metrics"].iloc[0].to_dict()
    fin = tables["financial_actuals"].iloc[0].to_dict()
    col = tables["collateral_data"].iloc[0].to_dict()

    industry_cfg = {
        "Technology": {"lev_mult": 1.10, "cov_mult": 0.95, "liq_mult": 1.0, "vol_penalty": 0.0},
        "Manufacturing": {"lev_mult": 1.00, "cov_mult": 1.00, "liq_mult": 1.0, "vol_penalty": 0.0},
        "Healthcare": {"lev_mult": 1.05, "cov_mult": 0.98, "liq_mult": 1.0, "vol_penalty": 0.0},
        "Retail": {"lev_mult": 0.90, "cov_mult": 1.05, "liq_mult": 1.0, "vol_penalty": 0.0},
        "Energy": {"lev_mult": 0.95, "cov_mult": 1.05, "liq_mult": 1.0, "vol_penalty": 8.0},
        "Financial Services": {"lev_mult": 1.00, "cov_mult": 1.00, "liq_mult": 1.10, "vol_penalty": 0.0},
        "Other": {"lev_mult": 1.00, "cov_mult": 1.00, "liq_mult": 1.0, "vol_penalty": 0.0},
    }.get(industry, {"lev_mult": 1.0, "cov_mult": 1.0, "liq_mult": 1.0, "vol_penalty": 0.0})

    tolerance_mult = 1.0 + (tolerance_pct / 100.0)

    ebitda = rep.get("reported_ebitda") if use_reported else fin.get("ebitda")
    interest = rep.get("reported_interest_expense") if use_reported else fin.get("interest_expense")
    debt = rep.get("reported_total_debt") if use_reported else fin.get("total_debt")
    cash = fin.get("cash")
    current_assets = fin.get("current_assets")
    current_liabilities = fin.get("current_liabilities")
    operating_cf = fin.get("operating_cash_flow")
    capex = fin.get("capital_expenditures")
    free_cf = fin.get("free_cash_flow")
    loan_amount = deal.get("loan_amount")
    collateral_value = col.get("collateral_value")

    ebitda = _to_numeric_financial(ebitda)
    interest = _to_numeric_financial(interest)
    debt = _to_numeric_financial(debt)
    cash = _to_numeric_financial(cash)
    current_assets = _to_numeric_financial(current_assets)
    current_liabilities = _to_numeric_financial(current_liabilities)
    operating_cf = _to_numeric_financial(operating_cf)
    capex = _to_numeric_financial(capex)
    free_cf = _to_numeric_financial(free_cf)
    loan_amount = _to_numeric_financial(loan_amount)
    collateral_value = _to_numeric_financial(collateral_value)

    if free_cf is None and operating_cf is not None and capex is not None:
        free_cf = operating_cf - capex

    if ebitda is not None:
        ebitda = ebitda * (1.0 - (stress_pct / 100.0))

    principal_payments = 0.0
    amort = _to_numeric_financial(deal.get("amortization_schedule"))
    if amort is not None:
        principal_payments = max(amort, 0.0)
    elif debt is not None:
        principal_payments = debt * 0.05

    min_icr = _to_numeric_financial(cov.get("min_interest_coverage_ratio"))
    max_lev = _to_numeric_financial(cov.get("max_total_leverage_ratio"))
    min_fccr = _to_numeric_financial(cov.get("min_fixed_charge_coverage_ratio"))
    min_liq = _to_numeric_financial(cov.get("min_liquidity_requirement"))

    if min_icr is not None:
        min_icr = (min_icr * industry_cfg["cov_mult"]) / tolerance_mult
    if min_fccr is not None:
        min_fccr = (min_fccr * industry_cfg["cov_mult"]) / tolerance_mult
    if max_lev is not None:
        max_lev = (max_lev * industry_cfg["lev_mult"]) * tolerance_mult
    if min_liq is not None:
        min_liq = (min_liq * industry_cfg["liq_mult"]) / tolerance_mult

    icr = _safe_div(ebitda, interest)
    total_lev = _safe_div(debt, ebitda)
    net_lev = _safe_div((debt - cash) if debt is not None and cash is not None else None, ebitda)
    fccr = _safe_div((ebitda - capex) if ebitda is not None and capex is not None else None, (interest + principal_payments) if interest is not None else None)
    liq_ratio = _safe_div(current_assets, current_liabilities)
    ltv = _safe_div(loan_amount, collateral_value)

    def cushion_pct(actual: Optional[float], threshold: Optional[float], kind: str) -> Optional[float]:
        if actual is None or threshold is None or threshold == 0:
            return None
        if kind == "coverage":
            return ((actual - threshold) / abs(threshold)) * 100.0
        return ((threshold - actual) / abs(threshold)) * 100.0

    cov_items: List[Dict[str, Any]] = []
    for name, actual, threshold, kind in [
        ("Interest Coverage", icr, min_icr, "coverage"),
        ("Total Leverage", total_lev, max_lev, "leverage"),
        ("Fixed Charge Coverage", fccr, min_fccr, "coverage"),
        ("Liquidity", liq_ratio, min_liq, "coverage"),
    ]:
        cp = cushion_pct(actual, threshold, kind)
        if cp is None:
            level = "Medium"
            breached = False
        elif cp < 0:
            level = "Critical"
            breached = True
        elif cp < 10:
            level = "High"
            breached = False
        elif cp <= 25:
            level = "Medium"
            breached = False
        else:
            level = "Low"
            breached = False
        cov_items.append(
            {
                "Covenant": name,
                "Actual": actual,
                "Threshold": threshold,
                "Cushion %": cp,
                "Risk": level,
                "Breached": breached,
            }
        )

    any_breach = any(x["Breached"] for x in cov_items if x["Threshold"] is not None)
    if any_breach:
        covenant_status = "Breach"
    elif any((x["Cushion %"] is not None and x["Cushion %"] < 10) for x in cov_items):
        covenant_status = "Warning"
    else:
        covenant_status = "Compliant"
    covenant_cushion = min([x["Cushion %"] for x in cov_items if x["Cushion %"] is not None], default=None)

    # Category bands
    lev_low = 3.0 * industry_cfg["lev_mult"] * tolerance_mult
    lev_high = 4.5 * industry_cfg["lev_mult"] * tolerance_mult
    if total_lev is None:
        lev_level = "Medium"
    elif total_lev < lev_low:
        lev_level = "Low"
    elif total_lev <= lev_high:
        lev_level = "Medium"
    else:
        lev_level = "High"

    cov_low = 4.0 * industry_cfg["cov_mult"] / tolerance_mult
    cov_mid = 2.0 * industry_cfg["cov_mult"] / tolerance_mult
    if icr is None:
        coverage_level = "Medium"
    elif icr > cov_low:
        coverage_level = "Low"
    elif icr >= cov_mid:
        coverage_level = "Medium"
    else:
        coverage_level = "High"

    if liq_ratio is None:
        liquidity_level = "Medium"
    elif liq_ratio > 1.5:
        liquidity_level = "Low"
    elif liq_ratio >= 1.0:
        liquidity_level = "Medium"
    else:
        liquidity_level = "High"

    if free_cf is None or operating_cf is None or operating_cf == 0:
        cf_level = "Medium"
    else:
        cf_cov = free_cf / operating_cf
        if cf_cov > 0.25:
            cf_level = "Low"
        elif cf_cov >= 0.0:
            cf_level = "Medium"
        else:
            cf_level = "High"

    if ltv is None:
        collateral_level = "Medium"
    elif ltv < 0.50:
        collateral_level = "Low"
    elif ltv <= 0.75:
        collateral_level = "Medium"
    else:
        collateral_level = "High"

    if str(col.get("lien_priority", "")).lower() not in {"first", "first lien", "1"}:
        if collateral_level == "Low":
            collateral_level = "Medium"
    gp = col.get("guarantee_presence")
    if gp is False and collateral_level == "Medium":
        collateral_level = "High"

    covenant_level = "Critical" if any_breach else max((x["Risk"] for x in cov_items if x["Risk"] in {"Low", "Medium", "High"}), default="Medium", key=lambda y: {"Low": 1, "Medium": 2, "High": 3}.get(y, 2))

    weights = {
        "Covenant Compliance": 0.30,
        "Leverage Risk": 0.20,
        "Coverage Risk": 0.20,
        "Liquidity Risk": 0.10,
        "Cash Flow Stability": 0.10,
        "Collateral Strength": 0.10,
    }
    components = {
        "Covenant Compliance": covenant_level if covenant_level != "Critical" else "High",
        "Leverage Risk": lev_level,
        "Coverage Risk": coverage_level,
        "Liquidity Risk": liquidity_level,
        "Cash Flow Stability": cf_level,
        "Collateral Strength": collateral_level,
    }

    weighted_score = sum(_score_from_level(level) * weights[name] for name, level in components.items())

    # Industry + region penalties
    weighted_score += industry_cfg["vol_penalty"]
    if industry == "Financial Services" and liquidity_level != "Low":
        weighted_score += 6.0
    if industry == "Manufacturing" and collateral_level == "Low":
        weighted_score -= 4.0

    if region == "Emerging Markets":
        weighted_score *= 1.10
    elif region == "Europe":
        weighted_score *= 1.05

    weighted_score = max(0.0, min(100.0, weighted_score))
    overall_level = _risk_level_from_score(weighted_score)
    if any_breach:
        overall_level = "Critical"
        weighted_score = 100.0

    top_drivers = sorted(
        [{"Driver": k, "Level": v, "Contribution": _score_from_level(v) * weights[k]} for k, v in components.items()],
        key=lambda x: x["Contribution"],
        reverse=True,
    )[:3]

    return {
        "ratios": {
            "Interest Coverage Ratio": icr,
            "Total Leverage": total_lev,
            "Net Leverage": net_lev,
            "Fixed Charge Coverage Ratio": fccr,
            "Liquidity Ratio": liq_ratio,
            "Free Cash Flow": free_cf,
            "Loan-to-Value": ltv,
        },
        "cov_items": pd.DataFrame(cov_items),
        "components": components,
        "weighted_score": round(weighted_score, 1),
        "overall_level": overall_level,
        "covenant_status": covenant_status,
        "covenant_cushion": covenant_cushion,
        "top_drivers": pd.DataFrame(top_drivers),
        "any_breach": any_breach,
    }


INDUSTRY_RISK_MULTIPLIER = {
    "Oil & Gas": 1.25,
    "Construction": 1.20,
    "Hospitality": 1.20,
    "Retail": 1.10,
    "Manufacturing": 1.05,
    "Transportation": 1.15,
    "Technology": 1.00,
    "Healthcare": 0.95,
    "Financial Services": 1.05,
    "Agriculture": 1.15,
    "Real Estate": 1.10,
}
INDUSTRY_STABILITY_FACTOR = {
    "Oil & Gas": 0.85,
    "Construction": 0.90,
    "Hospitality": 0.90,
    "Retail": 0.95,
    "Manufacturing": 1.00,
    "Technology": 1.00,
    "Healthcare": 1.05,
    "Agriculture": 0.95,
    "Real Estate": 0.95,
    "Transportation": 0.95,
    "Financial Services": 1.00,
}
GEOGRAPHY_MULTIPLIER = {
    "United States Tier 1": 1.00,
    "United States Tier 2": 1.05,
    "Canada": 1.00,
    "Emerging Market - Low Stability": 1.20,
    "Emerging Market - High Volatility": 1.35,
    "Sanctioned / High Risk Region": 1.50,
}
BUSINESS_STAGE_MULTIPLIER = {
    "Startup": 1.30,
    "Growth": 1.10,
    "Mature": 1.00,
    "Declining": 1.40,
}


def _extract_from_combined_df(
    combined_df: pd.DataFrame,
    candidates: List[str],
    preferred_doc_types: Optional[List[str]] = None,
    preferred_year: Optional[int] = None,
) -> Optional[float]:
    if combined_df is None or combined_df.empty:
        return None
    work = combined_df
    if preferred_doc_types and "Detected Type" in work.columns:
        allowed_types = {_norm_key(t) for t in preferred_doc_types}
        scoped = work[work["Detected Type"].astype(str).apply(lambda x: _norm_key(x) in allowed_types)]
        # Do not hard-fail when doc-type labels drift across modules/runs.
        # Fall back to unscoped rows so valid statement values can still be found.
        if not scoped.empty:
            work = scoped
    if "Selected Year" in work.columns:
        years = pd.to_numeric(work["Selected Year"], errors="coerce")
        if years.notna().any():
            work = work[years.notna()].copy()
            years = pd.to_numeric(work["Selected Year"], errors="coerce")
        used_preferred_year = False
        if preferred_year is not None:
            scoped_year = work[years == float(int(preferred_year))]
            if not scoped_year.empty:
                work = scoped_year
                used_preferred_year = True
        if not used_preferred_year:
            valid_years = years.dropna()
            if not valid_years.empty:
                latest_year = int(valid_years.max())
                scoped_latest = work[pd.to_numeric(work["Selected Year"], errors="coerce") == float(latest_year)]
                if not scoped_latest.empty:
                    work = scoped_latest
    work = work.copy()
    if "Confidence" in work.columns:
        work["_conf"] = pd.to_numeric(work["Confidence"], errors="coerce").fillna(-1.0)
    else:
        work["_conf"] = -1.0
    period_cols = ["Reporting Period Date", "Test Dates", "Maturity Date", "Amendment Effective Date", "effective_date"]
    work["_period"] = pd.NaT
    for pc in period_cols:
        if pc in work.columns:
            parsed = pd.to_datetime(work[pc], errors="coerce")
            work["_period"] = work["_period"].fillna(parsed)
    work["_period_epoch"] = work["_period"].apply(lambda x: x.value if pd.notna(x) else -1)
    if "Page" in work.columns:
        work["_page"] = pd.to_numeric(work["Page"], errors="coerce").fillna(-1.0)
    else:
        work["_page"] = -1.0
    work = work.sort_values(by=["_period_epoch", "_conf", "_page"], ascending=[False, False, False], kind="stable")
    keys = [_norm_key(c) for c in candidates if str(c).strip()]

    def _match_score(col_name: str) -> int:
        col_n = _norm_key(col_name)
        if not col_n:
            return 0
        if any(col_n == k for k in keys):
            return 3
        if any(re.search(rf"\b{re.escape(k)}\b", col_n) for k in keys):
            return 2
        if any(k in col_n for k in keys):
            return 1
        return 0

    for _, row in work.iterrows():
        best_value: Optional[float] = None
        best_score = 0
        for col, value in row.items():
            if col in META_COLUMNS or col in {"Source Document", "Detected Type", "Sheet"}:
                continue
            score = _match_score(str(col))
            if score <= 0:
                continue
            num = _to_numeric_financial(value)
            if num is None:
                continue
            if score > best_score:
                best_score = score
                best_value = num
        if best_value is not None:
            return best_value
    return None


def _extract_debt_components_sum_from_combined_df(
    combined_df: pd.DataFrame,
    preferred_doc_types: Optional[List[str]] = None,
    preferred_year: Optional[int] = None,
) -> Optional[float]:
    if combined_df is None or combined_df.empty:
        return None
    work = combined_df
    if preferred_doc_types and "Detected Type" in work.columns:
        allowed_types = {_norm_key(t) for t in preferred_doc_types}
        scoped = work[work["Detected Type"].astype(str).apply(lambda x: _norm_key(x) in allowed_types)]
        if not scoped.empty:
            work = scoped
    if "Selected Year" in work.columns:
        years = pd.to_numeric(work["Selected Year"], errors="coerce")
        if years.notna().any():
            work = work[years.notna()].copy()
            years = pd.to_numeric(work["Selected Year"], errors="coerce")
        used_preferred_year = False
        if preferred_year is not None:
            scoped_year = work[years == float(int(preferred_year))]
            if not scoped_year.empty:
                work = scoped_year
                used_preferred_year = True
        if not used_preferred_year:
            valid_years = years.dropna()
            if not valid_years.empty:
                latest_year = int(valid_years.max())
                scoped_latest = work[pd.to_numeric(work["Selected Year"], errors="coerce") == float(latest_year)]
                if not scoped_latest.empty:
                    work = scoped_latest
    work = work.copy()
    if "Confidence" in work.columns:
        work["_conf"] = pd.to_numeric(work["Confidence"], errors="coerce").fillna(-1.0)
    else:
        work["_conf"] = -1.0
    if "Page" in work.columns:
        work["_page"] = pd.to_numeric(work["Page"], errors="coerce").fillna(-1.0)
    else:
        work["_page"] = -1.0
    work = work.sort_values(by=["_conf", "_page"], ascending=[False, False], kind="stable")
    component_keys = [
        "short-term debt",
        "short term debt",
        "short-term borrowings",
        "current portion of long-term debt",
        "current portion of long-term obligations",
        "current maturities of long-term debt",
        "long-term debt",
        "long term debt",
        "long-term debt non-current",
        "long term debt non current",
        "long-term debt, net",
        "non-current borrowings",
        "notes payable",
    ]
    total_keys = ["total debt", "total borrowings", "interest bearing liabilities"]
    for _, row in work.iterrows():
        comp_sum = 0.0
        comp_found = 0
        direct_total = None
        for col, value in row.items():
            if col in META_COLUMNS or col in {"Source Document", "Detected Type", "Sheet", "_conf", "_page"}:
                continue
            col_l = str(col).lower()
            num = _to_numeric_financial(value)
            if num is None:
                continue
            if any(k in col_l for k in component_keys):
                comp_sum += float(num)
                comp_found += 1
            elif direct_total is None and any(k in col_l for k in total_keys):
                direct_total = float(num)
        if comp_found > 0:
            if direct_total is not None:
                return max(comp_sum, direct_total)
            return comp_sum
        if direct_total is not None:
            return direct_total
    return None


def _score_metric(value: Optional[float], threshold: Optional[float], higher_is_better: bool) -> tuple[Optional[int], str]:
    if value is None or threshold is None:
        return None, "Incomplete"
    if threshold <= 0:
        try:
            v = float(value)
        except Exception:
            return None, "Incomplete"
        if higher_is_better:
            score = 100 if v >= 0 else 0
        else:
            score = 100 if v <= 0 else 0
        return score, ("Green" if score >= 80 else "Red")
    try:
        if higher_is_better:
            raw_score = (float(value) / float(threshold)) * 100.0
        else:
            if float(value) == 0:
                raw_score = 100.0
            elif float(value) < 0:
                raw_score = 0.0
            else:
                raw_score = (float(threshold) / float(value)) * 100.0
    except Exception:
        return None, "Incomplete"
    bounded = max(0.0, min(100.0, raw_score))
    score = int(round(bounded))
    if score >= 80:
        return score, "Green"
    if score >= 60:
        return score, "Yellow"
    return score, "Red"


def _color_to_row_level(color: str) -> str:
    return {"Green": "Low", "Yellow": "Medium", "Red": "High", "Unknown": "Incomplete", "Incomplete": "Incomplete"}.get(color, "Medium")


def _risk_category_from_score(score: float) -> str:
    if score > 80:
        return "Low Risk"
    if score >= 60:
        return "Moderate Risk"
    if score >= 40:
        return "Elevated Risk"
    return "High Risk"


def _build_document_profile(
    combined_df: pd.DataFrame,
    doc_types: Optional[List[str]],
    analysis_mode: str,
    locked_year: Optional[int],
    available_years: Optional[List[int]],
) -> Dict[str, Any]:
    dtype = "unknown"
    if doc_types:
        dt = str(doc_types[0]).strip()
        if dt:
            dtype = dt
    elif combined_df is not None and not combined_df.empty and "Detected Type" in combined_df.columns:
        vals = [str(x).strip() for x in combined_df["Detected Type"].dropna().tolist() if str(x).strip()]
        if vals:
            dtype = vals[0]

    period_type = "annual"
    mode_l = str(analysis_mode or "").strip().lower()
    if "t12m" in mode_l:
        period_type = "quarterly"
    elif combined_df is not None and not combined_df.empty:
        blob_cols = [c for c in ["Snippet", "Source Document"] if c in combined_df.columns]
        blob = " ".join(str(v).lower() for c in blob_cols for v in combined_df[c].dropna().tolist())
        if any(k in blob for k in ["three months ended", "quarter ended", "quarterly", "10-q", "10 q"]):
            period_type = "quarterly"
    if period_type == "annual" and combined_df is not None and not combined_df.empty and "Source Document" in combined_df.columns:
        src_blob = " ".join(str(v).lower() for v in combined_df["Source Document"].dropna().tolist())
        if "quarter" in src_blob or "q1" in src_blob or "q2" in src_blob or "q3" in src_blob or "q4" in src_blob:
            period_type = "quarterly"

    selected_year = locked_year
    if selected_year is None and combined_df is not None and not combined_df.empty and "Selected Year" in combined_df.columns:
        yrs = [int(_to_numeric_financial(v)) for v in combined_df["Selected Year"].tolist() if _to_numeric_financial(v) is not None]
        if yrs:
            selected_year = max(yrs)
    if selected_year is None and available_years:
        selected_year = max(int(y) for y in available_years)

    detected_scale = _detect_scale_from_combined_df(combined_df)
    reporting_unit = {
        "Units": "whole_dollars",
        "Thousands": "thousands",
        "Millions": "millions",
        "Billions": "billions",
        "Trillions": "trillions",
    }.get(str(detected_scale), "whole_dollars")

    multi_year_columns = [int(y) for y in (available_years or []) if y is not None]
    if not multi_year_columns and combined_df is not None and not combined_df.empty and "Available Years" in combined_df.columns:
        years_found: List[int] = []
        for raw in combined_df["Available Years"].dropna().tolist():
            years_found.extend([int(y) for y in re.findall(r"\b(19\d{2}|20\d{2})\b", str(raw))])
        multi_year_columns = sorted(set(years_found))

    fiscal_year_end = "Unknown"
    if combined_df is not None and not combined_df.empty and "Snippet" in combined_df.columns:
        snip = " ".join(str(x) for x in combined_df["Snippet"].dropna().tolist())
        m = re.search(
            r"(?:year(?:s)? ended|as of)\s+([A-Za-z]+\s+\d{1,2}(?:,\s*\d{4})?)",
            snip,
            flags=re.IGNORECASE,
        )
        if m:
            fiscal_year_end = m.group(1)

    return {
        "doc_type": dtype,
        "period_type": period_type,
        "fiscal_year_end": fiscal_year_end,
        "selected_year": selected_year,
        "reporting_unit": reporting_unit,
        "currency": "USD",
        "multi_year_columns": multi_year_columns,
    }


def _precalc_input_gate(
    combined_df: pd.DataFrame,
    profile: Dict[str, Any],
    raw_inputs: Dict[str, Optional[float]],
) -> Tuple[List[str], List[str]]:
    errors: List[str] = []
    missing_required: List[str] = []

    # Core fields that should block scoring when absent. COGS / interest are ratio inputs—many filers omit
    # explicit lines (aggregate in other income, no gross margin, immaterial interest).
    # Cash flow lines are not required for underwriting halt — DSCR/FCF/OCF ratio rows stay Incomplete without them.
    required_manifest: Dict[str, List[str]] = {
        "income_statement": ["revenue", "ebit", "net_income"],
        "balance_sheet": ["current_assets", "current_liabilities", "total_assets", "equity"],
    }
    for section, fields in required_manifest.items():
        for f in fields:
            if raw_inputs.get(f) is None:
                missing_required.append(f"{section}.{f}")
    # Debt split (ST/LT) is often absent when filers only show rolled-up total debt or are debt-free.
    if raw_inputs.get("st_debt") is None and raw_inputs.get("lt_debt") is None:
        if raw_inputs.get("total_debt") is None:
            missing_required.append("balance_sheet.st_debt_or_lt_debt")

    revenue = raw_inputs.get("revenue")
    if profile.get("reporting_unit") == "whole_dollars" and revenue is not None and revenue > 500_000_000:
        errors.append("Revenue > $500M in whole dollars — possible scaling error.")

    equity = raw_inputs.get("equity")
    if equity is not None and abs(float(equity)) < 1000:
        errors.append("Equity near zero — possible extraction error.")

    st_debt = raw_inputs.get("st_debt")
    lt_debt = raw_inputs.get("lt_debt")
    if (st_debt is not None and st_debt < 0) or (lt_debt is not None and lt_debt < 0):
        errors.append("Debt is negative — sign extraction error.")

    if combined_df is not None and not combined_df.empty and "Sheet" in combined_df.columns and "Selected Year" in combined_df.columns:
        stmt_years: Dict[str, int] = {}
        for sheet in ["Income Statement", "Balance Sheet", "Cash Flow"]:
            subset = combined_df[combined_df["Sheet"].astype(str).str.lower() == sheet.lower()]
            yrs = [int(_to_numeric_financial(v)) for v in subset["Selected Year"].tolist() if _to_numeric_financial(v) is not None]
            if yrs:
                stmt_years[sheet] = max(yrs)
        # Multi-column SEC tables often differ by one year between statements; do not hard-stop on that alone.
        if len(set(stmt_years.values())) > 1:
            yr_span = max(stmt_years.values()) - min(stmt_years.values())
            if yr_span > 1:
                errors.append("YEAR MISMATCH — statements from different periods.")

    total_assets = raw_inputs.get("total_assets")
    total_liabilities = raw_inputs.get("total_liabilities")
    if total_assets is not None and total_liabilities is not None and equity is not None and total_assets != 0:
        reconstructed = float(total_liabilities) + float(equity)
        # 10-K rounding, NCI, and equity adjustments often exceed 1%; use a looser tolerance.
        if abs(reconstructed - float(total_assets)) / abs(float(total_assets)) > 0.05:
            errors.append("Balance sheet doesn't balance within 5% tolerance.")

    return errors, missing_required


def _build_dynamic_credit_analysis(
    tables: Dict[str, pd.DataFrame],
    combined_df: pd.DataFrame,
    industry: str,
    geography: str,
    business_stage: str,
    company_size: str,
    years_in_operation: int,
    requested_amount: float,
    currency_scale: str = "Units",
    analysis_mode: str = "Latest Available",
    locked_year: Optional[int] = None,
    available_years: Optional[List[int]] = None,
    t12m_status: str = "n/a",
    cross_year_error_flag: bool = False,
) -> Dict[str, Any]:
    fin = tables["financial_actuals"].iloc[0].to_dict()
    deal = tables["deal_terms"].iloc[0].to_dict()
    collateral = tables["collateral_data"].iloc[0].to_dict()

    def _clean_num(v: Any) -> Optional[float]:
        out = _to_numeric_financial(v)
        if out is None or pd.isna(out):
            return None
        return float(out)

    financial_doc_scope = [
        "Financial Statements",
        "10-K",
        "10-Q",
        "Annual Report",
        "Other Financial Filing",
        "Unknown",
        "Covenant Compliance Certificate",
        "Compliance Certificate",
    ]

    # Normalize statement-unit extracted values to base amounts for absolute-money outputs
    # (policy limits, comparisons with requested amount). Ratios remain unchanged because
    # numerator/denominator scale consistently.
    detected_scale = _detect_scale_from_combined_df(combined_df)
    _, detected_factor = _scale_factor_from_label(detected_scale)
    sample_vals = []
    for cand in [
        fin.get("revenue"),
        fin.get("ebit"),
        fin.get("ebitda"),
        fin.get("net_income"),
        fin.get("total_assets"),
        fin.get("total_liabilities"),
        fin.get("equity"),
        fin.get("operating_cash_flow"),
        fin.get("total_debt"),
    ]:
        c = _clean_num(cand)
        if c is not None:
            sample_vals.append(abs(c))
    # Guard: if values are already absolute (legacy cached runs), don't multiply again.
    already_absolute = bool(sample_vals) and max(sample_vals) >= 10_000_000
    apply_unit_factor = detected_factor > 1.0 and (not already_absolute)

    def _scaled(v: Optional[float]) -> Optional[float]:
        if v is None:
            return None
        base = float(v)
        return base * detected_factor if apply_unit_factor else base

    current_assets = _scaled(_clean_num(fin.get("current_assets")))
    current_liabilities = _scaled(_clean_num(fin.get("current_liabilities")))
    inventory = _scaled(_clean_num(fin.get("inventory")))
    cash = _scaled(_clean_num(fin.get("cash")))
    if cash is None:
        cash = _extract_from_combined_df(
            combined_df, ["cash", "cash and cash equivalents"], preferred_doc_types=financial_doc_scope, preferred_year=locked_year
        )
        cash = _scaled(_clean_num(cash))
    if current_liabilities is None:
        current_liabilities = _extract_from_combined_df(
            combined_df, ["total current liabilities", "current liabilities"], preferred_doc_types=financial_doc_scope, preferred_year=locked_year
        )
        current_liabilities = _scaled(_clean_num(current_liabilities))
    if inventory is None:
        inventory = _extract_from_combined_df(
            combined_df,
            ["inventory", "inventories", "total inventory"],
            preferred_doc_types=financial_doc_scope,
            preferred_year=locked_year,
        )
        inventory = _scaled(_clean_num(inventory))
    ar = _scaled(_clean_num(fin.get("accounts_receivable")))
    if ar is None:
        ar = _extract_from_combined_df(
            combined_df,
            ["accounts receivable", "trade receivables", "receivables, net", "receivables net", "a/r", "receivables"],
            preferred_doc_types=financial_doc_scope,
            preferred_year=locked_year,
        )
        ar = _scaled(_clean_num(ar))
    if cash is not None:
        cash = abs(float(cash))
    if current_assets is not None:
        current_assets = abs(float(current_assets))
    if current_liabilities is not None:
        current_liabilities = abs(float(current_liabilities))
    if inventory is not None:
        inventory = abs(float(inventory))
    if ar is not None:
        ar = abs(float(ar))
    total_debt = _scaled(_clean_num(fin.get("total_debt")))
    short_term_debt = _scaled(_clean_num(fin.get("short_term_debt")))
    current_portion_long_term_debt = _scaled(_clean_num(fin.get("current_portion_long_term_debt")))
    long_term_debt = _scaled(_clean_num(fin.get("long_term_debt")))
    if short_term_debt is None:
        short_term_debt = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    ["short-term debt", "short term debt", "short-term borrowings"],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    if current_portion_long_term_debt is None:
        current_portion_long_term_debt = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    ["current portion of long-term debt", "current portion of long-term obligations", "current maturities of long-term debt"],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    if long_term_debt is None:
        long_term_debt = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    ["long-term debt", "long term debt", "long-term debt non-current"],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    # Standardized debt definition: ST financial debt + LT financial debt.
    # Use current portion of LT debt as primary ST debt bucket, fallback to short-term debt.
    st_financial_debt = current_portion_long_term_debt if current_portion_long_term_debt is not None else short_term_debt
    debt_components = [x for x in [st_financial_debt, long_term_debt] if x is not None]
    if debt_components:
        # When ST+LT are available, treat this as authoritative and do not override
        # from broad combined_df scans that can include non-policy liabilities.
        total_debt = float(sum(debt_components))
    else:
        debt_from_combined = _extract_debt_components_sum_from_combined_df(
            combined_df,
            preferred_doc_types=financial_doc_scope,
            preferred_year=locked_year,
        )
        if debt_from_combined is not None:
            debt_from_combined = _scaled(_clean_num(debt_from_combined))
            if debt_from_combined is not None:
                total_debt = float(debt_from_combined)
    equity = _scaled(_clean_num(fin.get("equity")))
    if equity is None:
        equity = _extract_from_combined_df(
            combined_df, ["equity", "shareholder equity", "net worth"], preferred_doc_types=financial_doc_scope, preferred_year=locked_year
        )
    ebitda = _scaled(_clean_num(fin.get("ebitda")))
    ebit = _scaled(_clean_num(fin.get("ebit")))
    if ebit is None:
        ebit = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df, ["ebit", "operating income"], preferred_doc_types=financial_doc_scope, preferred_year=locked_year
                    
                )
            )
        )
    if ebit is None:
        ebit = ebitda
    depreciation_amortization = _scaled(_clean_num(fin.get("depreciation_amortization")))
    if depreciation_amortization is None:
        depreciation_amortization = _extract_from_combined_df(
            combined_df,
            ["depreciation and amortization", "depreciation", "amortization", "d&a"],
            preferred_doc_types=financial_doc_scope,
            preferred_year=locked_year,
        )
        depreciation_amortization = _scaled(_clean_num(depreciation_amortization))
    # Step 3: derive EBITDA when explicit EBITDA is missing.
    if ebitda is None and ebit is not None and depreciation_amortization is not None:
        ebitda = ebit + depreciation_amortization
    if ebitda is None:
        ebitda = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    ["ebitda", "adjusted ebitda", "operating profit before d&a", "operating income before depreciation"],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    if ebitda is None and ebit is not None:
        ebitda = ebit
    # Guard against overstated EBITDA picks (common bad map to revenue-like rows).
    if ebitda is not None and ebit is not None and ebit > 0:
        try:
            if float(ebitda) > float(ebit) * 3.0:
                if depreciation_amortization is not None:
                    ebitda = float(ebit) + float(depreciation_amortization)
                else:
                    ebitda = float(ebit)
        except Exception:
            pass
    # Sanity guard: explicit EBITDA far below EBIT is usually a bad map for public filings.
    if ebitda is not None and ebit is not None:
        try:
            if float(ebit) > 0 and float(ebitda) < (float(ebit) * 0.90):
                if depreciation_amortization is not None:
                    ebitda = float(ebit) + float(depreciation_amortization)
                else:
                    ebitda = float(ebit)
        except Exception:
            pass
    interest_expense = _scaled(_clean_num(fin.get("interest_expense")))
    if interest_expense is None:
        interest_expense = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    [
                        "interest expense",
                        "interest expense net",
                        "total interest expense",
                        "interest and financing costs",
                        "interest cost",
                        "cash paid for interest",
                        "cash paid for interest net of capitalized interest",
                        "interest paid",
                    ],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    revenue = _scaled(_clean_num(fin.get("revenue")))
    cogs = _scaled(_clean_num(fin.get("cogs")))
    if revenue is None:
        revenue = _extract_from_combined_df(
            combined_df,
            ["revenue", "total revenue", "net sales", "sales"],
            preferred_doc_types=financial_doc_scope + ["Covenant Compliance Certificate", "Compliance Certificate"],
            preferred_year=locked_year,
        )
        revenue = _scaled(_clean_num(revenue))
    if cogs is None:
        cogs = _extract_from_combined_df(
            combined_df,
            [
                "cogs",
                "cost of goods sold",
                "cost of revenue",
                "cost of revenues",
                "cost of sales",
                "total cost of revenues",
                "cost of revenue excluding traffic acquisition costs",
            ],
            preferred_doc_types=financial_doc_scope + ["Covenant Compliance Certificate", "Compliance Certificate"],
            preferred_year=locked_year,
        )
        cogs = _scaled(_clean_num(cogs))
    if cogs is not None:
        cogs = abs(float(cogs))
    net_income = _scaled(_clean_num(fin.get("net_income")))
    if net_income is None:
        net_income = _extract_from_combined_df(
            combined_df,
            ["net income", "net income (loss)", "net loss"],
            preferred_doc_types=financial_doc_scope + ["Covenant Compliance Certificate", "Compliance Certificate"],
            preferred_year=locked_year,
        )
        net_income = _scaled(_clean_num(net_income))
    total_assets = _scaled(_clean_num(fin.get("total_assets")))
    if total_assets is None:
        total_assets = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df, ["total assets"], preferred_doc_types=financial_doc_scope, preferred_year=locked_year
                )
            )
        )
    total_liabilities = _scaled(_clean_num(fin.get("total_liabilities")))
    if total_liabilities is None:
        total_liabilities = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df, ["total liabilities"], preferred_doc_types=financial_doc_scope, preferred_year=locked_year
                )
            )
        )
    operating_cf = _scaled(_clean_num(fin.get("operating_cash_flow")))
    if operating_cf is None and combined_df is not None and not combined_df.empty:
        operating_cf = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    [
                        "cash flows from operating activities",
                        "net cash provided by operating activities",
                        "net cash from operating activities",
                        "cash from operating activities",
                        "operating cash flows",
                    ],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    capex = _scaled(_clean_num(fin.get("capital_expenditures")))
    if capex is None:
        capex = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    ["capex", "capital expenditures", "purchase of property and equipment", "additions to property and equipment"],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
    capex_for_ratio = abs(capex) if capex is not None else None
    free_cf = _scaled(_clean_num(fin.get("free_cash_flow")))
    if free_cf is None and operating_cf is not None and capex_for_ratio is not None:
        free_cf = operating_cf - capex_for_ratio
    # Display monetary statement metrics in statement units (avoid double-scaling in UI table).
    free_cf_display = free_cf
    if free_cf_display is not None and apply_unit_factor and detected_factor > 1:
        free_cf_display = float(free_cf_display) / float(detected_factor)

    principal = _scaled(_clean_num(deal.get("amortization_schedule")))
    if principal is None:
        inferred_principal = _extract_principal_repayment(
            combined_df,
            preferred_doc_types=financial_doc_scope + ["Term Sheet", "Credit Agreement", "Amendment"],
            selection_rule="highest_confidence",
        )
        principal = _scaled(_clean_num(inferred_principal))
    if principal is None:
        principal = 0.0
    interest_for_coverage = abs(float(interest_expense)) if interest_expense is not None else None
    loan_amount = _scaled(_clean_num(deal.get("loan_amount")))
    collateral_value = _scaled(_clean_num(collateral.get("collateral_value")))
    document_profile = _build_document_profile(
        combined_df=combined_df,
        doc_types=None,
        analysis_mode=analysis_mode,
        locked_year=locked_year,
        available_years=available_years,
    )
    precheck_errors, missing_required_fields = _precalc_input_gate(
        combined_df=combined_df,
        profile=document_profile,
        raw_inputs={
            "revenue": revenue,
            "cogs": cogs,
            "ebit": ebit,
            "net_income": net_income,
            "interest_expense": interest_expense,
            "current_assets": current_assets,
            "current_liabilities": current_liabilities,
            "inventory": inventory,
            "total_assets": total_assets,
            "total_liabilities": total_liabilities,
            "equity": equity,
            "st_debt": st_financial_debt,
            "lt_debt": long_term_debt,
            "total_debt": total_debt,
            "operating_cf": operating_cf,
            "capex": capex_for_ratio,
            "depreciation_amortization": depreciation_amortization,
        },
    )

    def _sanitize_metric_value(v: Optional[float]) -> Optional[float]:
        if v is None:
            return None
        try:
            fv = float(v)
        except Exception:
            return None
        if abs(fv) < 0.0001:
            return 0.0
        return fv

    # Guard tiny EBIT/EBITDA artifacts that often come from mis-read rows in large filings.
    ebit_artifact_suppressed = False
    ebitda_artifact_suppressed = False
    if revenue is not None and abs(float(revenue)) > 1000:
        if ebit is not None and abs(float(ebit)) < 10:
            ebit = None
            ebit_artifact_suppressed = True
        if ebitda is not None and abs(float(ebitda)) < 10:
            ebitda = None
            ebitda_artifact_suppressed = True

    # Guard COGS == Revenue artifact (common false-map to total expense/summary lines).
    cogs_suspect_equal_revenue = False
    if revenue is not None and cogs is not None:
        try:
            if abs(float(cogs) - float(revenue)) <= max(1e-6, abs(float(revenue)) * 0.005):
                cogs = None
                cogs_suspect_equal_revenue = True
        except Exception:
            pass

    current_ratio = _sanitize_metric_value(_safe_div(current_assets, current_liabilities))
    quick_ratio_numerator: Optional[float] = None
    if current_assets is not None and inventory is not None:
        # Strict formula path.
        quick_ratio_numerator = current_assets - inventory
    elif cash is not None and ar is not None:
        # Fallback path when explicit inventory is unavailable but liquid-current assets are available.
        quick_ratio_numerator = cash + ar
    elif current_assets is not None:
        # Service / digital filers often omit inventory; treat missing as zero (CA − 0).
        inv_adj = float(inventory) if inventory is not None else 0.0
        quick_ratio_numerator = float(current_assets) - inv_adj
    quick_ratio = _sanitize_metric_value(_safe_div(quick_ratio_numerator, current_liabilities))
    debt_to_equity = _sanitize_metric_value(_safe_div(total_debt, equity))
    debt_to_ebitda = _sanitize_metric_value(_safe_div(total_debt, ebitda))
    debt_to_ebitda_suppressed_scale = False
    # Prevent misleading leverage output when EBITDA is non-positive or ratio is clearly implausible.
    if (ebitda is None) or (ebitda <= 0):
        debt_to_ebitda = None
    elif debt_to_ebitda is not None and abs(float(debt_to_ebitda)) > 100:
        debt_to_ebitda = None
        debt_to_ebitda_suppressed_scale = True
    dscr_numerator: Optional[float] = None
    if ebitda is not None and capex_for_ratio is not None:
        dscr_numerator = ebitda - capex_for_ratio
    # Cash-flow-available-for-debt-service / total debt service (interest + principal).
    debt_service: Optional[float] = None
    try:
        intr_cmp = float(interest_for_coverage) if interest_for_coverage is not None else 0.0
        prin_cmp = float(principal or 0.0)
        ds = intr_cmp + prin_cmp
        debt_service = ds if ds > 0 else None
    except Exception:
        debt_service = None
    dscr = _sanitize_metric_value(_safe_div(dscr_numerator, debt_service))
    # Standardized definition: Interest Coverage = EBIT / Interest Expense
    interest_coverage = (
        _sanitize_metric_value(_safe_div(ebit, interest_for_coverage))
        if interest_for_coverage is not None and float(interest_for_coverage) > 0
        else None
    )
    gross_margin = _sanitize_metric_value(_safe_div((revenue - cogs) if (revenue is not None and cogs is not None) else None, revenue))
    if gross_margin is None and revenue is not None and abs(float(revenue)) > 0 and combined_df is not None and not combined_df.empty:
        gp_abs = _scaled(
            _clean_num(
                _extract_from_combined_df(
                    combined_df,
                    [
                        "gross profit",
                        "gross profit (loss)",
                        "gross margin",
                        "total gross profit",
                    ],
                    preferred_doc_types=financial_doc_scope,
                    preferred_year=locked_year,
                )
            )
        )
        if gp_abs is not None:
            gross_margin = _sanitize_metric_value(_safe_div(float(gp_abs), float(revenue)))
    net_margin = _sanitize_metric_value(_safe_div(net_income, revenue))
    roa = _safe_div(net_income, total_assets)
    roe = _safe_div(net_income, equity)
    ocf_ratio = _sanitize_metric_value(_safe_div(operating_cf, current_liabilities))
    ltv_numerator = loan_amount if loan_amount is not None else total_debt
    ltv = _safe_div(ltv_numerator, collateral_value)

    def _field_scale_from_snippet(metric_terms: List[str]) -> Optional[str]:
        if combined_df is None or combined_df.empty:
            return None
        if "Snippet" not in combined_df.columns:
            return None
        terms = [t.lower() for t in metric_terms]
        snippets = combined_df["Snippet"].dropna().astype(str).tolist()
        for sn in snippets:
            low = sn.lower()
            if not any(t in low for t in terms):
                continue
            m = re.search(r"unit:\s*(thousands|millions|billions|trillions|units)", low)
            if m:
                return m.group(1).capitalize()
        return None

    revenue_scale_hint = _field_scale_from_snippet(["revenue", "sales"])
    net_income_scale_hint = _field_scale_from_snippet(["net income", "net loss", "net earnings", "profit"])
    scale_mismatch_warning = None
    if revenue_scale_hint and net_income_scale_hint and revenue_scale_hint != net_income_scale_hint:
        net_margin = None
        scale_mismatch_warning = "Inconsistent currency scale detected in Income Statement extraction."

    detected_types = set()
    if combined_df is not None and not combined_df.empty and "Detected Type" in combined_df.columns:
        detected_types = {_canonical_doc_type(str(x)) for x in combined_df["Detected Type"].dropna().tolist()}
    sheet_names = [str(x).strip().lower() for x in (combined_df.get("Sheet", pd.Series(dtype=str)).dropna().tolist() if combined_df is not None and "Sheet" in combined_df.columns else [])]
    source_names = [str(x).strip().lower() for x in (combined_df.get("Source Document", pd.Series(dtype=str)).dropna().tolist() if combined_df is not None and "Source Document" in combined_df.columns else [])]
    snippets = [str(x).strip().lower() for x in (combined_df.get("Snippet", pd.Series(dtype=str)).dropna().tolist() if combined_df is not None and "Snippet" in combined_df.columns else [])]
    scan_text = " ".join(source_names + snippets)

    income_markers = [
        "income statement",
        "statement of operations",
        "statements of operations",
        "statement of earnings",
        "profit and loss",
        "p&l",
        "comprehensive income",
        "consolidated statements of operations",
        "consolidated statement of operations",
    ]
    balance_markers = [
        "balance sheet",
        "statement of financial position",
        "consolidated balance sheets",
    ]
    cashflow_markers = [
        "cash flow statement",
        "statement of cash flows",
        "consolidated statements of cash flows",
        "consolidated statement of cash flows",
        "cash flows",
    ]
    debt_schedule_markers = [
        "debt schedule",
        "notes payable",
        "long-term debt",
        "long term debt",
        "maturities of debt",
        "debt obligations",
        "schedule of indebtedness",
    ]

    def _sheet_has_real_values(sheet_name: str) -> bool:
        if combined_df is None or combined_df.empty or "Sheet" not in combined_df.columns:
            return False
        rows = combined_df[combined_df["Sheet"].astype(str).str.lower() == sheet_name.lower()]
        if rows.empty:
            return False
        for _, r in rows.iterrows():
            for c, v in r.items():
                if c in META_COLUMNS or c in {"Source Document", "Detected Type", "Sheet"}:
                    continue
                if v in (None, "", "null", "None"):
                    continue
                if pd.isna(v):
                    continue
                if isinstance(v, str) and v.strip().lower() in {"", "none", "null", "nan"}:
                    continue
                # Any remaining non-empty value means this sheet has real extracted data.
                return True
        return False

    # Step 7: completeness is based on post-mapping/post-derivation field availability.
    has_income = (
        any(v is not None for v in [ebitda, ebit, revenue, net_income, interest_expense])
        or _sheet_has_real_values("Income Statement")
    )
    has_balance = (
        any(v is not None for v in [total_debt, current_assets, current_liabilities, cash, equity, total_assets])
        or _sheet_has_real_values("Balance Sheet")
    )
    has_cashflow = (
        any(v is not None for v in [operating_cf, capex, free_cf])
        or _sheet_has_real_values("Cash Flow")
    )
    has_debt_schedule = (
        ("Term Sheet" in detected_types)
        or ("Credit Agreement" in detected_types)
        or ("Amendment" in detected_types)
        or (principal is not None and principal > 0)
        or (total_debt is not None)
        or any(k in scan_text for k in debt_schedule_markers)
    )

    # Detection must be based on extracted evidence, not table-of-contents/header mentions.
    income_statement_detected = has_income
    balance_sheet_detected = has_balance
    cash_flow_detected = has_cashflow
    debt_schedule_detected = has_debt_schedule

    missing_document_types: List[str] = []
    if not income_statement_detected:
        missing_document_types.append("Income Statement")
    if not balance_sheet_detected:
        missing_document_types.append("Balance Sheet")
    if not cash_flow_detected:
        missing_document_types.append("Cash Flow Statement")
    if not debt_schedule_detected:
        missing_document_types.append("Debt Schedule")

    documents_detected = [
        {"name": "Income Statement", "present": income_statement_detected},
        {"name": "Balance Sheet", "present": balance_sheet_detected},
        {"name": "Cash Flow Statement", "present": cash_flow_detected},
        {"name": "Debt Schedule", "present": debt_schedule_detected},
    ]

    validation_warnings: List[str] = []
    is_public_like = any(tok in scan_text for tok in ["sec", "10-k", "10 k", "10-q", "10 q", "edgar"])
    if is_public_like and revenue is not None and revenue < 1000:
        validation_warnings.append("Revenue sanity check failed: extracted revenue is below 1,000 for a public filing.")
    if revenue is not None and net_income is not None and net_income > revenue:
        validation_warnings.append("Net income consistency check failed: Net Income is greater than Revenue.")
    revenue_for_unit_check = revenue
    if revenue is not None and apply_unit_factor and detected_factor > 1:
        revenue_for_unit_check = float(revenue) / float(detected_factor)
    if revenue_for_unit_check is not None and str(currency_scale).strip().lower() == "millions" and revenue_for_unit_check > 500_000:
        validation_warnings.append("Scaling validation error: Revenue exceeds 500,000 while unit is Millions. Check currency normalization.")
    if revenue is not None and total_debt is not None and abs(total_debt - revenue) < 1e-9:
        validation_warnings.append("Debt consistency check failed: Total Debt equals Revenue exactly. Re-validate mapping.")
    if debt_to_equity is not None and debt_to_equity > 10:
        validation_warnings.append("Scaling validation error: Debt to Equity exceeds 10. Check currency normalization.")
    if roa is not None and roa < 0:
        validation_warnings.append("Scaling validation error: ROA is below 0. Check extraction scale/sign consistency.")
    if roe is not None and roe < 0:
        validation_warnings.append("Scaling validation error: ROE is below 0. Check extraction scale/sign consistency.")
    if interest_coverage is not None and interest_coverage < 0 and ebit is not None and ebit > 0:
        validation_warnings.append("Sign normalization error: Interest Coverage is negative while EBIT is positive.")
    if debt_to_ebitda_suppressed_scale:
        validation_warnings.append("Debt to EBITDA suppressed due to likely scale mismatch (abs ratio > 100).")
    if ebit_artifact_suppressed:
        validation_warnings.append("EBIT suppressed due to tiny-value extraction artifact relative to revenue.")
    if ebitda_artifact_suppressed:
        validation_warnings.append("EBITDA suppressed due to tiny-value extraction artifact relative to revenue.")
    if cogs_suspect_equal_revenue:
        validation_warnings.append("COGS suppressed because extracted value is effectively equal to Revenue (likely mapping artifact).")

    industry_thresholds = {
        "Current Ratio": 1.25,
        "Quick Ratio": 1.00,
        "Debt to Equity": 2.00,
        "Debt to EBITDA": 3.50,
        "DSCR": 1.25,
        "Interest Coverage": 2.00,
        "Gross Margin": 0.25,
        "Net Margin": 0.08,
        "ROA": 0.04,
        "ROE": 0.12,
        "Operating Cash Flow Ratio": 0.20,
    }
    industry_overrides = {
        "Oil & Gas": {"DSCR": 1.75, "Debt to EBITDA": 2.50, "Gross Margin": 0.20},
        "Construction": {"DSCR": 1.50, "Debt to EBITDA": 3.00, "Gross Margin": 0.18},
        "Technology": {"DSCR": 1.25, "Debt to EBITDA": 4.00, "Gross Margin": 0.40},
        "Healthcare": {"DSCR": 1.30, "Debt to EBITDA": 3.50, "Gross Margin": 0.35},
    }
    if industry in industry_overrides:
        industry_thresholds.update(industry_overrides[industry])

    ratio_rows = [
        (
            "Liquidity",
            "Current Ratio",
            current_ratio,
            industry_thresholds["Current Ratio"],
            True,
            "current_assets/current_liabilities",
            (current_assets is not None and current_liabilities not in (None, 0)),
            "Missing Balance Sheet Data",
        ),
        (
            "Liquidity",
            "Quick Ratio",
            quick_ratio,
            industry_thresholds["Quick Ratio"],
            True,
            "(current_assets-inventory)/current_liabilities OR (cash+accounts_receivable)/current_liabilities",
            (quick_ratio_numerator is not None and current_liabilities not in (None, 0)),
            "Missing Current Assets/Inventory/Current Liabilities",
        ),
        (
            "Leverage",
            "Debt to Equity",
            debt_to_equity,
            industry_thresholds["Debt to Equity"],
            False,
            "total_debt/equity",
            (total_debt is not None and equity not in (None, 0)),
            "Missing Balance Sheet Data",
        ),
        (
            "Leverage",
            "Debt to EBITDA",
            debt_to_ebitda,
            industry_thresholds["Debt to EBITDA"],
            False,
            "total_debt/ebitda",
            (total_debt is not None and ebitda is not None and ebitda > 0),
            "Required fields 'Total Debt' or positive 'EBITDA' not found. Ratio cannot be calculated.",
        ),
        (
            "Coverage",
            "DSCR",
            dscr,
            industry_thresholds["DSCR"],
            True,
            "(ebitda-capex)/(interest_expense+principal_repayment)",
            (dscr_numerator is not None and debt_service is not None and debt_service > 0),
            "Required fields 'EBITDA', 'CapEx', or positive debt service (interest + principal) not found. Ratio cannot be calculated.",
        ),
        (
            "Coverage",
            "Interest Coverage",
            interest_coverage,
            industry_thresholds["Interest Coverage"],
            True,
            "ebit/interest_expense",
            (ebit is not None and interest_for_coverage not in (None, 0)),
            "Required field 'Interest Expense' not found in uploaded document. Ratio cannot be calculated.",
        ),
        (
            "Profitability",
            "Gross Margin",
            gross_margin,
            industry_thresholds["Gross Margin"],
            True,
            "(revenue-cogs)/revenue",
            (revenue not in (None, 0) and cogs is not None),
            "Required fields 'Revenue' or 'COGS' not found. Ratio cannot be calculated.",
        ),
        (
            "Profitability",
            "Net Margin",
            net_margin,
            industry_thresholds["Net Margin"],
            True,
            "net_income/revenue",
            (net_income is not None and revenue not in (None, 0)),
            "Income Statement document not detected. Please upload a valid Consolidated Statement of Operations.",
        ),
        (
            "Profitability",
            "ROA",
            roa,
            industry_thresholds["ROA"],
            True,
            "net_income/total_assets",
            (net_income is not None and total_assets not in (None, 0)),
            "Missing Income Statement + Balance Sheet Data",
        ),
        (
            "Profitability",
            "ROE",
            roe,
            industry_thresholds["ROE"],
            True,
            "net_income/equity",
            (net_income is not None and equity not in (None, 0)),
            "Missing Income Statement + Balance Sheet Data",
        ),
        (
            "Cash Flow",
            "Operating Cash Flow Ratio",
            ocf_ratio,
            industry_thresholds["Operating Cash Flow Ratio"],
            True,
            "operating_cash_flow/current_liabilities",
            (operating_cf is not None and current_liabilities not in (None, 0)),
            "Missing Cash Flow Statement + Balance Sheet Data",
        ),
        ("Cash Flow", "Free Cash Flow", free_cf_display, 0.0, True, "operating_cash_flow-capex", (operating_cf is not None and capex_for_ratio is not None) or (free_cf is not None), "Required fields 'Operating Cash Flow' or 'Capital Expenditures' not found."),
    ]

    ind_mult = INDUSTRY_RISK_MULTIPLIER.get(industry, 1.0)
    geo_mult = GEOGRAPHY_MULTIPLIER.get(geography, 1.0)
    stage_mult = BUSINESS_STAGE_MULTIPLIER.get(business_stage, 1.0)
    combined_multiplier = ind_mult * geo_mult * stage_mult

    rows: List[Dict[str, Any]] = []
    calc_trace: List[Dict[str, Any]] = []
    category_scores: Dict[str, List[float]] = {k: [] for k in ["Liquidity", "Leverage", "Coverage", "Profitability", "Cash Flow"]}
    strengths: List[str] = []
    weaknesses: List[str] = []
    incomplete_metrics: List[str] = []
    missing_reasons: List[str] = []
    complete_metric_count = 0

    for category, metric, value, threshold, hib, formula, deps_ok, missing_reason in ratio_rows:
        if not deps_ok:
            base_score, color = (None, "Incomplete")
            value = None
            incomplete_metrics.append(metric)
            missing_reasons.append(missing_reason)
            metric_status = "Incomplete"
        else:
            base_score, color = _score_metric(value, threshold, hib)
            if base_score is None:
                incomplete_metrics.append(metric)
                missing_reasons.append("Missing required financial data")
                metric_status = "Incomplete"
            else:
                complete_metric_count += 1
                category_scores[category].append(base_score)
                metric_status = "Calculated"

        adjusted = (base_score / combined_multiplier) if (base_score is not None and combined_multiplier) else None
        if color == "Green":
            strengths.append(metric)
        elif color == "Red":
            weaknesses.append(metric)

        out_value = None if value is None or pd.isna(value) else value
        out_threshold = None if threshold is None or (isinstance(threshold, float) and pd.isna(threshold)) else threshold
        out_base = None if base_score is None else float(base_score)
        out_adjusted = None if adjusted is None else round(float(adjusted), 2)

        rows.append(
            {
                "Category": category,
                "Metric": metric,
                "Calculated Value": out_value,
                "Industry Threshold": out_threshold,
                "Base Score": out_base,
                "Adjusted Score": out_adjusted,
                "Status": metric_status,
                "Risk": _color_to_row_level(color),
                "_row_level": _color_to_row_level(color),
                "_risk_score": (100.0 - out_base) if out_base is not None else None,
                "Source Trace": formula if out_value is not None else (missing_reason if not deps_ok else "Missing required financial data"),
            }
        )
        calc_trace.append(
            {
                "Metric": metric,
                "Extracted Value": out_value,
                "Threshold": out_threshold,
                "Formula": formula,
                "Intermediate": {"base_score": out_base, "multiplier": round(combined_multiplier, 4)},
                "Final Output": {"adjusted_score": out_adjusted, "risk": _color_to_row_level(color)},
            }
        )

    qualitative = 80.0
    qualitative -= {"Small": 12.0, "Medium": 6.0, "Large": 0.0}.get(company_size, 6.0)
    if years_in_operation < 3:
        qualitative -= 25.0
    elif years_in_operation < 5:
        qualitative -= 12.0
    qualitative -= {"Startup": 15.0, "Growth": 6.0, "Mature": 0.0, "Declining": 20.0}.get(business_stage, 6.0)
    qualitative = max(20.0, min(100.0, qualitative))

    def _avg(xs: List[float]) -> Optional[float]:
        return (sum(xs) / len(xs)) if xs else None

    component_quality = {
        "Liquidity": _avg(category_scores["Liquidity"]),
        "Leverage": _avg(category_scores["Leverage"]),
        "Coverage": _avg(category_scores["Coverage"]),
        "Profitability": _avg(category_scores["Profitability"]),
        "Cash Flow": _avg(category_scores["Cash Flow"]),
        "Qualitative": qualitative,
    }
    component_weights = {
        "Liquidity": 0.20,
        "Leverage": 0.20,
        "Coverage": 0.25,
        "Profitability": 0.15,
        "Cash Flow": 0.15,
        "Qualitative": 0.05,
    }

    numer = 0.0
    denom = 0.0
    for name, score in component_quality.items():
        if score is None:
            continue
        w = component_weights[name]
        numer += score * w
        denom += w
    quality_score = (numer / denom) if denom > 0 else 50.0
    adjusted_quality = quality_score / combined_multiplier if combined_multiplier else quality_score
    adjusted_quality = max(0.0, min(100.0, adjusted_quality))

    # Single consistent scoring model:
    # final_score is credit quality where higher = better.
    final_score = max(0.0, min(100.0, adjusted_quality))
    total_metrics = len(ratio_rows)
    missing_count = len(incomplete_metrics)
    missing_ratio = (missing_count / total_metrics) if total_metrics > 0 else 1.0
    required_liquidity_missing = current_ratio is None and quick_ratio is None
    required_leverage_missing = debt_to_equity is None and debt_to_ebitda is None
    provisional = (missing_ratio > 0.40) or required_liquidity_missing or required_leverage_missing
    grade = "Provisional" if provisional else _risk_category_from_score(final_score)
    overall_color = "Unknown" if provisional else ("Green" if final_score > 80 else "Yellow" if final_score >= 60 else "Red")

    # Policy capacity is independent from requested amount; final approved is capped by request.
    industry_ebitda_multiple = {
        "Oil & Gas": 3.0,
        "Construction": 3.25,
        "Hospitality": 3.0,
        "Retail": 2.75,
        "Manufacturing": 3.5,
        "Transportation": 3.0,
        "Technology": 4.0,
        "Healthcare": 3.75,
        "Financial Services": 3.0,
        "Agriculture": 2.75,
        "Real Estate": 3.25,
    }
    ltv_policy_limit = {
        "Oil & Gas": 0.60,
        "Construction": 0.65,
        "Hospitality": 0.65,
        "Retail": 0.65,
        "Manufacturing": 0.70,
        "Transportation": 0.65,
        "Technology": 0.60,
        "Healthcare": 0.70,
        "Financial Services": 0.60,
        "Agriculture": 0.65,
        "Real Estate": 0.75,
    }
    # Dynamic internal exposure cap based on financial capacity (not static company-size buckets).
    dynamic_exposure_caps: List[float] = []
    if revenue is not None and revenue > 0:
        dynamic_exposure_caps.append(revenue * 0.35)
    if total_assets is not None and total_assets > 0:
        dynamic_exposure_caps.append(total_assets * 0.25)
    if equity is not None and equity > 0:
        dynamic_exposure_caps.append(equity * 0.60)
    if collateral_value is not None and collateral_value > 0:
        dynamic_exposure_caps.append(collateral_value * ltv_policy_limit.get(industry, 0.70))
    if ebitda is not None and ebitda > 0:
        dynamic_exposure_caps.append(ebitda * industry_ebitda_multiple.get(industry, 3.0) * 1.25)
    internal_max_exposure = min([x for x in dynamic_exposure_caps if x is not None and x > 0], default=250_000_000.0)
    risk_haircut = 1.0
    if provisional:
        risk_haircut *= 0.80
    elif final_score < 40:
        risk_haircut *= 0.75
    elif final_score < 60:
        risk_haircut *= 0.90
    internal_max_exposure = max(5_000_000.0, min(internal_max_exposure * risk_haircut, 2_500_000_000.0))
    policy_candidates: List[float] = []
    ebitda_capacity_used = False
    dscr_capacity_used = False
    icr_capacity_used = False
    if ebitda is not None and ebitda > 0:
        policy_candidates.append(ebitda * industry_ebitda_multiple.get(industry, 3.0))
        ebitda_capacity_used = True
    target_dscr = industry_thresholds["DSCR"]
    if dscr is not None and ebitda is not None and target_dscr and target_dscr > 0:
        earnings_for_capacity = ebitda
        annual_debt_service_capacity = earnings_for_capacity / target_dscr
        assumed_debt_service_rate = 0.15
        dscr_capacity_limit = annual_debt_service_capacity / assumed_debt_service_rate
        policy_candidates.append(dscr_capacity_limit)
        dscr_capacity_used = True
    if collateral_value is not None:
        policy_candidates.append(collateral_value * ltv_policy_limit.get(industry, 0.70))
    if not policy_candidates and revenue is not None and float(revenue) > 0:
        # EBITDA/DSCR paths empty (common when EBITDA not mapped) — rough revenue-based capacity floor.
        policy_candidates.append(max(1_000_000.0, float(revenue) * 0.10))
    base_policy_limit = min([c for c in policy_candidates if c is not None and c >= 0], default=0.0)

    # Transparent adjustment breakdown for underwriting UI.
    industry_limit_adjustment = {
        "Technology": 0.15,
        "Healthcare": 0.10,
        "Manufacturing": 0.05,
        "Financial Services": 0.05,
        "Retail": -0.05,
        "Hospitality": -0.05,
        "Construction": -0.03,
        "Transportation": -0.03,
        "Oil & Gas": -0.08,
        "Agriculture": -0.03,
        "Real Estate": 0.00,
    }
    geography_limit_adjustment = {
        "United States Tier 1": 0.10,
        "United States Tier 2": 0.05,
        "Canada": 0.08,
        "Emerging Market - Low Stability": -0.08,
        "Emerging Market - High Volatility": -0.12,
        "Sanctioned / High Risk Region": -0.20,
    }
    maturity_adjustment = {
        "Mature": 0.05,
        "Growth": 0.02,
        "Startup": -0.05,
        "Declining": -0.10,
    }
    ind_adj = industry_limit_adjustment.get(industry, 0.0)
    geo_adj = geography_limit_adjustment.get(geography, 0.0)
    mat_adj = maturity_adjustment.get(business_stage, 0.0)
    adjusted_policy_limit = base_policy_limit
    adjusted_policy_limit *= (1.0 + ind_adj)
    adjusted_policy_limit *= (1.0 + geo_adj)
    adjusted_policy_limit *= (1.0 + mat_adj)
    policy_capacity_limit = min(max(0.0, adjusted_policy_limit), internal_max_exposure)
    # Bank-style final approval cannot exceed borrower requested amount.
    policy_approved_limit = min(policy_capacity_limit, max(0.0, float(requested_amount)))

    table_df = pd.DataFrame(rows).where(pd.notna(pd.DataFrame(rows)), None)
    top_risk_df = (
        table_df.dropna(subset=["_risk_score"])
        .sort_values("_risk_score", ascending=False)
        .head(3)[["Metric", "Risk", "_risk_score"]]
        .rename(columns={"_risk_score": "Risk Score"})
        .where(pd.notna, None)
    )

    data_completeness_pct = round((complete_metric_count / total_metrics) * 100.0, 1) if total_metrics > 0 else 0.0
    confidence_level = "Full Review" if data_completeness_pct >= 85 else ("Enhanced" if data_completeness_pct >= 60 else "Basic")

    hard_validation_errors: List[str] = []
    revenue_in_statement_units = revenue
    if revenue is not None and apply_unit_factor and detected_factor > 1:
        revenue_in_statement_units = float(revenue) / float(detected_factor)
    # Mega-cap revenue can exceed 500k "millions" units; only flag extreme scale contradictions.
    if detected_scale == "Millions" and revenue_in_statement_units is not None and revenue_in_statement_units > 2_000_000:
        hard_validation_errors.append("Validation halt: Revenue exceeds 2,000,000 while statement unit is Millions (check units).")
    # Net-cash / very low-leverage issuers have tiny positive D/E; do not treat as extraction failure.
    if debt_to_equity is not None and debt_to_equity < 0 and equity is not None and float(equity) > 0:
        hard_validation_errors.append("Validation halt: Negative Debt to Equity with positive equity (sign extraction error).")
    if interest_coverage is not None and interest_coverage < 0 and ebit is not None and ebit > 0:
        hard_validation_errors.append("Validation halt: Interest Coverage is negative with positive EBIT (sign normalization error).")
    if base_policy_limit > 0 and ebitda is not None and ebitda > 0 and base_policy_limit > (50.0 * ebitda):
        hard_validation_errors.append("Validation halt: Base policy limit exceeds 50x EBITDA.")
    if roa is not None and roa > 1:
        hard_validation_errors.append("Validation halt: ROA above 1.0 indicates likely asset mapping/scaling error.")
    for e in precheck_errors:
        hard_validation_errors.append(f"Validation halt: {e}")
    has_statement_context = False
    if combined_df is not None and not combined_df.empty and "Sheet" in combined_df.columns:
        sheet_vals = {str(x).strip().lower() for x in combined_df["Sheet"].dropna().tolist()}
        has_statement_context = bool(sheet_vals.intersection({"income statement", "balance sheet", "cash flow"}))
    # Some manifest fields are quality/completeness signals but should not hard-stop
    # underwriting when core ratio inputs are already present.
    non_blocking_missing_fields = {
        "cash_flow.depreciation_amortization",
        "cash_flow.operating_cf",
        "cash_flow.capex",
        "balance_sheet.inventory",
        "balance_sheet.st_debt",
    }
    blocking_missing_fields = [f for f in missing_required_fields if f not in non_blocking_missing_fields]
    if blocking_missing_fields and has_statement_context:
        hard_validation_errors.append(
            "Validation halt: Missing required input fields: " + ", ".join(sorted(set(blocking_missing_fields)))
        )
    halted_for_validation = len(hard_validation_errors) > 0

    recommendation = "Decline"
    if provisional and not halted_for_validation:
        recommendation = "Conditional Approval"
    elif not halted_for_validation:
        if policy_approved_limit <= 0:
            recommendation = "Decline"
        elif final_score > 80 and requested_amount <= policy_approved_limit:
            recommendation = "Approval"
        elif final_score >= 60:
            recommendation = "Conditional Approval"

    if halted_for_validation:
        grade = "Validation Error"
        overall_color = "Red"
        policy_approved_limit = 0.0

    covenant_cushions: List[float] = []
    covenant_breach_flag = False
    if debt_to_ebitda is not None and industry_thresholds["Debt to EBITDA"] is not None and industry_thresholds["Debt to EBITDA"] != 0:
        covenant_cushions.append(((industry_thresholds["Debt to EBITDA"] - debt_to_ebitda) / industry_thresholds["Debt to EBITDA"]) * 100.0)
    if dscr is not None and industry_thresholds["DSCR"] is not None and industry_thresholds["DSCR"] != 0:
        if dscr <= 0:
            covenant_breach_flag = True
        else:
            covenant_cushions.append(((dscr - industry_thresholds["DSCR"]) / industry_thresholds["DSCR"]) * 100.0)
    if interest_coverage is not None and industry_thresholds["Interest Coverage"] is not None and industry_thresholds["Interest Coverage"] != 0:
        if interest_coverage <= 0:
            covenant_breach_flag = True
        else:
            covenant_cushions.append(((interest_coverage - industry_thresholds["Interest Coverage"]) / industry_thresholds["Interest Coverage"]) * 100.0)
    covenant_cushion_pct = min(covenant_cushions) if covenant_cushions else None
    covenant_cushion_display_pct = None
    covenant_cushion_note = None
    if covenant_cushion_pct is not None:
        covenant_cushion_display_pct = max(-100.0, min(100.0, float(covenant_cushion_pct)))
        if abs(float(covenant_cushion_pct)) > 100.0:
            covenant_cushion_note = "Cushion capped at +/-100% for readability."

    explanation_lines: List[str] = []
    for r in rows:
        metric_name = str(r.get("Metric", "Metric"))
        rv = r.get("Calculated Value")
        thr = r.get("Industry Threshold")
        if rv is None or thr is None:
            explanation_lines.append(f"{metric_name}: insufficient data to evaluate against threshold.")
            continue
        relation = "meets"
        try:
            rvf = float(rv)
            thrf = float(thr)
            higher_is_better = metric_name in {"Current Ratio", "Quick Ratio", "DSCR", "Interest Coverage", "Gross Margin", "Net Margin", "ROA", "ROE", "Operating Cash Flow Ratio", "Free Cash Flow"}
            if higher_is_better and rvf < thrf:
                relation = "falls below"
            elif (not higher_is_better) and rvf > thrf:
                relation = "exceeds"
            elif higher_is_better and rvf >= thrf:
                relation = "exceeds"
            elif (not higher_is_better) and rvf <= thrf:
                relation = "is below"
        except Exception:
            relation = "is compared against"
        explanation_lines.append(f"{metric_name} of {rv} {relation} adjusted {industry} threshold of {thr}.")

    executive_summary = (
        f"Final Risk Rating: {grade} | Weighted Score: {final_score:.1f} | "
        f"Approved Limit: {policy_approved_limit:,.0f} vs Requested: {requested_amount:,.0f} | "
        f"Policy Capacity: {policy_capacity_limit:,.0f} | "
        f"Industry Multiplier: {ind_mult:.2f}, Geography Multiplier: {geo_mult:.2f}, Stage Multiplier: {stage_mult:.2f} | Currency Scale: {currency_scale}."
    )
    strengths_line = ", ".join(strengths[:6]) if strengths else "No metrics currently exceed adjusted thresholds."
    gaps = sorted(set(incomplete_metrics + weaknesses))
    gaps_line = ", ".join(gaps[:8]) if gaps else "No major gaps detected."
    if covenant_breach_flag:
        cushion_line = "Breach detected (one or more coverage covenants at or below zero)."
    elif covenant_cushion_display_pct is None:
        cushion_line = "Not available (missing covenant inputs)."
    else:
        cushion_line = f"{covenant_cushion_display_pct:.1f}%"
        if covenant_cushion_note:
            cushion_line += f" ({covenant_cushion_note})"
    recommendation_line = recommendation

    memo = (
        "Executive Summary: " + executive_summary + "\n\n"
        f"Analysis Mode: {analysis_mode} | Fiscal Year Used: {locked_year if locked_year is not None else 'N/A'} | Metrics Completed: {complete_metric_count}/{total_metrics} | Missing Metrics: {missing_count}\n\n"
        "Financial Strengths: " + strengths_line + "\n\n"
        "Key Risks & Data Gaps: " + gaps_line + "\n\n"
        "Covenant Cushion Analysis: " + cushion_line + "\n\n"
        "Recommendation: " + recommendation_line
    )
    if incomplete_metrics:
        if recommendation == "Approval":
            memo += "\n\nCertain ratios could not be calculated due to missing documentation. Approval granted based on available evidence; collect missing statements for post-approval monitoring."
        else:
            memo += "\n\nCertain ratios could not be calculated due to missing documentation. Final recommendation is conditional upon receipt of complete statements."

    missing_data_warning = "Some ratios cannot be calculated due to missing documents." if incomplete_metrics else None
    if validation_warnings:
        prefix = missing_data_warning + " " if missing_data_warning else ""
        missing_data_warning = prefix + " | ".join(validation_warnings)
    if hard_validation_errors:
        prefix = missing_data_warning + " " if missing_data_warning else ""
        missing_data_warning = prefix + " | ".join(hard_validation_errors)
    if scale_mismatch_warning:
        prefix = missing_data_warning + " " if missing_data_warning else ""
        missing_data_warning = prefix + scale_mismatch_warning
    if quick_ratio is not None and quick_ratio < 0.1 and is_public_like:
        prefix = missing_data_warning + " " if missing_data_warning else ""
        missing_data_warning = prefix + "Possible cash or AR extraction error — ratio unusually low."

    return {
        "table": table_df,
        "final_score": round(final_score, 1),
        "grade": grade,
        "overall_color": overall_color,
        "approved_limit": round(policy_approved_limit, 0),
        "policy_approved_limit": round(policy_approved_limit, 0),
        "policy_limit_breakdown": {
            "base_policy_limit": base_policy_limit,
            "industry_adjustment_pct": ind_adj,
            "geography_adjustment_pct": geo_adj,
            "maturity_adjustment_pct": mat_adj,
            "final_policy_limit": policy_capacity_limit,
            "approved_limit_final": policy_approved_limit,
            "internal_cap": internal_max_exposure,
        },
        "recommendation": recommendation,
        "requested_amount": requested_amount,
        "currency_scale": currency_scale,
        "top_risk": top_risk_df,
        "memo": memo,
        "incomplete_metrics": incomplete_metrics,
        "missing_data_warning": missing_data_warning,
        "missing_data_reasons": sorted(set(missing_reasons)),
        "validation_warnings": validation_warnings,
        "hard_validation_errors": hard_validation_errors,
        "halted_for_validation": halted_for_validation,
        "documents_detected": documents_detected,
        "missing_document_types": missing_document_types,
        "data_completeness_pct": data_completeness_pct,
        "confidence_level": confidence_level,
        "scoring_basis": f"{complete_metric_count}/{total_metrics} metrics scored (incomplete metrics excluded from denominator)",
        "threshold_explanations": explanation_lines,
        "covenant_cushion_pct": covenant_cushion_pct,
        "covenant_cushion_display_pct": covenant_cushion_display_pct,
        "covenant_cushion_note": covenant_cushion_note,
        "covenant_breach_flag": covenant_breach_flag,
        "combined_multiplier": combined_multiplier,
        "multiplier_components": {
            "industry": ind_mult,
            "geography": geo_mult,
            "business_stage": stage_mult,
        },
        "ratios": {
            "Current Ratio": current_ratio,
            "Quick Ratio": quick_ratio,
            "Debt to Equity": debt_to_equity,
            "Debt to EBITDA": debt_to_ebitda,
            "DSCR": dscr,
            "Interest Coverage": interest_coverage,
            "Gross Margin": gross_margin,
            "LTV": ltv,
        },
        "capacity_components_used": {
            "ebitda_leverage": ebitda_capacity_used,
            "dscr_capacity": dscr_capacity_used,
            "interest_coverage_capacity": icr_capacity_used,
        },
        "calc_trace": calc_trace,
        "analysis_mode": analysis_mode,
        "locked_year": locked_year,
        "available_years": available_years or [],
        "cross_year_error_flag": bool(cross_year_error_flag),
        "t12m_status": t12m_status,
        "analysis_audit": {
            "analysis_mode": analysis_mode,
            "locked_year": locked_year,
            "available_years": available_years or [],
            "incomplete_metric_count": missing_count,
            "t12m_status": t12m_status,
            "cross_year_error_flag": bool(cross_year_error_flag),
        },
        "document_profile": document_profile,
        "required_fields_manifest_missing": sorted(set(missing_required_fields)),
    }


def _apply_filters(df: pd.DataFrame, show_nulls: bool, min_conf: float, search: str) -> pd.DataFrame:
    out = df.copy()
    if out.empty:
        return out

    if not show_nulls:
        business_cols = [c for c in out.columns if c not in {"Sheet", "Confidence", "Page", "Open PDF Page", "Snippet"} and not str(c).startswith("__meta__")]
        mask = pd.Series(False, index=out.index)
        for c in business_cols:
            mask = mask | out[c].notna()
        out = out[mask]

    if min_conf > 0 and "Confidence" in out.columns:
        out = out[(out["Confidence"].isna()) | (out["Confidence"] >= min_conf)]

    if search:
        s = search.lower().strip()
        mask = pd.Series(False, index=out.index)
        for c in out.columns:
            mask = mask | out[c].astype(str).str.lower().str.contains(s, na=False)
        out = out[mask]

    return out


def _ui_null_df(df: pd.DataFrame) -> pd.DataFrame:
    return _ui_null_df_with_labels(df, null_label="–", derived_null_label="N/A")


def _ui_null_df_with_labels(
    df: pd.DataFrame,
    *,
    null_label: str = "–",
    derived_null_label: str = "N/A",
    derived_cols: Optional[set] = None,
) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    # Ensure UI never shows NaN/None; render missing values explicitly and format numbers with commas.
    out = df.copy()
    def _sanitize_cell_text(v: Any) -> Any:
        if v is None:
            return v
        s = str(v)
        if "<" in s and ">" in s:
            # Guard against accidental HTML/styler leakage into dataframe cells.
            s = re.sub(r"<[^>]+>", "", s)
        return html.unescape(s)
    derived_cols = derived_cols or DERIVED_COLUMNS
    no_comma_cols = {"Selected Year"}
    for c in out.columns:
        if c in {"Confidence"}:
            continue
        if pd.api.types.is_numeric_dtype(out[c]):
            out[c] = out[c].apply(
                lambda v: (derived_null_label if c in derived_cols else null_label)
                if v is None or pd.isna(v)
                else (
                    f"{int(v)}"
                    if (c in no_comma_cols and float(v).is_integer())
                    else (f"{int(v):,}" if float(v).is_integer() else f"{float(v):,.3f}".rstrip("0").rstrip("."))
                )
            )
        else:
            out[c] = out[c].apply(
                lambda v: (derived_null_label if c in derived_cols else null_label)
                if v is None or (isinstance(v, float) and pd.isna(v))
                else _sanitize_cell_text(v)
            )
    if "Confidence" in out.columns:
        out["Confidence"] = out["Confidence"].apply(
            lambda v: null_label if v is None or (isinstance(v, float) and pd.isna(v)) else f"{float(v):.2f}"
        )
    return out


def _confidence_tag_for_meta(meta: Optional[Dict[str, Any]]) -> str:
    if not isinstance(meta, dict):
        return ""
    match_type = str(meta.get("match_type") or meta.get("source") or "").lower()
    if match_type in {"exact"}:
        return "conf-exact"
    if match_type in {"fuzzy", "regex"}:
        return "conf-fuzzy"
    if match_type in {"derived", "computed", "estimated"}:
        return "conf-derived"
    return "conf-fuzzy"


def _apply_confidence_cell_classes(view: pd.DataFrame) -> pd.DataFrame:
    if view is None or view.empty:
        return view
    out = view.copy()
    for col in list(out.columns):
        meta_col = f"__meta__{col}"
        if meta_col not in out.columns:
            continue
        for idx in out.index:
            val = out.at[idx, col]
            if val in ("–", "N/A") or val is None:
                continue
            tag = _confidence_tag_for_meta(out.at[idx, meta_col])
            if tag:
                out.at[idx, col] = f"{val}||{tag}"
    return out


def _style_with_tooltips(df: pd.DataFrame, *, missing_tooltip: str, na_tooltip: str) -> pd.io.formats.style.Styler:
    tooltip_df = pd.DataFrame("", index=df.index, columns=df.columns)
    for c in df.columns:
        for idx in df.index:
            v = df.at[idx, c]
            if v == "–":
                tooltip_df.at[idx, c] = missing_tooltip
            elif v == "N/A":
                tooltip_df.at[idx, c] = na_tooltip
    return df.style.set_tooltips(tooltip_df)


def _format_abbrev_number(value: Any) -> str:
    num = _to_numeric_financial(value)
    if num is None:
        return "NULL"
    n = float(num)
    sign = "-" if n < 0 else ""
    abs_n = abs(n)
    if abs_n >= 1_000_000_000_000:
        return f"{sign}{abs_n/1_000_000_000_000:.2f}T"
    if abs_n >= 1_000_000_000:
        return f"{sign}{abs_n/1_000_000_000:.2f}B"
    if abs_n >= 1_000_000:
        return f"{sign}{abs_n/1_000_000:.2f}M"
    if abs_n >= 1_000:
        return f"{sign}{abs_n/1_000:.2f}K"
    if float(abs_n).is_integer():
        return f"{sign}{int(abs_n):,}"
    return f"{sign}{abs_n:,.2f}"


def _format_scaled_with_unit(value: Any, unit: str) -> str:
    num = _to_numeric_financial(value)
    if num is None:
        return "NULL"
    try:
        txt = f"{float(num):,.2f}".rstrip("0").rstrip(".")
    except Exception:
        return "NULL"
    return f"{txt}{unit or ''}"


def main() -> None:
    qp = st.query_params
    qp_doc = qp.get("open_doc")
    qp_page = qp.get("open_page")
    qp_module = qp.get("module")
    qp_selected = qp.get("selected")
    if isinstance(qp_doc, list):
        qp_doc = qp_doc[0] if qp_doc else None
    if isinstance(qp_page, list):
        qp_page = qp_page[0] if qp_page else None
    if isinstance(qp_module, list):
        qp_module = qp_module[0] if qp_module else None
    if isinstance(qp_selected, list):
        qp_selected = qp_selected[0] if qp_selected else None
    if qp_doc:
        candidate = Path(unquote(str(qp_doc)))
        if _is_allowed_pdf_path(candidate):
            st.session_state["selected_file"] = str(candidate.resolve())
    if qp_selected:
        selected_candidate = Path(unquote(str(qp_selected)))
        if _is_allowed_pdf_path(selected_candidate):
            st.session_state["selected_file"] = str(selected_candidate.resolve())
    if qp_page:
        try:
            st.session_state["viewer_page"] = max(1, int(str(qp_page)))
        except ValueError:
            st.session_state["viewer_page"] = 1

    st.markdown(
        f"""
        <div class="app-header">
          <div class="app-head-left">
            <div class="app-icon"><span class="icon-glyph">shield</span></div>
            <div>
              <h1 class="app-title">Covenant & Credit Risk Analyzer</h1>
              <p class="app-subtitle">Professional Financial Analysis | {APP_BUILD_ID}</p>
            </div>
          </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    files = _list_library_files()

    idx = 0
    if "selected_file" in st.session_state:
        for i, p in enumerate(files):
            if str(p) == st.session_state["selected_file"]:
                idx = i + 1
                break

    valid_modules = {"upload", "extract", "graph", "risk", "repo"}
    if "active_module" not in st.session_state:
        st.session_state["active_module"] = qp_module if qp_module in valid_modules else "upload"
    elif qp_module in valid_modules and qp_module != st.session_state["active_module"]:
        st.session_state["active_module"] = qp_module
    active_module = st.session_state["active_module"]
    prev_module = st.session_state.get("_prev_main_module")
    if prev_module == "risk" and active_module != "risk" and st.session_state.get("risk_has_unsaved_changes"):
        st.warning("You have unsaved risk input changes from Credit Risk Dashboard.")
    st.session_state["_prev_main_module"] = active_module
    left_col, right_col = st.columns([0.9, 2.8], vertical_alignment="top")
    with left_col:
        nav_items = [
            ("upload", "📄", "Upload Document"),
            ("extract", "🧾", "Data Extraction"),
            ("graph", "📊", "Data Visualization"),
            ("risk", "🛡", "Credit Risk Dashboard"),
            ("repo", "🗂", "Document Repository"),
        ]
        nav_links = []
        selected_param = ""
        current_selected = st.session_state.get("selected_file")
        if current_selected:
            selected_param = f"&selected={quote(str(current_selected), safe='')}"
        for module_id, icon_txt, label in nav_items:
            active_cls = "active" if active_module == module_id else ""
            nav_links.append(
                f'<a class="left-nav-link {active_cls}" href="?module={module_id}{selected_param}" target="_self" rel="noopener noreferrer">'
                f'<span class="nav-icon">{icon_txt}</span>'
                f'<span class="nav-label">{label}</span>'
                "</a>"
            )
        st.markdown(
            f"""
            <div class="left-nav-panel">
            <div class="left-nav-brand">
              <h3>Covenant & Credit Risk Analyzer</h3>
              <p>Professional Financial Analysis</p>
            </div>
            <div class="left-nav-modules">Modules</div>
            <div class="left-nav-links">
              {''.join(nav_links)}
            </div>
            </div>
            """,
            unsafe_allow_html=True,
        )
    with right_col:
        selected_pdf: Optional[Path] = None
        extract_requested = False
        mode = "Auto-detect"
        manual = ""

        def _run_extraction_for_selected(pdf_path: Path) -> bool:
            if not pdf_path.exists() or pdf_path.stat().st_size <= 0:
                st.error("Selected document is empty or missing. Please re-upload and try again.")
                return False
            if _looks_like_search_results_pdf(pdf_path):
                st.error(
                    "This PDF looks like a search-results page, not a financial filing statement. "
                    "Please open the actual 10-K/10-Q report and upload that PDF."
                )
                return False
            with st.spinner(f"Extracting from {pdf_path.name}..."):
                try:
                    result = extract_fields(pdf_path)
                except Exception as exc:
                    st.error(f"Extractor error for `{pdf_path.name}`: {str(exc)}")
                    if st.session_state.get("pdf_path") == str(pdf_path):
                        st.session_state.pop("result", None)
                        st.session_state["extracted_ready"] = False
                    return False
                st.session_state["result"] = result
                st.session_state["pdf_path"] = str(pdf_path)
                st.session_state["selected_file"] = str(pdf_path.resolve())
                st.session_state["extracted_ready"] = True
                st.session_state["extract_seq"] = int(st.session_state.get("extract_seq", 0)) + 1
                st.session_state["extract_token"] = f"{pdf_path.name}:{st.session_state['extract_seq']}"
                st.session_state["extract_timestamp"] = datetime.now().isoformat(timespec="seconds")
                _save_cached_extraction(pdf_path, result)
                st.session_state["document_metadata"] = {
                    "document_type": _resolve_detected_type(pdf_path, result.get("summary", {}).get("document_type")),
                    "company_name": _extract_company_name_from_pdf(pdf_path),
                    "fiscal_years_detected": _extract_fiscal_years_detected(pdf_path),
                }
                _post_extraction_to_backend(pdf_path, result)
                return True

        def _open_non_financial_confirm_dialog() -> None:
            pending = st.session_state.get("pending_non_financial_pdf")
            if not pending:
                st.session_state["show_non_fin_dialog"] = False
                return

            pending_path = Path(str(pending))

            if hasattr(st, "dialog"):
                @st.dialog("Document check required")
                def _confirm_non_fin() -> None:
                    st.write("This file may not be a financial document. Continue extraction?")
                    c1, c2 = st.columns(2)
                    if c1.button("Cancel", key="nonfin_cancel_modal", use_container_width=True):
                        st.session_state["pending_non_financial_pdf"] = None
                        st.session_state["show_non_fin_dialog"] = False
                        st.rerun()
                    if c2.button("Yes, continue", key="nonfin_yes_modal", type="primary", use_container_width=True):
                        st.session_state["pending_non_financial_pdf"] = None
                        st.session_state["show_non_fin_dialog"] = False
                        if pending_path.exists():
                            pending_key = str(pending_path.resolve())
                            st.session_state["selected_file"] = pending_key
                            ok = _run_extraction_for_selected(pending_path)
                            if ok:
                                st.session_state["active_module"] = "extract"
                                st.query_params["module"] = "extract"
                                st.query_params["selected"] = pending_key
                                st.rerun()
                            return
                        st.error("Selected pending document was not found. Please select and retry.")
                _confirm_non_fin()
            else:
                st.markdown(
                    """
                    <div class="warn-pop">
                      <p class="warn-pop-title">Document check required</p>
                      <p class="warn-pop-body">This file may not be a financial document. Continue extraction?</p>
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                c1, c2, _ = st.columns([1, 1, 2])
                if c1.button("Cancel", key="nonfin_cancel_fallback", use_container_width=True):
                    st.session_state["pending_non_financial_pdf"] = None
                    st.session_state["show_non_fin_dialog"] = False
                    st.rerun()
                if c2.button("Yes, continue", key="nonfin_yes_fallback", type="primary", use_container_width=True):
                    st.session_state["pending_non_financial_pdf"] = None
                    st.session_state["show_non_fin_dialog"] = False
                    if pending_path.exists():
                        pending_key = str(pending_path.resolve())
                        st.session_state["selected_file"] = pending_key
                        ok = _run_extraction_for_selected(pending_path)
                        if ok:
                            st.session_state["active_module"] = "extract"
                            st.query_params["module"] = "extract"
                            st.query_params["selected"] = pending_key
                            st.rerun()
                        return
                    st.error("Selected pending document was not found. Please select and retry.")

        module_titles = {
            "upload": ("Upload Financial Documents", "Upload statements, transaction reports, and legal financial files for extraction."),
            "extract": ("Data Extraction", "Review parsed fields by sheet and open cited pages directly from the extracted table."),
            "graph": ("Data Visualization", "Visualize extracted financial fields using configurable chart views."),
            "risk": ("Credit Risk Dashboard", "Dynamic covenant and structural risk scoring with industry and region calibration."),
            "repo": ("Document Repository", "Browse uploaded files and launch extraction runs from one place."),
        }
        page_title, page_sub = module_titles.get(active_module, module_titles["upload"])
        if active_module in {"upload", "extract", "graph", "repo"}:
            # Data extraction/visualization default must be Latest Available.
            st.session_state["analysis_year_mode"] = "latest_available"
            st.session_state["analysis_specific_year"] = ""
        st.markdown(
            f"""
            <div class="upload-shell">
              <h2>{page_title}</h2>
              <p>{page_sub}</p>
            """,
            unsafe_allow_html=True,
        )

        if active_module in {"upload", "extract", "graph"}:
            up_col, sel_col, act_col = st.columns([2.2, 2.0, 1.2], vertical_alignment="bottom")
            if "main_upload_nonce" not in st.session_state:
                st.session_state["main_upload_nonce"] = 0
            upload_key = f"main_upload_{int(st.session_state.get('main_upload_nonce', 0))}"
            uploaded_main = up_col.file_uploader("Drop your document here or click to browse", type=["pdf"], key=upload_key)
            uploader_has_file = uploaded_main is not None
            file_option_map: Dict[str, Path] = {}
            for p in files:
                try:
                    file_option_map[str(p.resolve())] = p
                except Exception:
                    file_option_map[str(p)] = p
            file_option_values = [""] + list(file_option_map.keys())
            preferred_selected = st.session_state.get("selected_file")
            if preferred_selected:
                try:
                    preferred_selected = str(Path(str(preferred_selected)).resolve())
                except Exception:
                    preferred_selected = str(preferred_selected)
            prefill_selected = st.session_state.pop("selected_file_prefill", None)
            forced_selected = st.session_state.pop("force_selected_file_once", None)
            if forced_selected:
                prefill_selected = forced_selected
            if prefill_selected:
                try:
                    prefill_selected = str(Path(str(prefill_selected)).resolve())
                except Exception:
                    prefill_selected = str(prefill_selected)
            if "selected_select_nonce" not in st.session_state:
                st.session_state["selected_select_nonce"] = 0
            if "selected_file_option_value" not in st.session_state:
                st.session_state["selected_file_option_value"] = ""
            if prefill_selected and prefill_selected in file_option_values:
                st.session_state["selected_file_option_value"] = prefill_selected
                st.session_state["selected_select_nonce"] = int(st.session_state.get("selected_select_nonce", 0)) + 1
            elif preferred_selected and preferred_selected in file_option_values:
                current_opt = st.session_state.get("selected_file_option_value")
                if current_opt == "" or current_opt not in file_option_values:
                    st.session_state["selected_file_option_value"] = preferred_selected
                    st.session_state["selected_select_nonce"] = int(st.session_state.get("selected_select_nonce", 0)) + 1
            elif (
                st.session_state.get("selected_file_option_value") == ""
                and st.session_state.get("pdf_path")
            ):
                try:
                    pdf_ctx = str(Path(str(st.session_state.get("pdf_path"))).resolve())
                except Exception:
                    pdf_ctx = str(st.session_state.get("pdf_path"))
                if pdf_ctx in file_option_values:
                    st.session_state["selected_file_option_value"] = pdf_ctx
                    st.session_state["selected_select_nonce"] = int(st.session_state.get("selected_select_nonce", 0)) + 1
            elif st.session_state.get("selected_file_option_value") not in file_option_values:
                st.session_state["selected_file_option_value"] = ""
            selected_key = f"selected_file_option_{int(st.session_state.get('selected_select_nonce', 0))}"
            selected_default = st.session_state.get("selected_file_option_value", "")
            selected_index = file_option_values.index(selected_default) if selected_default in file_option_values else 0
            # Session state only — combining key + index= triggers Streamlit warnings and can blank modules.
            if selected_key not in st.session_state:
                st.session_state[selected_key] = (
                    file_option_values[selected_index] if file_option_values else ""
                )
            elif st.session_state.get(selected_key) not in file_option_values:
                st.session_state[selected_key] = (
                    file_option_values[selected_index] if file_option_values else ""
                )
            selected_value = sel_col.selectbox(
                "Choose document",
                file_option_values,
                key=selected_key,
                format_func=lambda p: "Select document..." if p == "" else file_option_map.get(p, Path(p)).name,
                disabled=uploader_has_file,
            )
            if not uploader_has_file:
                st.session_state["selected_file_option_value"] = selected_value
            selected_pdf = file_option_map.get(selected_value) if selected_value else None
            if uploader_has_file:
                selected_pdf = None
            if selected_pdf is not None and not uploader_has_file:
                # Keep selected-file state in sync with the visible selectbox across modules.
                selected_key = selected_value if selected_value else str(selected_pdf.resolve())
                st.session_state["selected_file"] = selected_key
                st.query_params["selected"] = selected_key
            extract_requested = act_col.button(
                "Extract Data",
                type="primary",
                use_container_width=True,
                key="extract_top",
                disabled=(selected_pdf is None and uploaded_main is None),
            )
            extracted_pdf_ctx = st.session_state.get("pdf_path")
            extracted_name = None
            if extracted_pdf_ctx:
                try:
                    extracted_name = Path(str(extracted_pdf_ctx)).name
                except Exception:
                    extracted_name = str(extracted_pdf_ctx)
            if selected_pdf is not None:
                st.markdown(f'<div class="selected-file">Selected file: {selected_pdf.name}</div>', unsafe_allow_html=True)
                if extracted_name and extracted_name != selected_pdf.name:
                    st.info(f"Currently showing extracted data from: {extracted_name}. Click Extract Data to load the selected file.")
            else:
                st.markdown("<div class='selected-file'>Selected file: none</div>", unsafe_allow_html=True)

            if uploaded_main is not None:
                mc1, mc2 = st.columns([1.35, 2.65], vertical_alignment="center")
                if mc1.button("Add Uploaded File", key="add_upload_main", type="primary", use_container_width=True):
                    try:
                        saved = _save_uploaded(uploaded_main)
                    except Exception as exc:
                        st.error(f"Unable to add uploaded file: {str(exc)}")
                    else:
                        saved_key = str(saved.resolve())
                        st.session_state["selected_file"] = saved_key
                        st.session_state["selected_file_prefill"] = saved_key
                        st.session_state["selected_file_option_value"] = saved_key
                        st.session_state["force_selected_file_once"] = saved_key
                        st.session_state["main_upload_nonce"] = int(st.session_state.get("main_upload_nonce", 0)) + 1
                        st.session_state["selected_select_nonce"] = int(st.session_state.get("selected_select_nonce", 0)) + 1
                        st.query_params["selected"] = saved_key
                        # Straight-through flow: add and move directly to extracted view for this file.
                        if _run_extraction_for_selected(saved):
                            st.session_state["active_module"] = "extract"
                            st.query_params["module"] = "extract"
                            st.query_params["selected"] = saved_key
                            st.success(f"Added and extracted {saved.name}")
                            st.rerun()
                        st.success(f"Added {saved.name}")
                        st.rerun()
                mc2.markdown(
                    f"""
                    <div style='display:flex;align-items:center;min-height:44px;padding:0 10px;border:1px solid #b9d5f1;border-radius:10px;background:#eaf4ff;color:#183b63;font-weight:700;'>
                      Ready to add: {html.escape(uploaded_main.name)}
                    </div>
                    """,
                    unsafe_allow_html=True,
                )
                sel_col.caption("Dropdown is disabled while uploaded file is active. Add or clear upload to re-enable.")
            st.markdown("</div>", unsafe_allow_html=True)

            if active_module in {"upload", "extract"}:
                st.markdown(
                    """
                    <div class="supported-box">
                      <strong>Supported documents</strong><br/>
                      • Credit agreements and term sheets<br/>
                      • Amendments and compliance certificates<br/>
                      • Financial statements and forecast models<br/>
                      • Security agreements, borrowing base, and fee letters
                    </div>
                    """,
                    unsafe_allow_html=True,
                )

            st.markdown("<div class='mode-row'></div>", unsafe_allow_html=True)
            mode_col, manual_col = st.columns([1.2, 1.8], vertical_alignment="bottom")
            mode = mode_col.radio("Document type mode", ["Auto-detect", "Manual"], horizontal=True, index=0)
            manual = manual_col.selectbox(
                "Manual document type",
                [""] + MANUAL_DOC_TYPE_OPTIONS,
                index=0,
                disabled=mode != "Manual",
                format_func=lambda x: "Select document type..." if x == "" else x,
            )
        else:
            st.markdown("</div>", unsafe_allow_html=True)
            if active_module == "risk":
                fallback_pdf = st.session_state.get("selected_file")
                if fallback_pdf and Path(str(fallback_pdf)).exists():
                    selected_pdf = Path(str(fallback_pdf))
                else:
                    selected_pdf = files[0] if files else None

        if (
            active_module not in {"risk", "repo"}
            and selected_pdf is None
            and not st.session_state.get("show_non_fin_dialog")
            and not extract_requested
        ):
            if active_module == "upload":
                return
            st.info("Select a document to continue.")
            return

        if active_module == "repo":
            if not files:
                st.info("No documents in repository yet. Upload a document to get started.")
                return
            st.markdown("<div class='panel'><strong>Library Files</strong></div>", unsafe_allow_html=True)
            repo_rows: List[Dict[str, Any]] = []
            for doc in files:
                stat = doc.stat()
                repo_rows.append(
                    {
                        "Document": doc.name,
                        "Size (MB)": round(stat.st_size / (1024 * 1024), 2),
                        "Modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
                        "Suggested Type": _guess_type_from_filename(doc),
                    }
                )
            repo_df = pd.DataFrame(repo_rows)
            with st.expander("Advanced Filters", expanded=False):
                r1, r2 = st.columns([2, 1])
                repo_search = r1.text_input("Search document")
                all_types = ["All"] + sorted(repo_df["Suggested Type"].dropna().unique().tolist())
                repo_type = r2.selectbox("Document type", all_types, index=0)
            filtered_repo = repo_df.copy()
            if repo_search:
                q = repo_search.strip().lower()
                filtered_repo = filtered_repo[filtered_repo["Document"].str.lower().str.contains(q, na=False)]
            if repo_type != "All":
                filtered_repo = filtered_repo[filtered_repo["Suggested Type"] == repo_type]
            st.dataframe(filtered_repo, width="stretch", height=400, hide_index=True)
            return

        if extract_requested:
            if uploaded_main is not None:
                try:
                    saved = _save_uploaded(uploaded_main)
                except Exception as exc:
                    st.error(f"Unable to add uploaded file for extraction: {str(exc)}")
                    return
                saved_key = str(saved.resolve())
                st.session_state["selected_file"] = saved_key
                st.session_state["selected_file_prefill"] = saved_key
                st.session_state["selected_file_option_value"] = saved_key
                st.session_state["force_selected_file_once"] = saved_key
                st.session_state["main_upload_nonce"] = int(st.session_state.get("main_upload_nonce", 0)) + 1
                st.session_state["selected_select_nonce"] = int(st.session_state.get("selected_select_nonce", 0)) + 1
                st.query_params["selected"] = saved_key
                selected_pdf = saved
            if selected_pdf is None:
                st.warning("Select a document or upload a file to continue.")
                return
            if _non_financial_dialog_needed_before_extract(selected_pdf, mode, manual):
                st.session_state["pending_non_financial_pdf"] = str(selected_pdf)
                st.session_state["show_non_fin_dialog"] = True
                st.rerun()
            if _run_extraction_for_selected(selected_pdf):
                # Keep flow in one tab and move user directly to extracted data view.
                st.session_state["active_module"] = "extract"
                st.query_params["module"] = "extract"
                st.query_params["selected"] = str(selected_pdf.resolve())
                st.rerun()

        if st.session_state.get("show_non_fin_dialog"):
            _open_non_financial_confirm_dialog()

        if "result" in st.session_state:
            detected = _resolve_detected_type(selected_pdf, st.session_state["result"]["summary"].get("document_type"))
        else:
            detected = _guess_type_from_filename(selected_pdf)
        filing_hint_name = selected_pdf.name.lower() if selected_pdf else ""
        if (
            mode == "Auto-detect"
            and detected == "Credit Agreement"
            and any(k in filing_hint_name for k in ["10-k", "10k", "10-q", "10q", "annual", "financial", "statement"])
        ):
            st.warning("Document type could not be determined. Please select document type manually.")
        manual_norm = _normalize_manual_doc_type(manual)
        applied = detected if mode == "Auto-detect" else (manual_norm if manual_norm in DOC_TYPE_CONFIG else detected)
        if applied not in DOC_TYPE_CONFIG:
            applied = "Financial Statements"
        if active_module in {"upload", "extract", "graph"}:
            st.session_state["last_applied_doc_type"] = applied
            if selected_pdf is not None:
                by_file = dict(st.session_state.get("last_applied_doc_type_by_file", {}))
                by_file[str(selected_pdf.resolve())] = applied
                st.session_state["last_applied_doc_type_by_file"] = by_file
        active_sheets = _active_sheets_for_doc_type(applied)
        tab_names = ["Consolidated"] + active_sheets

        # Keep extraction context stable across module switches, but never override
        # an explicit user selection. Only fall back to the last extracted file
        # when nothing is selected.
        if active_module in {"extract", "graph"}:
            active_pdf_ctx = st.session_state.get("pdf_path")
            if selected_pdf is None and active_pdf_ctx:
                selected_pdf = Path(str(active_pdf_ctx))
                st.session_state["selected_file"] = str(selected_pdf.resolve())
                st.session_state["selected_file_option_value"] = str(selected_pdf.resolve())

        if active_module != "risk":
            # Strict per-file context sync: never render extract/graph using another file's in-memory result.
            active_pdf_ctx = st.session_state.get("pdf_path")
            same_selected_ctx = False
            if selected_pdf is not None and active_pdf_ctx:
                try:
                    same_selected_ctx = Path(str(active_pdf_ctx)).resolve() == selected_pdf.resolve()
                except Exception:
                    same_selected_ctx = str(active_pdf_ctx) == str(selected_pdf)
            if selected_pdf is not None and not same_selected_ctx:
                cached_selected = _load_cached_extraction(selected_pdf)
                if cached_selected is not None:
                    st.session_state["result"] = cached_selected
                    st.session_state["pdf_path"] = str(selected_pdf.resolve())
                    st.session_state["selected_file"] = str(selected_pdf.resolve())
                    st.session_state["extracted_ready"] = True
                    st.session_state["extract_seq"] = int(st.session_state.get("extract_seq", 0)) + 1
                    st.session_state["extract_token"] = f"{selected_pdf.name}:{st.session_state['extract_seq']}"
                    _post_extraction_to_backend(selected_pdf, cached_selected)
                else:
                    st.session_state.pop("result", None)
                    st.session_state["pdf_path"] = str(selected_pdf.resolve())
                    st.session_state["selected_file"] = str(selected_pdf.resolve())
                    st.session_state["extracted_ready"] = False

            if "result" not in st.session_state:
                cached_result = _load_cached_extraction(selected_pdf) if selected_pdf is not None else None
                if cached_result:
                    st.session_state["result"] = cached_result
                    st.session_state["pdf_path"] = str(selected_pdf.resolve())
                    st.session_state["selected_file"] = str(selected_pdf.resolve())
                    st.session_state["extracted_ready"] = True
                    st.session_state["extract_seq"] = int(st.session_state.get("extract_seq", 0)) + 1
                    st.session_state["extract_token"] = f"{selected_pdf.name}:{st.session_state['extract_seq']}"
                    _post_extraction_to_backend(selected_pdf, cached_result)
                else:
                    if active_module == "upload":
                        st.markdown(
                            """
                            <div class="cap-grid">
                              <div class="cap-card"><p class="cap-title">Data Extraction</p><p class="cap-text">Automatically parse legal and financial fields into structured tables by document type.</p></div>
                              <div class="cap-card"><p class="cap-title">Data Visualization</p><p class="cap-text">Generate field-level charts with unit scaling (K/M/B), color legend, and configurable field selection.</p></div>
                              <div class="cap-card"><p class="cap-title">Risk Analysis</p><p class="cap-text">Review covenant posture, leverage signals, and trigger flags after each extraction run.</p></div>
                              <div class="cap-card"><p class="cap-title">Repository</p><p class="cap-text">Maintain a central library of uploaded files and run extractions on demand.</p></div>
                            </div>
                            """,
                            unsafe_allow_html=True,
                        )
                        return
                    if active_module == "graph" and selected_pdf is not None:
                        if _non_financial_dialog_needed_before_extract(selected_pdf, mode, manual):
                            st.session_state["pending_non_financial_pdf"] = str(selected_pdf)
                            st.session_state["show_non_fin_dialog"] = True
                            st.rerun()
                        if _run_extraction_for_selected(selected_pdf):
                            st.rerun()
                        st.warning(
                            "No saved extraction found for this document. Click **Extract Data**, wait for it to finish, then return here."
                        )
                        return
                    st.info("Run Extract Data to populate this module.")
                    return

        display_cons = pd.DataFrame()
        display_sheets: Dict[str, pd.DataFrame] = {}
        all_sheet_map: Dict[str, pd.DataFrame] = {}
        summary = {"filename": (selected_pdf.name if selected_pdf is not None else "")}
        needs_sheet_views = active_module in {"upload", "extract", "graph"}
        if "result" in st.session_state:
            result = st.session_state["result"]
            summary = result["summary"]
            if needs_sheet_views:
                extraction = result["extraction"]
                if not st.session_state.get("pdf_path"):
                    if selected_pdf is not None:
                        st.session_state["pdf_path"] = str(selected_pdf.resolve())
                if not st.session_state.get("pdf_path"):
                    st.warning("Session is missing the PDF path. Click **Extract Data** again for this file.")
                    return
                current_pdf = Path(str(st.session_state["pdf_path"]))
                if (
                    selected_pdf is not None
                    and not _pdf_paths_equal(current_pdf, selected_pdf)
                    and active_module in {"extract", "graph"}
                ):
                    cached = _load_cached_extraction(selected_pdf) if selected_pdf else None
                    if cached is None:
                        if active_module == "graph" and selected_pdf is not None:
                            if _non_financial_dialog_needed_before_extract(selected_pdf, mode, manual):
                                st.session_state["pending_non_financial_pdf"] = str(selected_pdf)
                                st.session_state["show_non_fin_dialog"] = True
                                st.rerun()
                            if _run_extraction_for_selected(selected_pdf):
                                st.rerun()
                            st.warning(
                                "Extraction for the selected file is not available yet. Click **Extract Data**, then return to Data Visualization."
                            )
                            return
                        st.info("Selected file changed. Click Extract Data to load values for this file.")
                        return
                    # Swap to the cached extraction for the selected file to keep modules in sync.
                    st.session_state["result"] = cached
                    st.session_state["pdf_path"] = str(selected_pdf.resolve())
                    st.session_state["selected_file"] = str(selected_pdf.resolve())
                    st.session_state["extract_seq"] = int(st.session_state.get("extract_seq", 0)) + 1
                    st.session_state["extract_token"] = f"{selected_pdf.name}:{st.session_state['extract_seq']}"
                    st.session_state["extract_timestamp"] = datetime.now().isoformat(timespec="seconds")
                    result = cached
                    extraction = result["extraction"]
                    summary = result["summary"]
                    current_pdf = selected_pdf
                    _post_extraction_to_backend(selected_pdf, cached)

                # Recompute applied doc type/sheets after any context swap so
                # extract/graph never render using stale (previous-file) sheet config.
                detected = _resolve_detected_type(current_pdf, summary.get("document_type"))
                manual_norm = _normalize_manual_doc_type(manual)
                applied = detected if mode == "Auto-detect" else (manual_norm if manual_norm in DOC_TYPE_CONFIG else detected)
                if applied not in DOC_TYPE_CONFIG:
                    applied = "Financial Statements"
                active_sheets = _active_sheets_for_doc_type(applied)
                tab_names = ["Consolidated"] + active_sheets
                if active_module in {"upload", "extract", "graph"}:
                    st.session_state["last_applied_doc_type"] = applied
                    by_file = dict(st.session_state.get("last_applied_doc_type_by_file", {}))
                    by_file[str(current_pdf.resolve())] = applied
                    st.session_state["last_applied_doc_type_by_file"] = by_file

                sheet_cache_key = "|".join(
                    [
                        str(st.session_state.get("extract_token") or ""),
                        str(current_pdf.resolve()),
                        str(applied),
                        ",".join(active_sheets),
                    ]
                )
                sheet_render_cache = st.session_state.get("sheet_render_cache")
                if isinstance(sheet_render_cache, dict) and sheet_render_cache.get("key") == sheet_cache_key:
                    display_cons = sheet_render_cache.get("display_cons", pd.DataFrame())
                    display_sheets = sheet_render_cache.get("display_sheets", {})
                    all_sheet_map = dict(display_sheets)
                else:
                    all_sheet_map = {s: _build_sheet_cached(extraction, s, str(current_pdf)) for s in active_sheets}
                    consolidated_df = _build_consolidated_view(all_sheet_map)
                    display_cons = _order_display_columns(consolidated_df)
                    display_sheets = {s: _order_display_columns(df) for s, df in all_sheet_map.items()}
                    st.session_state["sheet_render_cache"] = {
                        "key": sheet_cache_key,
                        "display_cons": display_cons,
                        "display_sheets": display_sheets,
                    }
                resolved_currency = "USD"
                display_cons["Currency"] = resolved_currency
                for s in list(display_sheets.keys()):
                    display_sheets[s]["Currency"] = resolved_currency

        if active_module == "upload":
            st.markdown(
                f"""
                <div class='panel'>
                  <div class='info-grid'>
                    <div class='info-card'><div class='info-label'>Document</div><div class='info-value'>{summary.get('filename')}</div></div>
                    <div class='info-card'><div class='info-label'>Suggested Type (Auto-detect)</div><div class='info-value'>{detected}</div></div>
                    <div class='info-card'><div class='info-label'>Type Currently Applied</div><div class='info-value'>{applied}</div></div>
                    <div class='info-card'><div class='info-label'>Sheets Configured</div><div class='info-value'>{', '.join(active_sheets)}</div></div>
                  </div>
                </div>
                """,
                unsafe_allow_html=True,
            )
            return

        if active_module == "extract":
            st.markdown(f"<div class='recommendation'>{_recommendation(_display_doc_type_label(applied), mode, detected)}</div>", unsafe_allow_html=True)
            if not display_cons.empty and "Currency" in display_cons.columns:
                st.caption(f"Detected reporting unit: {display_cons['Currency'].iloc[0]}")
            if _is_filing_doc_type(applied):
                tab_names = ["Income Statement", "Balance Sheet", "Cash Flow", "Derived Metrics", "Extraction Confidence"]
            elif _is_standalone_financial_doc_type(applied):
                mapped = "Income Statement"
                if applied == "Balance Sheet (Standalone)":
                    mapped = "Balance Sheet"
                elif applied == "Cash Flow Statement (Standalone)":
                    mapped = "Cash Flow"
                tab_names = [mapped]
            else:
                tab_names = ["Consolidated"] + active_sheets
            tabs = st.tabs(tab_names)
            show_adjusted = st.toggle("Show adjusted (non-GAAP) values", value=False, key="show_adjusted_toggle")
            def _render_table_with_filters(df: pd.DataFrame, filter_key: str) -> pd.DataFrame:
                with st.expander("Advanced Filters", expanded=False):
                    s1, s2, s3 = st.columns([1, 1, 2])
                    show_nulls = s1.toggle("Show null values", value=True, key=f"flt_null_{filter_key}")
                    if "Confidence" in df.columns:
                        min_conf = s2.slider("Min confidence", 0.0, 1.0, 0.0, 0.05, key=f"flt_conf_{filter_key}")
                    else:
                        s2.caption("No confidence column")
                        min_conf = 0.0
                    search_q = s3.text_input("Search", key=f"flt_search_{filter_key}")
                view = _apply_filters(df, show_nulls, min_conf, search_q)
                if show_adjusted:
                    view = view.copy()
                    for base, adj in [("EBIT", "EBIT (Adjusted)"), ("EBITDA", "EBITDA (Adjusted)"), ("Net Income", "Net Income (Adjusted)")]:
                        if adj in view.columns:
                            adj_vals = view[adj]
                            base_vals = view.get(base)
                            if base_vals is not None:
                                view[base] = adj_vals.where(adj_vals.notna(), base_vals)
                            # Swap confidence metadata to the adjusted series when present.
                            meta_base = f"__meta__{base}"
                            meta_adj = f"__meta__{adj}"
                            if meta_adj in view.columns and meta_base in view.columns:
                                view[meta_base] = view[meta_adj]
                # Strip metadata columns before rendering.
                meta_cols = [c for c in view.columns if str(c).startswith("__meta__")]
                view_ui = _ui_null_df_with_labels(view.drop(columns=meta_cols, errors="ignore"), derived_cols=DERIVED_COLUMNS)
                # Format monetary fields using detected unit (display-only).
                unit_label = "units"
                if "Detected Unit" in view.columns:
                    vals = [str(x).strip().lower() for x in view["Detected Unit"].dropna().tolist() if str(x).strip()]
                    if vals:
                        unit_label = vals[0]
                unit_suffix = {"thousands": "K", "millions": "M", "billions": "B", "trillions": "T"}.get(unit_label, "")
                money_cols = {
                    "Revenue", "COGS", "Total Expense", "Employee Compensation", "EBITDA", "EBIT", "Net Income",
                    "Interest Expense", "Total Assets", "Total Liabilities", "Shareholders' Equity",
                    "Total Debt", "Short-term Debt", "Current Portion of Long-term Debt", "Long-term Debt",
                    "Cash", "Current Assets", "Inventory", "Accounts Receivable", "Operating Cash Flow", "CapEx",
                    "Free Cash Flow", "Loan Amount", "Collateral Value",
                }
                def _fmt_money(v: Any) -> Any:
                    if v is None or (isinstance(v, float) and pd.isna(v)):
                        return v
                    try:
                        fv = float(v)
                    except Exception:
                        return v
                    if unit_suffix:
                        prefix = "$"
                    else:
                        prefix = "$"
                    if abs(fv) >= 1:
                        txt = f"{fv:,.0f}"
                    else:
                        txt = f"{fv:,.2f}"
                    return f"{prefix}{txt}{unit_suffix}"
                for col in view_ui.columns:
                    if col in money_cols:
                        view_ui[col] = view_ui[col].apply(_fmt_money)
                # Apply per-field confidence cell tags (used for background coloring).
                view_badged = _apply_confidence_cell_classes(view_ui.join(view[meta_cols], how="left") if meta_cols else view_ui)
                view_badged = view_badged.drop(columns=meta_cols, errors="ignore")
                # Render with cell background colors based on confidence tags.
                view_display = view_badged.copy()
                cell_style: Dict[Tuple[int, str], Dict[str, str]] = {}
                for col in list(view_display.columns):
                    for ridx in view_display.index:
                        v = view_display.at[ridx, col]
                        if not isinstance(v, str) or "||" not in v:
                            continue
                        raw, tag = v.split("||", 1)
                        view_display.at[ridx, col] = raw.strip()
                        if tag == "conf-exact":
                            cell_style[(ridx, col)] = {"backgroundColor": "#E6F4EA"}
                        elif tag == "conf-fuzzy":
                            cell_style[(ridx, col)] = {"backgroundColor": "#FFF4E5"}
                        elif tag == "conf-derived":
                            cell_style[(ridx, col)] = {"backgroundColor": "#FDECEA"}
                if cell_style:
                    style_df = pd.DataFrame("", index=view_display.index, columns=view_display.columns)
                    for (ridx, c), styles in cell_style.items():
                        style_df.at[ridx, c] = "; ".join(f"{k}: {v}" for k, v in styles.items())
                    st.dataframe(
                        view_display.style.apply(lambda _: style_df, axis=None),
                        width="stretch",
                        height=_table_height(view_display),
                        hide_index=True,
                        column_config={
                            "Open PDF Page": st.column_config.TextColumn("Open PDF Page", width="small"),
                            "Confidence": st.column_config.NumberColumn("Confidence", format="%.2f", width="small"),
                            "Page": st.column_config.NumberColumn("Page", width="small"),
                            "Snippet": st.column_config.TextColumn("Snippet", width="large"),
                        },
                    )
                else:
                    st.dataframe(
                        view_display,
                        width="stretch",
                        height=_table_height(view_display),
                        hide_index=True,
                        column_config={
                            "Open PDF Page": st.column_config.TextColumn("Open PDF Page", width="small"),
                            "Confidence": st.column_config.NumberColumn("Confidence", format="%.2f", width="small"),
                            "Page": st.column_config.NumberColumn("Page", width="small"),
                            "Snippet": st.column_config.TextColumn("Snippet", width="large"),
                        },
                    )
                return view

            if _is_filing_doc_type(applied):
                for i, sheet in enumerate(["Income Statement", "Balance Sheet", "Cash Flow"]):
                    with tabs[i]:
                        sdf = display_sheets.get(sheet, pd.DataFrame())
                        if sdf.empty or all(
                            (v in (None, "", "None", "null") or pd.isna(v))
                            for c in sdf.columns if c not in META_COLUMNS
                            for v in sdf[c].tolist()
                        ):
                            st.warning("Statement not detected in document.")
                        else:
                            filtered_df = _render_table_with_filters(sdf, f"filing_{i}_{applied}")
                            _render_page_jump_controls(filtered_df, f"sheet_{i}")
                with tabs[3]:
                    derived_df = _build_derived_metrics_table(display_sheets)
                    derived_ui = _ui_null_df_with_labels(derived_df, derived_cols={"Value"})
                    st.dataframe(derived_ui, width="stretch", hide_index=True)
                with tabs[4]:
                    conf_df = _build_extraction_confidence_table(display_sheets)
                    st.dataframe(_ui_null_df(conf_df), width="stretch", hide_index=True)
            elif _is_standalone_financial_doc_type(applied):
                mapped_sheet = tab_names[0]
                with tabs[0]:
                    sdf = display_sheets.get(mapped_sheet, pd.DataFrame())
                    if sdf.empty:
                        st.warning("Statement not detected in document.")
                    else:
                        filtered_df = _render_table_with_filters(sdf, f"standalone_{mapped_sheet}_{applied}")
                        _render_page_jump_controls(filtered_df, f"sheet_{mapped_sheet}")
            else:
                with tabs[0]:
                    cons_view = _render_table_with_filters(display_cons, f"cons_{applied}")
                    _render_page_jump_controls(cons_view, "consolidated")
                for i, sheet in enumerate(active_sheets, start=1):
                    with tabs[i]:
                        filtered_df = _render_table_with_filters(display_sheets[sheet], f"{i}_{applied}")
                        _render_page_jump_controls(filtered_df, f"sheet_{i}")

            viewer_page = st.session_state.get("viewer_page")
            if viewer_page:
                v1, v2 = st.columns([0.85, 0.15], vertical_alignment="center")
                v1.markdown("<div class='panel'><strong>Source Document Viewer</strong></div>", unsafe_allow_html=True)
                if v2.button("Close", key="close_viewer_btn", use_container_width=True):
                    st.session_state["viewer_page"] = None
                    st.rerun()
                st.markdown(f"<div class='viewer-label'>Showing page {viewer_page} of {selected_pdf.name}</div>", unsafe_allow_html=True)
                png = _pdf_page_png(str(selected_pdf), int(viewer_page))
                if png:
                    st.image(png, use_container_width=True)
                else:
                    st.warning("Unable to render this page.")

            all_active_excel = _to_excel_bytes(display_sheets if display_sheets else all_sheet_map)
            consolidated_excel = _to_excel_bytes({"Consolidated": display_cons})
            consolidated_csv = display_cons.to_csv(index=False)
            ex1, ex2, ex3 = st.columns(3)
            ex1.download_button("Export This Consolidated CSV", data=consolidated_csv, file_name=f"{applied.lower().replace(' ', '_')}_consolidated.csv", mime="text/csv", use_container_width=True, key="export_consolidated_csv_bottom")
            ex2.download_button("Export Consolidated Excel", data=consolidated_excel, file_name=f"{applied.lower().replace(' ', '_')}_consolidated.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="export_consolidated_excel_bottom")
            ex3.download_button("Export All Active Sheets (Excel)", data=all_active_excel, file_name=f"{applied.lower().replace(' ', '_')}_workbook.xlsx", mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", use_container_width=True, key="export_all_active_excel_bottom")
            return

        if active_module == "graph":
            if "result" not in st.session_state:
                st.warning("No extraction is loaded. Choose a document and click **Extract Data**, then open Data Visualization again.")
                return
            # Stable key suffix for graph widgets (must NOT include `src` — changing source recreated keys and blanked the panel).
            _graph_ui_key = re.sub(r"[^a-zA-Z0-9_]", "_", str(st.session_state.get("extract_token") or "ctx"))[:120]
            st.markdown(
                "<div class='panel'><strong>Charts &amp; filters</strong> — same numbers as Data Extraction for this file.</div>",
                unsafe_allow_html=True,
            )
            try:
                with st.container(border=True):
                    # Reset graph UI state when extraction context changes so fields don't stay hidden
                    # from a prior document/source selection.
                    current_graph_ctx = str(st.session_state.get("extract_token") or "")
                    last_graph_ctx = str(st.session_state.get("graph_ctx_token") or "")
                    if current_graph_ctx != last_graph_ctx:
                        st.session_state["graph_source"] = "Consolidated"
                        st.session_state["show_graph_field_menu"] = False
                        for _k in [k for k in list(st.session_state.keys()) if k.startswith("graph_include_")]:
                            st.session_state.pop(_k, None)
                        st.session_state["graph_ctx_token"] = current_graph_ctx

                    g_h1, g_h2 = st.columns([0.9, 0.1], vertical_alignment="center")
                    g_h1.markdown("**Data Visualization Title**")
                    derived_graph_df = pd.DataFrame()
                    if _is_filing_doc_type(applied) and display_sheets:
                        sm = {k: display_sheets[k] for k in ("Income Statement", "Balance Sheet", "Cash Flow") if k in display_sheets}
                        if sm:
                            derived_graph_df = _build_derived_metrics_table(sm)
                    source_options = ["Consolidated"] + list(active_sheets)
                    if not derived_graph_df.empty:
                        source_options = source_options + ["Derived Metrics"]
                    # Stale session_state (e.g. "Derived Metrics" removed) breaks selectbox and can blank the whole module.
                    if not source_options:
                        source_options = ["Consolidated"]
                    if "graph_source" not in st.session_state or st.session_state.get("graph_source") not in source_options:
                        st.session_state["graph_source"] = source_options[0]
                    if "graph_type" not in st.session_state or st.session_state.get("graph_type") not in ("Bar", "Line", "Pie"):
                        st.session_state["graph_type"] = "Bar"

                    src = st.selectbox("Graph source", source_options, key="graph_source")
                    chart_type = st.selectbox("Chart type", ["Bar", "Line", "Pie"], key="graph_type")
                    base_df = pd.DataFrame()
                    if src == "Consolidated":
                        base_df = display_cons if display_cons is not None else pd.DataFrame()
                    elif src == "Derived Metrics":
                        base_df = derived_graph_df
                    else:
                        base_df = display_sheets.get(src, pd.DataFrame())
                        if base_df.empty and src:
                            st.caption(f"No table data for **{src}**. Pick another graph source or re-run extraction.")
                    with st.expander("Advanced Filters", expanded=False):
                        gf1, gf2, gf3 = st.columns([1, 1, 2])
                        g_show_nulls = gf1.toggle("Show null values", value=True, key=f"gflt_null_{_graph_ui_key}")
                        if "Confidence" in base_df.columns:
                            g_min_conf = gf2.slider("Min confidence", 0.0, 1.0, 0.0, 0.05, key=f"gflt_conf_{_graph_ui_key}")
                        else:
                            gf2.caption("No confidence column")
                            g_min_conf = 0.0
                        g_search = gf3.text_input("Search", key=f"gflt_search_{_graph_ui_key}")
                    base_df = _apply_filters(base_df, g_show_nulls, g_min_conf, g_search)
                    base_df_num = _coerce_object_cols_for_chart(base_df)
                    chart_df = _build_sheet_financial_chart_df(base_df_num)
                    chart_base_df = base_df_num
                    chart_note = ""
                    if chart_df.empty and src == "Consolidated" and display_sheets:
                        preferred = ["Income Statement", "Balance Sheet", "Cash Flow"]
                        order = [a for a in preferred if a in display_sheets] + [s for s in display_sheets if s not in preferred]
                        for alt in order:
                            alt_df = display_sheets.get(alt)
                            if alt_df is None or alt_df.empty:
                                continue
                            alt_f = _apply_filters(alt_df, g_show_nulls, g_min_conf, g_search)
                            alt_num = _coerce_object_cols_for_chart(alt_f)
                            cand = _build_sheet_financial_chart_df(alt_num)
                            if not cand.empty:
                                chart_df = cand
                                chart_base_df = alt_num
                                chart_note = f"Chart uses **{alt}** (Consolidated had no plottable numeric columns for this view)."
                                break
                    if chart_note:
                        st.caption(chart_note)
                    # Never auto-switch source: keep graph bound to user-selected source/document context.
                    # Keep graph currency consistent with extracted document unit, even if filters drop unit hints.
                    preferred_scale = _preferred_scale_from_df(chart_base_df)
                    if preferred_scale == "Units":
                        preferred_scale = _preferred_scale_from_df(display_cons)
                    _pdf_for_scale = selected_pdf
                    if _pdf_for_scale is None and st.session_state.get("pdf_path"):
                        try:
                            _pdf_for_scale = Path(str(st.session_state["pdf_path"]))
                        except Exception:
                            _pdf_for_scale = None
                    if _pdf_for_scale is not None:
                        doc_scale_graph = _scale_label_from_factor(_statement_unit_factor(_pdf_for_scale))
                        if doc_scale_graph != "Units":
                            preferred_scale = doc_scale_graph
                    trend_cols = [c for c in chart_base_df.columns if _column_name_calendar_year(c) is not None]
                    has_multi_year = len(trend_cols) >= 2
                    if chart_df.empty:
                        st.caption("Data availability: ⚪ No plottable financial values")
                    else:
                        preview_fields = chart_df["Field"].tolist()[:6]
                        st.caption("Data availability: " + " | ".join([f"✅ {f}" for f in preview_fields]))
                    available_fields = chart_df["Field"].tolist() if not chart_df.empty else []
                    selected_fields: List[str] = []
                    if available_fields:
                        if "show_graph_field_menu" not in st.session_state:
                            st.session_state["show_graph_field_menu"] = False
                        with g_h2:
                            st.markdown("<div class='graph-menu-dot'>", unsafe_allow_html=True)
                            if st.button("⋮", key=f"graph_menu_toggle_{_graph_ui_key}", use_container_width=True):
                                st.session_state["show_graph_field_menu"] = not st.session_state["show_graph_field_menu"]
                            st.markdown("</div>", unsafe_allow_html=True)
                        if st.session_state.get("show_graph_field_menu"):
                            st.markdown("<div class='graph-menu-panel'>", unsafe_allow_html=True)
                            st.markdown("**Select fields**")
                            for field_name in available_fields:
                                include_key = f"graph_include_{current_graph_ctx}_{src}_{re.sub(r'[^a-zA-Z0-9_]', '_', field_name)}"
                                if include_key not in st.session_state:
                                    st.session_state[include_key] = True
                                st.checkbox(field_name, key=include_key)
                            st.markdown("</div>", unsafe_allow_html=True)
                        for field_name in available_fields:
                            include_key = f"graph_include_{current_graph_ctx}_{src}_{re.sub(r'[^a-zA-Z0-9_]', '_', field_name)}"
                            if st.session_state.get(include_key, True):
                                selected_fields.append(field_name)
                    if chart_df.empty:
                        st.info("No numeric values found for selected source.")
                    else:
                        if not selected_fields:
                            selected_fields = available_fields
                        selected_df = chart_df[chart_df["Field"].isin(selected_fields)].copy()
                        scaled_df = selected_df.copy()
                        unit, factor = "", 1.0
                        if has_multi_year:
                            trend_rows: List[Dict[str, Any]] = []
                            for _, r in chart_base_df.iterrows():
                                metric_name = str(
                                    r.get("Metric")
                                    or r.get("Line Item")
                                    or r.get("Field")
                                    or r.get("Sheet")
                                    or "Metric"
                                )
                                for yc in trend_cols:
                                    yv = _to_numeric_financial(r.get(yc))
                                    yr = _column_name_calendar_year(yc)
                                    if yv is not None and yr is not None:
                                        trend_rows.append({"Year": yr, "Metric": metric_name, "Value": yv})
                            trend_df = pd.DataFrame(trend_rows)
                            if trend_df.empty:
                                scaled_df, unit, factor = _apply_preferred_scale(selected_df, preferred_scale)
                                _render_aura_grid_chart(scaled_df, chart_type, unit)
                                if unit:
                                    st.caption(f"Converted by ÷ {factor:,.0f}")
                            else:
                                piv = trend_df.pivot_table(index="Year", columns="Metric", values="Value", aggfunc="sum")
                                if piv.empty or not piv.notna().any().any():
                                    scaled_df, unit, factor = _apply_preferred_scale(selected_df, preferred_scale)
                                    _render_aura_grid_chart(scaled_df, chart_type, unit)
                                    if unit:
                                        st.caption(f"Converted by ÷ {factor:,.0f}")
                                    st.caption("Multi-year pivot was empty; showing field totals instead.")
                                    scaled_df = selected_df.copy()
                                else:
                                    if piv.shape[1] > 35:
                                        st.caption("Showing the 30 largest series by total |value| (readability).")
                                        tot = piv.abs().sum(axis=0).sort_values(ascending=False)
                                        piv = piv[list(tot.head(30).index)]
                                    st.line_chart(piv, height=360)
                                    st.caption(f"Trend view from multi-year columns ({preferred_scale}).")
                                    scaled_df = selected_df.copy()
                        else:
                            scaled_df, unit, factor = _apply_preferred_scale(selected_df, preferred_scale)
                            _render_aura_grid_chart(scaled_df, chart_type, unit)
                            cards = st.columns(min(4, max(1, len(selected_df))))
                            for idx, (_, rr) in enumerate(scaled_df.iterrows()):
                                card = cards[idx % len(cards)]
                                raw_row = selected_df[selected_df["Field"] == rr["Field"]]
                                raw_val = _to_numeric_financial(raw_row.iloc[0]["Value"]) if not raw_row.empty else None
                                compact_val = _format_scaled_with_unit(rr.get("Value"), unit)
                                raw_txt = "NULL" if raw_val is None else f"{float(raw_val):,.0f}"
                                card.metric(str(rr["Field"]), compact_val)
                                card.caption(f"Full: {raw_txt}")
                            if unit:
                                st.caption(f"Converted by ÷ {factor:,.0f}")
                        st.caption("Values shown in graph")
                        table_df = scaled_df.copy() if not scaled_df.empty else selected_df.copy()
                        table_df["Currency"] = "USD"
                        st.dataframe(_ui_null_df(table_df), width="stretch", hide_index=True)
                st.markdown("<div class='panel'><strong>Source Data</strong></div>", unsafe_allow_html=True)
                st.dataframe(_ui_null_df(base_df.assign(Currency="USD")), width="stretch", height=320, hide_index=True)
            except Exception as _graph_err:
                st.error("Data Visualization failed to render. Check the log or re-run extraction.")
                st.exception(_graph_err)
            return

        def _single_bundle_from_latest_extraction(update_state: bool = True, strict_selected: bool = False) -> Optional[Dict[str, Any]]:
            # Prefer currently selected file when it already has extraction cache;
            # otherwise fall back to the last extracted file context.
            selected_path = st.session_state.get("selected_file")
            pdf_ctx_path = st.session_state.get("pdf_path")
            preferred_path = None
            selected_exists = False
            selected_has_cache = False
            if selected_path:
                try:
                    selected_candidate = Path(str(selected_path))
                    selected_exists = selected_candidate.exists()
                    selected_has_cache = selected_exists and (_load_cached_extraction(selected_candidate) is not None)
                    if selected_has_cache:
                        preferred_path = str(selected_candidate)
                except Exception:
                    preferred_path = None
            if strict_selected and selected_path:
                # Strict mode: prefer currently selected file; if stale, fall back to active
                # extraction context only when it has a valid cache entry.
                if selected_exists and selected_has_cache:
                    if preferred_path is None:
                        preferred_path = str(Path(str(selected_path)))
                else:
                    active_candidate = st.session_state.get("pdf_path")
                    if not active_candidate:
                        return None
                    try:
                        active_path = Path(str(active_candidate))
                        if not active_path.exists() or _load_cached_extraction(active_path) is None:
                            return None
                        preferred_path = str(active_path)
                        if update_state:
                            st.session_state["selected_file"] = str(active_path.resolve())
                    except Exception:
                        return None
            if not preferred_path:
                preferred_path = pdf_ctx_path or selected_path
            if not preferred_path:
                return None
            latest_pdf = Path(str(preferred_path))
            if not latest_pdf.exists():
                return None

            latest_result = st.session_state.get("result")
            active_pdf_path = st.session_state.get("pdf_path")
            # If in-memory extraction context is stale/missing, recover from on-disk
            # extraction cache for the same file so "Use Current Extraction Context"
            # remains reliable across module switches and reruns.
            same_active_pdf = bool(active_pdf_path and str(active_pdf_path) == str(latest_pdf))
            if latest_result is None or not same_active_pdf:
                cached = _load_cached_extraction(latest_pdf)
                if cached is not None:
                    latest_result = cached
                    if update_state:
                        st.session_state["result"] = cached
                        st.session_state["pdf_path"] = str(latest_pdf)
                else:
                    if not same_active_pdf and latest_result is not None and update_state:
                        # If an extraction exists but points to another file, do not
                        # silently reuse it for this action.
                        return None
                    if latest_result is None:
                        return None

            latest_detected = _resolve_detected_type(latest_pdf, latest_result.get("summary", {}).get("document_type"))
            by_file = dict(st.session_state.get("last_applied_doc_type_by_file", {}))
            preferred_applied = by_file.get(str(latest_pdf.resolve()))
            guessed_applied = _guess_type_from_filename(latest_pdf)

            def _score_sheet_output(df_in: pd.DataFrame) -> int:
                if df_in is None or df_in.empty:
                    return 0
                meta_cols = set(META_COLUMNS) | {"Sheet", "Currency"}
                score = 0
                for c in df_in.columns:
                    if c in meta_cols:
                        continue
                    col_vals = df_in[c].tolist()
                    for v in col_vals:
                        if _to_numeric_financial(v) is not None:
                            score += 1
                financial_priority = {
                    "Revenue",
                    "EBITDA",
                    "EBIT",
                    "Net Income",
                    "Interest Expense",
                    "Total Assets",
                    "Total Liabilities",
                    "Shareholders' Equity",
                    "Current Assets",
                    "Current Liabilities",
                    "Operating Cash Flow",
                    "CapEx",
                }
                score += sum(2 for c in df_in.columns if c in financial_priority)
                return score

            candidates: List[str] = []
            for cand in [preferred_applied, latest_detected, guessed_applied]:
                if cand in DOC_TYPE_CONFIG and cand not in candidates:
                    candidates.append(cand)
            if not candidates:
                candidates = ["Financial Statements"]

            best_applied = candidates[0]
            best_cons = pd.DataFrame()
            best_score = -1
            for cand in candidates:
                try:
                    cand_sheets = DOC_TYPE_CONFIG[cand]["sheets"]
                    cand_map = {s: _build_sheet_cached(latest_result["extraction"], s, str(latest_pdf)) for s in cand_sheets}
                    cand_cons = _order_display_columns(_build_consolidated_view(cand_map))
                    cand_score = _score_sheet_output(cand_cons)
                    if cand_score > best_score:
                        best_applied = cand
                        best_cons = cand_cons
                        best_score = cand_score
                except Exception:
                    continue

            latest_applied = best_applied
            ctx_hash = _single_context_hash(str(latest_pdf), latest_result.get("extraction", {}), latest_applied)
            ctx_key = f"{latest_pdf.resolve()}::{int(latest_pdf.stat().st_mtime)}::{latest_applied}"
            cached_bundle = st.session_state.get("single_bundle_cache")
            if isinstance(cached_bundle, dict) and cached_bundle.get("context_hash") == ctx_hash:
                # Backfill older cache payloads so downstream runtime context is never None.
                cached_bundle["source"] = "single"
                cached_bundle["context_key"] = cached_bundle.get("context_key") or ctx_key
                cached_bundle["extracted_at"] = cached_bundle.get("extracted_at") or (
                    st.session_state.get("extract_timestamp")
                    or datetime.fromtimestamp(latest_pdf.stat().st_mtime).isoformat(timespec="seconds")
                )
                return cached_bundle
            latest_cons = best_cons
            if latest_cons is None or latest_cons.empty:
                latest_sheets = DOC_TYPE_CONFIG[latest_applied]["sheets"]
                latest_sheet_map = {s: _build_sheet_cached(latest_result["extraction"], s, str(latest_pdf)) for s in latest_sheets}
                latest_cons = _order_display_columns(_build_consolidated_view(latest_sheet_map))
            latest_cons.insert(0, "Source Document", latest_pdf.name)
            latest_cons.insert(1, "Detected Type", _canonical_doc_type(latest_applied))
            latest_detected_scale = _scale_label_from_factor(_statement_unit_factor(latest_pdf))
            latest_display_scale = _preferred_scale_from_df(latest_cons)
            latest_currency = "USD"
            latest_cons["Currency"] = latest_currency
            canonical_single = _canonical_doc_type(latest_applied)
            bundle = {
                "combined_df": latest_cons,
                "doc_types": [canonical_single],
                "processed_docs": [latest_pdf.name],
                "confidence_level": _confidence_level_from_doc_types([canonical_single]),
                "source": "single",
                "context_key": ctx_key,
                "context_hash": ctx_hash,
                "extracted_at": st.session_state.get("extract_timestamp") or datetime.fromtimestamp(latest_pdf.stat().st_mtime).isoformat(timespec="seconds"),
                "logic_version": RISK_BUNDLE_LOGIC_VERSION,
            }
            st.session_state["single_bundle_cache"] = bundle
            return bundle

        head_c1, head_c2 = st.columns([0.62, 0.38], vertical_alignment="center")
        head_c1.markdown(
            """
            <div class='panel' style='padding:1rem 1.05rem;'>
              <div style='font-size:1.42rem;font-weight:800;color:#0f2740;line-height:1.15;'>Credit Risk Dashboard</div>
              <div style='margin-top:0.24rem;color:#4e6178;font-weight:600;'>Institutional credit analysis and underwriting</div>
            </div>
            """,
            unsafe_allow_html=True,
        )
        hbtn1, hbtn2 = head_c2.columns([1, 1], vertical_alignment="center")
        if hbtn1.button("Upload Document", key="risk_go_upload_btn", use_container_width=True):
            st.session_state["active_module"] = "upload"
            st.query_params["module"] = "upload"
            st.rerun()
        if hbtn2.button("Edit Data", key="open_risk_input_modal_top", type="primary", use_container_width=True):
            st.session_state["risk_show_input_modal"] = True
        st.markdown(
            "<div class='panel'><strong>Credit Analysis Upload</strong><br/><span style='color:#4e6178'>Upload all relevant borrower documents, then generate one unified risk table.</span></div>",
            unsafe_allow_html=True,
        )
        risk_docs = st.file_uploader(
            "Upload credit analysis documents",
            type=["pdf"],
            accept_multiple_files=True,
            key="risk_multi_docs",
        )
        current_upload_key = (
            "|".join(sorted([f"{getattr(f, 'name', '')}:{getattr(f, 'size', 0)}" for f in risk_docs]))
            if risk_docs
            else ""
        )
        uploaded_count = len(risk_docs or [])
        has_current_context = _single_bundle_from_latest_extraction(update_state=False) is not None
        ract1, ract2 = st.columns([1, 1], vertical_alignment="bottom")
        run_multi = False
        use_single = False
        use_uploaded_single = False
        if uploaded_count == 1:
            use_uploaded_single = ract1.button("Use This New Document", type="primary", use_container_width=True, key="use_uploaded_single")
            use_single = ract2.button("Use Current Extraction Context", use_container_width=True, key="use_current_extract")
        elif uploaded_count > 1:
            run_multi = ract1.button("Generate Unified Risk Table", type="primary", use_container_width=True, key="run_multi_credit")
            use_single = ract2.button("Use Current Extraction Context", use_container_width=True, key="use_current_extract")
        else:
            use_single = ract1.button("Use Current Extraction Context", type="primary", use_container_width=True, key="use_current_extract")
            run_multi = ract2.button("Generate Unified Risk Table", use_container_width=True, key="run_multi_credit")

        if run_multi:
            if not risk_docs:
                st.warning("Upload at least one document to run credit analysis.")
            else:
                st.session_state["risk_bundle"] = _build_credit_risk_bundle(risk_docs)
                st.session_state["risk_bundle"]["source"] = "multi"
                st.session_state["risk_bundle"]["context_key"] = current_upload_key
                st.session_state["risk_bundle"]["extracted_at"] = datetime.now().isoformat(timespec="seconds")
                st.session_state["risk_bundle"]["analysis_mode"] = st.session_state.get("analysis_year_mode", "latest_available")
                st.session_state["risk_bundle"]["analysis_specific_year"] = st.session_state.get("analysis_specific_year", "")
                st.session_state["risk_mode"] = "multi"
                st.session_state["risk_recompute_reason"] = "dataset regenerate"
        if use_single:
            latest_bundle = _single_bundle_from_latest_extraction(update_state=True, strict_selected=True)
            # If user uploaded exactly one file in risk tab, allow direct extraction context creation here.
            if not latest_bundle and risk_docs and len(risk_docs) == 1:
                try:
                    saved = _save_uploaded(risk_docs[0])
                    if _run_extraction_for_selected(saved):
                        latest_bundle = _single_bundle_from_latest_extraction(update_state=True)
                except Exception:
                    latest_bundle = None
            if latest_bundle:
                latest_bundle["source"] = "single"
                latest_bundle["context_key"] = latest_bundle.get("context_key") or latest_bundle.get("context_hash") or "single-context"
                context_changed = (
                    st.session_state.get("risk_mode") == "single"
                    and st.session_state.get("risk_saved_context_hash")
                    and st.session_state.get("risk_saved_context_hash") != latest_bundle.get("context_hash")
                )
                if context_changed:
                    st.info("Extraction context changed. Reloaded latest extracted document context.")
                st.session_state["risk_bundle"] = latest_bundle
                st.session_state["risk_bundle"]["analysis_mode"] = st.session_state.get("analysis_year_mode", "latest_available")
                st.session_state["risk_bundle"]["analysis_specific_year"] = st.session_state.get("analysis_specific_year", "")
                st.session_state["risk_mode"] = "single"
                st.session_state["risk_recompute_reason"] = "context save"
                # Clear stale multi-document state when explicitly switching to current extraction.
                st.session_state.pop("risk_conflict_rule", None)
                st.session_state.pop("risk_bundle_context_key", None)
                # Clear stale multi-run gating/override state when switching to single context.
                for ss_key in list(st.session_state.keys()):
                    if str(ss_key).startswith("risk_borrower_mismatch_override::") or str(ss_key).startswith("risk_missing_docs_override::"):
                        st.session_state.pop(ss_key, None)
            else:
                # Clear stale multi/single bundle so only one warning path is shown.
                st.session_state.pop("risk_bundle", None)
                st.session_state["risk_mode"] = "single"
                st.warning("No extracted document context found. Extract a document first, then retry.")
                return
        if use_uploaded_single:
            if not risk_docs or len(risk_docs) != 1:
                st.warning("Upload exactly one document to use this action.")
                return
            try:
                saved = _save_uploaded(risk_docs[0])
            except Exception:
                st.error("Unable to save uploaded document. Please retry.")
                return
            if not _run_extraction_for_selected(saved):
                st.error("Extraction failed for uploaded document. Please review file and retry.")
                return
            latest_bundle = _single_bundle_from_latest_extraction(update_state=True)
            if not latest_bundle:
                st.error("Uploaded document extracted, but risk context was not created. Retry once.")
                return
            st.session_state["risk_bundle"] = latest_bundle
            st.session_state["risk_bundle"]["analysis_mode"] = st.session_state.get("analysis_year_mode", "latest_available")
            st.session_state["risk_bundle"]["analysis_specific_year"] = st.session_state.get("analysis_specific_year", "")
            st.session_state["risk_mode"] = "single"
            st.session_state["risk_recompute_reason"] = "dataset regenerate"
            st.rerun()

        # Prevent stale multi-document bundle from leaking into single-file analysis view.
        if (
            uploaded_count <= 1
            and not run_multi
            and st.session_state.get("risk_mode") == "multi"
            and not use_single
            and not use_uploaded_single
        ):
            st.session_state.pop("risk_bundle", None)
            st.session_state["risk_mode"] = "single"

        bundle = st.session_state.get("risk_bundle")
        # In single mode, force-refresh from currently selected extraction context so
        # risk never runs on stale or context-less payloads.
        if st.session_state.get("risk_mode", "single") == "single":
            strict_single = _single_bundle_from_latest_extraction(update_state=True, strict_selected=True)
            if strict_single:
                strict_single["source"] = "single"
                strict_single["context_key"] = strict_single.get("context_key") or strict_single.get("context_hash") or "single-context"
                st.session_state["risk_bundle"] = strict_single
                bundle = strict_single
        # If user is in single mode, never keep an old multi-source bundle.
        if st.session_state.get("risk_mode") == "single" and isinstance(bundle, dict) and bundle.get("source") != "single":
            refreshed_single = _single_bundle_from_latest_extraction(update_state=True)
            if refreshed_single:
                refreshed_single["source"] = "single"
                refreshed_single["context_key"] = refreshed_single.get("context_key") or refreshed_single.get("context_hash") or "single-context"
                st.session_state["risk_bundle"] = refreshed_single
                bundle = refreshed_single
        if bundle and bundle.get("logic_version") != RISK_BUNDLE_LOGIC_VERSION:
            st.session_state.pop("risk_bundle", None)
            bundle = None
        invalidated_multi_bundle = False
        if _should_invalidate_multi_bundle(bundle, current_upload_key, st.session_state.get("risk_mode", "single")):
            st.session_state.pop("risk_bundle", None)
            bundle = None
            invalidated_multi_bundle = True
        if bundle and bundle.get("source") == "single":
            live_bundle = _single_bundle_from_latest_extraction(update_state=False)
            if not live_bundle:
                st.session_state.pop("risk_bundle", None)
                bundle = None
            elif bundle.get("context_hash") != live_bundle.get("context_hash"):
                st.session_state["risk_bundle"] = live_bundle
                bundle = live_bundle
        if not bundle and st.session_state.get("risk_mode") != "multi" and not invalidated_multi_bundle and not run_multi and not use_single:
            latest_bundle = _single_bundle_from_latest_extraction(update_state=True)
            if latest_bundle:
                st.session_state["risk_bundle"] = latest_bundle
                st.session_state["risk_bundle"]["analysis_mode"] = st.session_state.get("analysis_year_mode", "latest_available")
                st.session_state["risk_bundle"]["analysis_specific_year"] = st.session_state.get("analysis_specific_year", "")
                st.session_state["risk_mode"] = "single"
                bundle = st.session_state.get("risk_bundle")
        if not bundle:
            if invalidated_multi_bundle:
                st.error("Dataset changed. Regenerate Unified Risk Table.")
                return
            st.info("Upload documents and click Generate Unified Risk Table.")
            return

        combined_df = bundle["combined_df"]
        confidence_level = bundle.get("confidence_level", "Basic")
        if combined_df is None or combined_df.empty:
            st.warning("No analyzable values found in uploaded documents.")
            return
        available_years_set: set[int] = set()
        for v in combined_df.get("Selected Year", pd.Series(dtype=float)).tolist():
            n = _to_numeric_financial(v)
            if n is not None:
                available_years_set.add(int(n))
        if "Available Years" in combined_df.columns:
            for raw in combined_df["Available Years"].dropna().tolist():
                for y in re.findall(r"\b(19\d{2}|20\d{2})\b", str(raw)):
                    try:
                        available_years_set.add(int(y))
                    except Exception:
                        pass
        available_years = sorted(available_years_set)
        # If a previously selected specific year is no longer available, fall back to latest.
        try:
            if available_years:
                if st.session_state.get("analysis_year_mode") == "specific_year" and st.session_state.get("analysis_specific_year") not in available_years:
                    st.session_state["analysis_year_mode"] = "latest_available"
                    st.session_state["analysis_specific_year"] = ""
                    st.session_state["analysis_year_mode_selector"] = "Latest Available"
        except Exception:
            pass
        ym1, ym2, ym3 = st.columns([1, 1, 1], vertical_alignment="bottom")
        analysis_mode = ym1.selectbox(
            "Analysis Mode",
            ["Latest Available", "Specific Year", "T12M"],
            key="analysis_year_mode_selector",
        )
        specific_year = None
        if analysis_mode == "Specific Year":
            if available_years:
                specific_year = ym2.selectbox("Specific Fiscal Year", available_years, index=len(available_years) - 1, key="analysis_year_specific_selector")
            else:
                ym2.selectbox("Specific Fiscal Year", ["No years found"], disabled=True, key="analysis_year_specific_selector_disabled")
        mode_key = "latest_available" if analysis_mode == "Latest Available" else ("specific_year" if analysis_mode == "Specific Year" else "t12m")
        st.session_state["analysis_year_mode"] = mode_key
        st.session_state["analysis_specific_year"] = int(specific_year) if specific_year is not None else ""
        bundle_mode = str(bundle.get("analysis_mode") or "")
        bundle_year = bundle.get("analysis_specific_year")
        if bundle_mode != mode_key or bundle_year != st.session_state["analysis_specific_year"]:
            if bundle.get("source") == "single":
                refreshed = _single_bundle_from_latest_extraction(update_state=True)
                if refreshed:
                    refreshed["analysis_mode"] = mode_key
                    refreshed["analysis_specific_year"] = st.session_state["analysis_specific_year"]
                    st.session_state["risk_bundle"] = refreshed
                    st.session_state["risk_recompute_reason"] = "analysis mode change"
                    st.rerun()
            elif bundle.get("source") == "multi" and risk_docs:
                refreshed = _build_credit_risk_bundle(risk_docs)
                refreshed["source"] = "multi"
                refreshed["context_key"] = current_upload_key
                refreshed["extracted_at"] = datetime.now().isoformat(timespec="seconds")
                refreshed["analysis_mode"] = mode_key
                refreshed["analysis_specific_year"] = st.session_state["analysis_specific_year"]
                st.session_state["risk_bundle"] = refreshed
                st.session_state["risk_recompute_reason"] = "analysis mode change"
                st.rerun()
        lock_result = _lock_analysis_dataset(
            combined_df,
            mode=mode_key,
            specific_year=(int(specific_year) if specific_year is not None else None),
        )
        locked_combined_df = lock_result.get("locked_df", combined_df)
        locked_year = lock_result.get("locked_year")
        ym3.markdown(
            f"<div class='panel'><strong>Active Fiscal Year</strong><br/><span style='color:#4e6178'>{html.escape(str(locked_year) if locked_year is not None else 'N/A')}</span></div>",
            unsafe_allow_html=True,
        )
        st.markdown(
            f"<div class='panel'><strong>Analysis Context</strong><br/><span style='color:#4e6178'>Mode: <strong>{html.escape(analysis_mode)}</strong> | Locked Year: <strong>{html.escape(str(locked_year) if locked_year is not None else 'N/A')}</strong> | Available Years: {', '.join([str(y) for y in lock_result.get('available_years', [])]) if lock_result.get('available_years') else 'None'}</span></div>",
            unsafe_allow_html=True,
        )
        for w in lock_result.get("warnings", []):
            st.warning(w)
        if lock_result.get("critical_year_selection_failure"):
            st.error("Year selection engine failure.")
            return
        if lock_result.get("cross_year_error_flag"):
            st.error("Cross-year data mismatch detected. Please reselect analysis year.")
            return
        combined_df = locked_combined_df
        normalized_metric_df = _normalized_financial_metric_dataset(combined_df)
        st.session_state["risk_normalized_metric_df"] = normalized_metric_df
        if normalized_metric_df.empty:
            st.warning("No normalized financial metrics available for the locked analysis year.")
        if bundle.get("failed_docs"):
            st.warning(f"Some files failed and were skipped: {', '.join(bundle.get('failed_docs', []))}")
        block_reason = _multi_block_reason(bundle)
        bundle_scope_key = str(bundle.get("context_hash") or bundle.get("context_key") or "")
        mismatch_override_key = f"risk_borrower_mismatch_override::{bundle_scope_key}"
        missing_docs_override_key = f"risk_missing_docs_override::{bundle_scope_key}"
        if block_reason:
            if bundle.get("borrower_mismatch"):
                if not st.session_state.get(mismatch_override_key, False):
                    st.error(block_reason)
                    borrower_map = bundle.get("borrower_by_doc", {})
                    mismatch_list = ", ".join([f"{doc}: {name}" for doc, name in borrower_map.items()])
                    if mismatch_list:
                        st.caption(f"Detected borrowers: {mismatch_list}")
                    b1, b2 = st.columns([1, 1], vertical_alignment="bottom")
                    if b1.button("Proceed anyway", key=f"risk_proceed_anyway_{bundle_scope_key}", type="primary", use_container_width=True):
                        st.session_state[mismatch_override_key] = True
                        st.rerun()
                    if b2.button("Cancel", key=f"risk_cancel_mismatch_{bundle_scope_key}", use_container_width=True):
                        st.info("Upload documents for one borrower to continue with strict validation.")
                    return
                st.warning("Proceeding with mixed borrower documents by user override.")
            elif bundle.get("source") == "multi" and bundle.get("completeness", {}).get("missing_buckets"):
                if not st.session_state.get(missing_docs_override_key, False):
                    st.error(block_reason)
                    c1, c2 = st.columns([1, 1], vertical_alignment="bottom")
                    if c1.button("Proceed anyway", key=f"risk_proceed_missing_docs_{bundle_scope_key}", type="primary", use_container_width=True):
                        st.session_state[missing_docs_override_key] = True
                        st.rerun()
                    if c2.button("Cancel", key=f"risk_cancel_missing_docs_{bundle_scope_key}", use_container_width=True):
                        st.info("Upload missing documents for full analysis, or proceed with partial scoring.")
                    return
                st.warning("Proceeding with missing document buckets by user override.")
            else:
                st.error(block_reason)
                return

        conflict_items: List[Dict[str, Any]] = []
        # Always auto-apply a deterministic rule. Conflicts are informational only.
        rule_opt = "highest_confidence"
        st.session_state["risk_conflict_rule"] = "highest_confidence"
        if str(bundle.get("source", "single")) == "multi":
            conflict_items = _detect_metric_conflicts(combined_df, bundle.get("doc_types", []))
            if conflict_items:
                st.caption("Conflict resolution rule: Highest Confidence (auto-applied)")
                preview = pd.DataFrame(conflict_items[:8])
                with st.expander("Conflicting metric values detected", expanded=False):
                    st.dataframe(preview, width="stretch", hide_index=True)

        st.markdown(
            f"<div class='panel'><strong>Document Check</strong><br/><span style='color:#4e6178'>Confidence: <strong>{confidence_level}</strong> | Files Uploaded: {len(bundle.get('processed_docs', []))}</span></div>",
            unsafe_allow_html=True,
        )
        st.caption(
            f"Runtime: {APP_BUILD_ID} | Source: {bundle.get('source', 'single')} | Context: {bundle.get('context_key', 'n/a')}"
        )

        def _infer_risk_driver_suggestions() -> Dict[str, Any]:
            out: Dict[str, Any] = {
                "industry": "Technology",
                "geography": "United States Tier 1",
                "business_stage": "Mature",
                "company_size": "Medium",
                "loan_type": "Term Loan",
                "currency_scale": "Units",
                "years_in_operation": 5,
                "requested_amount": 10_000_000.0,
            }
            industry_guess = _extract_first_value(combined_df, ["Industry", "Sector", "Business Type"])
            geo_guess = _extract_first_value(combined_df, ["Region", "Geography", "Country"])
            years_guess = _extract_num(combined_df, ["Years in Operation", "Operating History"])
            req_guess = _extract_num(combined_df, ["Loan Amount", "Facility Amount", "Requested Amount"])

            if industry_guess:
                norm_ind = str(industry_guess).strip().lower()
                for opt in INDUSTRY_RISK_MULTIPLIER.keys():
                    if opt.lower() in norm_ind or norm_ind in opt.lower():
                        out["industry"] = opt
                        break
            if geo_guess:
                norm_geo = str(geo_guess).strip().lower()
                for opt in GEOGRAPHY_MULTIPLIER.keys():
                    if any(tok in norm_geo for tok in [opt.lower(), "united states", "canada", "emerging", "sanctioned"]):
                        if "united states" in norm_geo:
                            out["geography"] = "United States Tier 1"
                        elif "canada" in norm_geo:
                            out["geography"] = "Canada"
                        elif "high volatility" in norm_geo:
                            out["geography"] = "Emerging Market - High Volatility"
                        elif "emerging" in norm_geo:
                            out["geography"] = "Emerging Market - Low Stability"
                        elif "sanction" in norm_geo or "high risk" in norm_geo:
                            out["geography"] = "Sanctioned / High Risk Region"
                        break
            if years_guess is not None:
                out["years_in_operation"] = int(max(0, round(years_guess)))
                if out["years_in_operation"] < 3:
                    out["business_stage"] = "Startup"
                elif out["years_in_operation"] < 7:
                    out["business_stage"] = "Growth"
            # Requested amount must be a positive principal-like number.
            # Ignore zero/negative or tiny ratio-like values from noisy extraction.
            if req_guess is not None and float(req_guess) > 1_000:
                out["requested_amount"] = float(req_guess)
            detected_scale_local = _preferred_scale_from_df(combined_df)
            if detected_scale_local in {"Units", "Thousands", "Millions", "Billions", "Trillions"}:
                out["currency_scale"] = detected_scale_local
            elif bundle.get("processed_docs"):
                max_factor = 1.0
                for doc_name in bundle.get("processed_docs", []):
                    p = LIB_DIR / str(doc_name)
                    if not p.exists():
                        candidates = [x for x in WORKSPACE_PDFS if x.name == str(doc_name)]
                        p = candidates[0] if candidates else p
                    if p.exists():
                        try:
                            max_factor = max(max_factor, _statement_unit_factor(p))
                        except Exception:
                            pass
                scale_from_docs = _scale_label_from_factor(max_factor)
                if scale_from_docs in {"Units", "Thousands", "Millions", "Billions", "Trillions"}:
                    out["currency_scale"] = scale_from_docs
            return out

        # Use stable hash-based context signature to avoid resetting user-edited inputs on reruns.
        context_sig = f"{bundle.get('source')}|{bundle.get('context_hash') or bundle.get('context_key')}"
        suggestions = _infer_risk_driver_suggestions()
        if st.session_state.get("risk_driver_context") != context_sig:
            st.session_state["risk_driver_context"] = context_sig
            st.session_state["risk_driver_inputs"] = suggestions.copy()
            st.session_state["risk_driver_saved"] = suggestions.copy()
            st.session_state["risk_applied_inputs"] = suggestions.copy()
            st.session_state["risk_show_input_modal"] = True
            for k in list(st.session_state.keys()):
                if k.startswith("dlg_val_"):
                    st.session_state.pop(k, None)
            st.session_state["cra2_industry"] = suggestions["industry"]
            st.session_state["cra2_geo"] = suggestions["geography"]
            st.session_state["cra2_stage"] = suggestions["business_stage"]
            st.session_state["cra2_size"] = suggestions["company_size"]
            st.session_state["cra2_loan_type"] = suggestions.get("loan_type", "Term Loan")
            st.session_state["cra2_currency_scale"] = suggestions.get("currency_scale", "Units")
            st.session_state["cra2_years"] = int(suggestions["years_in_operation"])
            st.session_state["cra2_req"] = float(suggestions["requested_amount"])
            st.session_state["risk_data_version"] = 1
            st.session_state["risk_last_saved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state["risk_saved_context_hash"] = bundle.get("context_hash")
            st.session_state["risk_recompute_reason"] = "dataset regenerate"
            st.session_state["risk_has_unsaved_changes"] = False

        if st.session_state.get("risk_show_input_modal", False) and hasattr(st, "dialog"):
            @st.dialog("Suggested Risk Driver Inputs")
            def _risk_input_dialog() -> None:
                st.caption("Suggested values are preselected. Update any value and click Save.")
                fields = [
                    ("Industry Type", "industry", list(INDUSTRY_RISK_MULTIPLIER.keys())),
                    ("Geographic Risk", "geography", list(GEOGRAPHY_MULTIPLIER.keys())),
                    ("Business Stage", "business_stage", list(BUSINESS_STAGE_MULTIPLIER.keys())),
                    ("Company Size", "company_size", ["Small", "Medium", "Large"]),
                    ("Loan Type", "loan_type", ["Term Loan", "Revolving Credit", "Bridge Loan", "Working Capital", "Asset-Based"]),
                    ("Currency Scale", "currency_scale", ["Units", "Thousands", "Millions", "Billions", "Trillions"]),
                ]
                for label, key, options in fields:
                    dc1, dc2 = st.columns([1.2, 1.8])
                    dc1.markdown(f"**{label} :red[*]**")
                    dlg_key = f"dlg_val_{key}"
                    if dlg_key not in st.session_state:
                        default_v = st.session_state["risk_driver_inputs"].get(key, options[0])
                        st.session_state[dlg_key] = default_v if default_v in options else options[0]
                    dc2.selectbox("Value", options, key=dlg_key, label_visibility="collapsed")

                n1, n2 = st.columns([1.2, 1.8])
                n1.markdown("**Years in Operation :red[*]**")
                if "dlg_val_years_in_operation" not in st.session_state:
                    st.session_state["dlg_val_years_in_operation"] = int(st.session_state["risk_driver_inputs"].get("years_in_operation", 5))
                n2.number_input("Value", min_value=0, max_value=200, step=1, key="dlg_val_years_in_operation", label_visibility="collapsed")

                r1, r2 = st.columns([1.2, 1.8])
                r1.markdown("**Requested Amount :red[*]**")
                if "dlg_val_requested_amount" not in st.session_state:
                    st.session_state["dlg_val_requested_amount"] = float(st.session_state["risk_driver_inputs"].get("requested_amount", 10_000_000.0))
                r2.number_input("Value", min_value=0.0, step=100_000.0, key="dlg_val_requested_amount", label_visibility="collapsed")

                s1, s2 = st.columns(2)
                if s1.button("Cancel", use_container_width=True):
                    st.session_state["risk_show_input_modal"] = False
                    st.rerun()
                if s2.button("Save", type="primary", use_container_width=True):
                    existing = st.session_state.get("risk_driver_inputs", suggestions.copy())
                    for _, key, options in fields:
                        existing[key] = st.session_state.get(f"dlg_val_{key}", options[0])
                    existing["years_in_operation"] = int(st.session_state.get("dlg_val_years_in_operation", 5))
                    existing["requested_amount"] = float(st.session_state.get("dlg_val_requested_amount", 10_000_000.0))
                    required_text_fields = ["industry", "geography", "business_stage", "company_size", "loan_type", "currency_scale"]
                    missing_required = [k for k in required_text_fields if not str(existing.get(k, "")).strip()]
                    if str(existing.get("currency_scale", "")).strip() not in {"Units", "Thousands", "Millions", "Billions", "Trillions"}:
                        missing_required.append("currency_scale")
                    if existing["years_in_operation"] <= 0:
                        missing_required.append("years_in_operation")
                    if existing["requested_amount"] <= 0:
                        missing_required.append("requested_amount")
                    if missing_required:
                        st.error("All fields are mandatory. Enter valid values for every field before saving.")
                        return
                    st.session_state["risk_driver_inputs"] = existing
                    st.session_state["risk_driver_saved"] = existing.copy()
                    st.session_state["risk_applied_inputs"] = existing.copy()
                    st.session_state["cra2_industry"] = existing["industry"]
                    st.session_state["cra2_geo"] = existing["geography"]
                    st.session_state["cra2_stage"] = existing["business_stage"]
                    st.session_state["cra2_size"] = existing["company_size"]
                    st.session_state["cra2_loan_type"] = existing.get("loan_type", "Term Loan")
                    st.session_state["cra2_currency_scale"] = existing.get("currency_scale", "Units")
                    st.session_state["cra2_years"] = int(existing["years_in_operation"])
                    st.session_state["cra2_req"] = float(existing["requested_amount"])
                    # Sync popup saves directly into inline controls.
                    st.session_state["risk_inline_industry"] = existing["industry"]
                    st.session_state["risk_inline_geography"] = existing["geography"]
                    st.session_state["risk_inline_stage"] = existing["business_stage"]
                    st.session_state["risk_inline_size"] = existing["company_size"]
                    st.session_state["risk_inline_loan_type"] = existing.get("loan_type", "Term Loan")
                    st.session_state["risk_inline_currency_scale"] = existing.get("currency_scale", "Units")
                    st.session_state["risk_inline_years"] = int(existing["years_in_operation"])
                    st.session_state["risk_inline_requested"] = float(existing["requested_amount"])
                    st.session_state["risk_data_version"] = int(st.session_state.get("risk_data_version", 0)) + 1
                    st.session_state["risk_last_saved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    st.session_state["risk_saved_context_hash"] = bundle.get("context_hash")
                    st.session_state["risk_recompute_reason"] = "context save"
                    st.session_state["risk_show_input_modal"] = False
                    st.session_state["risk_apply_refresh_nonce"] = int(st.session_state.get("risk_apply_refresh_nonce", 0)) + 1
                    st.rerun()
            _risk_input_dialog()

        if "cra2_loan_type" not in st.session_state:
            st.session_state["cra2_loan_type"] = suggestions.get("loan_type", "Term Loan")
        if "cra2_currency_scale" not in st.session_state:
            st.session_state["cra2_currency_scale"] = suggestions.get("currency_scale", "Units")
        applied_inputs = st.session_state.get("risk_applied_inputs", st.session_state.get("risk_driver_saved", suggestions.copy()))
        applied_inputs["loan_type"] = st.session_state.get("cra2_loan_type", applied_inputs.get("loan_type", "Term Loan"))
        applied_inputs["currency_scale"] = st.session_state.get("cra2_currency_scale", applied_inputs.get("currency_scale", "Units"))
        detected_scale = _preferred_scale_from_df(combined_df)
        if detected_scale == "Units":
            # Fallback: detect scale directly from uploaded PDFs in this bundle.
            for doc_name in bundle.get("processed_docs", []):
                doc_path = LIB_DIR / str(doc_name)
                if not doc_path.exists():
                    continue
                uf = _statement_unit_factor(doc_path)
                if uf == 1_000_000_000_000.0:
                    detected_scale = "Trillions"
                    break
                if uf == 1_000_000_000.0:
                    detected_scale = "Billions"
                    break
                if uf == 1_000_000.0:
                    detected_scale = "Millions"
                    break
                if uf == 1_000.0:
                    detected_scale = "Thousands"
                    break
        current_scale = str(applied_inputs.get("currency_scale", "Units"))
        if current_scale not in {"Units", "Thousands", "Millions", "Billions", "Trillions"}:
            current_scale = detected_scale if detected_scale in {"Units", "Thousands", "Millions", "Billions", "Trillions"} else "Units"
            applied_inputs["currency_scale"] = current_scale
            st.session_state["cra2_currency_scale"] = current_scale
        # Auto-correct stale session values when document units are explicit whole dollars.
        if detected_scale == "Units" and current_scale != "Units":
            current_scale = "Units"
            applied_inputs["currency_scale"] = current_scale
            st.session_state["cra2_currency_scale"] = current_scale
            st.session_state["risk_inline_currency_scale"] = current_scale
            st.session_state["risk_driver_inputs"] = {**st.session_state.get("risk_driver_inputs", {}), "currency_scale": current_scale}
            st.session_state["risk_driver_saved"] = {**st.session_state.get("risk_driver_saved", {}), "currency_scale": current_scale}
            st.session_state["risk_applied_inputs"] = {**st.session_state.get("risk_applied_inputs", {}), "currency_scale": current_scale}
        if current_scale != detected_scale:
            w1, w2 = st.columns([4, 1], vertical_alignment="center")
            w1.warning(f"Detected document unit is '{detected_scale}', but current Currency Scale is '{current_scale}'.")
            if w2.button("Use Detected Scale", key="risk_use_detected_scale", use_container_width=True):
                applied_inputs["currency_scale"] = detected_scale
                st.session_state["cra2_currency_scale"] = detected_scale
                st.session_state["risk_inline_currency_scale"] = detected_scale
                st.session_state["risk_driver_inputs"] = {**st.session_state.get("risk_driver_inputs", {}), "currency_scale": detected_scale}
                st.session_state["risk_driver_saved"] = {**st.session_state.get("risk_driver_saved", {}), "currency_scale": detected_scale}
                st.session_state["risk_applied_inputs"] = {**st.session_state.get("risk_applied_inputs", {}), "currency_scale": detected_scale}
                st.rerun()
        bundle_doc_types = bundle.get("doc_types", [])
        applied_for_tables = bundle_doc_types[0] if bundle_doc_types else "Financial Statements"
        if applied_for_tables == "Covenant Compliance Certificate":
            applied_for_tables = "Compliance Certificate"
        if applied_for_tables == "Forecast/Projections":
            applied_for_tables = "Forecast Model"
        if applied_for_tables not in DOC_TYPE_CONFIG:
            applied_for_tables = "Financial Statements"
        risk_compute_key = _stable_json_digest(
            {
                "context_hash": bundle.get("context_hash"),
                "context_key": bundle.get("context_key"),
                "logic_version": bundle.get("logic_version"),
                "doc_types": bundle_doc_types,
                "applied_for_tables": applied_for_tables,
                "rule_opt": rule_opt,
                "analysis_mode": analysis_mode,
                "locked_year": lock_result.get("locked_year"),
                "available_years": lock_result.get("available_years", []),
                "t12m_status": str(lock_result.get("t12m_status", "n/a")),
                "cross_year_error_flag": bool(lock_result.get("cross_year_error_flag", False)),
                "inputs": {
                    "industry": applied_inputs["industry"],
                    "geography": applied_inputs["geography"],
                    "business_stage": applied_inputs["business_stage"],
                    "company_size": applied_inputs["company_size"],
                    "loan_type": str(applied_inputs.get("loan_type", "Term Loan")),
                    "years_in_operation": int(applied_inputs["years_in_operation"]),
                    "requested_amount": float(applied_inputs["requested_amount"]),
                    "currency_scale": str(applied_inputs.get("currency_scale", "Units")),
                },
            }
        )
        risk_compute_cache = st.session_state.get("risk_compute_cache", {})
        cached_compute = risk_compute_cache.get(risk_compute_key) if isinstance(risk_compute_cache, dict) else None
        if isinstance(cached_compute, dict) and "tables" in cached_compute and "model" in cached_compute:
            tables = cached_compute["tables"]
            model = cached_compute["model"]
        else:
            tables = _build_credit_risk_tables(
                combined_df,
                applied_for_tables,
                bundle_doc_types,
                selection_rule=rule_opt,
                locked_year=lock_result.get("locked_year"),
            )
            model = _build_dynamic_credit_analysis(
                tables=tables,
                combined_df=combined_df,
                industry=applied_inputs["industry"],
                geography=applied_inputs["geography"],
                business_stage=applied_inputs["business_stage"],
                company_size=applied_inputs["company_size"],
                years_in_operation=int(applied_inputs["years_in_operation"]),
                requested_amount=float(applied_inputs["requested_amount"]),
                currency_scale=str(applied_inputs.get("currency_scale", "Units")),
                analysis_mode=analysis_mode,
                locked_year=lock_result.get("locked_year"),
                available_years=lock_result.get("available_years", []),
                t12m_status=str(lock_result.get("t12m_status", "n/a")),
                cross_year_error_flag=bool(lock_result.get("cross_year_error_flag", False)),
            )
            if not isinstance(risk_compute_cache, dict):
                risk_compute_cache = {}
            risk_compute_cache[risk_compute_key] = {"tables": tables, "model": model}
            if len(risk_compute_cache) > 8:
                # keep cache bounded to avoid session memory growth
                for k in list(risk_compute_cache.keys())[:-8]:
                    risk_compute_cache.pop(k, None)
            st.session_state["risk_compute_cache"] = risk_compute_cache
        risk_df = model["table"].copy()
        risk_df = risk_df.where(pd.notna(risk_df), None)
        _maybe_persist_credit_analysis_to_backend(model, applied_inputs, risk_compute_key)
        loan_type_display = str(applied_inputs.get("loan_type", "Term Loan"))
        currency_scale_display = str(applied_inputs.get("currency_scale", "Units"))
        industry_opts = list(INDUSTRY_RISK_MULTIPLIER.keys())
        geo_opts = list(GEOGRAPHY_MULTIPLIER.keys())
        stage_opts = list(BUSINESS_STAGE_MULTIPLIER.keys())
        size_opts = ["Small", "Medium", "Large"]
        loan_type_opts = ["Term Loan", "Revolving Credit", "Bridge Loan", "Working Capital", "Asset-Based"]
        currency_scale_opts = ["Units", "Thousands", "Millions", "Billions", "Trillions"]

        if "risk_inline_industry" not in st.session_state:
            st.session_state["risk_inline_industry"] = str(applied_inputs.get("industry", "Technology"))
        if "risk_inline_geography" not in st.session_state:
            st.session_state["risk_inline_geography"] = str(applied_inputs.get("geography", "United States Tier 1"))
        if "risk_inline_stage" not in st.session_state:
            st.session_state["risk_inline_stage"] = str(applied_inputs.get("business_stage", "Mature"))
        if "risk_inline_size" not in st.session_state:
            st.session_state["risk_inline_size"] = str(applied_inputs.get("company_size", "Medium"))
        if "risk_inline_loan_type" not in st.session_state:
            st.session_state["risk_inline_loan_type"] = loan_type_display
        if "risk_inline_currency_scale" not in st.session_state:
            st.session_state["risk_inline_currency_scale"] = currency_scale_display
        if "risk_inline_years" not in st.session_state:
            st.session_state["risk_inline_years"] = int(applied_inputs.get("years_in_operation", 5))
        if "risk_inline_requested" not in st.session_state:
            st.session_state["risk_inline_requested"] = float(applied_inputs.get("requested_amount", 10_000_000.0))

        st.markdown("<div class='panel'><strong>Underwriting Inputs</strong></div>", unsafe_allow_html=True)
        st.markdown("<div class='risk-inline-wrap'>", unsafe_allow_html=True)
        with st.container(border=True):
            ic1, ic2, ic3, ic4, ic5, ic6 = st.columns(6, vertical_alignment="bottom")
            industry_inline = ic1.selectbox("Industry :red[*]", industry_opts, key="risk_inline_industry")
            geography_inline = ic2.selectbox("Geographic Risk :red[*]", geo_opts, key="risk_inline_geography")
            business_stage_inline = ic3.selectbox("Business Stage :red[*]", stage_opts, key="risk_inline_stage")
            company_size_inline = ic4.selectbox("Company Size :red[*]", size_opts, key="risk_inline_size")
            loan_type_inline = ic5.selectbox("Loan Type :red[*]", loan_type_opts, key="risk_inline_loan_type")
            years_inline = int(ic6.number_input("Years in Operation :red[*]", min_value=0, max_value=200, step=1, key="risk_inline_years"))

            ir1, ir2, ir3, ir4 = st.columns([1.0, 1.0, 0.45, 2.0], vertical_alignment="bottom")
            requested_inline = float(ir1.number_input("Requested Amount :red[*]", min_value=0.0, step=100_000.0, key="risk_inline_requested"))
            currency_scale_inline = ir2.selectbox("Currency Scale :red[*]", currency_scale_opts, key="risk_inline_currency_scale")
            save_inline = ir3.button("Save", key="risk_inline_save", type="primary", use_container_width=True)
            draft_inputs = {
                "industry": industry_inline,
                "geography": geography_inline,
                "business_stage": business_stage_inline,
                "company_size": company_size_inline,
                "loan_type": loan_type_inline,
                "currency_scale": currency_scale_inline,
                "years_in_operation": years_inline,
                "requested_amount": requested_inline,
            }
            dirty_inline = draft_inputs != {
                "industry": str(applied_inputs.get("industry", "Technology")),
                "geography": str(applied_inputs.get("geography", "United States Tier 1")),
                "business_stage": str(applied_inputs.get("business_stage", "Mature")),
                "company_size": str(applied_inputs.get("company_size", "Medium")),
                "loan_type": str(applied_inputs.get("loan_type", "Term Loan")),
                "currency_scale": str(applied_inputs.get("currency_scale", "Units")),
                "years_in_operation": int(applied_inputs.get("years_in_operation", 5)),
                "requested_amount": float(applied_inputs.get("requested_amount", 10_000_000.0)),
            }
            st.session_state["risk_has_unsaved_changes"] = dirty_inline
            ir4.caption("Unsaved edits" if dirty_inline else "Saved. Click Save after changes.")

            long_labels: List[str] = []
            if len(str(geography_inline)) > 22:
                long_labels.append(f"Geographic Risk: {geography_inline}")
            if len(str(loan_type_inline)) > 22:
                long_labels.append(f"Loan Type: {loan_type_inline}")
            if len(str(industry_inline)) > 22:
                long_labels.append(f"Industry: {industry_inline}")
            if long_labels:
                st.caption(" | ".join(long_labels))
        st.markdown("</div>", unsafe_allow_html=True)

        if save_inline:
            updated = draft_inputs
            required_text_fields = ["industry", "geography", "business_stage", "company_size", "loan_type", "currency_scale"]
            missing_required = [k for k in required_text_fields if not str(updated.get(k, "")).strip()]
            if str(updated.get("currency_scale", "")).strip() not in {"Units", "Thousands", "Millions", "Billions", "Trillions"}:
                missing_required.append("currency_scale")
            if int(updated["years_in_operation"]) <= 0:
                missing_required.append("years_in_operation")
            if float(updated["requested_amount"]) <= 0:
                missing_required.append("requested_amount")
            if missing_required:
                st.error("All fields are mandatory. Enter valid values for every field before saving.")
                return
            st.session_state["risk_driver_inputs"] = updated.copy()
            st.session_state["risk_driver_saved"] = updated.copy()
            st.session_state["risk_applied_inputs"] = updated.copy()
            st.session_state["cra2_industry"] = updated["industry"]
            st.session_state["cra2_geo"] = updated["geography"]
            st.session_state["cra2_stage"] = updated["business_stage"]
            st.session_state["cra2_size"] = updated["company_size"]
            st.session_state["cra2_loan_type"] = updated["loan_type"]
            st.session_state["cra2_currency_scale"] = updated.get("currency_scale", "Units")
            st.session_state["cra2_years"] = int(updated["years_in_operation"])
            st.session_state["cra2_req"] = float(updated["requested_amount"])
            st.session_state["risk_data_version"] = int(st.session_state.get("risk_data_version", 0)) + 1
            st.session_state["risk_last_saved_at"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            st.session_state["risk_saved_context_hash"] = bundle.get("context_hash")
            st.session_state["risk_recompute_reason"] = "driver change"
            st.session_state["risk_has_unsaved_changes"] = False
            st.session_state["risk_apply_refresh_nonce"] = int(st.session_state.get("risk_apply_refresh_nonce", 0)) + 1
            applied_inputs = updated.copy()
            st.rerun()

        if model.get("grade") == "Provisional":
            detected = model.get("documents_detected", []) or []
            detected_line = " | ".join([f"{'YES' if d.get('present') else 'NO'} {d.get('name', '')}" for d in detected[:4]])
            missing_docs_list = model.get("missing_document_types", []) or []
            missing_docs = ", ".join(missing_docs_list) if missing_docs_list else "None"
            st.markdown(
                f"<div class='provisional-box'><strong>Provisional Analysis - Incomplete Financial Package:</strong> More than 40% of financial metrics are missing or critical liquidity/leverage data is unavailable."
                f"<br/><strong>Documents detected:</strong> {html.escape(detected_line)}"
                f"<br/><strong>Missing documents:</strong> {html.escape(missing_docs)}"
                f"<br/>Liquidity and leverage metrics require additional documents. Risk score may change upon full submission.</div>",
                unsafe_allow_html=True,
            )
        elif model.get("missing_data_warning"):
            st.markdown(
                f"<div class='provisional-box'><strong>Data Gap:</strong> {html.escape(str(model.get('missing_data_warning')))}</div>",
                unsafe_allow_html=True,
            )

        hdr_l, hdr_r = st.columns([0.75, 0.25], vertical_alignment="center")
        with hdr_l:
            st.markdown("<div class='panel'><strong>Unified Credit Analysis Table</strong></div>", unsafe_allow_html=True)
        with hdr_r:
            risk_pkg_excel = _to_credit_risk_package_excel_bytes(tables, model)
            st.download_button(
                "Download Full Credit Risk Package",
                data=risk_pkg_excel,
                file_name="credit_risk_package.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                key="download_full_credit_risk_package",
            )
        _render_risk_html_table(
            risk_df,
            "_row_level",
            [
                "Category",
                "Metric",
                "Calculated Value",
                "Industry Threshold",
                "Base Score",
                "Adjusted Score",
                "Status",
                "Risk",
            ],
        )
        combined_mult = float(model.get("combined_multiplier", 1.0) or 1.0)
        if abs(combined_mult - 1.0) < 1e-9:
            st.markdown(
                "<span style='color:#475569;font-size:0.86rem;font-weight:600;'>Scoring details</span>"
                "<span class='hover-help' title='Base and Adjusted are equal because selected multipliers are baseline: industry 1.00 × geography 1.00 × stage 1.00.'>i</span>",
                unsafe_allow_html=True,
            )
        else:
            mm = model.get("multiplier_components", {}) or {}
            st.markdown(
                "<span style='color:#475569;font-size:0.86rem;font-weight:600;'>Scoring details</span>"
                f"<span class='hover-help' title='Adjusted Score = Base Score ÷ {combined_mult:.2f} "
                f"(industry {float(mm.get('industry', 1.0) or 1.0):.2f} × "
                f"geography {float(mm.get('geography', 1.0) or 1.0):.2f} × "
                f"stage {float(mm.get('business_stage', 1.0) or 1.0):.2f}).'>i</span>",
                unsafe_allow_html=True,
            )
        with st.expander("Metric Source Trace", expanded=False):
            src_cols = [c for c in ["Metric", "Source Trace", "Calculated Value", "Industry Threshold"] if c in risk_df.columns]
            if src_cols:
                st.dataframe(_ui_null_df(risk_df[src_cols]), width="stretch", hide_index=True)

        overall_style = _risk_style(_color_to_row_level(model["overall_color"]))
        approved_pct = 0.0
        approval_status = "Partially Approved"
        excess_capacity = 0.0
        try:
            approved_limit_val = float(model.get("policy_approved_limit", 0) or 0)
            requested_val = float(model.get("requested_amount", 0) or 0)
            capacity_val = float((model.get("policy_limit_breakdown", {}) or {}).get("final_policy_limit", approved_limit_val) or approved_limit_val)
            if requested_val > 0 and approved_limit_val >= requested_val:
                approved_pct = 100.0
                approval_status = "Fully Approved"
                excess_capacity = max(0.0, capacity_val - requested_val)
            elif requested_val > 0 and approved_limit_val > 0:
                approved_pct = (approved_limit_val / requested_val) * 100.0
                approval_status = "Partially Approved"
                excess_capacity = 0.0
            else:
                approved_pct = 0.0
                approval_status = "Declined"
                excess_capacity = 0.0
        except Exception:
            approved_pct = 0.0
            approval_status = "Declined"
            excess_capacity = 0.0
        breakdown = model.get("policy_limit_breakdown", {}) or {}
        base_limit = float(breakdown.get("base_policy_limit", 0.0) or 0.0)
        policy_capacity_limit = float(breakdown.get("final_policy_limit", model.get("policy_approved_limit", 0.0)) or 0.0)
        ind_adj_pct = float(breakdown.get("industry_adjustment_pct", 0.0) or 0.0)
        geo_adj_pct = float(breakdown.get("geography_adjustment_pct", 0.0) or 0.0)
        mat_adj_pct = float(breakdown.get("maturity_adjustment_pct", 0.0) or 0.0)
        ind_amt = base_limit * ind_adj_pct
        geo_amt = base_limit * geo_adj_pct
        mat_amt = base_limit * mat_adj_pct
        policy_unit_label = "USD (base units)"
        cap_used = model.get("capacity_components_used", {}) or {}
        coverage_capacity_used = bool(cap_used.get("dscr_capacity")) or bool(cap_used.get("interest_coverage_capacity"))
        limit_basis_sentence = (
            "Based on EBITDA leverage capacity, coverage capacity, and policy caps. Independent of requested amount."
            if coverage_capacity_used
            else "Limit determined based on policy cap and available financial data."
        )
        approved_limit_abbrev = _format_abbrev_number(model.get("policy_approved_limit", 0))
        approved_limit_full = f"{float(model.get('policy_approved_limit', 0) or 0):,.0f}"
        score_card_class = "risk-score-card provisional" if model.get("grade") == "Provisional" else "risk-score-card"
        statement_scale_label = _currency_scale_display_label(model.get("currency_scale", "Units"))
        st.markdown(
            f"""
            <div class='risk-score-grid'>
              <div class='{score_card_class}' style='background:{overall_style['bg'] if model.get("grade") != "Provisional" else "#dcecff"};border-color:{overall_style['text'] if model.get("grade") != "Provisional" else "#7eb5ff"};'>
                <div style='font-size:1.85rem;font-weight:800;color:{overall_style['text']};line-height:1;'>{"—" if model.get("grade") == "Provisional" else f"{model['final_score']:.1f}"}</div>
                <div style='margin-top:0.5rem;display:inline-block;background:#ffffff;border:1px solid {overall_style['text']};color:{overall_style['text']};padding:0.32rem 0.72rem;border-radius:10px;font-weight:800;'>{html.escape(str(model['grade']))}</div>
                <div style='margin-top:0.7rem;color:{overall_style['text']};font-weight:700;'>Only {html.escape(str(model.get('scoring_basis', 'N/A')))}</div>
              </div>
              <div class='limit-card'>
                <div style='font-size:1.05rem;font-weight:800;color:#161616;'>Policy Approved Limit</div>
                <div style='font-size:2.45rem;font-weight:900;color:#111111;line-height:1.05;margin-top:0.38rem;'>${approved_limit_abbrev}</div>
                <div style='margin-top:0.2rem;color:#111111;font-weight:700;'>(${approved_limit_full})</div>
                <div style='margin-top:0.65rem;color:#1f1f1f;font-weight:700;'>Requested Amount: ${model['requested_amount']:,.0f}</div>
                <div style='margin-top:0.24rem;color:#1f1f1f;font-weight:700;'>Approval Status: {approval_status} ({approved_pct:.0f}%)</div>
                <div style='margin-top:0.24rem;color:#1f1f1f;font-weight:700;'>Excess Capacity: ${excess_capacity:,.0f}</div>
                <div style='margin-top:0.52rem;color:#161616;font-weight:700;'>Base Policy Limit: ${base_limit:,.0f}</div>
                <div style='margin-top:0.18rem;color:#161616;font-weight:700;'>Policy Capacity Limit: ${policy_capacity_limit:,.0f}</div>
                <div style='margin-top:0.18rem;color:#2a2a2a;font-weight:650;'>Industry Adjustment ({applied_inputs.get('industry', 'N/A')} {ind_adj_pct*100:+.0f}%): {ind_amt:+,.0f}</div>
                <div style='margin-top:0.12rem;color:#2a2a2a;font-weight:650;'>Geography Adjustment ({applied_inputs.get('geography', 'N/A')} {geo_adj_pct*100:+.0f}%): {geo_amt:+,.0f}</div>
                <div style='margin-top:0.12rem;color:#2a2a2a;font-weight:650;'>Maturity Adjustment ({mat_adj_pct*100:+.0f}%): {mat_amt:+,.0f}</div>
                <div style='margin-top:0.12rem;color:#2a2a2a;font-weight:650;'>Policy Limits Unit: {policy_unit_label}</div>
                <div style='margin-top:0.12rem;color:#2a2a2a;font-weight:650;'>Statement Currency Scale: {html.escape(statement_scale_label)}</div>
                <div style='margin-top:0.48rem;color:#2a2a2a;font-weight:600;font-size:0.88rem;'>{html.escape(limit_basis_sentence)}</div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        low_rows = risk_df[risk_df["Risk"] == "Low"]["Metric"].head(5).tolist() if "Risk" in risk_df.columns else []
        risk_gap_rows: List[str] = []
        if "Risk" in risk_df.columns:
            for _, rr in risk_df[risk_df["Risk"].isin(["High", "Incomplete"])].head(8).iterrows():
                metric_name = str(rr.get("Metric", "Metric"))
                status_name = str(rr.get("Status", "Incomplete"))
                if status_name == "Incomplete":
                    risk_gap_rows.append(f"{metric_name}: Data not available")
                else:
                    risk_gap_rows.append(f"{metric_name}: Below policy threshold")
        top_risk_rows = model.get("top_risk", pd.DataFrame())
        top_risk_lines: List[str] = []
        if isinstance(top_risk_rows, pd.DataFrame) and not top_risk_rows.empty:
            for _, tr in top_risk_rows.head(3).iterrows():
                risk_lbl = str(tr.get("Risk", "Medium"))
                if risk_lbl in {"High", "Medium"}:
                    top_risk_lines.append(f"{tr.get('Metric', 'Metric')} ({risk_lbl})")
        cushion_pct = model.get("covenant_cushion_display_pct")
        if bool(model.get("covenant_breach_flag", False)):
            cushion_txt = "Breach detected (one or more coverage covenants at or below zero)."
        elif cushion_pct is None:
            cushion_txt = "Not available (missing covenant inputs)."
        else:
            cushion_txt = f"{float(cushion_pct):.1f}% cushion"
            cushion_note = model.get("covenant_cushion_note")
            if cushion_note:
                cushion_txt += f" ({cushion_note})"
        missing_count = len(model.get("incomplete_metrics", []) or [])
        total_count = len(risk_df.index)
        followup_line = (
            "Approval granted based on available evidence; collect missing statements for post-approval monitoring."
            if (missing_count and str(model.get("recommendation", "")) == "Approval")
            else (
                "Final recommendation is conditional upon receipt of complete financial statements."
                if missing_count
                else "Current documentation supports a complete risk recommendation."
            )
        )
        st.markdown(
            f"""
            <div class='memo-card'>
              <div style='display:flex;align-items:center;gap:0.45rem;margin-bottom:0.3rem;'><span style='font-size:1.2rem;'>📄</span><div style='font-size:1.58rem;font-weight:800;color:#0f2740;'>Auto Credit Memo</div></div>
              <div class='memo-sec memo-exec'>
                <p class='memo-sec-title'>Executive Summary</p>
                <div style='color:#1f2937;font-weight:700;line-height:1.45;'>
                  Final Risk Rating: {html.escape(str(model.get('grade', 'N/A')))} | Weighted Score: {model.get('final_score', 0):.1f}<br/>
                  Approved Limit: {approved_limit_abbrev} ({approved_limit_full}) | Status: {approval_status} ({approved_pct:.0f}%) | Excess Capacity: {excess_capacity:,.0f}<br/>
                  Base Policy Limit: {base_limit:,.0f} | Policy Capacity: {policy_capacity_limit:,.0f} | Industry Adj: {ind_adj_pct*100:+.0f}% | Geography Adj: {geo_adj_pct*100:+.0f}% | Maturity Adj: {mat_adj_pct*100:+.0f}%<br/>
                  Industry/Geography Adjustments: {html.escape(str(applied_inputs['industry']))}, {html.escape(str(applied_inputs['geography']))}, {html.escape(str(applied_inputs['business_stage']))}.<br/>
                  Data Quality: {model.get('data_completeness_pct', 0):.1f}% complete ({html.escape(str(model.get('confidence_level', 'Basic')))} confidence)<br/>
                  Policy limits shown in USD base units; statement metrics shown in {html.escape(statement_scale_label)}.
                </div>
              </div>
              <div class='memo-sec memo-strength'>
                <p class='memo-sec-title' style='color:#161616;'>Financial Strengths</p>
                <ul class='memo-list'>
                  {''.join([f'<li>{html.escape(str(x))}</li>' for x in low_rows]) if low_rows else '<li>No metrics currently exceed adjusted thresholds.</li>'}
                </ul>
              </div>
              <div class='memo-sec memo-risk'>
                <p class='memo-sec-title' style='color:#202020;'>Key Risks & Data Gaps</p>
                <ul class='memo-list'>
                  {''.join([f'<li>{html.escape(str(x))}</li>' for x in risk_gap_rows]) if risk_gap_rows else '<li>No major risk gaps detected.</li>'}
                </ul>
                <div style='margin-top:0.35rem;color:#2c2c2c;font-weight:700;'>Top Risk Drivers: {html.escape(', '.join(top_risk_lines) if top_risk_lines else 'No material high-risk drivers')}</div>
              </div>
              <div class='memo-sec memo-cushion'>
                <p class='memo-sec-title' style='color:#1f1f1f;'>Covenant Cushion Analysis</p>
                <div style='color:#1f2937;font-weight:700;'>Interest Coverage Cushion: {html.escape(cushion_txt)}</div>
              </div>
              <div class='memo-sec memo-rec'>
                <p class='memo-sec-title' style='color:#1f1f1f;'>Recommendation</p>
                <div style='color:#1f2937;font-weight:700;line-height:1.45;'>
                  {html.escape(str(model.get('recommendation', 'Conditional Approval')))}. Missing metrics: {missing_count}/{total_count}.<br/>
                  {html.escape(followup_line)}
                </div>
              </div>
            </div>
            """,
            unsafe_allow_html=True,
        )

        trace_df = pd.DataFrame(
            [
                {
                    "dataset_mode": st.session_state.get("risk_mode", bundle.get("source", "single")),
                    "dataset_id": bundle.get("context_key"),
                    "dataset_hash": bundle.get("context_hash", ""),
                    "files": ", ".join(bundle.get("processed_docs", [])),
                    "borrower": ", ".join(sorted(set(bundle.get("borrower_by_doc", {}).values()))) if bundle.get("borrower_by_doc") else "N/A",
                    "extracted_at": bundle.get("extracted_at", "N/A"),
                    "driver_settings": json.dumps(applied_inputs, default=str),
                    "recompute_reason": st.session_state.get("risk_recompute_reason", "driver change"),
                    "data_version": int(st.session_state.get("risk_data_version", 1)),
                }
            ]
        )
        with st.expander("Underwriting Audit Trail", expanded=False):
            st.markdown("**Recalculation Trace**")
            st.dataframe(_ui_null_df(trace_df), width="stretch", hide_index=True)

            st.markdown("**Calculation Trace**")
            calc_trace_raw = model.get("calc_trace", [])
            calc_trace_rows: List[Dict[str, Any]] = []
            for item in calc_trace_raw:
                iv = item.get("Extracted Value")
                tv = item.get("Threshold")
                formula = item.get("Formula")
                fo = item.get("Final Output") if isinstance(item.get("Final Output"), dict) else {}
                calc_trace_rows.append(
                    {
                        "Metric": item.get("Metric"),
                        "Extracted Value": iv,
                        "Threshold": tv,
                        "Formula Used": formula,
                        "Adjusted Score": fo.get("adjusted_score"),
                        "Risk": fo.get("risk"),
                    }
                )
            calc_trace_df = pd.DataFrame(calc_trace_rows).where(pd.notna(pd.DataFrame(calc_trace_rows)), "NULL") if calc_trace_rows else pd.DataFrame()
            if not calc_trace_df.empty:
                st.dataframe(_ui_null_df(calc_trace_df), width="stretch", hide_index=True)
            else:
                st.caption("No calculation trace available.")

            st.markdown("**Threshold Explanations**")
            for line in model.get("threshold_explanations", []):
                st.caption(line)

        with st.expander("Source Trace Tables", expanded=False):
            for name, tdf in tables.items():
                st.markdown(f"**{name}**")
                st.dataframe(_ui_null_df(tdf.assign(Currency=model.get("currency_scale", "Units"))), width="stretch", hide_index=True)


if __name__ == "__main__":
    main()
